#!/usr/bin/env python3
"""
CPRD Extractor — Version 1
Compatible with BMRC Server
Developed by Milad Nazarzadeh, University of Oxford
"""

import streamlit as st
import pandas as pd
import numpy as np
import duckdb
import os
import io
import glob
import zipfile
import tempfile
import json
import hashlib
import time
import re
import sys
import argparse
import platform
import subprocess
import socket
import getpass
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import pyarrow  # noqa: F401 – needed for .to_parquet()
    HAS_PYARROW = True
except ImportError:
    HAS_PYARROW = False

try:
    from openpyxl import Workbook as _OpenpyxlWorkbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ══════════════════════════════════════════════════════════════════════════════
# BRANDING (removed logo)
# ══════════════════════════════════════════════════════════════════════════════

import threading
from datetime import datetime, date
from pathlib import Path
import plotly.express as px
import plotly.graph_objects as go

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION & PATHS (verified from FileZilla screenshots)
# ══════════════════════════════════════════════════════════════════════════════

# Default CPRD base — overridden at runtime by data_mode selection
_DEFAULT_CPRD_BASE = "/gpfs3/well/rahimi/projects/CPRD"
CPRD_BASE = _DEFAULT_CPRD_BASE


def _reconfigure_paths(new_base):
    """Reconfigure all paths when CPRD base changes (Live Any Server / Windows mode)."""
    global CPRD_BASE, LINKAGE_RESULTS_DIRS, PATHS
    CPRD_BASE = new_base
    LINKAGE_RESULTS_DIRS = [
        os.path.join(CPRD_BASE, "linkage", "20_095_Results", d, "Aurum_linked", "Final")
        for d in ("Results-one", "Results-two", "Results-three")
    ]
    PATHS["aurum_base"] = CPRD_BASE
    PATHS["emis_dictionary"] = os.path.join(CPRD_BASE, "202102_lookups", "202102_EMISMedicalDictionary.txt")
    PATHS["emis_product_dictionary"] = os.path.join(CPRD_BASE, "202102_lookups", "202102_EMISProductDictionary.txt")
    PATHS["linkage_base"] = LINKAGE_RESULTS_DIRS
    PATHS["linkage_eligibility"] = os.path.join(CPRD_BASE, "linkage_eligibility_Set21", "linkage_eligibility.txt")
    PATHS["patid_lists"] = os.path.join(CPRD_BASE, "patid_lists_dcranoy")
    for _key, _fname in _LINKAGE_FILES.items():
        PATHS[_key] = _resolve_linkage(_fname)

# ── Linkage file auto-resolver ──────────────────────────────────
# Files are spread across Results-one, Results-two, Results-three.
# This resolver searches all three directories for each file.
LINKAGE_RESULTS_DIRS = [
    os.path.join(CPRD_BASE, "linkage", "20_095_Results", d, "Aurum_linked", "Final")
    for d in ("Results-one", "Results-two", "Results-three")
]

def _resolve_linkage(filename):
    """Find a linkage file across Results-one/two/three. Returns first match or a default path."""
    for d in LINKAGE_RESULTS_DIRS:
        candidate = os.path.join(d, filename)
        if os.path.isfile(candidate):
            return candidate
    # Not found on this machine (e.g. dev/test) — return Results-one as placeholder
    return os.path.join(LINKAGE_RESULTS_DIRS[0], filename)

def _resolve_all_linkage(filename):
    """Find ALL copies of a linkage file across Results-one/two/three. Returns list of existing paths."""
    found = []
    for d in LINKAGE_RESULTS_DIRS:
        candidate = os.path.join(d, filename)
        if os.path.isfile(candidate):
            found.append(candidate)
    return found if found else [os.path.join(LINKAGE_RESULTS_DIRS[0], filename)]

# Map of linkage key → filename
_LINKAGE_FILES = {
    "hes_diagnosis_epi":     "hes_diagnosis_epi_20_095_DM.txt",
    "hes_diagnosis_hosp":    "hes_diagnosis_hosp_20_095_DM.txt",
    "hes_primary_diag":      "hes_primary_diag_hosp_20_095_DM.txt",
    "hes_episodes":          "hes_episodes_20_095_DM.txt",
    "hes_hospital":          "hes_hospital_20_095_DM.txt",
    "hes_procedures_epi":    "hes_procedures_epi_20_095_DM.txt",
    "hesop_clinical":        "hesop_clinical_20_095_DM.txt",
    "hesop_patient":         "hesop_patient_20_095_DM.txt",
    "death_patient":         "death_patient_20_095_DM.txt",
    "hes_patient":           "hes_patient_20_095_DM.txt",
    "hes_maternity":         "hes_maternity_20_095_DM.txt",
    "hes_ccare":             "hes_ccare_20_095_DM.txt",
    "hesae_attendance":      "hesae_attendance_20_095_DM.txt",
    "hesae_diagnosis":       "hesae_diagnosis_20_095_DM.txt",
    "hesae_investigation":   "hesae_investigation_20_095_DM.txt",
    "hesae_patient":         "hesae_patient_20_095_DM.txt",
    "hesae_treatment":       "hesae_treatment_20_095_DM.txt",
    "patient_imd":           "patient_imd2015_20_095.txt",
    "practice_imd":          "practice_imd_20_095.txt",
    "linkage_eligibility_aurum": "20_095_linkage_eligibility_aurum_set21.txt",
}

PATHS = {
    "aurum_base": os.path.join(CPRD_BASE),
    "emis_dictionary": os.path.join(CPRD_BASE, "202102_lookups", "202102_EMISMedicalDictionary.txt"),
    "emis_product_dictionary": os.path.join(CPRD_BASE, "202102_lookups", "202102_EMISProductDictionary.txt"),
    "linkage_base": LINKAGE_RESULTS_DIRS,  # list of all three base dirs
    "linkage_eligibility": os.path.join(CPRD_BASE, "linkage_eligibility_Set21", "linkage_eligibility.txt"),
    "patid_lists": os.path.join(CPRD_BASE, "patid_lists_dcranoy"),
}
# Auto-resolve each linkage file across Results-one/two/three
for _key, _fname in _LINKAGE_FILES.items():
    PATHS[_key] = _resolve_linkage(_fname)

AURUM_FILE_TYPES = {
    "Observation": {"pattern": "Extract_Observation", "key_col": "medcodeid", "date_col": "obsdate",
                    "cols": ["patid", "consid", "pracid", "obsid", "obsdate", "enterdate",
                             "staffid", "parentobsid", "medcodeid", "value", "numunitid",
                             "obstypeid", "numrangelow", "numrangehigh", "probobsid"]},
    "DrugIssue": {"pattern": "Extract_DrugIssue", "key_col": "prodcodeid", "date_col": "issuedate",
                  "cols": ["patid", "issueid", "pracid", "probobsid", "drugrecid", "issuedate",
                           "enterdate", "staffid", "prodcodeid", "dosageid", "quantity",
                           "quantunitid", "duration", "estnhscost"]},
    "Consultation": {"pattern": "Extract_Consultation", "key_col": "consmedcodeid", "date_col": "consdate",
                     "cols": ["patid", "consid", "pracid", "consdate", "enterdate", "staffid",
                              "conssourceid", "cprdconstype", "consmedcodeid"]},
    "Patient": {"pattern": "Extract_Patient", "key_col": "patid", "date_col": "regstartdate",
                "cols": ["patid", "pracid", "usualgpstaffid", "gender", "yob", "mob",
                         "regstartdate", "patienttypeid", "regenddate", "emession",
                         "acceptable", "cprd_ddate"]},
    "Practice": {"pattern": "Extract_Practice", "key_col": "pracid", "date_col": "lcd",
                 "cols": ["pracid", "lcd", "uts", "region"]},
    "Problem": {"pattern": "Extract_Problem", "key_col": "probmedcodeid", "date_col": "probenddate",
                "cols": ["patid", "obsid", "pracid", "parentprobobsid", "probenddate",
                         "expduration", "lastrevdate", "lastrevstaffid", "parentprobrelid",
                         "probmedcodeid", "probstatusid", "signid"]},
    "Referral": {"pattern": "Extract_Referral", "key_col": "medcodeid", "date_col": "obsdate",
                 "cols": ["patid", "obsid", "pracid", "refsourceorgid", "refsourceresid",
                          "reftargetorgid", "reftargetresid", "refurgencyid", "refdirectionid",
                          "refservicetypeid", "obsdate", "medcodeid"]},
    "Staff": {"pattern": "Extract_Staff", "key_col": "staffid", "date_col": None,
              "cols": ["staffid", "pracid", "jobcatid"]},
}

VHD_CODES = {
    "snomed": {
        "Mitral Stenosis": ["86466006", "31085000", "787001"],
        "Mitral Regurgitation": ["83916000", "72011007", "78031003"],
        "Mitral Valve Prolapse": ["17759006", "409712001"],
        "Mitral Valve Disease (Other)": ["48724000", "16440002"],
        "Aortic Stenosis": ["60573004", "836480008", "836481007", "836482000"],
        "Aortic Regurgitation": ["60234000", "8722008"],
        "Aortic Valve Disease (Other)": ["472847005", "233917008"],
        "Pulmonary Valve Stenosis": ["56786000", "91434003", "204357006"],
        "Pulmonary Regurgitation": ["63042009"],
        "Tricuspid Stenosis": ["194984004"],
        "Tricuspid Regurgitation": ["82458004", "79619009", "708966001"],
        "Tricuspid Valve Disease (Other)": ["78495000", "62067003"],
    },
    "icd10": {
        "Mitral Stenosis": ["I05.0"],
        "Mitral Regurgitation": ["I05.1"],
        "Mitral Stenosis and Regurgitation": ["I05.2"],
        "Mitral Valve Disease (Other)": ["I05.8", "I05.9"],
        "Aortic Stenosis": ["I06.0", "I35.0"],
        "Aortic Regurgitation": ["I06.1", "I35.1"],
        "Aortic Stenosis and Regurgitation": ["I06.2", "I35.2"],
        "Aortic Valve Disease (Other)": ["I06.8", "I06.9", "I35.8", "I35.9"],
        "Tricuspid Stenosis": ["I07.0", "I36.0"],
        "Tricuspid Regurgitation": ["I07.1", "I36.1"],
        "Tricuspid Valve Disease (Other)": ["I07.2", "I07.8", "I07.9", "I36.2", "I36.8", "I36.9"],
        "Pulmonary Valve Disease": ["I37.0", "I37.1", "I37.2", "I37.8", "I37.9"],
        "Multiple Valve Disease": ["I08.0", "I08.1", "I08.2", "I08.3", "I08.8", "I08.9"],
        "Endocarditis (Valve)": ["I09.1"],
        "Unspecified Valve Disease": ["I38"],
    },
}

COMORBIDITY_CODES_ICD10 = {
    "Hypertension": ["I10", "I11", "I12", "I13", "I15"],
    "Diabetes Type 2": ["E11"],
    "Diabetes Type 1": ["E10"],
    "Atrial Fibrillation": ["I48"],
    "Heart Failure": ["I50"],
    "Coronary Heart Disease": ["I20", "I21", "I22", "I23", "I24", "I25"],
    "Stroke": ["I60", "I61", "I62", "I63", "I64"],
    "COPD": ["J44"],
    "CKD": ["N18"],
}

# ══════════════════════════════════════════════════════════════════════════════
# DRUG LOOKUP LIBRARIES — BNF codes + Drug substance names
# ══════════════════════════════════════════════════════════════════════════════

DRUG_LIBRARIES = {
    "CVD Drugs": {
        "bnf_codes": {
            "ACE Inhibitor": ["0205051R0", "0205051C0", "0205051E0", "0205051F0",
                              "0205051L0", "0205051N0", "0205051P0", "0205051Q0",
                              "0205051S0", "0205051T0"],
            "ARB": ["0205052A0", "0205052AB", "0205052AD", "0205052AE",
                    "0205052AF", "0205052AH", "0205052AI", "0205052AK"],
            "Beta-blocker": ["0204000A0", "0204000B0", "0204000C0", "0204000E0",
                             "0204000F0", "0204000G0", "0204000H0", "0204000K0",
                             "0204000M0", "0204000N0", "0204000P0", "0204000Q0",
                             "0204000R0", "0204000S0", "0204000T0", "0204000U0",
                             "0204000V0", "0204000W0", "0204000Y0"],
            "CCB (dihydropyridine)": ["0206020A0", "0206020B0", "0206020C0",
                                      "0206020E0", "0206020F0", "0206020G0",
                                      "0206020H0", "0206020I0"],
            "CCB (rate-limiting)": ["0206020D0", "0206020J0"],
            "Thiazide diuretic": ["0202010B0", "0202010D0", "0202010F0",
                                  "0202010G0", "0202010J0", "0202010L0",
                                  "0202010N0", "0202010P0"],
            "Loop diuretic": ["0202020L0", "0202020M0", "0202020N0", "0202020T0"],
            "K-sparing diuretic": ["0202030D0", "0202030G0", "0202030L0",
                                   "0202030N0", "0202040D0"],
            "MRA": ["0202030D0", "0303020E0"],
            "Statin": ["0212000AC", "0212000B0", "0212000C0", "0212000G0",
                       "0212000X0", "0212000Y0", "0212000AA"],
            "Antiplatelet": ["0209000A0", "0209000C0", "0209000M0", "0209000N0",
                             "0209000P0", "0209000R0"],
            "Warfarin": ["0208020V0", "0208020W0"],
            "DOAC": ["0208020X0", "0208020Y0", "0208020Z0", "0208020AA"],
            "Nitrate": ["0206030A0", "0206030C0", "0206030D0", "0206030E0"],
            "Ivabradine": ["0206030F0"],
            "Digoxin": ["0201010D0"],
            "Amiodarone": ["0203020A0"],
            "ARNI (sacubitril/valsartan)": ["0205052AJ"],
            "Hydralazine": ["0205020B0"],
            "SGLT2i": ["0601023AW", "0601023AX", "0601023AY", "0601023AZ",
                       "0601023BA", "0601023BB"],
            "GLP-1 RA": ["0601023U0", "0601023V0", "0601023W0", "0601023X0",
                         "0601023Y0", "0601023Z0", "0601023AA"],
        },
        "drug_names": {
            # ACE inhibitors
            "ramipril": "ACE Inhibitor", "lisinopril": "ACE Inhibitor",
            "perindopril": "ACE Inhibitor", "enalapril": "ACE Inhibitor",
            "captopril": "ACE Inhibitor", "fosinopril": "ACE Inhibitor",
            "trandolapril": "ACE Inhibitor", "quinapril": "ACE Inhibitor",
            "imidapril": "ACE Inhibitor", "moexipril": "ACE Inhibitor",
            # ARBs
            "losartan": "ARB", "candesartan": "ARB", "irbesartan": "ARB",
            "valsartan": "ARB", "olmesartan": "ARB", "telmisartan": "ARB",
            "eprosartan": "ARB", "azilsartan": "ARB",
            # Beta-blockers
            "bisoprolol": "Beta-blocker", "atenolol": "Beta-blocker",
            "metoprolol": "Beta-blocker", "propranolol": "Beta-blocker",
            "carvedilol": "Beta-blocker", "nebivolol": "Beta-blocker",
            "sotalol": "Beta-blocker", "labetalol": "Beta-blocker",
            "nadolol": "Beta-blocker", "celiprolol": "Beta-blocker",
            # CCBs
            "amlodipine": "CCB", "felodipine": "CCB", "nifedipine": "CCB",
            "lercanidipine": "CCB", "lacidipine": "CCB", "nicardipine": "CCB",
            "diltiazem": "CCB", "verapamil": "CCB",
            # Diuretics
            "furosemide": "Loop diuretic", "bumetanide": "Loop diuretic",
            "torasemide": "Loop diuretic", "bendroflumethiazide": "Thiazide",
            "indapamide": "Thiazide", "chlortalidone": "Thiazide",
            "hydrochlorothiazide": "Thiazide", "metolazone": "Thiazide",
            "spironolactone": "MRA", "eplerenone": "MRA",
            "amiloride": "K-sparing diuretic",
            # Statins
            "atorvastatin": "Statin", "simvastatin": "Statin",
            "rosuvastatin": "Statin", "pravastatin": "Statin",
            "fluvastatin": "Statin",
            # Antiplatelets
            "aspirin": "Antiplatelet", "clopidogrel": "Antiplatelet",
            "ticagrelor": "Antiplatelet", "prasugrel": "Antiplatelet",
            "dipyridamole": "Antiplatelet",
            # Anticoagulants
            "warfarin": "Warfarin", "apixaban": "DOAC",
            "rivaroxaban": "DOAC", "edoxaban": "DOAC", "dabigatran": "DOAC",
            # HF drugs
            "sacubitril": "ARNI", "ivabradine": "Ivabradine",
            "hydralazine": "Vasodilator", "digoxin": "Digoxin",
            "amiodarone": "Amiodarone",
            # SGLT2i
            "dapagliflozin": "SGLT2i", "empagliflozin": "SGLT2i",
            "canagliflozin": "SGLT2i", "ertugliflozin": "SGLT2i",
            # GLP-1 RA
            "semaglutide": "GLP-1 RA", "liraglutide": "GLP-1 RA",
            "dulaglutide": "GLP-1 RA", "exenatide": "GLP-1 RA",
            # PAH
            "bosentan": "PAH drug", "ambrisentan": "PAH drug",
            "macitentan": "PAH drug", "sildenafil": "PAH drug",
            "tadalafil": "PAH drug", "riociguat": "PAH drug",
            "selexipag": "PAH drug", "iloprost": "PAH drug",
            # Other lipid-lowering
            "ezetimibe": "Lipid-lowering", "evolocumab": "PCSK9 inhibitor",
            "alirocumab": "PCSK9 inhibitor", "inclisiran": "Lipid-lowering",
            "bempedoic": "Lipid-lowering",
        },
    },
    "MDD Drugs (Antidepressants)": {
        "bnf_codes": {
            "SSRI": ["0403030D0", "0403030E0", "0403030P0", "0403030Q0",
                     "0403030R0", "0403030S0"],
            "SNRI": ["0403040U0", "0403040V0"],
            "TCA": ["0403010A0", "0403010B0", "0403010C0", "0403010E0",
                    "0403010F0", "0403010G0", "0403010H0", "0403010I0",
                    "0403010J0", "0403010K0", "0403010L0"],
            "Mirtazapine": ["0403040Q0"],
            "MAO Inhibitor": ["0403020D0", "0403020F0", "0403020H0", "0403020J0"],
            "Other antidepressant": ["0403040A0", "0403040B0", "0403040D0",
                                     "0403040G0", "0403040P0", "0403040T0"],
        },
        "drug_names": {
            "fluoxetine": "SSRI", "sertraline": "SSRI", "citalopram": "SSRI",
            "escitalopram": "SSRI", "paroxetine": "SSRI", "fluvoxamine": "SSRI",
            "venlafaxine": "SNRI", "duloxetine": "SNRI",
            "amitriptyline": "TCA", "nortriptyline": "TCA", "imipramine": "TCA",
            "clomipramine": "TCA", "dosulepin": "TCA", "doxepin": "TCA",
            "lofepramine": "TCA", "trimipramine": "TCA",
            "mirtazapine": "Mirtazapine", "trazodone": "Other antidepressant",
            "bupropion": "Other antidepressant", "vortioxetine": "Other antidepressant",
            "agomelatine": "Other antidepressant",
            "phenelzine": "MAO Inhibitor", "tranylcypromine": "MAO Inhibitor",
            "moclobemide": "MAO Inhibitor",
        },
    },
}

# ── Friendly column name mappings ──
FRIENDLY_NAMES = {
    "patid": "Patient ID",
    "pracid": "Practice ID",
    "obsid": "Observation ID",
    "obsdate": "Observation Date",
    "enterdate": "Entry Date",
    "medcodeid": "Medical Code ID",
    "prodcodeid": "Product Code ID",
    "consid": "Consultation ID",
    "staffid": "Staff ID",
    "value": "Value",
    "numunitid": "Numeric Unit ID",
    "obstypeid": "Observation Type ID",
    "numrangelow": "Range Low",
    "numrangehigh": "Range High",
    "probobsid": "Problem Observation ID",
    "SnomedCTConceptId": "SNOMED CT Code",
    "Term": "Clinical Term",
    "MedCodeId": "Medical Code ID",
    "issueid": "Issue ID",
    "issuedate": "Issue Date",
    "drugrecid": "Drug Record ID",
    "dosageid": "Dosage ID",
    "quantity": "Quantity",
    "quantunitid": "Quantity Unit",
    "duration": "Duration (days)",
    "estnhscost": "Est. NHS Cost",
    "drugname": "Drug Name",
    "gender": "Gender",
    "yob": "Year of Birth",
    "mob": "Month of Birth",
    "regstartdate": "Registration Start",
    "regenddate": "Registration End",
    "acceptable": "Acceptable Patient",
    "cprd_ddate": "CPRD Death Date",
    "spno": "Spell Number",
    "epikey": "Episode Key",
    "epistart": "Episode Start",
    "epiend": "Episode End",
    "ICD": "ICD-10 Code",
    "ICD_PRIMARY": "Primary ICD-10 Code",
    "d_order": "Diagnosis Position",
    "attendkey": "Attendance Key",
    "appointdt": "Appointment Date",
    "diag_01": "Diagnosis 1",
    "opertn_01": "Operation 1",
    "dod": "Date of Death",
    "cause": "Underlying Cause of Death",
    "cause_position": "Cause Position",
    "icd_code": "ICD-10 Code",
    "hes_e": "HES Eligible",
    "death_e": "Death Eligible",
    "lsoa_e": "LSOA Eligible",
    "imd2015_5": "IMD Quintile (1=most deprived)",
    "imd2015_10": "IMD Decile",
    "practice_folder": "Practice Folder",
    "source": "Data Source",
}

# ── Tooltip/help texts for every section ──
TOOLTIPS = {
    "nav_home": "Dashboard showing data environment status, available modules, and data paths.",
    "nav_newbie": "Simplified one-click workflow. Paste your codes, and the system searches everything automatically.",
    "nav_aurum": "Extract primary care records from CPRD Aurum practice files (SNOMED CT or MedCode based).",
    "nav_linkage": "Extract hospital, outpatient, and death records from linked NHS datasets (ICD-10 based).",
    "nav_multi": "Search across ALL data sources simultaneously with a single code list.",
    "nav_cohort": "Build patient cohorts step-by-step with inclusion/exclusion criteria and attrition tracking.",
    "nav_analytics": "Generate Table 1 baseline characteristics, Kaplan-Meier curves, and distribution plots.",
    "nav_config": "View and verify all data file paths, practice folders, and linkage coverage dates.",
    "snomed_codes": "SNOMED CT codes are long numeric identifiers (e.g. 60573004) used in UK primary care (GP records).",
    "icd10_codes": "ICD-10 codes are alphanumeric (e.g. I05.0) used in hospitals (HES) and death certificates.",
    "medcode_ids": "MedCodeIds are CPRD-internal numeric identifiers that map to SNOMED CT via the EMIS dictionary.",
    "hes_apc": "Hospital Episode Statistics - Admitted Patient Care. Contains all NHS hospital admission diagnoses.",
    "hes_op": "Hospital Episode Statistics - Outpatient. Contains NHS outpatient clinic visit diagnoses.",
    "ons_death": "Office for National Statistics death records. Contains up to 16 cause-of-death codes per patient.",
    "linkage_elig": "Patients must be linkage-eligible to appear in HES/Death data. Check this before linkage analysis.",
    "imd": "Index of Multiple Deprivation 2015. Quintile 1 = most deprived, Quintile 5 = least deprived.",
    "merge_mode": "Merge combines all sources into one file joined by Patient ID. Separate keeps each source as its own file.",
    "friendly_cols": "Friendly names replace cryptic column headers (e.g. 'patid' → 'Patient ID') for easier reading.",
}

# ══════════════════════════════════════════════════════════════════════════════
# DISEASE CODE LIBRARY — 63 cardiovascular diseases, 1052 codes
# Sources: SNOMED CT (Aurum primary care) + ICD-10 (HES/ONS)
# ══════════════════════════════════════════════════════════════════════════════

DISEASE_GROUPS = {
    "Ischaemic Heart Disease": ["Coronary Artery Disease", "Myocardial Infarction", "Stable and Unstable Angina", "Microvascular Angina", "Vasospastic Angina", "Silent Ischemia", "Spontaneous Coronary Artery Dissection"],
    "Valvular Heart Disease": ["Aortic Valve Stenosis", "Aortic Valve Regurgitation", "Bicuspid Aortic Valve", "Mitral Valve Stenosis", "Mitral Valve Regurgitation", "Mitral Valve Prolapse", "Mitral Annular Calcification", "Tricuspid Valve Stenosis", "Tricuspid Valve Regurgitation", "Ebstein Anomaly", "Pulmonary Valve Stenosis", "Pulmonary Valve Regurgitation", "Pulmonary Atresia", "Rheumatic Heart Disease", "Multi-valve Disease", "Prosthetic Valve Disease"],
    "Arrhythmia": ["Atrial Fibrillation", "Atrial Flutter", "Supraventricular Tachycardia", "Ventricular Tachycardia", "Ventricular Fibrillation", "Bradycardia and Sick Sinus Syndrome", "Heart Block", "Long QT Syndrome", "Brugada Syndrome", "Wolff-Parkinson-White Syndrome"],
    "Cardiomyopathy": ["Dilated Cardiomyopathy", "Hypertrophic Cardiomyopathy", "Restrictive Cardiomyopathy", "Arrhythmogenic Right Ventricular Dysplasia", "Takotsubo Cardiomyopathy", "Peripartum Cardiomyopathy", "Cardiac Amyloidosis", "Cardiac Sarcoidosis", "Cardiac Hemochromatosis"],
    "Vascular Disease": ["Peripheral Artery Disease", "Carotid Artery Disease", "Aortic Aneurysm", "Aortic Dissection", "Deep Vein Thrombosis", "Pulmonary Embolism", "Renal Artery Stenosis", "Chronic Venous Insufficiency", "Raynaud Phenomenon", "Vasculitis"],
    "Congenital Heart Disease": ["Atrial Septal Defect", "Ventricular Septal Defect", "Coarctation of the Aorta", "Tetralogy of Fallot", "Transposition of the Great Arteries", "Hypoplastic Left Heart Syndrome", "Patent Ductus Arteriosus"],
    "Infectious/Inflammatory Heart Disease": ["Infective Endocarditis", "Myocarditis", "Pericarditis", "Pericardial Effusion and Cardiac Tamponade", "Kawasaki Disease", "Chagas Disease"],
    "Heart Failure": ["Heart Failure"],
}

DISEASE_CODE_LIBRARY = {
    # ── Ischaemic Heart Disease ──
    "Coronary Artery Disease": {"short": "CAD", "snomed": ["194828000", "194842008", "194843003", "233817007", "233819005", "233821000", "233937004", "233970002", "251052003", "300995000", "308065005", "371806006", "371807002", "398274000", "413439005", "413838009", "413844008", "414545008", "415070008", "41702007", "426396005", "429559004", "443502000", "46109009", "46635009", "53741008", "59021001", "63739005", "64715009", "82523003", "89323001"], "icd10": ["I20", "I20.0", "I20.1", "I20.8", "I20.9", "I25", "I25.0", "I25.1", "I25.10", "I25.11", "I25.2", "I25.5", "I25.6", "I25.8", "I25.9"], "icd10_nodot": ["I200", "I201", "I208", "I209", "I20X", "I250", "I251", "I2510", "I2511", "I252", "I255", "I256", "I258", "I259", "I25X"]},
    "Myocardial Infarction": {"short": "MI", "snomed": ["129574000", "15712000", "194809007", "194856005", "22298006", "233837006", "233838001", "233839009", "233840006", "233841005", "233842003", "233843008", "25106000", "285981000", "304914007", "371068009", "394710008", "399211009", "401303003", "401314000", "413444003", "414795007", "426396005", "426979002", "52035003", "54329005", "57054005", "62695002", "698593002", "70211005", "703328002", "70422006", "73170004", "73795002", "76593002", "82523003"], "icd10": ["I21", "I21.0", "I21.1", "I21.2", "I21.3", "I21.4", "I21.9", "I21.A1", "I21.A9", "I21.B", "I22", "I22.0", "I22.1", "I22.8", "I22.9", "I23", "I23.0", "I23.1", "I23.2", "I23.3", "I23.4", "I23.5", "I23.6", "I23.8", "I24", "I24.1", "I25.2"], "icd10_nodot": ["I210", "I211", "I212", "I213", "I214", "I219", "I21A1", "I21A9", "I21B", "I21X", "I220", "I221", "I228", "I229", "I22X", "I230", "I231", "I232", "I233", "I234", "I235", "I236", "I238", "I23X", "I241", "I24X", "I252"]},
    "Stable and Unstable Angina": {"short": "ANG", "snomed": ["19057007", "194828000", "233819005", "233821000", "25106000", "300995000", "371806006", "371807002", "371809004", "413444003", "414545008", "426396005", "429559004", "4557003", "59021001", "712866001", "82523003", "87343002", "89323001"], "icd10": ["I20", "I20.0", "I20.1", "I20.8", "I20.81", "I20.89", "I20.9", "I24.0", "I24.8", "I24.9"], "icd10_nodot": ["I200", "I201", "I208", "I2081", "I2089", "I209", "I20X", "I240", "I248", "I249"]},
    "Microvascular Angina": {"short": "MVA", "snomed": ["194828000", "300995000", "371807002", "413666001", "429559004", "52674009", "698247007", "713405002", "82523003"], "icd10": ["I20.8", "I20.81", "I20.9", "I25.8"], "icd10_nodot": ["I208", "I2081", "I209", "I258"]},
    "Vasospastic Angina": {"short": "VSA", "snomed": ["194828000", "233818002", "300995000", "371809004", "429559004", "59021001", "87343002"], "icd10": ["I20.1", "I20.8", "I20.9"], "icd10_nodot": ["I201", "I208", "I209"]},
    "Silent Ischemia": {"short": "SIL", "snomed": ["233843008", "413838009", "413844008", "414545008", "414795007", "426396005"], "icd10": ["I25.6", "I25.9"], "icd10_nodot": ["I256", "I259"]},
    "Spontaneous Coronary Artery Dissection": {"short": "SCAD", "snomed": ["128053003", "128564005", "233936008", "413444003", "57054005", "840580004"], "icd10": ["I21.9", "I25.4", "I25.42"], "icd10_nodot": ["I219", "I254", "I2542"]},
    # ── Valvular Heart Disease ──
    "Aortic Valve Stenosis": {"short": "AVS", "snomed": ["155269004", "194983005", "204393006", "250977001", "250978003", "250978006", "250979003", "253545000", "424031003", "448613009", "44993000", "472111007", "472112000", "60573004", "72011007", "836480008", "836481007", "836482000", "86299006", "8722008"], "icd10": ["I06.0", "I06.2", "I35.0", "I35.2", "Q23.0", "Q25.3"], "icd10_nodot": ["I060", "I062", "I350", "I352", "Q230", "Q253"]},
    "Aortic Valve Regurgitation": {"short": "AVR", "snomed": ["16440002", "17759006", "194984004", "250980000", "250982008", "250984009", "472113005", "472847005", "48724000", "60234000", "78031003", "79619009", "838544003", "838545002", "838546001", "8722008"], "icd10": ["I06.1", "I06.2", "I06.8", "I06.9", "I35.1", "I35.2", "I35.8", "I35.9", "Q23.1"], "icd10_nodot": ["I061", "I062", "I068", "I069", "I351", "I352", "I358", "I359", "Q231"]},
    "Bicuspid Aortic Valve": {"short": "BAV", "snomed": ["204357006", "204381005", "204393006", "253451005", "253452003", "253547004", "253548009", "253549001", "472893007", "4962009"], "icd10": ["Q23.0", "Q23.1", "Q23.8", "Q23.81", "Q23.9"], "icd10_nodot": ["Q230", "Q231", "Q238", "Q2381", "Q239"]},
    "Mitral Valve Stenosis": {"short": "MVS", "snomed": ["111287006", "194706009", "253550001", "44241007", "472074006", "472220004", "472221000", "472222007", "48724000", "787001", "79619009", "82458004", "838448003", "838449006", "838450006", "86299006", "86466006"], "icd10": ["I05.0", "I05.2", "I34.2", "Q23.2"], "icd10_nodot": ["I050", "I052", "I342", "Q232"]},
    "Mitral Valve Regurgitation": {"short": "MVR", "snomed": ["111287006", "11851006", "194699005", "31085000", "472223002", "472224008", "472225009", "472226005", "48724000", "708966001", "715395008", "78495000", "79955004", "838451005", "838452003", "838453008", "838552006", "838553001", "838554007", "83916000"], "icd10": ["I05.1", "I05.2", "I05.8", "I05.9", "I34.0", "I34.8", "I34.89", "I34.9", "Q23.3"], "icd10_nodot": ["I051", "I052", "I058", "I059", "I340", "I348", "I3489", "I349", "Q233"]},
    "Mitral Valve Prolapse": {"short": "MVP", "snomed": ["1144928003", "11851006", "194694000", "253449001", "409712001", "445237003", "48724000"], "icd10": ["I05.8", "I34.0", "I34.1", "I34.8", "I34.81", "I34.89", "I34.9", "Q23.2"], "icd10_nodot": ["I058", "I340", "I341", "I348", "I3481", "I3489", "I349", "Q232"]},
    "Mitral Annular Calcification": {"short": "MAC", "snomed": ["111287006", "11851006", "472071003", "48724000", "86299006"], "icd10": ["I05.8", "I34.0", "I34.2", "I34.8", "I34.81", "I34.9"], "icd10_nodot": ["I058", "I340", "I342", "I348", "I3481", "I349"]},
    "Tricuspid Valve Stenosis": {"short": "TVS", "snomed": ["111290002", "194731009", "194733007", "204350006", "472076008", "49915006", "838535004", "838536003", "838537007"], "icd10": ["I07.0", "I07.2", "I07.8", "I07.9", "I36.0", "I36.2", "I36.8", "I36.9", "Q22.4", "Q22.8", "Q22.9"], "icd10_nodot": ["I070", "I072", "I078", "I079", "I360", "I362", "I368", "I369", "Q224", "Q228", "Q229"]},
    "Tricuspid Valve Regurgitation": {"short": "TVR", "snomed": ["111280001", "111287006", "111290002", "46931007", "472227001", "472228006", "472229003", "472230008", "715399002", "838454002", "838455001", "838456000", "838556009", "838557000", "838558005"], "icd10": ["I07.1", "I36.1", "I36.2", "I36.8", "I36.9"], "icd10_nodot": ["I071", "I361", "I362", "I368", "I369"]},
    "Ebstein Anomaly": {"short": "EBS", "snomed": ["204350006", "204351005", "204352003", "253366007", "253367003", "63042009"], "icd10": ["Q22.4", "Q22.5", "Q22.6", "Q22.8", "Q22.9"], "icd10_nodot": ["Q224", "Q225", "Q226", "Q228", "Q229"]},
    "Pulmonary Valve Stenosis": {"short": "PVS", "snomed": ["194740000", "204367002", "253309000", "253311009", "56786000", "836484004", "836485003", "836486002", "838538002", "838540007", "86299006"], "icd10": ["I37.0", "I37.2", "Q22.1", "Q25.6"], "icd10_nodot": ["I370", "I372", "Q221", "Q256"]},
    "Pulmonary Valve Regurgitation": {"short": "PVR", "snomed": ["472231007", "472232000", "472233005", "472234004", "60573004", "838548003", "838549006", "838550006", "91434003", "91438003"], "icd10": ["I37.1", "I37.2", "I37.8", "I37.9", "Q22.2", "Q22.3", "Q22.9"], "icd10_nodot": ["I371", "I372", "I378", "I379", "Q222", "Q223", "Q229"]},
    "Pulmonary Atresia": {"short": "PAT", "snomed": ["204367002", "204369004", "204372003", "233917008", "253297000", "253298005"], "icd10": ["Q22.0", "Q22.1", "Q22.3", "Q22.6", "Q25.5"], "icd10_nodot": ["Q220", "Q221", "Q223", "Q226", "Q255"]},
    "Rheumatic Heart Disease": {"short": "RHD", "snomed": ["111281005", "111290002", "155315007", "194699005", "194706009", "195010000", "195012008", "195528001", "195528004", "23685000", "266316001", "44241007", "44993000", "48724000"], "icd10": ["I00", "I01", "I01.0", "I01.1", "I01.2", "I01.8", "I01.9", "I02", "I05", "I05.0", "I05.1", "I05.2", "I05.8", "I05.9", "I06", "I06.0", "I06.1", "I06.2", "I06.8", "I06.9", "I07", "I08", "I09", "I09.0", "I09.1", "I09.2", "I09.8", "I09.9"], "icd10_nodot": ["I00X", "I010", "I011", "I012", "I018", "I019", "I01X", "I02X", "I050", "I051", "I052", "I058", "I059", "I05X", "I060", "I061", "I062", "I068", "I069", "I06X", "I07X", "I08X", "I090", "I091", "I092", "I098", "I099", "I09X"]},
    # ── Arrhythmia ──
    "Atrial Fibrillation": {"short": "AFI", "snomed": ["120041000119109", "15964901000119110", "195080001", "233910005", "282825002", "313217007", "314208002", "426749004", "440028005", "440059007", "49436004", "698247007", "706923002"], "icd10": ["I48.0", "I48.1", "I48.11", "I48.19", "I48.2", "I48.20", "I48.21", "I48.9", "I48.91"], "icd10_nodot": ["I480", "I481", "I4811", "I4819", "I482", "I4820", "I4821", "I489", "I4891"]},
    "Atrial Flutter": {"short": "AFL", "snomed": ["195080001", "195082009", "233910005", "233911009", "233912002", "5370000", "698247007"], "icd10": ["I48.3", "I48.4", "I48.9", "I48.92"], "icd10_nodot": ["I483", "I484", "I489", "I4892"]},
    "Supraventricular Tachycardia": {"short": "SVT", "snomed": ["11092001", "195068009", "195069001", "233892006", "233893001", "233896004", "233897008", "6456007", "698247007", "78849008"], "icd10": ["I47.0", "I47.1", "I47.10", "I47.11", "I47.19", "I47.9"], "icd10_nodot": ["I470", "I471", "I4710", "I4711", "I4719", "I479"]},
    "Ventricular Tachycardia": {"short": "VT", "snomed": ["233903009", "233904003", "233905002", "233906001", "233907005", "251162005", "25569003", "54016002", "698247007", "71908006"], "icd10": ["I47.2", "I47.20", "I47.21", "I47.29", "I47.9"], "icd10_nodot": ["I472", "I4720", "I4721", "I4729", "I479"]},
    "Ventricular Fibrillation": {"short": "VFI", "snomed": ["164889003", "397829000", "410429000", "48485001", "54016002", "698247007", "71908006"], "icd10": ["I46", "I46.0", "I46.1", "I46.9", "I49.0", "I49.01", "I49.02"], "icd10_nodot": ["I460", "I461", "I469", "I46X", "I490", "I4901", "I4902"]},
    "Bradycardia and Sick Sinus Syndrome": {"short": "BRA", "snomed": ["13640000", "195071004", "195072006", "233917008", "251161003", "29894001", "419752009", "426627000", "49710005", "74615001"], "icd10": ["I49.5", "I49.8", "R00.1"], "icd10_nodot": ["I495", "I498", "R001"]},
    "Heart Block": {"short": "HBL", "snomed": ["195042002", "195043007", "204384002", "233934001", "233935000", "233936004", "251120003", "251121004", "27885002", "28189009", "426183003", "54681002", "59118001", "63467002", "6374002"], "icd10": ["I44.0", "I44.1", "I44.2", "I44.3", "I44.30", "I44.39", "I44.4", "I44.5", "I44.6", "I44.7", "I45.0", "I45.1", "I45.10", "I45.19", "I45.2", "I45.3", "I45.4"], "icd10_nodot": ["I440", "I441", "I442", "I443", "I4430", "I4439", "I444", "I445", "I446", "I447", "I450", "I451", "I4510", "I4519", "I452", "I453", "I454"]},
    "Long QT Syndrome": {"short": "LQT", "snomed": ["233907005", "233917008", "253684004", "253685003", "418461001", "441836004", "698247007", "700212002", "9651007"], "icd10": ["I45.8", "I45.81", "I45.89", "R94.31"], "icd10_nodot": ["I458", "I4581", "I4589", "R9431"]},
    "Brugada Syndrome": {"short": "BRU", "snomed": ["418818005", "48485001", "698247007", "722950002", "722951003"], "icd10": ["I45.8", "I45.89", "I46.1", "I49.8", "I49.9"], "icd10_nodot": ["I458", "I4589", "I461", "I498", "I499"]},
    "Wolff-Parkinson-White Syndrome": {"short": "WPW", "snomed": ["11092001", "233897008", "29390004", "43880005", "49260003", "6456007", "74390002"], "icd10": ["I45.6", "I45.89", "I45.9", "I47.1", "I47.10"], "icd10_nodot": ["I456", "I4589", "I459", "I471", "I4710"]},
    # ── Cardiomyopathy ──
    "Dilated Cardiomyopathy": {"short": "DCM", "snomed": ["195020003", "195021004", "233872009", "233873004", "426212002", "609387002", "82251002", "85898001"], "icd10": ["I42.0", "I42.6", "I42.7", "I42.8", "I42.9"], "icd10_nodot": ["I420", "I426", "I427", "I428", "I429"]},
    "Hypertrophic Cardiomyopathy": {"short": "HCM", "snomed": ["195020003", "233873004", "233874005", "233875006", "233876007", "253537007", "45227007", "472201006", "83521008"], "icd10": ["I42.1", "I42.2", "I42.9"], "icd10_nodot": ["I421", "I422", "I429"]},
    "Restrictive Cardiomyopathy": {"short": "RCM", "snomed": ["195020003", "233877003", "233878008", "75543006", "78839003"], "icd10": ["I42.3", "I42.4", "I42.5", "I42.8", "I42.9", "I43"], "icd10_nodot": ["I423", "I424", "I425", "I428", "I429", "I43X"]},
    "Arrhythmogenic Right Ventricular Dysplasia": {"short": "ARVD", "snomed": ["195020003", "253528005", "25569003", "281170000"], "icd10": ["I42.0", "I42.8", "I42.9", "I47.2", "I49.0"], "icd10_nodot": ["I420", "I428", "I429", "I472", "I490"]},
    "Takotsubo Cardiomyopathy": {"short": "TTC", "snomed": ["195020003", "429098002", "700213007", "712853002", "840614005", "870538006"], "icd10": ["I42.8", "I42.9", "I51.81", "I51.89"], "icd10_nodot": ["I428", "I429", "I5181", "I5189"]},
    "Peripartum Cardiomyopathy": {"short": "PPC", "snomed": ["16253001", "195020003", "200113008", "34881000", "62377009", "84114007"], "icd10": ["I42.0", "I42.8", "I42.9", "O90.3", "O99.41"], "icd10_nodot": ["I420", "I428", "I429", "O903", "O9941"]},
    "Cardiac Amyloidosis": {"short": "CAM", "snomed": ["129161005", "17602002", "267426001", "371090009", "425439007", "68820003", "723333003", "723505009", "725417000", "774080007"], "icd10": ["E85.0", "E85.1", "E85.3", "E85.4", "E85.8", "E85.9", "I43"], "icd10_nodot": ["E850", "E851", "E853", "E854", "E858", "E859", "I43X"]},
    "Cardiac Sarcoidosis": {"short": "CSA", "snomed": ["233869006", "233870007", "31541009", "4597004", "59630003"], "icd10": ["D86.8", "D86.85", "D86.9", "I41", "I43"], "icd10_nodot": ["D868", "D8685", "D869", "I41X", "I43X"]},
    "Cardiac Hemochromatosis": {"short": "CHE", "snomed": ["127058000", "233870007", "31659003", "360043002", "399144008", "399187006"], "icd10": ["E83.1", "E83.10", "E83.11", "E83.19", "I42.8", "I43"], "icd10_nodot": ["E831", "E8310", "E8311", "E8319", "I428", "I43X"]},
    # ── Vascular Disease ──
    "Peripheral Artery Disease": {"short": "PAD", "snomed": ["127285001", "14637003", "195256008", "19948002", "233945004", "233946003", "233948004", "233949007", "233951006", "233960008", "371048003", "399957001", "404684003", "413758000", "70195001", "840580004", "87317003"], "icd10": ["I70.2", "I70.20", "I70.21", "I70.22", "I70.23", "I70.24", "I70.25", "I73.89", "I73.9", "I74.3", "I74.4"], "icd10_nodot": ["I702", "I7020", "I7021", "I7022", "I7023", "I7024", "I7025", "I7389", "I739", "I743", "I744"]},
    "Carotid Artery Disease": {"short": "CAS", "snomed": ["195190007", "195191006", "233940007", "300920004", "301764004", "420552002", "64572001", "64586002"], "icd10": ["I65.2", "I65.3", "I70.8"], "icd10_nodot": ["I652", "I653", "I708"]},
    "Aortic Aneurysm": {"short": "AAN", "snomed": ["10273003", "233983001", "233984007", "233985008", "233986009", "233987000", "233988005", "26788006", "59660003", "67362008", "698254001"], "icd10": ["I71.1", "I71.2", "I71.3", "I71.4", "I71.5", "I71.6", "I71.8", "I71.9"], "icd10_nodot": ["I711", "I712", "I713", "I714", "I715", "I716", "I718", "I719"]},
    "Aortic Dissection": {"short": "ADI", "snomed": ["233990006", "233991005", "233992003", "233993008", "308546005", "59660003", "67362008"], "icd10": ["I71.0", "I71.00", "I71.01", "I71.02", "I71.03"], "icd10_nodot": ["I710", "I7100", "I7101", "I7102", "I7103"]},
    "Deep Vein Thrombosis": {"short": "DVT", "snomed": ["128053003", "132251000119100", "132281000119108", "134399007", "195437003", "233935000", "233936004", "233938009", "26090000", "312584000", "404223003", "442190001", "709687000"], "icd10": ["I80.1", "I80.2", "I80.20", "I80.3", "I82.4", "I82.40", "I82.5", "I82.50", "I82.6", "I82.9"], "icd10_nodot": ["I801", "I802", "I8020", "I803", "I824", "I8240", "I825", "I8250", "I826", "I829"]},
    "Pulmonary Embolism": {"short": "PE", "snomed": ["133971000119108", "194883006", "233935000", "233935002", "233936004", "31542003", "39400006", "415225003", "438773007", "59282003", "706870000", "706871001", "74315008"], "icd10": ["I26", "I26.0", "I26.02", "I26.9", "I26.92", "I26.93", "I26.94", "I27.82"], "icd10_nodot": ["I260", "I2602", "I269", "I2692", "I2693", "I2694", "I26X", "I2782"]},
    "Renal Artery Stenosis": {"short": "RAS", "snomed": ["233947009", "234074008", "236466006", "32513000", "64572001", "68267002", "68820003"], "icd10": ["I15.0", "I70.1", "I77.3", "N28.0"], "icd10_nodot": ["I150", "I701", "I773", "N280"]},
    "Chronic Venous Insufficiency": {"short": "CVI", "snomed": ["127288002", "128060003", "195440000", "234042004", "40273006", "402862000", "71028006", "77567007", "95337007", "95345008"], "icd10": ["I83", "I83.0", "I83.1", "I83.2", "I83.9", "I87.0", "I87.1", "I87.2", "I87.8", "I87.9"], "icd10_nodot": ["I830", "I831", "I832", "I839", "I83X", "I870", "I871", "I872", "I878", "I879"]},
    "Raynaud Phenomenon": {"short": "RAY", "snomed": ["195325008", "195326009", "266261006", "443613007", "47933007"], "icd10": ["I73.0", "I73.00", "I73.01"], "icd10_nodot": ["I730", "I7300", "I7301"]},
    "Vasculitis": {"short": "VAS", "snomed": ["155441006", "195353004", "239934003", "277468003", "278936003", "312738009", "31996006", "400130008", "413758000", "416740007", "46956008", "75053002", "82275007"], "icd10": ["I73.1", "I77.6", "M30.0", "M30.1", "M31.3", "M31.4", "M31.5", "M31.6", "M31.7", "M35.2"], "icd10_nodot": ["I731", "I776", "M300", "M301", "M313", "M314", "M315", "M316", "M317", "M352"]},
    # ── Congenital Heart Disease ──
    "Atrial Septal Defect": {"short": "ASD", "snomed": ["204317008", "253341008", "253342009", "253343004", "253344005", "253345006", "253347003", "253416000", "70142008"], "icd10": ["I51.0", "Q21.1", "Q21.10", "Q21.11", "Q21.12", "Q21.13", "Q21.14", "Q21.15", "Q21.16", "Q21.19", "Q21.2", "Q21.9"], "icd10_nodot": ["I510", "Q211", "Q2110", "Q2111", "Q2112", "Q2113", "Q2114", "Q2115", "Q2116", "Q2119", "Q212", "Q219"]},
    "Ventricular Septal Defect": {"short": "VSD", "snomed": ["253356006", "253357002", "253358007", "253359004", "253360009", "253361008", "253362001", "253363006", "30288003", "399217008"], "icd10": ["I23.2", "I51.0", "Q21.0", "Q21.2", "Q21.8", "Q21.9"], "icd10_nodot": ["I232", "I510", "Q210", "Q212", "Q218", "Q219"]},
    "Coarctation of the Aorta": {"short": "COA", "snomed": ["12770006", "204432001", "253424008", "253425009", "253426005", "253427001", "253428006", "7305005"], "icd10": ["Q25.1", "Q25.2", "Q25.21", "Q25.29", "Q25.3", "Q25.4", "Q25.40"], "icd10_nodot": ["Q251", "Q252", "Q2521", "Q2529", "Q253", "Q254", "Q2540"]},
    "Tetralogy of Fallot": {"short": "TOF", "snomed": ["204296002", "204297006", "253286009", "253287000", "253288005", "373099003", "86299006"], "icd10": ["Q21.0", "Q21.3", "Q21.8", "Q22.0", "Q22.1"], "icd10_nodot": ["Q210", "Q213", "Q218", "Q220", "Q221"]},
    "Transposition of the Great Arteries": {"short": "TGA", "snomed": ["13213009", "204306007", "253274004", "253276002", "399216004", "48652001"], "icd10": ["Q20.0", "Q20.1", "Q20.2", "Q20.3", "Q20.4", "Q20.5", "Q20.8", "Q20.9"], "icd10_nodot": ["Q200", "Q201", "Q202", "Q203", "Q204", "Q205", "Q208", "Q209"]},
    "Hypoplastic Left Heart Syndrome": {"short": "HLHS", "snomed": ["253428006", "253548009", "253550001", "253564005", "61959006", "62067003"], "icd10": ["Q23.0", "Q23.2", "Q23.4", "Q23.8", "Q23.9", "Q25.2", "Q25.29"], "icd10_nodot": ["Q230", "Q232", "Q234", "Q238", "Q239", "Q252", "Q2529"]},
    "Patent Ductus Arteriosus": {"short": "PDA", "snomed": ["204447006", "276507000", "373945007", "83330001"], "icd10": ["P29.3", "Q21.4", "Q25.0", "Q25.4", "Q25.9"], "icd10_nodot": ["P293", "Q214", "Q250", "Q254", "Q259"]},
    # ── Infectious/Inflammatory Heart Disease ──
    "Infective Endocarditis": {"short": "IE", "snomed": ["112283007", "155315007", "194682008", "233854009", "233855005", "233856006", "233857002", "233860003", "233862006", "278474008", "428553009", "429071009", "56819008", "71863001"], "icd10": ["B37.6", "I33", "I33.0", "I33.9", "I38", "I39"], "icd10_nodot": ["B376", "I330", "I339", "I33X", "I38X", "I39X"]},
    "Myocarditis": {"short": "MYO", "snomed": ["195015006", "233861005", "233862003", "233863001", "233863008", "233864002", "233864007", "233865001", "233866000", "24979000", "398550007", "399045003", "50920009", "82290005", "84114007"], "icd10": ["B33.2", "B33.22", "I40", "I40.0", "I40.1", "I40.8", "I40.9", "I41", "I51.4"], "icd10_nodot": ["B332", "B3322", "I400", "I401", "I408", "I409", "I40X", "I41X", "I514"]},
    "Pericarditis": {"short": "PER", "snomed": ["129573006", "195016007", "233867009", "233868001", "233868004", "233869006", "23685000", "266316001", "3238004", "398623003", "417981000", "44993000", "50920009", "58880005", "72843002", "76232008"], "icd10": ["I09.2", "I30", "I30.0", "I30.1", "I30.8", "I30.9", "I31.0", "I31.1", "I32"], "icd10_nodot": ["I092", "I300", "I301", "I308", "I309", "I30X", "I310", "I311", "I32X"]},
    "Pericardial Effusion and Cardiac Tamponade": {"short": "PCE", "snomed": ["233867009", "309068002", "3238004", "35304003", "35436006", "373945007", "431082004"], "icd10": ["I31.2", "I31.3", "I31.4", "I31.8", "I31.9"], "icd10_nodot": ["I312", "I313", "I314", "I318", "I319"]},
    "Kawasaki Disease": {"short": "KAW", "snomed": ["195353004", "233944000", "398274000", "52674009", "75053002"], "icd10": ["I25.4", "I25.41", "M30.0", "M30.3"], "icd10_nodot": ["I254", "I2541", "M300", "M303"]},
    "Chagas Disease": {"short": "CHA", "snomed": ["16859001", "233870007", "34051009", "371129009", "426850004", "77506005"], "icd10": ["B57.0", "B57.1", "B57.2", "B57.5"], "icd10_nodot": ["B570", "B571", "B572", "B575"]},
    # ── Heart Failure ──
    "Heart Failure": {"short": "HFA", "snomed": ["10335000", "10633002", "1204200007", "1204203009", "1204204003", "1204206001", "128404006", "153931000119109", "153941000119100", "153951000119103", "15771000119104", "15781000119107", "194767001", "194779001", "194781004", "367363000", "417996009", "418304008", "42343007", "424404003", "426263006", "426611007", "43736008", "44088000", "441481004", "44313006", "443253003", "443254009", "446221000", "46113002", "471880001", "48447003", "5148006", "56675007", "67431000119105", "67441000119101", "6895651000006113", "698592004", "698594003", "703272007", "717840005", "7321121000006119", "741701000006114", "788950000", "82523003", "84114007", "85232009", "88805009", "92506005"], "icd10": ["I11.0", "I13.0", "I13.2", "I25.5", "I42.0", "I42.1", "I42.2", "I42.3", "I42.4", "I42.5", "I42.6", "I42.7", "I42.8", "I42.9", "I43", "I50", "I50.0", "I50.1", "I50.9", "O90.3"], "icd10_nodot": ["I110", "I130", "I132", "I255", "I420", "I421", "I422", "I423", "I424", "I425", "I426", "I427", "I428", "I429", "I43X", "I500", "I501", "I509", "I50X", "O903"]},
    # ── Additional Valvular (from VHD Master) ──
    "Multi-valve Disease": {"short": "MVD", "snomed": ["194733006", "368009", "398995000"], "icd10": ["I08.0", "I08.1", "I08.2", "I08.3", "I08.8", "I08.9"], "icd10_nodot": ["I080", "I081", "I082", "I083", "I088", "I089"]},
    "Prosthetic Valve Disease": {"short": "PVX", "snomed": ["24211005"], "icd10": ["T82.01", "T82.02", "T82.03", "T82.09", "T82.6", "Z95.2", "Z95.3", "Z95.4"], "icd10_nodot": ["T8201", "T8202", "T8203", "T8209", "T826", "Z952", "Z953", "Z954"]},
}


# ═══════════════════════════════════════════════════════════════════════
# DRUG CODE LIBRARY — 315 CV/Cardiometabolic drugs, 18 therapeutic classes
# Source: CPRD_Aurum_CV_Metabolic_Drug_CodeList.xlsx v3 (2026-02-18)
# Each entry has search terms for CPRD Product Dictionary matching
# ═══════════════════════════════════════════════════════════════════════

DRUG_GROUPS = {
    "Antiarrhythmics": [
        "Disopyramide",
        "Procainamide",
        "Quinidine",
        "Lidocaine",
        "Mexiletine",
        "Flecainide",
        "Propafenone",
        "Amiodarone",
        "Dronedarone",
        "Dofetilide",
        "Ibutilide",
        "Vernakalant",
        "Adenosine",
        "Atropine",
        "Magnesium Sulfate",
        "Phenytoin",
    ],
    "Anticoagulants": [
        "Warfarin",
        "Acenocoumarol",
        "Phenindione",
        "Apixaban",
        "Rivaroxaban",
        "Edoxaban",
        "Dabigatran",
        "Enoxaparin",
        "Dalteparin",
        "Tinzaparin",
        "Bemiparin",
        "Unfractionated Heparin",
        "Fondaparinux",
        "Argatroban",
        "Bivalirudin",
        "Idarucizumab",
        "Andexanet Alfa",
        "Protamine Sulfate",
        "Phytomenadione (Vitamin K1)",
        "Prothrombin Complex Concentrate",
    ],
    "Antifibrinolytics": ["Tranexamic Acid"],
    "Antihypertensives": [
        "Captopril",
        "Enalapril",
        "Lisinopril",
        "Ramipril",
        "Perindopril",
        "Quinapril",
        "Fosinopril",
        "Trandolapril",
        "Moexipril",
        "Benazepril",
        "Cilazapril",
        "Imidapril",
        "Losartan",
        "Candesartan",
        "Valsartan",
        "Irbesartan",
        "Telmisartan",
        "Olmesartan",
        "Eprosartan",
        "Azilsartan",
        "Losartan/HCTZ",
        "Candesartan/HCTZ",
        "Valsartan/HCTZ",
        "Irbesartan/HCTZ",
        "Telmisartan/HCTZ",
        "Olmesartan/HCTZ",
        "Enalapril/HCTZ",
        "Lisinopril/HCTZ",
        "Perindopril/Indapamide",
        "Perindopril/Amlodipine",
        "Perindopril/Indapamide/Amlodipine",
        "Valsartan/Amlodipine",
        "Valsartan/Amlodipine/HCTZ",
        "Ramipril/Felodipine",
        "Bisoprolol/HCTZ",
        "Atenolol/Chlortalidone",
        "Co-amilofruse",
        "Co-amilozide",
        "Co-triamterzide",
        "Co-flumactone",
        "Atenolol",
        "Bisoprolol",
        "Metoprolol Tartrate",
        "Metoprolol Succinate",
        "Propranolol",
        "Carvedilol",
        "Nebivolol",
        "Sotalol",
        "Labetalol",
        "Nadolol",
        "Pindolol",
        "Acebutolol",
        "Celiprolol",
        "Oxprenolol",
        "Timolol",
        "Betaxolol",
        "Esmolol",
        "Landiolol",
        "Amlodipine",
        "Nifedipine",
        "Felodipine",
        "Lercanidipine",
        "Lacidipine",
        "Nicardipine",
        "Nimodipine",
        "Isradipine",
        "Nisoldipine",
        "Clevidipine",
        "Diltiazem",
        "Verapamil",
        "Etripamil",
        "Bendroflumethiazide",
        "Hydrochlorothiazide",
        "Indapamide",
        "Chlorthalidone",
        "Cyclopenthiazide",
        "Metolazone",
        "Xipamide",
        "Furosemide",
        "Bumetanide",
        "Torsemide/Torasemide",
        "Ethacrynic Acid",
        "Spironolactone",
        "Eplerenone",
        "Finerenone",
        "Amiloride",
        "Triamterene",
        "Methyldopa",
        "Clonidine",
        "Moxonidine",
        "Guanfacine",
        "Doxazosin",
        "Prazosin",
        "Terazosin",
        "Phenoxybenzamine",
        "Phentolamine",
        "Hydralazine",
        "Minoxidil",
        "Sodium Nitroprusside",
        "Aliskiren",
        "Baxdrostat",
        "Lorundrostat",
        "Fenoldopam",
    ],
    "Antiplatelets": [
        "Aspirin (Antiplatelet)",
        "Clopidogrel",
        "Ticagrelor",
        "Prasugrel",
        "Ticlopidine",
        "Cangrelor",
        "Dipyridamole",
        "Cilostazol",
        "Abciximab",
        "Eptifibatide",
        "Tirofiban",
        "Vorapaxar",
    ],
    "Cardiac Amyloidosis": ["Tafamidis", "Acoramidis", "Patisiran", "Vutrisiran", "Inotersen", "Daratumumab"],
    "Critical Care & Vasoactive": [
        "Norepinephrine",
        "Epinephrine",
        "Dopamine",
        "Dobutamine",
        "Isoproterenol/Isoprenaline",
        "Vasopressin",
        "Phenylephrine",
        "Metaraminol",
        "Milrinone",
        "Levosimendan",
        "Angiotensin II",
    ],
    "Glucose-Lowering": [
        "Metformin",
        "Metformin/Sitagliptin",
        "Metformin/Vildagliptin",
        "Metformin/Linagliptin",
        "Metformin/Saxagliptin",
        "Metformin/Alogliptin",
        "Metformin/Pioglitazone",
        "Metformin/Dapagliflozin",
        "Metformin/Empagliflozin",
        "Metformin/Canagliflozin",
        "Gliclazide",
        "Glimepiride",
        "Glipizide",
        "Glyburide/Glibenclamide",
        "Tolbutamide",
        "Chlorpropamide",
        "Repaglinide",
        "Nateglinide",
        "Pioglitazone",
        "Rosiglitazone",
        "Acarbose",
        "Miglitol",
        "Sitagliptin",
        "Vildagliptin",
        "Saxagliptin",
        "Linagliptin",
        "Alogliptin",
        "Dapagliflozin",
        "Empagliflozin",
        "Canagliflozin",
        "Ertugliflozin",
        "Sotagliflozin",
        "Dapagliflozin/Saxagliptin",
        "Empagliflozin/Linagliptin",
        "Ertugliflozin/Sitagliptin",
        "Liraglutide",
        "Semaglutide (SC)",
        "Semaglutide (Oral)",
        "Dulaglutide",
        "Exenatide",
        "Lixisenatide",
        "Tirzepatide",
        "Survodutide",
        "Pramlintide",
        "Insulin Degludec/Liraglutide",
        "Insulin Glargine/Lixisenatide",
        "Insulin Aspart",
        "Insulin Lispro",
        "Insulin Lispro Biosimilar",
        "Insulin Glulisine",
        "Soluble Insulin (Regular)",
        "Isophane Insulin (NPH)",
        "Biphasic Insulin Mixtures",
        "Insulin Glargine",
        "Insulin Glargine U-300",
        "Insulin Detemir",
        "Insulin Degludec",
        "Technosphere Insulin",
    ],
    "Heart Failure": [
        "Sacubitril/Valsartan",
        "Ivabradine",
        "Vericiguat",
        "Digoxin",
        "Digitoxin",
        "Hydralazine/Isosorbide Dinitrate",
        "Nesiritide",
        "Omecamtiv Mecarbil",
        "Mavacamten",
        "Aficamten",
    ],
    "Lipid-Lowering": [
        "Atorvastatin",
        "Simvastatin",
        "Rosuvastatin",
        "Pravastatin",
        "Fluvastatin",
        "Pitavastatin",
        "Lovastatin",
        "Cerivastatin",
        "Simvastatin/Ezetimibe",
        "Atorvastatin/Ezetimibe",
        "Rosuvastatin/Ezetimibe",
        "Amlodipine/Atorvastatin",
        "Fenofibrate/Simvastatin",
        "Ezetimibe",
        "Fenofibrate",
        "Bezafibrate",
        "Gemfibrozil",
        "Ciprofibrate",
        "Cholestyramine",
        "Colestipol",
        "Colesevelam",
        "Evolocumab",
        "Alirocumab",
        "Inclisiran",
        "Bempedoic Acid",
        "Bempedoic Acid/Ezetimibe",
        "Icosapent Ethyl",
        "Omega-3 Acid Ethyl Esters",
        "Omega-3 Marine Triglycerides",
        "Niacin ER",
        "Acipimox",
        "Nicotinic Acid/Laropiprant",
        "Pelacarsen",
        "Olezarsen",
        "Evinacumab",
        "Lomitapide",
        "Mipomersen",
    ],
    "Local Sclerosants": ["Sodium Tetradecyl Sulfate", "Polidocanol"],
    "Nitrates & Antianginals": [
        "Glyceryl Trinitrate (NTG)",
        "Isosorbide Mononitrate",
        "Isosorbide Dinitrate",
        "Nicorandil",
        "Ranolazine",
        "Trimetazidine",
        "Perhexiline",
    ],
    "Obesity (CV-Relevant)": [
        "Liraglutide 3mg",
        "Semaglutide 2.4mg",
        "Tirzepatide (Obesity)",
        "Orlistat",
        "Naltrexone/Bupropion",
        "Phentermine/Topiramate",
    ],
    "Pericarditis & Inflammatory": ["Colchicine", "Anakinra", "Rilonacept"],
    "Peripheral Vascular Disease": ["Naftidrofuryl", "Pentoxifylline", "Inositol Nicotinate"],
    "Potassium Management": [
        "Sodium Zirconium Cyclosilicate",
        "Patiromer",
        "Calcium/Sodium Polystyrene Sulfonate",
        "Potassium Chloride",
    ],
    "Pulmonary Hypertension": [
        "Bosentan",
        "Ambrisentan",
        "Macitentan",
        "Macitentan/Tadalafil",
        "Sildenafil (PAH)",
        "Tadalafil (PAH)",
        "Riociguat",
        "Epoprostenol",
        "Iloprost",
        "Treprostinil",
        "Selexipag",
    ],
    "Thrombolytics": ["Alteplase", "Tenecteplase", "Reteplase", "Streptokinase", "Urokinase"],
}

DRUG_CODE_LIBRARY = {
    "Abciximab": {"class": "Antiplatelets", "cat": "GP IIb/IIIa Inhibitor", "bnf": "2.9", "route": "IV", "rel": "Secondary", "terms": ["abciximab", "reopro"]},
    "Acarbose": {"class": "Glucose-Lowering", "cat": "Alpha-Glucosidase Inhibitor", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["acarbose", "glucobay"]},
    "Acebutolol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral", "rel": "Primary", "terms": ["acebutolol", "sectral"]},
    "Acenocoumarol": {"class": "Anticoagulants", "cat": "Coumarin", "bnf": "2.8.2", "route": "Oral", "rel": "Primary",
        "terms": ["acenocoumarol", "sinthrome", "nicoumalone"]},
    "Acipimox": {"class": "Lipid-Lowering", "cat": "Nicotinic Acid Derivative", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["acipimox", "olbetam"]},
    "Acoramidis": {"class": "Cardiac Amyloidosis", "cat": "TTR Stabiliser", "bnf": "—", "route": "Oral", "rel": "Secondary",
        "terms": ["acoramidis", "attruby"]},
    "Adenosine": {"class": "Antiarrhythmics", "cat": "Other", "bnf": "2.3.2", "route": "IV", "rel": "Secondary", "terms": ["adenosine", "adenocor"]},
    "Aficamten": {"class": "Heart Failure", "cat": "Cardiac Myosin Inhibitor", "bnf": "2.1", "route": "Oral", "rel": "Secondary", "terms": ["aficamten"]},
    "Alirocumab": {"class": "Lipid-Lowering", "cat": "PCSK9 Inhibitor (mAb)", "bnf": "2.12", "route": "SC", "rel": "Primary",
        "terms": ["alirocumab", "praluent"]},
    "Aliskiren": {"class": "Antihypertensives", "cat": "Direct Renin Inhibitor", "bnf": "2.5.5.3", "route": "Oral", "rel": "Primary",
        "terms": ["aliskiren", "rasilez"]},
    "Alogliptin": {"class": "Glucose-Lowering", "cat": "DPP-4 Inhibitor", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["alogliptin", "vipidia"]},
    "Alteplase": {"class": "Thrombolytics", "cat": "tPA", "bnf": "2.10.2", "route": "IV", "rel": "Secondary", "terms": ["alteplase", "actilyse", "rt-pa"]},
    "Ambrisentan": {"class": "Pulmonary Hypertension", "cat": "ERA", "bnf": "2.5.1.2", "route": "Oral", "rel": "Primary", "terms": ["ambrisentan", "volibris"]},
    "Amiloride": {"class": "Antihypertensives", "cat": "K+-Sparing Diuretic", "bnf": "2.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["amiloride", "midamor"]},
    "Amiodarone": {"class": "Antiarrhythmics", "cat": "Class III", "bnf": "2.3.2", "route": "Oral/IV", "rel": "Primary", "terms": ["amiodarone", "cordarone"]},
    "Amlodipine": {"class": "Antihypertensives", "cat": "CCB (Dihydropyridine)", "bnf": "2.6.2", "route": "Oral", "rel": "Primary",
        "terms": ["amlodipine", "istin"]},
    "Amlodipine/Atorvastatin": {"class": "Lipid-Lowering", "cat": "Statin+CCB", "bnf": "2.12/2.6", "route": "Oral", "rel": "Primary",
        "terms": ["amlodipine/atorvastatin", "caduet", "amlodipine atorvastatin"]},
    "Anakinra": {"class": "Pericarditis & Inflammatory", "cat": "IL-1 Antagonist", "bnf": "10.1.3", "route": "SC", "rel": "Secondary",
        "terms": ["anakinra", "kineret"]},
    "Andexanet Alfa": {"class": "Anticoagulants", "cat": "Reversal Agent", "bnf": "2.8.2", "route": "IV", "rel": "Secondary",
        "terms": ["andexanet alfa", "ondexxya"]},
    "Angiotensin II": {"class": "Critical Care & Vasoactive", "cat": "Vasopressor", "bnf": "2.7.2", "route": "IV", "rel": "Secondary",
        "terms": ["angiotensin ii injection", "giapreza"]},
    "Apixaban": {"class": "Anticoagulants", "cat": "DOAC", "bnf": "2.8.2", "route": "Oral", "rel": "Primary", "terms": ["apixaban", "eliquis"]},
    "Argatroban": {"class": "Anticoagulants", "cat": "Direct Thrombin Inhibitor", "bnf": "2.8.1", "route": "IV", "rel": "Secondary",
        "terms": ["argatroban", "exembol"]},
    "Aspirin (Antiplatelet)": {"class": "Antiplatelets", "cat": "COX Inhibitor", "bnf": "2.9", "route": "Oral", "rel": "Primary",
        "terms": ["aspirin 75", "aspirin 150", "aspirin 300", "dispersible aspirin", "nu-seals aspirin", "micropirin", "disprin cv", "acetylsalicylic acid"]},
    "Atenolol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral", "rel": "Primary", "terms": ["atenolol", "tenormin"]},
    "Atenolol/Chlortalidone": {"class": "Antihypertensives", "cat": "BB/Diuretic Combination", "bnf": "2.4/2.2", "route": "Oral", "rel": "Primary",
        "terms": ["atenolol/chlortalidone", "co-tenidone", "tenoretic", "atenolol chlorthalidone"]},
    "Atorvastatin": {"class": "Lipid-Lowering", "cat": "Statin", "bnf": "2.12", "route": "Oral", "rel": "Primary", "terms": ["atorvastatin", "lipitor"]},
    "Atorvastatin/Ezetimibe": {"class": "Lipid-Lowering", "cat": "Statin+Ezetimibe", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["atorvastatin/ezetimibe", "atorvastatin ezetimibe"]},
    "Atropine": {"class": "Antiarrhythmics", "cat": "Other", "bnf": "2.3.2", "route": "IV", "rel": "Secondary",
        "terms": ["atropine injection", "atropine sulfate inj"]},
    "Azilsartan": {"class": "Antihypertensives", "cat": "ARB", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary", "terms": ["azilsartan", "edarbi"]},
    "Baxdrostat": {"class": "Antihypertensives", "cat": "Aldosterone Synthase Inhibitor", "bnf": "2.5", "route": "Oral", "rel": "Secondary",
        "terms": ["baxdrostat"]},
    "Bemiparin": {"class": "Anticoagulants", "cat": "LMWH", "bnf": "2.8.1", "route": "SC", "rel": "Secondary", "terms": ["bemiparin", "zibor"]},
    "Bempedoic Acid": {"class": "Lipid-Lowering", "cat": "ACL Inhibitor", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["bempedoic acid", "nilemdo"]},
    "Bempedoic Acid/Ezetimibe": {"class": "Lipid-Lowering", "cat": "ACL Inhibitor+Ezetimibe", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["bempedoic acid/ezetimibe", "nustendi", "bempedoic ezetimibe"]},
    "Benazepril": {"class": "Antihypertensives", "cat": "ACE Inhibitor", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["benazepril", "lotensin"]},
    "Bendroflumethiazide": {"class": "Antihypertensives", "cat": "Thiazide Diuretic", "bnf": "2.2.1", "route": "Oral", "rel": "Primary",
        "terms": ["bendroflumethiazide", "aprinox", "bendrofluazide"]},
    "Betaxolol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral", "rel": "Primary", "terms": ["betaxolol oral", "kerlone"]},
    "Bezafibrate": {"class": "Lipid-Lowering", "cat": "Fibrate", "bnf": "2.12", "route": "Oral", "rel": "Primary", "terms": ["bezafibrate", "bezalip"]},
    "Biphasic Insulin Mixtures": {"class": "Glucose-Lowering", "cat": "Pre-mixed Insulin", "bnf": "6.1.1.2", "route": "SC", "rel": "Primary",
        "terms": ["biphasic insulin", "novomix", "humalog mix", "humulin m3", "insuman comb", "mixtard"]},
    "Bisoprolol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral", "rel": "Primary",
        "terms": ["bisoprolol", "cardicor", "emcor"]},
    "Bisoprolol/HCTZ": {"class": "Antihypertensives", "cat": "ACEi/Diuretic Combination", "bnf": "2.4/2.2", "route": "Oral", "rel": "Primary",
        "terms": ["bisoprolol/hydrochlorothiazide", "lodoz", "bisoprolol hctz"]},
    "Bivalirudin": {"class": "Anticoagulants", "cat": "Direct Thrombin Inhibitor", "bnf": "2.8.1", "route": "IV", "rel": "Secondary",
        "terms": ["bivalirudin", "angiox"]},
    "Bosentan": {"class": "Pulmonary Hypertension", "cat": "ERA", "bnf": "2.5.1.2", "route": "Oral", "rel": "Primary", "terms": ["bosentan", "tracleer"]},
    "Bumetanide": {"class": "Antihypertensives", "cat": "Loop Diuretic", "bnf": "2.2.2", "route": "Oral/IV", "rel": "Primary",
        "terms": ["bumetanide", "burinex"]},
    "Calcium/Sodium Polystyrene Sulfonate": {"class": "Potassium Management", "cat": "K+ Binder", "bnf": "9.2.1.1", "route": "Oral/Rectal", "rel": "Primary",
        "terms": ["calcium polystyrene sulfonate", "calcium resonium", "sodium polystyrene", "resonium a", "kayexalate", "polystyrene sulphonate"]},
    "Canagliflozin": {"class": "Glucose-Lowering", "cat": "SGLT2 Inhibitor", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["canagliflozin", "invokana"]},
    "Candesartan": {"class": "Antihypertensives", "cat": "ARB", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary",
        "terms": ["candesartan", "amias", "atacand"]},
    "Candesartan/HCTZ": {"class": "Antihypertensives", "cat": "ARB/Diuretic Combination", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary",
        "terms": ["candesartan/hydrochlorothiazide", "atacand plus", "candesartan hctz"]},
    "Cangrelor": {"class": "Antiplatelets", "cat": "P2Y12 Inhibitor (IV)", "bnf": "2.9", "route": "IV", "rel": "Secondary",
        "terms": ["cangrelor", "kengrexal"]},
    "Captopril": {"class": "Antihypertensives", "cat": "ACE Inhibitor", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary", "terms": ["captopril", "capoten"]},
    "Carvedilol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral", "rel": "Primary", "terms": ["carvedilol", "eucardic"]},
    "Celiprolol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral", "rel": "Primary", "terms": ["celiprolol", "celectol"]},
    "Cerivastatin": {"class": "Lipid-Lowering", "cat": "Statin (Withdrawn)", "bnf": "2.12", "route": "Oral", "rel": "Historical",
        "terms": ["cerivastatin", "lipobay", "baycol"]},
    "Chlorpropamide": {"class": "Glucose-Lowering", "cat": "Sulfonylurea", "bnf": "6.1.2.1", "route": "Oral", "rel": "Historical",
        "terms": ["chlorpropamide", "diabinese"]},
    "Chlorthalidone": {"class": "Antihypertensives", "cat": "Thiazide-like Diuretic", "bnf": "2.2.1", "route": "Oral", "rel": "Primary",
        "terms": ["chlorthalidone", "chlortalidone", "hygroton"]},
    "Cholestyramine": {"class": "Lipid-Lowering", "cat": "Bile Acid Sequestrant", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["cholestyramine", "colestyramine", "questran"]},
    "Cilazapril": {"class": "Antihypertensives", "cat": "ACE Inhibitor", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["cilazapril", "vascace"]},
    "Cilostazol": {"class": "Antiplatelets", "cat": "PDE3 Inhibitor", "bnf": "2.9", "route": "Oral", "rel": "Primary", "terms": ["cilostazol", "pletal"]},
    "Ciprofibrate": {"class": "Lipid-Lowering", "cat": "Fibrate", "bnf": "2.12", "route": "Oral", "rel": "Primary", "terms": ["ciprofibrate", "modalim"]},
    "Clevidipine": {"class": "Antihypertensives", "cat": "CCB (Dihydropyridine)", "bnf": "2.6.2", "route": "IV", "rel": "Secondary",
        "terms": ["clevidipine", "cleviprex"]},
    "Clonidine": {"class": "Antihypertensives", "cat": "Centrally-Acting", "bnf": "2.5.2", "route": "Oral", "rel": "Primary",
        "terms": ["clonidine", "catapres", "dixarit"]},
    "Clopidogrel": {"class": "Antiplatelets", "cat": "P2Y12 Inhibitor", "bnf": "2.9", "route": "Oral", "rel": "Primary",
        "terms": ["clopidogrel", "plavix", "grepid"]},
    "Co-amilofruse": {"class": "Antihypertensives", "cat": "Diuretic Combination", "bnf": "2.2.4", "route": "Oral", "rel": "Primary",
        "terms": ["co-amilofruse", "frumil", "amiloride/furosemide", "amiloride furosemide"]},
    "Co-amilozide": {"class": "Antihypertensives", "cat": "Diuretic Combination", "bnf": "2.2.4", "route": "Oral", "rel": "Primary",
        "terms": ["co-amilozide", "moduretic", "amiloride/hydrochlorothiazide", "amiloride hctz"]},
    "Co-flumactone": {"class": "Antihypertensives", "cat": "Diuretic Combination", "bnf": "2.2.4", "route": "Oral", "rel": "Primary",
        "terms": ["co-flumactone", "aldactide", "spironolactone/hydroflumethiazide"]},
    "Co-triamterzide": {"class": "Antihypertensives", "cat": "Diuretic Combination", "bnf": "2.2.4", "route": "Oral", "rel": "Primary",
        "terms": ["co-triamterzide", "dyazide", "triamterene/hydrochlorothiazide", "triamterene hctz"]},
    "Colchicine": {"class": "Pericarditis & Inflammatory", "cat": "Anti-inflammatory", "bnf": "10.1.4/2.12", "route": "Oral", "rel": "Primary",
        "terms": ["colchicine"]},
    "Colesevelam": {"class": "Lipid-Lowering", "cat": "Bile Acid Sequestrant", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["colesevelam", "cholestagel"]},
    "Colestipol": {"class": "Lipid-Lowering", "cat": "Bile Acid Sequestrant", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["colestipol", "colestid"]},
    "Cyclopenthiazide": {"class": "Antihypertensives", "cat": "Thiazide Diuretic", "bnf": "2.2.1", "route": "Oral", "rel": "Primary",
        "terms": ["cyclopenthiazide", "navidrex"]},
    "Dabigatran": {"class": "Anticoagulants", "cat": "DOAC", "bnf": "2.8.2", "route": "Oral", "rel": "Primary", "terms": ["dabigatran", "pradaxa"]},
    "Dalteparin": {"class": "Anticoagulants", "cat": "LMWH", "bnf": "2.8.1", "route": "SC", "rel": "Primary", "terms": ["dalteparin", "fragmin"]},
    "Dapagliflozin": {"class": "Glucose-Lowering", "cat": "SGLT2 Inhibitor", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["dapagliflozin", "forxiga"]},
    "Dapagliflozin/Saxagliptin": {"class": "Glucose-Lowering", "cat": "SGLT2i+DPP4i Combo", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["dapagliflozin/saxagliptin", "qtern", "dapagliflozin saxagliptin"]},
    "Daratumumab": {"class": "Cardiac Amyloidosis", "cat": "Anti-CD38", "bnf": "—", "route": "IV/SC", "rel": "Secondary", "terms": ["daratumumab", "darzalex"]},
    "Digitoxin": {"class": "Heart Failure", "cat": "Cardiac Glycoside", "bnf": "2.1.1", "route": "Oral", "rel": "Secondary", "terms": ["digitoxin", "digitek"]},
    "Digoxin": {"class": "Heart Failure", "cat": "Cardiac Glycoside", "bnf": "2.1.1", "route": "Oral", "rel": "Primary",
        "terms": ["digoxin", "lanoxin", "digoxin-nativelle"]},
    "Diltiazem": {"class": "Antihypertensives", "cat": "CCB (Non-dihydropyridine)", "bnf": "2.6.2", "route": "Oral", "rel": "Primary",
        "terms": ["diltiazem", "tildiem", "adizem", "angitil", "calcicard", "dilzem", "slozem", "viazem", "zemtard", "alizem"]},
    "Dipyridamole": {"class": "Antiplatelets", "cat": "PDE3 Inhibitor", "bnf": "2.9", "route": "Oral", "rel": "Primary",
        "terms": ["dipyridamole", "persantin", "asasantin"]},
    "Disopyramide": {"class": "Antiarrhythmics", "cat": "Class Ia", "bnf": "2.3.1", "route": "Oral", "rel": "Primary", "terms": ["disopyramide", "rythmodan"]},
    "Dobutamine": {"class": "Critical Care & Vasoactive", "cat": "Catecholamine", "bnf": "2.7.2", "route": "IV", "rel": "Secondary",
        "terms": ["dobutamine", "dobutrex"]},
    "Dofetilide": {"class": "Antiarrhythmics", "cat": "Class III", "bnf": "2.3.2", "route": "Oral", "rel": "Secondary", "terms": ["dofetilide", "tikosyn"]},
    "Dopamine": {"class": "Critical Care & Vasoactive", "cat": "Catecholamine", "bnf": "2.7.2", "route": "IV", "rel": "Secondary",
        "terms": ["dopamine infusion", "dopamine hydrochloride"]},
    "Doxazosin": {"class": "Antihypertensives", "cat": "Alpha-Blocker", "bnf": "2.5.4", "route": "Oral", "rel": "Primary", "terms": ["doxazosin", "cardura"]},
    "Dronedarone": {"class": "Antiarrhythmics", "cat": "Class III", "bnf": "2.3.2", "route": "Oral", "rel": "Primary", "terms": ["dronedarone", "multaq"]},
    "Dulaglutide": {"class": "Glucose-Lowering", "cat": "GLP-1 RA", "bnf": "6.1.2.3", "route": "SC", "rel": "Primary", "terms": ["dulaglutide", "trulicity"]},
    "Edoxaban": {"class": "Anticoagulants", "cat": "DOAC", "bnf": "2.8.2", "route": "Oral", "rel": "Primary", "terms": ["edoxaban", "lixiana"]},
    "Empagliflozin": {"class": "Glucose-Lowering", "cat": "SGLT2 Inhibitor", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["empagliflozin", "jardiance"]},
    "Empagliflozin/Linagliptin": {"class": "Glucose-Lowering", "cat": "SGLT2i+DPP4i Combo", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["empagliflozin/linagliptin", "glyxambi", "empagliflozin linagliptin"]},
    "Enalapril": {"class": "Antihypertensives", "cat": "ACE Inhibitor", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["enalapril", "innovace"]},
    "Enalapril/HCTZ": {"class": "Antihypertensives", "cat": "ACEi/Diuretic Combination", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["enalapril/hydrochlorothiazide", "innozide", "enalapril hctz"]},
    "Enoxaparin": {"class": "Anticoagulants", "cat": "LMWH", "bnf": "2.8.1", "route": "SC", "rel": "Primary", "terms": ["enoxaparin", "clexane", "inhixa"]},
    "Epinephrine": {"class": "Critical Care & Vasoactive", "cat": "Catecholamine", "bnf": "2.7.2", "route": "IV/IM", "rel": "Secondary",
        "terms": ["epinephrine", "adrenaline", "epipen"]},
    "Eplerenone": {"class": "Antihypertensives", "cat": "K+-Sparing Diuretic", "bnf": "2.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["eplerenone", "inspra"]},
    "Epoprostenol": {"class": "Pulmonary Hypertension", "cat": "Prostacyclin Analogue", "bnf": "2.5.1.2", "route": "IV", "rel": "Secondary",
        "terms": ["epoprostenol", "flolan", "veletri"]},
    "Eprosartan": {"class": "Antihypertensives", "cat": "ARB", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary", "terms": ["eprosartan", "teveten"]},
    "Eptifibatide": {"class": "Antiplatelets", "cat": "GP IIb/IIIa Inhibitor", "bnf": "2.9", "route": "IV", "rel": "Secondary",
        "terms": ["eptifibatide", "integrilin"]},
    "Ertugliflozin": {"class": "Glucose-Lowering", "cat": "SGLT2 Inhibitor", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["ertugliflozin", "steglatro"]},
    "Ertugliflozin/Sitagliptin": {"class": "Glucose-Lowering", "cat": "SGLT2i+DPP4i Combo", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["ertugliflozin/sitagliptin", "steglujan", "ertugliflozin sitagliptin"]},
    "Esmolol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "IV", "rel": "Secondary", "terms": ["esmolol", "brevibloc"]},
    "Ethacrynic Acid": {"class": "Antihypertensives", "cat": "Loop Diuretic", "bnf": "2.2.2", "route": "Oral", "rel": "Secondary",
        "terms": ["ethacrynic acid", "etacrynic acid", "edecrin"]},
    "Etripamil": {"class": "Antihypertensives", "cat": "CCB (Non-dihydropyridine)", "bnf": "2.6.2", "route": "Intranasal", "rel": "Secondary",
        "terms": ["etripamil"]},
    "Evinacumab": {"class": "Lipid-Lowering", "cat": "ANGPTL3 Inhibitor", "bnf": "2.12", "route": "IV", "rel": "Secondary", "terms": ["evinacumab", "evkeeza"]},
    "Evolocumab": {"class": "Lipid-Lowering", "cat": "PCSK9 Inhibitor (mAb)", "bnf": "2.12", "route": "SC", "rel": "Primary",
        "terms": ["evolocumab", "repatha"]},
    "Exenatide": {"class": "Glucose-Lowering", "cat": "GLP-1 RA", "bnf": "6.1.2.3", "route": "SC", "rel": "Primary",
        "terms": ["exenatide", "byetta", "bydureon"]},
    "Ezetimibe": {"class": "Lipid-Lowering", "cat": "Cholesterol Absorption Inhibitor", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["ezetimibe", "ezetrol"]},
    "Felodipine": {"class": "Antihypertensives", "cat": "CCB (Dihydropyridine)", "bnf": "2.6.2", "route": "Oral", "rel": "Primary",
        "terms": ["felodipine", "plendil"]},
    "Fenofibrate": {"class": "Lipid-Lowering", "cat": "Fibrate", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["fenofibrate", "lipantil", "supralip"]},
    "Fenofibrate/Simvastatin": {"class": "Lipid-Lowering", "cat": "Statin+Fibrate (Withdrawn)", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["fenofibrate/simvastatin", "cholib"]},
    "Fenoldopam": {"class": "Antihypertensives", "cat": "Vasodilator", "bnf": "2.5.1", "route": "IV", "rel": "Secondary", "terms": ["fenoldopam", "corlopam"]},
    "Finerenone": {"class": "Antihypertensives", "cat": "Non-steroidal MRA", "bnf": "2.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["finerenone", "kerendia"]},
    "Flecainide": {"class": "Antiarrhythmics", "cat": "Class Ic", "bnf": "2.3.1", "route": "Oral/IV", "rel": "Primary", "terms": ["flecainide", "tambocor"]},
    "Fluvastatin": {"class": "Lipid-Lowering", "cat": "Statin", "bnf": "2.12", "route": "Oral", "rel": "Primary", "terms": ["fluvastatin", "lescol"]},
    "Fondaparinux": {"class": "Anticoagulants", "cat": "Indirect Factor Xa", "bnf": "2.8.1", "route": "SC", "rel": "Primary",
        "terms": ["fondaparinux", "arixtra"]},
    "Fosinopril": {"class": "Antihypertensives", "cat": "ACE Inhibitor", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["fosinopril", "staril"]},
    "Furosemide": {"class": "Antihypertensives", "cat": "Loop Diuretic", "bnf": "2.2.2", "route": "Oral/IV", "rel": "Primary",
        "terms": ["furosemide", "frusemide", "lasix", "frusol"]},
    "Gemfibrozil": {"class": "Lipid-Lowering", "cat": "Fibrate", "bnf": "2.12", "route": "Oral", "rel": "Primary", "terms": ["gemfibrozil", "lopid"]},
    "Gliclazide": {"class": "Glucose-Lowering", "cat": "Sulfonylurea", "bnf": "6.1.2.1", "route": "Oral", "rel": "Primary",
        "terms": ["gliclazide", "diamicron"]},
    "Glimepiride": {"class": "Glucose-Lowering", "cat": "Sulfonylurea", "bnf": "6.1.2.1", "route": "Oral", "rel": "Primary",
        "terms": ["glimepiride", "amaryl"]},
    "Glipizide": {"class": "Glucose-Lowering", "cat": "Sulfonylurea", "bnf": "6.1.2.1", "route": "Oral", "rel": "Primary",
        "terms": ["glipizide", "minodiab", "glibenese"]},
    "Glyburide/Glibenclamide": {"class": "Glucose-Lowering", "cat": "Sulfonylurea", "bnf": "6.1.2.1", "route": "Oral", "rel": "Primary",
        "terms": ["glibenclamide", "glyburide", "daonil", "semi-daonil"]},
    "Glyceryl Trinitrate (NTG)": {"class": "Nitrates & Antianginals", "cat": "Nitrate", "bnf": "2.6.1", "route": "SL/Spray/TD/IV", "rel": "Primary",
        "terms": ["glyceryl trinitrate", "gtn", "nitrolingual", "nitromin", "nitro-dur", "deponit", "transiderm-nitro", "minitran", "percutol", "rectogesic", "suscard", "trintek"]},
    "Guanfacine": {"class": "Antihypertensives", "cat": "Centrally-Acting", "bnf": "2.5.2", "route": "Oral", "rel": "Secondary",
        "terms": ["guanfacine", "tenex", "intuniv"]},
    "Hydralazine": {"class": "Antihypertensives", "cat": "Direct Vasodilator", "bnf": "2.5.1", "route": "Oral/IV", "rel": "Primary",
        "terms": ["hydralazine", "apresoline"]},
    "Hydralazine/Isosorbide Dinitrate": {"class": "Heart Failure", "cat": "Vasodilator Combination", "bnf": "2.5.1/2.6.1", "route": "Oral", "rel": "Primary",
        "terms": ["hydralazine/isosorbide dinitrate", "bidil", "hydralazine isdn"]},
    "Hydrochlorothiazide": {"class": "Antihypertensives", "cat": "Thiazide Diuretic", "bnf": "2.2.1", "route": "Oral", "rel": "Primary",
        "terms": ["hydrochlorothiazide", "hctz"]},
    "Ibutilide": {"class": "Antiarrhythmics", "cat": "Class III", "bnf": "2.3.2", "route": "IV", "rel": "Secondary", "terms": ["ibutilide", "corvert"]},
    "Icosapent Ethyl": {"class": "Lipid-Lowering", "cat": "Omega-3 Fatty Acid", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["icosapent ethyl", "vazkepa"]},
    "Idarucizumab": {"class": "Anticoagulants", "cat": "Reversal Agent", "bnf": "2.8.2", "route": "IV", "rel": "Secondary",
        "terms": ["idarucizumab", "praxbind"]},
    "Iloprost": {"class": "Pulmonary Hypertension", "cat": "Prostacyclin Analogue", "bnf": "2.5.1.2", "route": "Inhaled", "rel": "Secondary",
        "terms": ["iloprost", "ventavis", "ilomedin"]},
    "Imidapril": {"class": "Antihypertensives", "cat": "ACE Inhibitor", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["imidapril", "tanatril"]},
    "Inclisiran": {"class": "Lipid-Lowering", "cat": "PCSK9 Inhibitor (siRNA)", "bnf": "2.12", "route": "SC", "rel": "Primary",
        "terms": ["inclisiran", "leqvio"]},
    "Indapamide": {"class": "Antihypertensives", "cat": "Thiazide-like Diuretic", "bnf": "2.2.1", "route": "Oral", "rel": "Primary",
        "terms": ["indapamide", "natrilix"]},
    "Inositol Nicotinate": {"class": "Peripheral Vascular Disease", "cat": "Vasodilator", "bnf": "2.6.4", "route": "Oral", "rel": "Primary",
        "terms": ["inositol nicotinate", "hexopal"]},
    "Inotersen": {"class": "Cardiac Amyloidosis", "cat": "TTR Gene Silencer", "bnf": "—", "route": "SC", "rel": "Secondary", "terms": ["inotersen", "tegsedi"]},
    "Insulin Aspart": {"class": "Glucose-Lowering", "cat": "Rapid-Acting Insulin", "bnf": "6.1.1.1", "route": "SC/IV", "rel": "Primary",
        "terms": ["insulin aspart", "novorapid", "fiasp"]},
    "Insulin Degludec": {"class": "Glucose-Lowering", "cat": "Ultra-Long Insulin", "bnf": "6.1.1.2", "route": "SC", "rel": "Primary",
        "terms": ["insulin degludec", "tresiba"]},
    "Insulin Degludec/Liraglutide": {"class": "Glucose-Lowering", "cat": "Insulin+GLP-1 RA", "bnf": "6.1.1/6.1.2", "route": "SC", "rel": "Primary",
        "terms": ["insulin degludec/liraglutide", "xultophy", "degludec liraglutide"]},
    "Insulin Detemir": {"class": "Glucose-Lowering", "cat": "Long-Acting Insulin", "bnf": "6.1.1.2", "route": "SC", "rel": "Primary",
        "terms": ["insulin detemir", "levemir"]},
    "Insulin Glargine": {"class": "Glucose-Lowering", "cat": "Long-Acting Insulin", "bnf": "6.1.1.2", "route": "SC", "rel": "Primary",
        "terms": ["insulin glargine", "lantus", "semglee", "abasaglar", "optisulin"]},
    "Insulin Glargine U-300": {"class": "Glucose-Lowering", "cat": "Long-Acting Insulin", "bnf": "6.1.1.2", "route": "SC", "rel": "Primary",
        "terms": ["insulin glargine u-300", "toujeo"]},
    "Insulin Glargine/Lixisenatide": {"class": "Glucose-Lowering", "cat": "Insulin+GLP-1 RA", "bnf": "6.1.1/6.1.2", "route": "SC", "rel": "Primary",
        "terms": ["insulin glargine/lixisenatide", "suliqua", "glargine lixisenatide"]},
    "Insulin Glulisine": {"class": "Glucose-Lowering", "cat": "Rapid-Acting Insulin", "bnf": "6.1.1.1", "route": "SC", "rel": "Primary",
        "terms": ["insulin glulisine", "apidra"]},
    "Insulin Lispro": {"class": "Glucose-Lowering", "cat": "Rapid-Acting Insulin", "bnf": "6.1.1.1", "route": "SC/IV", "rel": "Primary",
        "terms": ["insulin lispro", "humalog", "lyumjev"]},
    "Insulin Lispro Biosimilar": {"class": "Glucose-Lowering", "cat": "Rapid-Acting Insulin", "bnf": "6.1.1.1", "route": "SC", "rel": "Primary",
        "terms": ["insulin lispro biosimilar", "admelog", "insulin lispro sanofi"]},
    "Irbesartan": {"class": "Antihypertensives", "cat": "ARB", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary", "terms": ["irbesartan", "aprovel"]},
    "Irbesartan/HCTZ": {"class": "Antihypertensives", "cat": "ARB/Diuretic Combination", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary",
        "terms": ["irbesartan/hydrochlorothiazide", "coaprovel", "irbesartan hctz"]},
    "Isophane Insulin (NPH)": {"class": "Glucose-Lowering", "cat": "Intermediate Insulin", "bnf": "6.1.1.2", "route": "SC", "rel": "Primary",
        "terms": ["isophane insulin", "insulatard", "humulin i", "insuman basal", "nph insulin", "protaphane"]},
    "Isoproterenol/Isoprenaline": {"class": "Critical Care & Vasoactive", "cat": "Beta Agonist", "bnf": "2.7.1", "route": "IV", "rel": "Secondary",
        "terms": ["isoprenaline", "isoproterenol", "isoproterenol hydrochloride"]},
    "Isosorbide Dinitrate": {"class": "Nitrates & Antianginals", "cat": "Nitrate", "bnf": "2.6.1", "route": "Oral/IV/SL", "rel": "Primary",
        "terms": ["isosorbide dinitrate", "isdn", "angitaka", "isoket", "cedocard"]},
    "Isosorbide Mononitrate": {"class": "Nitrates & Antianginals", "cat": "Nitrate", "bnf": "2.6.1", "route": "Oral", "rel": "Primary",
        "terms": ["isosorbide mononitrate", "ismn", "elantan", "imdur", "isib", "isotard", "chemydur", "monomax", "monosorb", "ismo", "moni"]},
    "Isradipine": {"class": "Antihypertensives", "cat": "CCB (Dihydropyridine)", "bnf": "2.6.2", "route": "Oral", "rel": "Primary",
        "terms": ["isradipine", "prescal"]},
    "Ivabradine": {"class": "Heart Failure", "cat": "If Channel Blocker", "bnf": "2.6.3", "route": "Oral", "rel": "Primary",
        "terms": ["ivabradine", "procoralan"]},
    "Labetalol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral/IV", "rel": "Primary", "terms": ["labetalol", "trandate"]},
    "Lacidipine": {"class": "Antihypertensives", "cat": "CCB (Dihydropyridine)", "bnf": "2.6.2", "route": "Oral", "rel": "Primary",
        "terms": ["lacidipine", "motens"]},
    "Landiolol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "IV", "rel": "Secondary", "terms": ["landiolol", "rapibloc"]},
    "Lercanidipine": {"class": "Antihypertensives", "cat": "CCB (Dihydropyridine)", "bnf": "2.6.2", "route": "Oral", "rel": "Primary",
        "terms": ["lercanidipine", "zanidip"]},
    "Levosimendan": {"class": "Critical Care & Vasoactive", "cat": "Ca2+ Sensitiser", "bnf": "2.1.2", "route": "IV", "rel": "Secondary",
        "terms": ["levosimendan", "simdax"]},
    "Lidocaine": {"class": "Antiarrhythmics", "cat": "Class Ib", "bnf": "2.3.1", "route": "IV", "rel": "Secondary",
        "terms": ["lidocaine", "lignocaine", "xylocard"]},
    "Linagliptin": {"class": "Glucose-Lowering", "cat": "DPP-4 Inhibitor", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["linagliptin", "trajenta"]},
    "Liraglutide": {"class": "Glucose-Lowering", "cat": "GLP-1 RA", "bnf": "6.1.2.3", "route": "SC", "rel": "Primary", "terms": ["liraglutide", "victoza"]},
    "Liraglutide 3mg": {"class": "Obesity (CV-Relevant)", "cat": "GLP-1 RA (Obesity)", "bnf": "4.5.1", "route": "SC", "rel": "Primary",
        "terms": ["liraglutide 3mg", "saxenda"]},
    "Lisinopril": {"class": "Antihypertensives", "cat": "ACE Inhibitor", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["lisinopril", "zestril", "carace"]},
    "Lisinopril/HCTZ": {"class": "Antihypertensives", "cat": "ACEi/Diuretic Combination", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["lisinopril/hydrochlorothiazide", "zestoretic", "lisinopril hctz"]},
    "Lixisenatide": {"class": "Glucose-Lowering", "cat": "GLP-1 RA", "bnf": "6.1.2.3", "route": "SC", "rel": "Primary", "terms": ["lixisenatide", "lyxumia"]},
    "Lomitapide": {"class": "Lipid-Lowering", "cat": "MTP Inhibitor", "bnf": "2.12", "route": "Oral", "rel": "Secondary",
        "terms": ["lomitapide", "juxtapid", "lojuxta"]},
    "Lorundrostat": {"class": "Antihypertensives", "cat": "Aldosterone Synthase Inhibitor", "bnf": "2.5", "route": "Oral", "rel": "Secondary",
        "terms": ["lorundrostat"]},
    "Losartan": {"class": "Antihypertensives", "cat": "ARB", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary", "terms": ["losartan", "cozaar"]},
    "Losartan/HCTZ": {"class": "Antihypertensives", "cat": "ARB/Diuretic Combination", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary",
        "terms": ["losartan/hydrochlorothiazide", "cozaar comp", "hyzaar", "losartan hctz"]},
    "Lovastatin": {"class": "Lipid-Lowering", "cat": "Statin", "bnf": "2.12", "route": "Oral", "rel": "Primary", "terms": ["lovastatin", "mevacor"]},
    "Macitentan": {"class": "Pulmonary Hypertension", "cat": "ERA", "bnf": "2.5.1.2", "route": "Oral", "rel": "Primary", "terms": ["macitentan", "opsumit"]},
    "Macitentan/Tadalafil": {"class": "Pulmonary Hypertension", "cat": "ERA+PDE5i Combination", "bnf": "2.5.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["macitentan/tadalafil", "opsynvi", "macitentan tadalafil"]},
    "Magnesium Sulfate": {"class": "Antiarrhythmics", "cat": "Other", "bnf": "2.3.2", "route": "IV", "rel": "Secondary",
        "terms": ["magnesium sulfate", "magnesium sulphate inj"]},
    "Mavacamten": {"class": "Heart Failure", "cat": "Cardiac Myosin Inhibitor", "bnf": "2.1", "route": "Oral", "rel": "Secondary",
        "terms": ["mavacamten", "camzyos"]},
    "Metaraminol": {"class": "Critical Care & Vasoactive", "cat": "Alpha Agonist", "bnf": "2.7.2", "route": "IV", "rel": "Secondary",
        "terms": ["metaraminol injection", "metaraminol"]},
    "Metformin": {"class": "Glucose-Lowering", "cat": "Biguanide", "bnf": "6.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["metformin", "glucophage", "sukkarto", "diagemet", "bolamyn"]},
    "Metformin/Alogliptin": {"class": "Glucose-Lowering", "cat": "Metformin Combination", "bnf": "6.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["metformin/alogliptin", "vipdomet", "metformin alogliptin"]},
    "Metformin/Canagliflozin": {"class": "Glucose-Lowering", "cat": "Metformin Combination", "bnf": "6.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["metformin/canagliflozin", "vokanamet", "metformin canagliflozin"]},
    "Metformin/Dapagliflozin": {"class": "Glucose-Lowering", "cat": "Metformin Combination", "bnf": "6.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["metformin/dapagliflozin", "xigduo", "metformin dapagliflozin"]},
    "Metformin/Empagliflozin": {"class": "Glucose-Lowering", "cat": "Metformin Combination", "bnf": "6.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["metformin/empagliflozin", "synjardy", "metformin empagliflozin"]},
    "Metformin/Linagliptin": {"class": "Glucose-Lowering", "cat": "Metformin Combination", "bnf": "6.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["metformin/linagliptin", "jentadueto", "metformin linagliptin"]},
    "Metformin/Pioglitazone": {"class": "Glucose-Lowering", "cat": "Metformin Combination", "bnf": "6.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["metformin/pioglitazone", "competact", "metformin pioglitazone"]},
    "Metformin/Saxagliptin": {"class": "Glucose-Lowering", "cat": "Metformin Combination", "bnf": "6.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["metformin/saxagliptin", "komboglyze", "metformin saxagliptin"]},
    "Metformin/Sitagliptin": {"class": "Glucose-Lowering", "cat": "Metformin Combination", "bnf": "6.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["metformin/sitagliptin", "janumet", "metformin sitagliptin"]},
    "Metformin/Vildagliptin": {"class": "Glucose-Lowering", "cat": "Metformin Combination", "bnf": "6.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["metformin/vildagliptin", "eucreas", "metformin vildagliptin"]},
    "Methyldopa": {"class": "Antihypertensives", "cat": "Centrally-Acting", "bnf": "2.5.2", "route": "Oral", "rel": "Primary",
        "terms": ["methyldopa", "aldomet"]},
    "Metolazone": {"class": "Antihypertensives", "cat": "Thiazide Diuretic", "bnf": "2.2.1", "route": "Oral", "rel": "Primary",
        "terms": ["metolazone", "metenix"]},
    "Metoprolol Succinate": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral", "rel": "Primary",
        "terms": ["metoprolol succinate", "betaloc sa", "metoprolol xl"]},
    "Metoprolol Tartrate": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral", "rel": "Primary",
        "terms": ["metoprolol tartrate", "lopresor", "betaloc"]},
    "Mexiletine": {"class": "Antiarrhythmics", "cat": "Class Ib", "bnf": "2.3.1", "route": "Oral", "rel": "Secondary",
        "terms": ["mexiletine", "mexitil", "namuscla"]},
    "Miglitol": {"class": "Glucose-Lowering", "cat": "Alpha-Glucosidase Inhibitor", "bnf": "6.1.2.3", "route": "Oral", "rel": "Secondary",
        "terms": ["miglitol", "glyset"]},
    "Milrinone": {"class": "Critical Care & Vasoactive", "cat": "PDE3 Inhibitor", "bnf": "2.1.2", "route": "IV", "rel": "Secondary",
        "terms": ["milrinone", "primacor"]},
    "Minoxidil": {"class": "Antihypertensives", "cat": "Direct Vasodilator", "bnf": "2.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["minoxidil oral", "loniten"]},
    "Mipomersen": {"class": "Lipid-Lowering", "cat": "ApoB Antisense", "bnf": "2.12", "route": "SC", "rel": "Secondary", "terms": ["mipomersen", "kynamro"]},
    "Moexipril": {"class": "Antihypertensives", "cat": "ACE Inhibitor", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary", "terms": ["moexipril", "perdix"]},
    "Moxonidine": {"class": "Antihypertensives", "cat": "Centrally-Acting", "bnf": "2.5.2", "route": "Oral", "rel": "Primary",
        "terms": ["moxonidine", "physiotens"]},
    "Nadolol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral", "rel": "Primary", "terms": ["nadolol", "corgard"]},
    "Naftidrofuryl": {"class": "Peripheral Vascular Disease", "cat": "Vasodilator", "bnf": "2.6.4", "route": "Oral", "rel": "Primary",
        "terms": ["naftidrofuryl", "praxilene"]},
    "Naltrexone/Bupropion": {"class": "Obesity (CV-Relevant)", "cat": "Combination", "bnf": "4.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["naltrexone/bupropion", "mysimba", "contrave"]},
    "Nateglinide": {"class": "Glucose-Lowering", "cat": "Meglitinide", "bnf": "6.1.2.1", "route": "Oral", "rel": "Primary",
        "terms": ["nateglinide", "starlix"]},
    "Nebivolol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral", "rel": "Primary", "terms": ["nebivolol", "nebilet"]},
    "Nesiritide": {"class": "Heart Failure", "cat": "BNP Analogue", "bnf": "2.5", "route": "IV", "rel": "Secondary", "terms": ["nesiritide", "natrecor"]},
    "Niacin ER": {"class": "Lipid-Lowering", "cat": "Nicotinic Acid", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["niacin", "nicotinic acid", "niaspan"]},
    "Nicardipine": {"class": "Antihypertensives", "cat": "CCB (Dihydropyridine)", "bnf": "2.6.2", "route": "Oral/IV", "rel": "Primary",
        "terms": ["nicardipine", "cardene"]},
    "Nicorandil": {"class": "Nitrates & Antianginals", "cat": "K+ Channel Activator", "bnf": "2.6.3", "route": "Oral", "rel": "Primary",
        "terms": ["nicorandil", "ikorel"]},
    "Nicotinic Acid/Laropiprant": {"class": "Lipid-Lowering", "cat": "Nicotinic Acid (Withdrawn)", "bnf": "2.12", "route": "Oral", "rel": "Historical",
        "terms": ["nicotinic acid/laropiprant", "tredaptive", "laropiprant"]},
    "Nifedipine": {"class": "Antihypertensives", "cat": "CCB (Dihydropyridine)", "bnf": "2.6.2", "route": "Oral", "rel": "Primary",
        "terms": ["nifedipine", "adalat", "adipine", "coracten"]},
    "Nimodipine": {"class": "Antihypertensives", "cat": "CCB (Dihydropyridine)", "bnf": "2.6.2", "route": "Oral/IV", "rel": "Secondary",
        "terms": ["nimodipine", "nimotop"]},
    "Nisoldipine": {"class": "Antihypertensives", "cat": "CCB (Dihydropyridine)", "bnf": "2.6.2", "route": "Oral", "rel": "Primary",
        "terms": ["nisoldipine", "syscor"]},
    "Norepinephrine": {"class": "Critical Care & Vasoactive", "cat": "Catecholamine", "bnf": "2.7.2", "route": "IV", "rel": "Secondary",
        "terms": ["norepinephrine", "noradrenaline", "levophed"]},
    "Olezarsen": {"class": "Lipid-Lowering", "cat": "Lp(a) Targeted", "bnf": "2.12", "route": "SC", "rel": "Secondary", "terms": ["olezarsen"]},
    "Olmesartan": {"class": "Antihypertensives", "cat": "ARB", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary", "terms": ["olmesartan", "olmetec"]},
    "Olmesartan/HCTZ": {"class": "Antihypertensives", "cat": "ARB/Diuretic Combination", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary",
        "terms": ["olmesartan/hydrochlorothiazide", "olmetec plus", "olmesartan hctz"]},
    "Omecamtiv Mecarbil": {"class": "Heart Failure", "cat": "Cardiac Myosin Activator", "bnf": "2.1", "route": "Oral", "rel": "Secondary",
        "terms": ["omecamtiv mecarbil"]},
    "Omega-3 Acid Ethyl Esters": {"class": "Lipid-Lowering", "cat": "Omega-3 Fatty Acid", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["omega-3-acid ethyl esters", "omacor", "teromeg", "omega 3 acid"]},
    "Omega-3 Marine Triglycerides": {"class": "Lipid-Lowering", "cat": "Omega-3 Fatty Acid", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["omega-3 marine triglycerides", "maxepa", "marine triglycerides"]},
    "Orlistat": {"class": "Obesity (CV-Relevant)", "cat": "Lipase Inhibitor", "bnf": "4.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["orlistat", "xenical", "alli"]},
    "Oxprenolol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral", "rel": "Primary",
        "terms": ["oxprenolol", "trasicor", "slow-trasicor"]},
    "Patiromer": {"class": "Potassium Management", "cat": "K+ Binder", "bnf": "9.2.1.1", "route": "Oral", "rel": "Primary", "terms": ["patiromer", "veltassa"]},
    "Patisiran": {"class": "Cardiac Amyloidosis", "cat": "TTR Gene Silencer", "bnf": "—", "route": "IV", "rel": "Secondary",
        "terms": ["patisiran", "onpattro"]},
    "Pelacarsen": {"class": "Lipid-Lowering", "cat": "Lp(a) Targeted", "bnf": "2.12", "route": "SC", "rel": "Secondary", "terms": ["pelacarsen"]},
    "Pentoxifylline": {"class": "Peripheral Vascular Disease", "cat": "Xanthine", "bnf": "2.6.4", "route": "Oral", "rel": "Primary",
        "terms": ["pentoxifylline", "oxpentifylline", "trental"]},
    "Perhexiline": {"class": "Nitrates & Antianginals", "cat": "Metabolic Modulator", "bnf": "2.6.3", "route": "Oral", "rel": "Secondary",
        "terms": ["perhexiline", "pexid"]},
    "Perindopril": {"class": "Antihypertensives", "cat": "ACE Inhibitor", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["perindopril", "coversyl"]},
    "Perindopril/Amlodipine": {"class": "Antihypertensives", "cat": "ACEi/CCB Combination", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["perindopril/amlodipine", "coveram"]},
    "Perindopril/Indapamide": {"class": "Antihypertensives", "cat": "ACEi/Diuretic Combination", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["perindopril/indapamide", "coversyl plus"]},
    "Perindopril/Indapamide/Amlodipine": {"class": "Antihypertensives", "cat": "Triple Combination", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["perindopril/indapamide/amlodipine", "triplixam"]},
    "Phenindione": {"class": "Anticoagulants", "cat": "Indanedione", "bnf": "2.8.2", "route": "Oral", "rel": "Secondary", "terms": ["phenindione"]},
    "Phenoxybenzamine": {"class": "Antihypertensives", "cat": "Alpha-Blocker", "bnf": "2.5.4", "route": "Oral", "rel": "Secondary",
        "terms": ["phenoxybenzamine", "dibenzyline"]},
    "Phentermine/Topiramate": {"class": "Obesity (CV-Relevant)", "cat": "Sympathomimetic/Anticonvulsant", "bnf": "4.5.1", "route": "Oral", "rel": "Secondary",
        "terms": ["phentermine/topiramate", "qsymia"]},
    "Phentolamine": {"class": "Antihypertensives", "cat": "Alpha-Blocker", "bnf": "2.5.4", "route": "IV", "rel": "Secondary",
        "terms": ["phentolamine", "rogitine"]},
    "Phenylephrine": {"class": "Critical Care & Vasoactive", "cat": "Alpha Agonist", "bnf": "2.7.2", "route": "IV", "rel": "Secondary",
        "terms": ["phenylephrine injection", "phenylephrine iv"]},
    "Phenytoin": {"class": "Antiarrhythmics", "cat": "Other", "bnf": "2.3.2/4.8", "route": "IV/Oral", "rel": "Secondary", "terms": ["phenytoin", "epanutin"]},
    "Phytomenadione (Vitamin K1)": {"class": "Anticoagulants", "cat": "Reversal Agent", "bnf": "2.8.2/9.6.6", "route": "Oral/IV", "rel": "Secondary",
        "terms": ["phytomenadione", "vitamin k1", "konakion", "menadiol"]},
    "Pindolol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral", "rel": "Primary", "terms": ["pindolol", "visken"]},
    "Pioglitazone": {"class": "Glucose-Lowering", "cat": "TZD", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary", "terms": ["pioglitazone", "actos"]},
    "Pitavastatin": {"class": "Lipid-Lowering", "cat": "Statin", "bnf": "2.12", "route": "Oral", "rel": "Primary", "terms": ["pitavastatin", "livazo"]},
    "Polidocanol": {"class": "Local Sclerosants", "cat": "Sclerosant", "bnf": "2.13", "route": "IV", "rel": "Secondary",
        "terms": ["polidocanol", "aethoxysklerol", "lauromacrogol"]},
    "Potassium Chloride": {"class": "Potassium Management", "cat": "K+ Supplement", "bnf": "9.2.1.1", "route": "Oral", "rel": "Primary",
        "terms": ["potassium chloride", "sando-k", "kay-cee-l", "slow-k", "potassium effervescent"]},
    "Pramlintide": {"class": "Glucose-Lowering", "cat": "Amylin Analogue", "bnf": "6.1.2.3", "route": "SC", "rel": "Secondary",
        "terms": ["pramlintide", "symlin"]},
    "Prasugrel": {"class": "Antiplatelets", "cat": "P2Y12 Inhibitor", "bnf": "2.9", "route": "Oral", "rel": "Primary", "terms": ["prasugrel", "efient"]},
    "Pravastatin": {"class": "Lipid-Lowering", "cat": "Statin", "bnf": "2.12", "route": "Oral", "rel": "Primary", "terms": ["pravastatin", "lipostat"]},
    "Prazosin": {"class": "Antihypertensives", "cat": "Alpha-Blocker", "bnf": "2.5.4", "route": "Oral", "rel": "Primary", "terms": ["prazosin", "hypovase"]},
    "Procainamide": {"class": "Antiarrhythmics", "cat": "Class Ia", "bnf": "2.3.1", "route": "IV", "rel": "Secondary", "terms": ["procainamide", "pronestyl"]},
    "Propafenone": {"class": "Antiarrhythmics", "cat": "Class Ic", "bnf": "2.3.1", "route": "Oral", "rel": "Primary", "terms": ["propafenone", "arythmol"]},
    "Propranolol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral", "rel": "Primary",
        "terms": ["propranolol", "inderal", "half inderal"]},
    "Protamine Sulfate": {"class": "Anticoagulants", "cat": "Reversal Agent", "bnf": "2.8.1", "route": "IV", "rel": "Secondary",
        "terms": ["protamine sulfate", "protamine sulphate"]},
    "Prothrombin Complex Concentrate": {"class": "Anticoagulants", "cat": "Reversal Agent", "bnf": "2.8.2", "route": "IV", "rel": "Secondary",
        "terms": ["prothrombin complex", "beriplex", "octaplex", "pcc", "4-factor pcc"]},
    "Quinapril": {"class": "Antihypertensives", "cat": "ACE Inhibitor", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary", "terms": ["quinapril", "accupro"]},
    "Quinidine": {"class": "Antiarrhythmics", "cat": "Class Ia", "bnf": "2.3.1", "route": "Oral", "rel": "Secondary", "terms": ["quinidine"]},
    "Ramipril": {"class": "Antihypertensives", "cat": "ACE Inhibitor", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary", "terms": ["ramipril", "tritace"]},
    "Ramipril/Felodipine": {"class": "Antihypertensives", "cat": "ACEi/CCB Combination", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["ramipril/felodipine", "triapin"]},
    "Ranolazine": {"class": "Nitrates & Antianginals", "cat": "Late Na+ Channel Blocker", "bnf": "2.6.3", "route": "Oral", "rel": "Primary",
        "terms": ["ranolazine", "ranexa"]},
    "Repaglinide": {"class": "Glucose-Lowering", "cat": "Meglitinide", "bnf": "6.1.2.1", "route": "Oral", "rel": "Primary",
        "terms": ["repaglinide", "prandin", "novonorm"]},
    "Reteplase": {"class": "Thrombolytics", "cat": "tPA", "bnf": "2.10.2", "route": "IV", "rel": "Secondary", "terms": ["reteplase", "rapilysin"]},
    "Rilonacept": {"class": "Pericarditis & Inflammatory", "cat": "IL-1 Trap", "bnf": "10.1.3", "route": "SC", "rel": "Secondary",
        "terms": ["rilonacept", "arcalyst"]},
    "Riociguat": {"class": "Pulmonary Hypertension", "cat": "sGC Stimulator", "bnf": "2.5.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["riociguat", "adempas"]},
    "Rivaroxaban": {"class": "Anticoagulants", "cat": "DOAC", "bnf": "2.8.2", "route": "Oral", "rel": "Primary", "terms": ["rivaroxaban", "xarelto"]},
    "Rosiglitazone": {"class": "Glucose-Lowering", "cat": "TZD (Withdrawn)", "bnf": "6.1.2.3", "route": "Oral", "rel": "Historical",
        "terms": ["rosiglitazone", "avandia"]},
    "Rosuvastatin": {"class": "Lipid-Lowering", "cat": "Statin", "bnf": "2.12", "route": "Oral", "rel": "Primary", "terms": ["rosuvastatin", "crestor"]},
    "Rosuvastatin/Ezetimibe": {"class": "Lipid-Lowering", "cat": "Statin+Ezetimibe", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["rosuvastatin/ezetimibe", "rosuvastatin ezetimibe"]},
    "Sacubitril/Valsartan": {"class": "Heart Failure", "cat": "ARNI", "bnf": "2.5.5", "route": "Oral", "rel": "Primary",
        "terms": ["sacubitril/valsartan", "entresto", "sacubitril valsartan"]},
    "Saxagliptin": {"class": "Glucose-Lowering", "cat": "DPP-4 Inhibitor", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["saxagliptin", "onglyza"]},
    "Selexipag": {"class": "Pulmonary Hypertension", "cat": "IP Receptor Agonist", "bnf": "2.5.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["selexipag", "uptravi"]},
    "Semaglutide (Oral)": {"class": "Glucose-Lowering", "cat": "GLP-1 RA", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["semaglutide oral", "rybelsus"]},
    "Semaglutide (SC)": {"class": "Glucose-Lowering", "cat": "GLP-1 RA", "bnf": "6.1.2.3", "route": "SC", "rel": "Primary",
        "terms": ["semaglutide injection", "ozempic"]},
    "Semaglutide 2.4mg": {"class": "Obesity (CV-Relevant)", "cat": "GLP-1 RA (Obesity)", "bnf": "4.5.1", "route": "SC", "rel": "Primary",
        "terms": ["semaglutide 2.4mg", "wegovy"]},
    "Sildenafil (PAH)": {"class": "Pulmonary Hypertension", "cat": "PDE5 Inhibitor", "bnf": "2.5.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["sildenafil pah", "revatio", "sildenafil pulmonary"]},
    "Simvastatin": {"class": "Lipid-Lowering", "cat": "Statin", "bnf": "2.12", "route": "Oral", "rel": "Primary", "terms": ["simvastatin", "zocor"]},
    "Simvastatin/Ezetimibe": {"class": "Lipid-Lowering", "cat": "Statin+Ezetimibe", "bnf": "2.12", "route": "Oral", "rel": "Primary",
        "terms": ["simvastatin/ezetimibe", "inegy", "simvastatin ezetimibe"]},
    "Sitagliptin": {"class": "Glucose-Lowering", "cat": "DPP-4 Inhibitor", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["sitagliptin", "januvia"]},
    "Sodium Nitroprusside": {"class": "Antihypertensives", "cat": "Direct Vasodilator", "bnf": "2.5.1", "route": "IV", "rel": "Secondary",
        "terms": ["sodium nitroprusside", "nitroprusside", "nipride"]},
    "Sodium Tetradecyl Sulfate": {"class": "Local Sclerosants", "cat": "Sclerosant", "bnf": "2.13", "route": "IV", "rel": "Secondary",
        "terms": ["sodium tetradecyl", "fibrovein", "std pharmaceutical"]},
    "Sodium Zirconium Cyclosilicate": {"class": "Potassium Management", "cat": "K+ Binder", "bnf": "9.2.1.1", "route": "Oral", "rel": "Primary",
        "terms": ["sodium zirconium cyclosilicate", "lokelma"]},
    "Soluble Insulin (Regular)": {"class": "Glucose-Lowering", "cat": "Short-Acting Insulin", "bnf": "6.1.1.1", "route": "SC/IV", "rel": "Primary",
        "terms": ["soluble insulin", "actrapid", "humulin s", "insuman rapid", "human actrapid"]},
    "Sotagliflozin": {"class": "Glucose-Lowering", "cat": "SGLT1/2 Inhibitor", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["sotagliflozin", "inpefa"]},
    "Sotalol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.3.2", "route": "Oral", "rel": "Primary",
        "terms": ["sotalol", "beta-cardone", "sotacor"]},
    "Spironolactone": {"class": "Antihypertensives", "cat": "K+-Sparing Diuretic", "bnf": "2.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["spironolactone", "aldactone"]},
    "Streptokinase": {"class": "Thrombolytics", "cat": "Streptokinase", "bnf": "2.10.2", "route": "IV", "rel": "Secondary",
        "terms": ["streptokinase", "streptase"]},
    "Survodutide": {"class": "Glucose-Lowering", "cat": "Triple Agonist", "bnf": "6.1.2.3", "route": "SC", "rel": "Secondary", "terms": ["survodutide"]},
    "Tadalafil (PAH)": {"class": "Pulmonary Hypertension", "cat": "PDE5 Inhibitor", "bnf": "2.5.1.2", "route": "Oral", "rel": "Primary",
        "terms": ["tadalafil pah", "adcirca", "tadalafil pulmonary"]},
    "Tafamidis": {"class": "Cardiac Amyloidosis", "cat": "TTR Stabiliser", "bnf": "—", "route": "Oral", "rel": "Secondary",
        "terms": ["tafamidis", "vyndaqel", "vyndamax"]},
    "Technosphere Insulin": {"class": "Glucose-Lowering", "cat": "Inhaled Insulin", "bnf": "6.1.1.1", "route": "Inhaled", "rel": "Secondary",
        "terms": ["technosphere insulin", "afrezza"]},
    "Telmisartan": {"class": "Antihypertensives", "cat": "ARB", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary", "terms": ["telmisartan", "micardis"]},
    "Telmisartan/HCTZ": {"class": "Antihypertensives", "cat": "ARB/Diuretic Combination", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary",
        "terms": ["telmisartan/hydrochlorothiazide", "micardis plus", "telmisartan hctz"]},
    "Tenecteplase": {"class": "Thrombolytics", "cat": "tPA", "bnf": "2.10.2", "route": "IV", "rel": "Secondary", "terms": ["tenecteplase", "metalyse"]},
    "Terazosin": {"class": "Antihypertensives", "cat": "Alpha-Blocker", "bnf": "2.5.4", "route": "Oral", "rel": "Primary", "terms": ["terazosin", "hytrin"]},
    "Ticagrelor": {"class": "Antiplatelets", "cat": "P2Y12 Inhibitor", "bnf": "2.9", "route": "Oral", "rel": "Primary", "terms": ["ticagrelor", "brilique"]},
    "Ticlopidine": {"class": "Antiplatelets", "cat": "P2Y12 Inhibitor", "bnf": "2.9", "route": "Oral", "rel": "Secondary", "terms": ["ticlopidine", "ticlid"]},
    "Timolol": {"class": "Antihypertensives", "cat": "Beta-Blocker", "bnf": "2.4", "route": "Oral", "rel": "Primary",
        "terms": ["timolol oral", "timolol tablet"]},
    "Tinzaparin": {"class": "Anticoagulants", "cat": "LMWH", "bnf": "2.8.1", "route": "SC", "rel": "Primary", "terms": ["tinzaparin", "innohep"]},
    "Tirofiban": {"class": "Antiplatelets", "cat": "GP IIb/IIIa Inhibitor", "bnf": "2.9", "route": "IV", "rel": "Secondary",
        "terms": ["tirofiban", "aggrastat"]},
    "Tirzepatide": {"class": "Glucose-Lowering", "cat": "Dual GIP/GLP-1 RA", "bnf": "6.1.2.3", "route": "SC", "rel": "Primary",
        "terms": ["tirzepatide", "mounjaro"]},
    "Tirzepatide (Obesity)": {"class": "Obesity (CV-Relevant)", "cat": "GIP/GLP-1 RA (Obesity)", "bnf": "4.5.1", "route": "SC", "rel": "Primary",
        "terms": ["tirzepatide obesity", "zepbound"]},
    "Tolbutamide": {"class": "Glucose-Lowering", "cat": "Sulfonylurea", "bnf": "6.1.2.1", "route": "Oral", "rel": "Primary", "terms": ["tolbutamide"]},
    "Torsemide/Torasemide": {"class": "Antihypertensives", "cat": "Loop Diuretic", "bnf": "2.2.2", "route": "Oral", "rel": "Primary",
        "terms": ["torasemide", "torsemide", "torem"]},
    "Trandolapril": {"class": "Antihypertensives", "cat": "ACE Inhibitor", "bnf": "2.5.5.1", "route": "Oral", "rel": "Primary",
        "terms": ["trandolapril", "gopten", "odrik"]},
    "Tranexamic Acid": {"class": "Antifibrinolytics", "cat": "Antifibrinolytic", "bnf": "2.11", "route": "Oral/IV", "rel": "Secondary",
        "terms": ["tranexamic acid", "cyklokapron", "tranexamic"]},
    "Treprostinil": {"class": "Pulmonary Hypertension", "cat": "Prostacyclin Analogue", "bnf": "2.5.1.2", "route": "SC/IV/Inhaled", "rel": "Secondary",
        "terms": ["treprostinil", "remodulin", "tyvaso"]},
    "Triamterene": {"class": "Antihypertensives", "cat": "K+-Sparing Diuretic", "bnf": "2.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["triamterene", "dytac"]},
    "Trimetazidine": {"class": "Nitrates & Antianginals", "cat": "Metabolic Modulator", "bnf": "2.6.3", "route": "Oral", "rel": "Secondary",
        "terms": ["trimetazidine", "vastarel"]},
    "Unfractionated Heparin": {"class": "Anticoagulants", "cat": "UFH", "bnf": "2.8.1", "route": "IV/SC", "rel": "Primary",
        "terms": ["heparin sodium", "heparin calcium", "unfractionated heparin", "hep-flush"]},
    "Urokinase": {"class": "Thrombolytics", "cat": "tPA", "bnf": "2.10.2", "route": "IV", "rel": "Secondary", "terms": ["urokinase", "syner-kinase"]},
    "Valsartan": {"class": "Antihypertensives", "cat": "ARB", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary", "terms": ["valsartan", "diovan"]},
    "Valsartan/Amlodipine": {"class": "Antihypertensives", "cat": "ARB/CCB Combination", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary",
        "terms": ["valsartan/amlodipine", "exforge"]},
    "Valsartan/Amlodipine/HCTZ": {"class": "Antihypertensives", "cat": "Triple Combination", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary",
        "terms": ["valsartan/amlodipine/hydrochlorothiazide", "exforge hct"]},
    "Valsartan/HCTZ": {"class": "Antihypertensives", "cat": "ARB/Diuretic Combination", "bnf": "2.5.5.2", "route": "Oral", "rel": "Primary",
        "terms": ["valsartan/hydrochlorothiazide", "co-diovan", "valsartan hctz"]},
    "Vasopressin": {"class": "Critical Care & Vasoactive", "cat": "Vasopressin Analogue", "bnf": "2.7.2", "route": "IV", "rel": "Secondary",
        "terms": ["vasopressin", "argipressin", "pitressin"]},
    "Verapamil": {"class": "Antihypertensives", "cat": "CCB (Non-dihydropyridine)", "bnf": "2.6.2", "route": "Oral", "rel": "Primary",
        "terms": ["verapamil", "securon", "cordilox", "univer", "verapress", "half securon"]},
    "Vericiguat": {"class": "Heart Failure", "cat": "sGC Stimulator", "bnf": "2.5.5", "route": "Oral", "rel": "Primary", "terms": ["vericiguat", "verquvo"]},
    "Vernakalant": {"class": "Antiarrhythmics", "cat": "Class III", "bnf": "2.3.2", "route": "IV", "rel": "Secondary", "terms": ["vernakalant", "brinavess"]},
    "Vildagliptin": {"class": "Glucose-Lowering", "cat": "DPP-4 Inhibitor", "bnf": "6.1.2.3", "route": "Oral", "rel": "Primary",
        "terms": ["vildagliptin", "galvus"]},
    "Vorapaxar": {"class": "Antiplatelets", "cat": "PAR-1 Antagonist", "bnf": "2.9", "route": "Oral", "rel": "Secondary", "terms": ["vorapaxar", "zontivity"]},
    "Vutrisiran": {"class": "Cardiac Amyloidosis", "cat": "TTR Gene Silencer", "bnf": "—", "route": "SC", "rel": "Secondary",
        "terms": ["vutrisiran", "amvuttra"]},
    "Warfarin": {"class": "Anticoagulants", "cat": "Coumarin", "bnf": "2.8.2", "route": "Oral", "rel": "Primary", "terms": ["warfarin", "marevan"]},
    "Xipamide": {"class": "Antihypertensives", "cat": "Thiazide Diuretic", "bnf": "2.2.1", "route": "Oral", "rel": "Primary",
        "terms": ["xipamide", "diurexan"]},
}


# ── ICD-10 Code → Description mapping (for Linkage extraction enrichment) ──
ICD10_DESCRIPTIONS = {
    # Valvular Heart Disease (I05-I09, I34-I39)
    "I05": "Rheumatic mitral valve diseases",
    "I05.0": "Rheumatic mitral stenosis",
    "I05.1": "Rheumatic mitral insufficiency",
    "I05.2": "Rheumatic mitral stenosis with insufficiency",
    "I05.8": "Other rheumatic mitral valve diseases",
    "I05.9": "Rheumatic mitral valve disease, unspecified",
    "I06": "Rheumatic aortic valve diseases",
    "I06.0": "Rheumatic aortic stenosis",
    "I06.1": "Rheumatic aortic insufficiency",
    "I06.2": "Rheumatic aortic stenosis with insufficiency",
    "I06.8": "Other rheumatic aortic valve diseases",
    "I06.9": "Rheumatic aortic valve disease, unspecified",
    "I07": "Rheumatic tricuspid valve diseases",
    "I07.0": "Rheumatic tricuspid stenosis",
    "I07.1": "Rheumatic tricuspid insufficiency",
    "I07.2": "Rheumatic tricuspid stenosis and insufficiency",
    "I07.8": "Other rheumatic tricuspid valve diseases",
    "I07.9": "Rheumatic tricuspid valve disease, unspecified",
    "I08": "Multiple valve diseases",
    "I08.0": "Disorders of both mitral and aortic valves",
    "I08.1": "Disorders of both mitral and tricuspid valves",
    "I08.2": "Disorders of both aortic and tricuspid valves",
    "I08.3": "Combined disorders of mitral, aortic and tricuspid valves",
    "I08.8": "Other multiple valve diseases",
    "I08.9": "Multiple valve disease, unspecified",
    "I09.1": "Rheumatic diseases of endocardium, valve unspecified",
    "I34": "Nonrheumatic mitral valve disorders",
    "I34.0": "Nonrheumatic mitral (valve) insufficiency",
    "I34.1": "Nonrheumatic mitral (valve) prolapse",
    "I34.2": "Nonrheumatic mitral (valve) stenosis",
    "I34.8": "Other nonrheumatic mitral valve disorders",
    "I34.9": "Nonrheumatic mitral valve disorder, unspecified",
    "I35": "Nonrheumatic aortic valve disorders",
    "I35.0": "Nonrheumatic aortic (valve) stenosis",
    "I35.1": "Nonrheumatic aortic (valve) insufficiency",
    "I35.2": "Nonrheumatic aortic (valve) stenosis with insufficiency",
    "I35.8": "Other nonrheumatic aortic valve disorders",
    "I35.9": "Nonrheumatic aortic valve disorder, unspecified",
    "I36": "Nonrheumatic tricuspid valve disorders",
    "I36.0": "Nonrheumatic tricuspid (valve) stenosis",
    "I36.1": "Nonrheumatic tricuspid (valve) insufficiency",
    "I36.2": "Nonrheumatic tricuspid (valve) stenosis with insufficiency",
    "I36.8": "Other nonrheumatic tricuspid valve disorders",
    "I36.9": "Nonrheumatic tricuspid valve disorder, unspecified",
    "I37": "Nonrheumatic pulmonary valve disorders",
    "I37.0": "Nonrheumatic pulmonary valve stenosis",
    "I37.1": "Nonrheumatic pulmonary valve insufficiency",
    "I37.2": "Nonrheumatic pulmonary valve stenosis with insufficiency",
    "I37.8": "Other nonrheumatic pulmonary valve disorders",
    "I37.9": "Nonrheumatic pulmonary valve disorder, unspecified",
    "I38": "Endocarditis, valve unspecified",
    "I39": "Endocarditis and heart valve disorders in diseases classified elsewhere",
    # Common comorbidity codes
    "I10": "Essential (primary) hypertension",
    "I11": "Hypertensive heart disease",
    "I12": "Hypertensive chronic kidney disease",
    "I13": "Hypertensive heart and chronic kidney disease",
    "I15": "Secondary hypertension",
    "E10": "Type 1 diabetes mellitus",
    "E11": "Type 2 diabetes mellitus",
    "I48": "Atrial fibrillation and flutter",
    "I50": "Heart failure",
    "I20": "Angina pectoris",
    "I21": "Acute myocardial infarction",
    "I22": "Subsequent myocardial infarction",
    "I23": "Certain current complications following acute MI",
    "I24": "Other acute ischaemic heart diseases",
    "I25": "Chronic ischaemic heart disease",
    "I60": "Nontraumatic subarachnoid haemorrhage",
    "I61": "Nontraumatic intracerebral haemorrhage",
    "I62": "Other and unspecified nontraumatic intracranial haemorrhage",
    "I63": "Cerebral infarction",
    "I64": "Stroke, not specified as haemorrhage or infarction",
    "J44": "Other chronic obstructive pulmonary disease",
    "N18": "Chronic kidney disease (CKD)",
}

# ── CPRD Term Definitions (for Definitions tab) ──
CPRD_DEFINITIONS = [
    {"Term": "Observation", "Definition": "Clinical findings, diagnoses, symptoms, test results, and measurements recorded in GP consultations.",
     "Use Case": "Medical history & test results. Primary source for identifying disease diagnoses in primary care."},
    {"Term": "DrugIssue", "Definition": "GP-issued prescriptions including drug name, quantity, dose, and estimated NHS cost.",
     "Use Case": "Medication adherence studies. Tracking prescribing patterns and treatment sequences."},
    {"Term": "Consultation", "Definition": "Interaction metadata including date, type of consultation, and the consulting staff member.",
     "Use Case": "Healthcare utilisation tracking. Counting GP visits and understanding care patterns."},
    {"Term": "Patient", "Definition": "Demographics including year/month of birth, gender, registration dates, and acceptable patient flag.",
     "Use Case": "Cohort defining & age-matching. Core table for patient selection and eligibility checks."},
    {"Term": "Practice", "Definition": "Practice-level information including last collection date (lcd) and region.",
     "Use Case": "Defining data completeness windows. Ensuring observations fall within the practice's active period."},
    {"Term": "Problem", "Definition": "Long-term conditions and problem status including expected duration and last review date.",
     "Use Case": "Identifying chronic conditions and active vs resolved diagnoses."},
    {"Term": "Referral", "Definition": "GP referrals to specialist services including source, target, urgency, and direction.",
     "Use Case": "Tracking specialist referral pathways and waiting times."},
    {"Term": "Staff", "Definition": "Staff identifiers and job categories for anonymised tracking of care providers.",
     "Use Case": "Attributing care to provider types (GP, nurse, etc.)."},
    {"Term": "Linkage Eligibility", "Definition": "Flag indicating whether a patient's records can be linked to external NHS datasets (HES, ONS Death, IMD). Requires valid NHS number and consent.",
     "Use Case": "Mandatory check before using any linked data. Patients without linkage eligibility will not appear in HES or Death datasets."},
    {"Term": "HES APC", "Definition": "Hospital Episode Statistics — Admitted Patient Care. Records all NHS hospital admissions including inpatient stays, day cases, and surgical procedures. Uses ICD-10 diagnosis codes and OPCS-4 procedure codes.",
     "Use Case": "Identifying hospital diagnoses, surgical interventions, and inpatient episodes."},
    {"Term": "HES OP", "Definition": "Hospital Episode Statistics — Outpatient. Records all NHS outpatient specialist appointments. Uses ICD-10 codes without decimal points.",
     "Use Case": "Tracking specialist outpatient visits, diagnoses made in secondary care."},
    {"Term": "HES A&E", "Definition": "Hospital Episode Statistics — Accident & Emergency. Records urgent and emergency care visit details including attendance, diagnosis, investigations, and treatments.",
     "Use Case": "Studying emergency presentations, A&E utilisation, and acute care pathways."},
    {"Term": "ONS Death", "Definition": "Office for National Statistics mortality records. Contains official date of death plus underlying and up to 15 contributory cause-of-death ICD-10 codes.",
     "Use Case": "Mortality studies, survival analysis, and cause-of-death ascertainment."},
    {"Term": "IMD", "Definition": "Index of Multiple Deprivation (2015). Neighbourhood-level socio-economic deprivation rankings mapped to patient or practice. Quintile 1 = most deprived, Quintile 5 = least deprived.",
     "Use Case": "Adjusting for socio-economic confounding. Studying health inequalities."},
    {"Term": "SNOMED CT", "Definition": "Systematized Nomenclature of Medicine — Clinical Terms. International clinical terminology standard used in UK primary care (GP records). Codes are long numeric identifiers (e.g. 60573004).",
     "Use Case": "All primary care diagnoses and findings in CPRD Aurum are coded with SNOMED CT via MedCodeId mapping."},
    {"Term": "ICD-10", "Definition": "International Classification of Diseases, 10th Revision. WHO standard used in UK hospitals (HES) and death certificates (ONS). Format: letter + 2 digits + optional decimal (e.g. I05.0).",
     "Use Case": "All hospital and death record diagnoses use ICD-10 coding."},
    {"Term": "MedCodeId", "Definition": "CPRD Aurum internal numeric identifier linking to SNOMED CT concepts via the EMIS Medical Dictionary lookup table.",
     "Use Case": "Bridge between CPRD's internal coding and the clinical SNOMED CT terminology."},
    {"Term": "EMIS Dictionary", "Definition": "Lookup table mapping MedCodeId to SNOMED CT codes, clinical terms, and Read codes. Provided by CPRD with each data release.",
     "Use Case": "Required for translating MedCodeIds into clinically meaningful SNOMED CT concepts."},
]

# ── Linkage page abbreviation footnotes ──
LINKAGE_ABBREVIATIONS = """
**Abbreviations:**
**HES** = Hospital Episode Statistics · **APC** = Admitted Patient Care · **OP** = Outpatient ·
**A&E** = Accident & Emergency · **ONS** = Office for National Statistics ·
**IMD** = Index of Multiple Deprivation · **ICD-10** = International Classification of Diseases, 10th Revision ·
**OPCS-4** = Office of Population Censuses and Surveys Classification of Interventions and Procedures, Version 4 ·
**NHS** = National Health Service · **CPRD** = Clinical Practice Research Datalink ·
**SNOMED CT** = Systematized Nomenclature of Medicine — Clinical Terms ·
**LSOA** = Lower Layer Super Output Area · **GP** = General Practitioner
"""

# ── SSH Tunnel Configuration ──
SSH_CONFIG = {
    "cluster_hosts": [
        "cluster1.bmrc.ox.ac.uk",
        "cluster2.bmrc.ox.ac.uk",
        "cluster3.bmrc.ox.ac.uk",
        "cluster4.bmrc.ox.ac.uk",
    ],
    "login_nodes": [
        "rescomp1.well.ox.ac.uk",
        "rescomp2.well.ox.ac.uk",
    ],
    "cluster_host": "cluster1.bmrc.ox.ac.uk",  # fallback default
    "default_port": 8501,
    "heartbeat_interval": 30,   # seconds between keepalive pings
    "reconnect_max_retries": 5,
    "reconnect_backoff_base": 2,  # exponential backoff base seconds
}


# ══════════════════════════════════════════════════════════════════════════════
# MOCK DATA GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

def generate_mock_data():
    """Generate comprehensive mock data for testing all modules."""
    np.random.seed(42)
    n_patients = 500
    patids = [str(100000 + i) for i in range(n_patients)]
    genders = np.random.choice([1, 2], size=n_patients)
    yobs = np.random.randint(1930, 2000, size=n_patients)

    patient_df = pd.DataFrame({
        "patid": patids,
        "pracid": np.random.randint(1, 50, size=n_patients),
        "gender": genders,
        "yob": yobs,
        "mob": np.random.randint(1, 13, size=n_patients),
        "regstartdate": pd.to_datetime(
            [f"{np.random.randint(1998,2015)}-{np.random.randint(1,13):02d}-{np.random.randint(1,28):02d}"
             for _ in range(n_patients)]),
        "regenddate": pd.to_datetime(
            [f"{np.random.randint(2016,2021)}-{np.random.randint(1,13):02d}-{np.random.randint(1,28):02d}"
             for _ in range(n_patients)]),
        "acceptable": 1,
        "cprd_ddate": [None] * n_patients,
    })

    n_obs = 2000
    all_snomed = []
    for entry in DISEASE_CODE_LIBRARY.values():
        all_snomed.extend(entry.get("snomed", []))
    obs_patids = np.random.choice(patids[:300], size=n_obs)
    obs_df = pd.DataFrame({
        "patid": obs_patids,
        "obsid": range(1, n_obs + 1),
        "pracid": np.random.randint(1, 50, size=n_obs),
        "obsdate": pd.to_datetime(
            [f"{np.random.randint(2000,2020)}-{np.random.randint(1,13):02d}-{np.random.randint(1,28):02d}"
             for _ in range(n_obs)]),
        "medcodeid": [str(np.random.randint(100000, 999999)) for _ in range(n_obs)],
        "value": np.random.uniform(0, 200, size=n_obs).round(1),
        "numunitid": np.random.randint(1, 100, size=n_obs),
        "enterdate": pd.to_datetime("2020-01-01"),
        "staffid": np.random.randint(1, 100, size=n_obs),
        "SnomedCTConceptId": np.random.choice(all_snomed, size=n_obs),
        "Term": np.random.choice([
            "Aortic valve stenosis", "Mitral regurgitation",
            "Mitral valve prolapse", "Tricuspid regurgitation",
            "Aortic regurgitation", "Pulmonary stenosis",
            "Calcified aortic valve", "Mitral stenosis"
        ], size=n_obs),
    })

    n_drugs = 1500
    drug_patids = np.random.choice(patids[:300], size=n_drugs)
    drug_df = pd.DataFrame({
        "patid": drug_patids,
        "issueid": range(1, n_drugs + 1),
        "pracid": np.random.randint(1, 50, size=n_drugs),
        "issuedate": pd.to_datetime(
            [f"{np.random.randint(2000,2020)}-{np.random.randint(1,13):02d}-{np.random.randint(1,28):02d}"
             for _ in range(n_drugs)]),
        "prodcodeid": [str(np.random.randint(1000000, 9999999)) for _ in range(n_drugs)],
        "quantity": np.random.randint(1, 100, size=n_drugs),
        "duration": np.random.randint(1, 365, size=n_drugs),
        "drugname": np.random.choice([
            "Amlodipine 5mg", "Ramipril 5mg", "Bisoprolol 2.5mg",
            "Warfarin 3mg", "Aspirin 75mg", "Atorvastatin 20mg",
            "Furosemide 40mg", "Digoxin 125mcg", "Metformin 500mg",
            "Lisinopril 10mg"
        ], size=n_drugs),
    })

    n_hes = 800
    all_icd = []
    for entry in DISEASE_CODE_LIBRARY.values():
        all_icd.extend(entry.get("icd10", []))
    for codes in COMORBIDITY_CODES_ICD10.values():
        all_icd.extend(codes)
    hes_patids = np.random.choice(patids[:250], size=n_hes)
    hes_df = pd.DataFrame({
        "patid": hes_patids,
        "spno": np.random.randint(1, 1000, size=n_hes),
        "epikey": np.random.randint(10000, 99999, size=n_hes),
        "epistart": pd.to_datetime(
            [f"{np.random.randint(2000,2020)}-{np.random.randint(1,13):02d}-{np.random.randint(1,28):02d}"
             for _ in range(n_hes)]),
        "epiend": pd.to_datetime(
            [f"{np.random.randint(2000,2020)}-{np.random.randint(1,13):02d}-{np.random.randint(1,28):02d}"
             for _ in range(n_hes)]),
        "ICD": np.random.choice(all_icd, size=n_hes),
        "ICD_PRIMARY": np.random.choice(all_icd[:10], size=n_hes),
        "d_order": np.random.randint(1, 15, size=n_hes),
    })

    n_op = 400
    op_patids = np.random.choice(patids[:200], size=n_op)
    op_icd_no_dots = [c.replace(".", "") for c in all_icd]
    op_df = pd.DataFrame({
        "patid": op_patids,
        "attendkey": np.random.randint(10000, 99999, size=n_op),
        "appointdt": pd.to_datetime(
            [f"{np.random.randint(2003,2020)}-{np.random.randint(1,13):02d}-{np.random.randint(1,28):02d}"
             for _ in range(n_op)]),
        "diag_01": np.random.choice(op_icd_no_dots, size=n_op),
        "opertn_01": ["X" + str(np.random.randint(10, 99)) for _ in range(n_op)],
    })

    n_dead = 150
    dead_patids = np.random.choice(patids[:150], size=n_dead, replace=False)
    death_dict = {
        "patid": dead_patids,
        "dod": pd.to_datetime(
            [f"{np.random.randint(2005,2020)}-{np.random.randint(1,13):02d}-{np.random.randint(1,28):02d}"
             for _ in range(n_dead)]),
        "cause": np.random.choice(all_icd[:15], size=n_dead),
    }
    for i in range(1, 16):
        col = f"cause{i}"
        death_dict[col] = np.where(
            np.random.random(n_dead) > 0.5,
            np.random.choice(all_icd[:20], size=n_dead),
            ""
        )
    death_df = pd.DataFrame(death_dict)

    mock_medcodes = obs_df["medcodeid"].unique().tolist()
    dict_entries = []
    for mc in mock_medcodes:
        snomed = np.random.choice(all_snomed)
        term = np.random.choice([
            "Aortic valve stenosis", "Mitral regurgitation", "Mitral valve prolapse",
            "Tricuspid regurgitation", "Aortic regurgitation", "Pulmonary stenosis",
            "Calcified aortic valve", "Mitral stenosis", "Bicuspid aortic valve",
        ])
        dict_entries.append({
            "MedCodeId": mc,
            "Term": term,
            "OriginalReadCode": f"G5{np.random.randint(10,99)}.",
            "CleansedReadCode": f"G5{np.random.randint(10,99)}.",
            "SnomedCTConceptId": snomed,
            "SnomedCTDescriptionId": str(np.random.randint(100000000, 999999999)),
            "Release": "v37.0",
            "EmisCodeCategoryId": np.random.choice([1, 2, 3]),
        })
    emis_dict_df = pd.DataFrame(dict_entries)

    # ── Mock Product Dictionary ──
    mock_prodcodes = drug_df["prodcodeid"].unique().tolist()
    prod_entries = []
    drug_names = [
        "Amlodipine 5mg tablets", "Ramipril 5mg capsules", "Bisoprolol 2.5mg tablets",
        "Warfarin 3mg tablets", "Aspirin 75mg tablets", "Atorvastatin 20mg tablets",
        "Furosemide 40mg tablets", "Digoxin 125microgram tablets", "Metformin 500mg tablets",
        "Lisinopril 10mg tablets", "Clopidogrel 75mg tablets", "Lansoprazole 30mg capsules",
    ]
    drug_substances = [
        "Amlodipine", "Ramipril", "Bisoprolol", "Warfarin", "Aspirin",
        "Atorvastatin", "Furosemide", "Digoxin", "Metformin", "Lisinopril",
        "Clopidogrel", "Lansoprazole",
    ]
    for pc in mock_prodcodes:
        idx = np.random.randint(0, len(drug_names))
        prod_entries.append({
            "ProdCodeId": pc,
            "Term": drug_names[idx],
            "ProductName": drug_names[idx],
            "DrugSubstanceName": drug_substances[idx],
            "DMD_ID": str(np.random.randint(100000000, 999999999)),
            "FormulationName": np.random.choice(["Tablet", "Capsule", "Oral solution", "Injection"]),
            "RouteName": np.random.choice(["Oral", "Intravenous", "Subcutaneous"]),
            "BNFChapter": f"{np.random.randint(1,16):02d}.{np.random.randint(1,13):02d}",
        })
    emis_prod_dict_df = pd.DataFrame(prod_entries)

    elig_df = pd.DataFrame({
        "patid": patids[:400],
        "pracid": np.random.randint(1, 50, size=400),
        "linkdate": pd.to_datetime("2021-01-01"),
        "hes_e": np.random.choice([0, 1], size=400, p=[0.1, 0.9]),
        "death_e": np.random.choice([0, 1], size=400, p=[0.1, 0.9]),
        "lsoa_e": np.random.choice([0, 1], size=400, p=[0.2, 0.8]),
    })

    imd_df = pd.DataFrame({
        "patid": patids[:400],
        "pracid": np.random.randint(1, 50, size=400),
        "imd2015_5": np.random.randint(1, 6, size=400),
        "imd2015_10": np.random.randint(1, 11, size=400),
    })

    return {
        "patient": patient_df,
        "observation": obs_df,
        "drug": drug_df,
        "hes_apc": hes_df,
        "hes_op": op_df,
        "death": death_df,
        "emis_dict": emis_dict_df,
        "emis_prod_dict": emis_prod_dict_df,
        "linkage_elig": elig_df,
        "imd": imd_df,
    }


# ══════════════════════════════════════════════════════════════════════════════
# HELPER: Code-type auto-detection
# ══════════════════════════════════════════════════════════════════════════════

def classify_codes(codes):
    """Auto-detect whether codes are SNOMED CT, ICD-10, or MedCodeId.
    Returns dict: {"snomed": [...], "icd10": [...], "unknown": [...]}
    """
    snomed, icd10, unknown = [], [], []
    icd_pattern = re.compile(r'^[A-Z]\d{2}(\.\d{1,2})?$', re.IGNORECASE)
    for code in codes:
        code = code.strip()
        if not code:
            continue
        if icd_pattern.match(code):
            icd10.append(code)
        elif code.isdigit() and len(code) >= 5:
            # Purely numeric codes ≥5 digits → SNOMED CT
            snomed.append(code)
        elif re.match(r'^[A-Z]\d{2,3}$', code, re.IGNORECASE):
            # 3-4 char alphanumeric like I05 or I050 → ICD-10 truncated
            icd10.append(code)
        else:
            unknown.append(code)
    return {"snomed": snomed, "icd10": icd10, "unknown": unknown}


def rename_columns_friendly(df):
    """Rename columns using FRIENDLY_NAMES mapping."""
    return df.rename(columns={k: v for k, v in FRIENDLY_NAMES.items() if k in df.columns})


def add_icd10_descriptions(df):
    """Add ICD10_Description column(s) mapping ICD codes to their descriptions.

    Handles multiple possible column layouts across HES/Death datasets:
    - Single ICD column (HES APC): ICD, ICD_PRIMARY, icd_code → one description column
    - Multiple diag columns (HES OP): diag_01..diag_12 → one description column per diag
    - Cause columns (Death): cause, cause1..cause15 → one description column per cause
    """
    def _lookup(code):
        if pd.isna(code) or not str(code).strip():
            return ""
        code_str = str(code).strip()
        # Direct match
        if code_str in ICD10_DESCRIPTIONS:
            return ICD10_DESCRIPTIONS[code_str]
        # Try with dot (e.g. I050 → I05.0)
        if len(code_str) >= 4 and "." not in code_str:
            dotted = code_str[:3] + "." + code_str[3:]
            if dotted in ICD10_DESCRIPTIONS:
                return ICD10_DESCRIPTIONS[dotted]
        # Try parent (3-char prefix)
        parent = code_str[:3].replace(".", "")
        if parent in ICD10_DESCRIPTIONS:
            return ICD10_DESCRIPTIONS[parent]
        return ""

    result = df.copy()

    # ── Strategy 1: Single primary ICD column ──
    single_candidates = ["ICD", "ICD_PRIMARY", "icd_code"]
    for col in single_candidates:
        if col in result.columns:
            result["ICD10_Description"] = result[col].apply(_lookup)
            return result

    # ── Strategy 2: Multiple diag columns (HES OP: diag_01 .. diag_12) ──
    diag_cols = sorted([c for c in result.columns if re.match(r"diag_\d+", c)])
    if diag_cols:
        for col in diag_cols:
            result[f"ICD10_Description_{col}"] = result[col].apply(_lookup)
        return result

    # ── Strategy 3: Cause columns (Death: cause, cause1..cause15) ──
    cause_cols = sorted([c for c in result.columns if re.match(r"cause\d*$", c)])
    if cause_cols:
        for col in cause_cols:
            result[f"ICD10_Description_{col}"] = result[col].apply(_lookup)
        return result

    # No ICD-like columns found
    return result


# ══════════════════════════════════════════════════════════════════════════════
# ENRICHMENT: Merge code descriptions into extraction results
# ══════════════════════════════════════════════════════════════════════════════

def _load_emis_medical_dict():
    """Load EMIS Medical Dictionary (medcodeid → Term, SnomedCTConceptId).
    Returns DataFrame or None. Uses session-state caching."""
    cache_key = "_cached_emis_medical_dict"
    if cache_key in st.session_state:
        return st.session_state[cache_key]
    engine = st.session_state.get("engine")
    if engine and engine.is_mock() and engine.mock_data:
        df = engine.mock_data.get("emis_dict")
        if df is not None:
            st.session_state[cache_key] = df
            return df
    path = PATHS.get("emis_dictionary", "")
    if os.path.exists(path):
        try:
            df = pd.read_csv(path, sep='\t', dtype=str,
                             usecols=["MedCodeId", "Term", "SnomedCTConceptId"])
            st.session_state[cache_key] = df
            return df
        except Exception:
            pass
    return None


def _load_emis_product_dict():
    """Load EMIS Product Dictionary (prodcodeid → ProductName, DrugSubstanceName, etc.).
    Returns DataFrame or None. Uses session-state caching."""
    cache_key = "_cached_emis_product_dict"
    if cache_key in st.session_state:
        return st.session_state[cache_key]
    engine = st.session_state.get("engine")
    if engine and engine.is_mock() and engine.mock_data:
        df = engine.mock_data.get("emis_prod_dict")
        if df is not None:
            st.session_state[cache_key] = df
            return df
    path = PATHS.get("emis_product_dictionary", "")
    if os.path.exists(path):
        try:
            cols_to_try = ["ProdCodeId", "Term", "ProductName", "DrugSubstanceName",
                           "DMD_ID", "FormulationName", "RouteName", "BNFChapter"]
            # Read header to detect available columns
            header_df = pd.read_csv(path, sep='\t', dtype=str, nrows=0)
            avail = [c for c in cols_to_try if c in header_df.columns]
            if not avail:
                avail = None  # read all
            df = pd.read_csv(path, sep='\t', dtype=str, usecols=avail)
            st.session_state[cache_key] = df
            return df
        except Exception:
            pass
    return None


def _search_product_dict_by_terms(search_terms):
    """Search EMIS Product Dictionary by drug search terms (generic + brand names).

    Returns list of prodcodeid strings, or empty list if dict unavailable.
    """
    prod_dict = _load_emis_product_dict()
    if prod_dict is None:
        return [], pd.DataFrame()

    available_cols = prod_dict.columns.tolist()
    search_cols = [c for c in ["Term", "ProductName", "DrugSubstanceName",
                                "FormulationName"] if c in available_cols]
    if not search_cols:
        search_cols = available_cols[:3]

    pid_col = "ProdCodeId" if "ProdCodeId" in available_cols else available_cols[0]

    mask = pd.Series(False, index=prod_dict.index)
    for term in search_terms:
        t_lower = term.lower()
        for col in search_cols:
            mask = mask | prod_dict[col].fillna("").str.lower().str.contains(t_lower, regex=False)

    matches = prod_dict[mask].copy()
    if len(matches) == 0:
        return [], pd.DataFrame()

    codes = matches[pid_col].astype(str).unique().tolist()
    return codes, matches


def enrich_with_code_details(df):
    """Auto-detect data type and merge descriptive code-detail columns.

    Enrichment rules:
    - medcodeid present → merge EMIS Medical Dict (adds Term, SnomedCTConceptId)
    - prodcodeid present → merge EMIS Product Dict (adds ProductName, DrugSubstanceName, etc.)
    - ICD/diag/cause columns present → add ICD-10 descriptions
    - consmedcodeid/probmedcodeid → merge EMIS Medical Dict on those columns too

    Returns enriched DataFrame (original columns preserved, new columns appended).
    """
    if df is None or df.empty:
        return df

    result = df.copy()
    enriched_cols = []

    # ── 1. Observation / MedCode data → EMIS Medical Dictionary ──
    medcode_cols = []
    for col in ["medcodeid", "consmedcodeid", "probmedcodeid"]:
        if col in result.columns and f"Term_{col}" not in result.columns and "Term" not in result.columns:
            medcode_cols.append(col)

    if medcode_cols:
        emis_dict = _load_emis_medical_dict()
        if emis_dict is not None:
            for col in medcode_cols:
                suffix = f"_{col}" if len(medcode_cols) > 1 else ""
                merge_cols = ["MedCodeId"]
                new_cols = {}
                if "Term" in emis_dict.columns:
                    merge_cols.append("Term")
                    new_cols["Term"] = f"Term{suffix}"
                if "SnomedCTConceptId" in emis_dict.columns:
                    merge_cols.append("SnomedCTConceptId")
                    new_cols["SnomedCTConceptId"] = f"SnomedCTConceptId{suffix}"

                lookup = emis_dict[merge_cols].drop_duplicates(subset=["MedCodeId"])
                result = result.merge(lookup, left_on=col, right_on="MedCodeId",
                                      how="left", suffixes=("", "_lookup"))
                # Rename merged columns if needed
                for old_name, new_name in new_cols.items():
                    if old_name in result.columns and new_name != old_name:
                        # Only rename if we used a suffix
                        if suffix:
                            result.rename(columns={old_name: new_name}, inplace=True)
                            enriched_cols.append(new_name)
                        else:
                            enriched_cols.append(old_name)
                # Clean up merge key
                if "MedCodeId" in result.columns and "MedCodeId" != col:
                    result.drop(columns=["MedCodeId"], errors="ignore", inplace=True)
                # Also clean any _lookup suffixed duplicates
                for c in list(result.columns):
                    if c.endswith("_lookup"):
                        result.drop(columns=[c], errors="ignore", inplace=True)

    # ── 2. Drug Issue data → EMIS Product Dictionary ──
    if "prodcodeid" in result.columns and "ProductName" not in result.columns:
        prod_dict = _load_emis_product_dict()
        if prod_dict is not None:
            prod_merge_cols = ["ProdCodeId"]
            for c in ["Term", "ProductName", "DrugSubstanceName", "FormulationName",
                       "RouteName", "BNFChapter", "DMD_ID"]:
                if c in prod_dict.columns:
                    prod_merge_cols.append(c)

            lookup = prod_dict[prod_merge_cols].drop_duplicates(subset=["ProdCodeId"])
            # Handle case where Term already exists from medcode enrichment
            if "Term" in result.columns and "Term" in lookup.columns:
                lookup = lookup.rename(columns={"Term": "DrugTerm"})
                prod_merge_cols = [c if c != "Term" else "DrugTerm" for c in prod_merge_cols]

            result = result.merge(lookup, left_on="prodcodeid", right_on="ProdCodeId",
                                  how="left", suffixes=("", "_prod"))
            if "ProdCodeId" in result.columns:
                result.drop(columns=["ProdCodeId"], errors="ignore", inplace=True)
            for c in list(result.columns):
                if c.endswith("_prod"):
                    result.drop(columns=[c], errors="ignore", inplace=True)
            enriched_cols.extend([c for c in ["ProductName", "DrugSubstanceName", "DrugTerm",
                                              "FormulationName", "RouteName", "BNFChapter"]
                                  if c in result.columns])

    # ── 3. HES / Death data → ICD-10 descriptions ──
    icd_candidates = ["ICD", "ICD_PRIMARY", "icd_code"]
    diag_cols = [c for c in result.columns if re.match(r"diag_\d+", c)]
    cause_cols = [c for c in result.columns if re.match(r"cause\d*$", c)]
    has_icd = any(c in result.columns for c in icd_candidates) or diag_cols or cause_cols

    if has_icd and "ICD10_Description" not in result.columns:
        result = add_icd10_descriptions(result)
        enriched_cols.extend([c for c in result.columns if "ICD10_Description" in c])

    return result


def get_current_username():
    """Detect the current system username for SSH commands."""
    try:
        return getpass.getuser()
    except Exception:
        return os.environ.get("USER", os.environ.get("USERNAME", "username"))


def get_local_hostname():
    """Detect hostname of the current node (for BMRC compute node detection)."""
    try:
        return socket.gethostname()
    except Exception:
        return "unknown"


def detect_login_node():
    """Auto-detect which BMRC login node the user connected through.

    Detection strategy (in priority order):
      1. $SLURM_SUBMIT_HOST — set by Slurm for both srun and sbatch
      2. $SSH_CONNECTION     — source IP of the SSH session
      3. Fall back to SSH_CONFIG default

    Returns the external-facing address (e.g. 'cluster4.bmrc.ox.ac.uk').
    """
    raw = ""

    # Method 1: Slurm knows which node submitted the job/session
    raw = os.environ.get("SLURM_SUBMIT_HOST", "")

    # Method 2: Reverse-resolve SSH source IP
    if not raw:
        ssh_conn = os.environ.get("SSH_CONNECTION", "")
        if ssh_conn:
            src_ip = ssh_conn.split()[0]
            try:
                raw = socket.gethostbyaddr(src_ip)[0]
            except Exception:
                pass

    # Resolve to external address — always extract short name first
    # because internal FQDNs like "cluster4.hpc.in.bmrc.ox.ac.uk"
    # also contain ".bmrc.ox.ac.uk" and would pass a suffix check
    if raw:
        short = raw.split(".")[0].lower()
        if short.startswith("cluster"):
            return f"{short}.bmrc.ox.ac.uk"
        if short.startswith("rescomp"):
            return f"{short}.well.ox.ac.uk"

    # Fallback
    return SSH_CONFIG["cluster_host"]


def generate_ssh_snippets(username, remote_port=8501):
    """Generate cross-platform SSH tunnel command snippets.

    Auto-detects:
      - compute_node: from hostname (FQDN for tunnel resolvability)
      - login_node:   from Slurm/SSH environment (correct cluster address)
    """
    login_node = detect_login_node()
    local_port = remote_port

    # Detect compute node if we're running on BMRC
    # Use FQDN so the tunnel resolves correctly from the login node
    hostname_short = get_local_hostname().split(".")[0].lower()
    hostname_full = get_local_hostname()
    is_compute = any(hostname_short.startswith(p) for p in ["comp", "node"])
    compute_node = hostname_full if is_compute else "<compute_node>"

    snippets = {}

    # ── SSH tunnel commands ──
    snippets["windows_powershell"] = {
        "title": "Windows (PowerShell / CMD)",
        "commands": [
            f"ssh -L {local_port}:{compute_node}:{remote_port} {username}@{login_node}",
        ],
        "note": "Then open your browser to: http://localhost:{port}".format(port=local_port),
    }
    snippets["macos"] = {
        "title": "macOS (Terminal)",
        "commands": [
            f"ssh -L {local_port}:{compute_node}:{remote_port} -o ServerAliveInterval=30 -o ServerAliveCountMax=3 {username}@{login_node}",
        ],
        "note": "Then open: http://localhost:{port}".format(port=local_port),
    }
    snippets["linux"] = {
        "title": "Linux (Terminal)",
        "commands": [
            f"ssh -L {local_port}:{compute_node}:{remote_port} -o ServerAliveInterval=30 -o ServerAliveCountMax=3 -o TCPKeepAlive=yes {username}@{login_node}",
        ],
        "note": "Then open: http://localhost:{port}".format(port=local_port),
    }

    # ── Heartbeat / keepalive flags explanation ──
    snippets["keepalive_info"] = (
        "`ServerAliveInterval=30` sends a keepalive packet every 30 seconds. "
        "`ServerAliveCountMax=3` closes the connection only after 3 missed replies (90s of silence). "
        "`TCPKeepAlive=yes` enables TCP-level keepalive as a secondary safeguard. "
        "This prevents the 'Connection refused (Channel 2/3)' tunnel crashes."
    )

    return snippets, compute_node, remote_port


# ══════════════════════════════════════════════════════════════════════════════
# CPRD ENGINE
# ══════════════════════════════════════════════════════════════════════════════

class CPRDEngine:
    def __init__(self):
        self._using_mock = not os.path.exists(CPRD_BASE)
        if self._using_mock:
            self.mock_data = generate_mock_data()
        else:
            self.mock_data = None

    def use_mock_data(self):
        self._using_mock = True
        if self.mock_data is None:
            self.mock_data = generate_mock_data()

    def is_mock(self):
        return self._using_mock

    def find_practice_folders(self, base_path=None):
        """Find all practice_* directories that contain zip files.

        Only returns directories whose name starts with 'practice_' to avoid
        scanning linkage/, 202102_lookups/, patid_lists_dcanoy/, etc.
        """
        if self._using_mock:
            return [f"practice_{i:03d}" for i in range(1, 21)]
        base = base_path or PATHS["aurum_base"]
        folders = []
        for item in sorted(glob.glob(os.path.join(base, "practice_*"))):
            if os.path.isdir(item):
                # Quick check: does it have at least one zip?
                has_zip = any(True for _ in glob.iglob(os.path.join(item, "*.zip")))
                if has_zip:
                    folders.append(item)
        return folders

    def find_zip_files(self, practice_folder, file_type):
        if self._using_mock:
            return []
        pattern_str = AURUM_FILE_TYPES[file_type]["pattern"]
        zips = []
        for zf in glob.glob(os.path.join(practice_folder, "*.zip")):
            if pattern_str.lower() in os.path.basename(zf).lower():
                zips.append(zf)
        return sorted(zips)

    def extract_from_zip(self, zip_path, file_type, filter_col=None, filter_values=None, select_cols=None):
        """Fast extraction using DuckDB with filter pushdown.

        Strategy 1 (Linux):  unzip -p → temp file → DuckDB WHERE scan
        Strategy 2 (any OS): Python zipfile → temp file → DuckDB WHERE scan
        Strategy 3 (fallback): pandas with usecols + isin filter

        DuckDB pushes the WHERE filter into the CSV scan, so only matching
        rows are ever materialised — typically <0.01% of the file.
        """
        # ── Identify inner file ──
        try:
            with zipfile.ZipFile(zip_path, 'r') as z:
                txt_files = [f for f in z.namelist() if f.endswith('.txt')]
                if not txt_files:
                    return None
                inner = txt_files[0]
        except Exception:
            return None

        # ── Build DuckDB SQL ──
        if select_cols:
            need = list(dict.fromkeys(([filter_col] if filter_col else []) + list(select_cols)))
            cols_sql = ", ".join(f'"{c}"' for c in need)
        else:
            cols_sql = "*"

        where_sql = ""
        if filter_col and filter_values:
            # Use a set-based IN clause for fast matching
            vals = ", ".join(f"'{v}'" for v in filter_values)
            where_sql = f'WHERE "{filter_col}" IN ({vals})'

        query = f"""
            SELECT {cols_sql}
            FROM read_csv('{{path}}',
                delim       = '\t',
                header      = true,
                all_varchar = true,
                null_padding = true,
                ignore_errors = true
            )
            {where_sql}
        """

        # Unique temp filename per thread
        tid = f"{os.getpid()}_{threading.get_ident() & 0xFFFFFF}"

        # ── Strategy 1 (Linux): unzip -p → temp file → DuckDB ──
        if platform.system() != "Windows":
            tmp_path = f"/tmp/cprd_{tid}.tsv"
            try:
                with open(tmp_path, 'wb') as tmp_f:
                    ret = subprocess.run(
                        ["unzip", "-p", zip_path, inner],
                        stdout=tmp_f, stderr=subprocess.DEVNULL,
                        timeout=600,  # 10 min per zip max
                    )
                if ret.returncode == 0 and os.path.getsize(tmp_path) > 0:
                    con = duckdb.connect()
                    df = con.execute(query.format(path=tmp_path)).fetchdf()
                    con.close()
                    try:
                        os.unlink(tmp_path)
                    except OSError:
                        pass
                    return df if len(df) > 0 else None
            except (subprocess.TimeoutExpired, Exception):
                pass
            finally:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

        # ── Strategy 2: Python zipfile → temp file → DuckDB ──
        tmp_path = f"/tmp/cprd_py_{tid}.tsv"
        try:
            with zipfile.ZipFile(zip_path, 'r') as z:
                with z.open(inner) as src, open(tmp_path, 'wb') as dst:
                    # Stream in 8 MB chunks to avoid loading entire file in RAM
                    while True:
                        chunk = src.read(8 * 1024 * 1024)
                        if not chunk:
                            break
                        dst.write(chunk)

            con = duckdb.connect()
            df = con.execute(query.format(path=tmp_path)).fetchdf()
            con.close()
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
            return df if len(df) > 0 else None
        except Exception:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

        # ── Strategy 3 (fallback): pandas ──
        try:
            with zipfile.ZipFile(zip_path, 'r') as z:
                with z.open(inner) as f:
                    kw = {"sep": '\t', "dtype": str, "low_memory": False}
                    if select_cols:
                        available_cols = pd.read_csv(f, sep='\t', nrows=0).columns.tolist()
                        f.seek(0)
                        kw["usecols"] = [c for c in select_cols if c in available_cols]
                    df = pd.read_csv(f, **kw)
            if filter_col and filter_values and filter_col in df.columns:
                df = df[df[filter_col].isin(set(filter_values))]
            return df if len(df) > 0 else None
        except Exception:
            return None

    # ── Parallel folder processing ──────────────────────────────────────

    def _get_max_workers(self):
        """Number of parallel workers — match Slurm CPU allocation."""
        try:
            slurm_cpus = int(os.environ.get("SLURM_CPUS_PER_TASK", "0"))
            if slurm_cpus > 0:
                return max(1, slurm_cpus - 1)  # leave 1 core for main thread
        except ValueError:
            pass
        return min(8, max(1, (os.cpu_count() or 4) - 1))

    def _process_one_zip(self, zip_path, file_type, filter_col, filter_values, select_cols):
        """Process a single zip file. Thread-safe, returns (folder_name, df|None)."""
        df = self.extract_from_zip(
            zip_path, file_type,
            filter_col=filter_col, filter_values=filter_values,
            select_cols=select_cols,
        )
        if df is not None:
            df["practice_folder"] = os.path.basename(os.path.dirname(zip_path))
        return df

    def _parallel_extract(self, file_type, filter_col, filter_values,
                          select_cols=None, practice_folders=None,
                          progress_callback=None):
        """Extract from all practice folders in parallel using ThreadPoolExecutor.

        Submits individual zip files as tasks (not whole folders) so the
        progress bar updates after every ~190 MB zip rather than every
        ~2.3 GB folder.  On 16 CPUs this processes 15 zips simultaneously.
        """
        folders = practice_folders or self.find_practice_folders()

        # ── Enumerate all matching zip files across all folders ──
        all_zips = []
        for folder in folders:
            zips = self.find_zip_files(folder, file_type)
            all_zips.extend(zips)

        if not all_zips:
            if progress_callback:
                progress_callback(1.0, "no files found", 0, 0)
            return pd.DataFrame()

        total = len(all_zips)
        results = []
        completed = 0
        max_w = self._get_max_workers()
        t_start = time.time()

        with ThreadPoolExecutor(max_workers=max_w) as pool:
            future_to_zip = {
                pool.submit(
                    self._process_one_zip,
                    zp, file_type, filter_col, filter_values, select_cols
                ): zp
                for zp in all_zips
            }

            for future in as_completed(future_to_zip):
                zp = future_to_zip[future]
                completed += 1
                try:
                    df = future.result()
                    if df is not None:
                        results.append(df)
                except Exception:
                    pass

                if progress_callback:
                    elapsed = time.time() - t_start
                    folder_name = os.path.basename(os.path.dirname(zp))
                    zip_name = os.path.basename(zp)
                    # Show short label: folder / zip filename
                    label = f"{folder_name}/{zip_name[:30]}"
                    progress_callback(
                        completed / total,
                        label,
                        completed, total,
                        elapsed,
                    )

        return pd.concat(results, ignore_index=True) if results else pd.DataFrame()

    def extract_observation_by_snomed(self, snomed_codes, practice_folders=None, progress_callback=None):
        """Extract observations matching SNOMED CT codes — parallel + DuckDB streaming."""
        if self._using_mock:
            if progress_callback:
                folders = self.find_practice_folders()
                for i, f in enumerate(folders):
                    progress_callback(i / len(folders), f, i, len(folders))
                    time.sleep(0.02)
            dict_df = self.mock_data["emis_dict"]
            matching_dict = dict_df[dict_df["SnomedCTConceptId"].isin(snomed_codes)]
            matching_medcodes = matching_dict["MedCodeId"].unique().tolist()
            obs_df = self.mock_data["observation"]
            result = obs_df[obs_df["medcodeid"].isin(matching_medcodes)].copy()
            result = result.merge(matching_dict[["MedCodeId", "Term", "SnomedCTConceptId"]],
                                  left_on="medcodeid", right_on="MedCodeId", how="left",
                                  suffixes=("_obs", "_dict"))
            if "Term_dict" in result.columns:
                result["Term"] = result["Term_dict"]
                result.drop(columns=["Term_obs", "Term_dict"], errors="ignore", inplace=True)
            result.drop(columns=["MedCodeId"], errors="ignore", inplace=True)
            return result

        # ── Load EMIS dictionary & map SNOMED → MedCodeId ──
        emis_path = PATHS["emis_dictionary"]
        if not os.path.exists(emis_path):
            st.error(f"EMIS Dictionary not found: {emis_path}")
            return pd.DataFrame()
        emis_df = pd.read_csv(emis_path, sep='\t', dtype=str)
        matching = emis_df[emis_df["SnomedCTConceptId"].isin(snomed_codes)]
        medcode_list = matching["MedCodeId"].unique().tolist()
        if not medcode_list:
            st.warning("No MedCodeIds found for the given SNOMED codes.")
            return pd.DataFrame()

        # ── Parallel extraction across practice folders ──
        combined = self._parallel_extract(
            file_type="Observation",
            filter_col="medcodeid",
            filter_values=medcode_list,
            select_cols=["patid", "obsid", "obsdate", "medcodeid", "value", "numunitid", "pracid"],
            practice_folders=practice_folders,
            progress_callback=progress_callback,
        )

        if combined.empty:
            return pd.DataFrame()

        # ── Enrich with SNOMED terms ──
        combined = combined.merge(
            matching[["MedCodeId", "Term", "SnomedCTConceptId"]],
            left_on="medcodeid", right_on="MedCodeId", how="left"
        )
        combined.drop(columns=["MedCodeId"], errors="ignore", inplace=True)
        return combined

    def extract_observation_by_medcode(self, medcode_list, practice_folders=None, progress_callback=None):
        if self._using_mock:
            if progress_callback:
                folders = self.find_practice_folders()
                for i, f in enumerate(folders):
                    progress_callback(i / len(folders), f, i, len(folders))
                    time.sleep(0.02)
            obs_df = self.mock_data["observation"]
            return obs_df[obs_df["medcodeid"].isin(medcode_list)].copy()

        return self._parallel_extract(
            file_type="Observation",
            filter_col="medcodeid",
            filter_values=medcode_list,
            select_cols=["patid", "obsid", "obsdate", "medcodeid", "value", "numunitid", "pracid"],
            practice_folders=practice_folders,
            progress_callback=progress_callback,
        )

    def extract_drugs_by_prodcode(self, prodcode_list, practice_folders=None, progress_callback=None):
        if self._using_mock:
            if progress_callback:
                folders = self.find_practice_folders()
                for i, f in enumerate(folders):
                    progress_callback(i / len(folders), f, i, len(folders))
                    time.sleep(0.02)
            drug_df = self.mock_data["drug"]
            return drug_df[drug_df["prodcodeid"].isin(prodcode_list)].copy()

        return self._parallel_extract(
            file_type="DrugIssue",
            filter_col="prodcodeid",
            filter_values=prodcode_list,
            select_cols=["patid", "issueid", "issuedate", "prodcodeid", "quantity", "duration", "pracid"],
            practice_folders=practice_folders,
            progress_callback=progress_callback,
        )

    def extract_any_filetype(self, file_type, filter_col=None, filter_values=None,
                             practice_folders=None, progress_callback=None):
        if file_type not in AURUM_FILE_TYPES:
            st.error(f"Unknown file type: {file_type}")
            return pd.DataFrame()
        if self._using_mock:
            if progress_callback:
                folders = self.find_practice_folders()
                for i, f in enumerate(folders):
                    progress_callback(i / len(folders), f, i, len(folders))
                    time.sleep(0.02)
            if file_type == "Observation":
                return self.mock_data["observation"].copy()
            elif file_type == "DrugIssue":
                return self.mock_data["drug"].copy()
            elif file_type == "Patient":
                return self.mock_data["patient"].copy()
            else:
                return pd.DataFrame({
                    "patid": [str(100000 + i) for i in range(100)],
                    "data_type": file_type,
                })

        ftype_info = AURUM_FILE_TYPES[file_type]
        fc = filter_col or ftype_info.get("key_col")

        return self._parallel_extract(
            file_type=file_type,
            filter_col=fc,
            filter_values=filter_values,
            practice_folders=practice_folders,
            progress_callback=progress_callback,
        )

    @staticmethod
    def _read_linkage_file(filepath, label="file"):
        """Robustly read a tab-delimited CPRD linkage file.

        Handles malformed rows (missing/extra columns) that are common in
        large HES and ONS files by trying progressively more tolerant
        read strategies.
        """
        if not filepath or not os.path.exists(filepath):
            st.error(f"{label} not found: {filepath}")
            return pd.DataFrame()

        # Strategy 1: pandas with on_bad_lines='warn' (skip malformed rows)
        try:
            df = pd.read_csv(filepath, sep='\t', dtype=str,
                             low_memory=False, on_bad_lines='warn')
            return df
        except TypeError:
            # Older pandas (<1.3) uses error_bad_lines instead
            pass
        except Exception as e:
            st.warning(f"Standard read failed for {label}, trying fallback: {e}")

        # Strategy 2: pandas with error_bad_lines=False (older pandas)
        try:
            df = pd.read_csv(filepath, sep='\t', dtype=str,
                             low_memory=False, error_bad_lines=False,
                             warn_bad_lines=True)
            return df
        except Exception:
            pass

        # Strategy 3: DuckDB with null_padding (handles missing columns)
        try:
            con = duckdb.connect()
            df = con.execute(f"""
                SELECT * FROM read_csv('{filepath}',
                    delim = '\\t',
                    header = true,
                    all_varchar = true,
                    null_padding = true,
                    ignore_errors = true
                )
            """).fetchdf()
            con.close()
            skipped_note = "(some malformed rows were null-padded or skipped)"
            st.info(f"Loaded {label} via DuckDB fallback {skipped_note}")
            return df
        except Exception as e2:
            st.error(f"All read strategies failed for {label}: {e2}")
            return pd.DataFrame()

    def extract_hes_apc(self, icd_codes, source="diagnosis_hosp"):
        if self._using_mock:
            hes = self.mock_data["hes_apc"]
            clean_codes = set()
            for c in icd_codes:
                clean_codes.add(c)
                clean_codes.add(c.replace(".", ""))
                if len(c) >= 3:
                    clean_codes.add(c[:3])
                    clean_codes.add(c[:3].replace(".", ""))
            result = hes[hes["ICD"].isin(clean_codes) | hes["ICD"].apply(
                lambda x: any(x.startswith(c.replace(".", "")) for c in icd_codes if len(c) >= 3)
            )].copy()
            return result
        source_map = {
            "diagnosis_hosp": PATHS["hes_diagnosis_hosp"],
            "diagnosis_epi": PATHS["hes_diagnosis_epi"],
            "primary_diag": PATHS["hes_primary_diag"],
        }
        filepath = source_map.get(source)
        df = self._read_linkage_file(filepath, label=f"HES APC ({source})")
        if df.empty:
            return df
        clean_codes = set()
        for c in icd_codes:
            clean_codes.add(c)
            clean_codes.add(c.replace(".", ""))
        icd_col = "ICD" if "ICD" in df.columns else next(
            (c for c in df.columns if "icd" in c.lower()), None)
        if icd_col is None:
            st.error("Cannot find ICD column in HES data")
            return pd.DataFrame()
        mask = df[icd_col].isin(clean_codes)
        short_codes = set(c[:3].replace(".", "") for c in icd_codes if len(c) >= 3)
        mask |= df[icd_col].apply(lambda x: str(x)[:3] in short_codes if pd.notna(x) else False)
        return df[mask].copy()

    def extract_hes_op(self, icd_codes):
        if self._using_mock:
            op = self.mock_data["hes_op"]
            codes_no_dots = set(c.replace(".", "") for c in icd_codes)
            codes_no_dots.update(c[:3].replace(".", "") for c in icd_codes if len(c) >= 3)
            diag_cols = [c for c in op.columns if c.startswith("diag_")]
            mask = pd.Series(False, index=op.index)
            for col in diag_cols:
                mask |= op[col].isin(codes_no_dots)
                mask |= op[col].apply(lambda x: str(x)[:3] in codes_no_dots if pd.notna(x) else False)
            return op[mask].copy()
        filepath = PATHS["hesop_clinical"]
        df = self._read_linkage_file(filepath, label="HES OP")
        if df.empty:
            return df
        codes_no_dots = set(c.replace(".", "") for c in icd_codes)
        diag_cols = [c for c in df.columns if "diag" in c.lower() or "icd" in c.lower()]
        mask = pd.Series(False, index=df.index)
        for col in diag_cols:
            mask |= df[col].isin(codes_no_dots)
        return df[mask].copy()

    def extract_death_records(self, icd_codes):
        if self._using_mock:
            death = self.mock_data["death"]
            codes = set()
            for c in icd_codes:
                codes.add(c)
                codes.add(c.replace(".", ""))
            cause_cols = ["cause"] + [f"cause{i}" for i in range(1, 16)]
            available_cols = [c for c in cause_cols if c in death.columns]
            mask = pd.Series(False, index=death.index)
            for col in available_cols:
                mask |= death[col].isin(codes)
                mask |= death[col].apply(lambda x: any(str(x).startswith(c.replace(".", "")[:3]) for c in icd_codes if len(c) >= 3) if pd.notna(x) and str(x).strip() else False)
            result = death[mask].copy()
            id_cols = ["patid", "dod"]
            melted = result.melt(id_vars=id_cols, value_vars=available_cols,
                                 var_name="cause_position", value_name="icd_code")
            melted = melted[melted["icd_code"].isin(codes) | melted["icd_code"].apply(
                lambda x: any(str(x).startswith(c.replace(".", "")[:3]) for c in icd_codes) if pd.notna(x) and str(x).strip() else False
            )]
            return melted
        filepath = PATHS["death_patient"]
        df = self._read_linkage_file(filepath, label="ONS Death records")
        if df.empty:
            return df
        codes = set()
        for c in icd_codes:
            codes.add(c)
            codes.add(c.replace(".", ""))
        cause_cols = [c for c in df.columns if c.startswith("cause")]
        id_cols = [c for c in ["patid", "dod", "dor"] if c in df.columns]
        mask = pd.Series(False, index=df.index)
        for col in cause_cols:
            mask |= df[col].isin(codes)
        result = df[mask].copy()
        melted = result.melt(id_vars=id_cols, value_vars=cause_cols,
                             var_name="cause_position", value_name="icd_code")
        melted = melted[melted["icd_code"].isin(codes)]
        return melted

    def get_linkage_eligibility(self):
        if self._using_mock:
            return self.mock_data["linkage_elig"]
        filepath = PATHS["linkage_eligibility_aurum"]
        if not os.path.exists(filepath):
            filepath = PATHS["linkage_eligibility"]
        return self._read_linkage_file(filepath, label="Linkage eligibility")

    def get_patient_imd(self):
        if self._using_mock:
            return self.mock_data["imd"]
        filepath = PATHS["patient_imd"]
        return self._read_linkage_file(filepath, label="Patient IMD")


# ══════════════════════════════════════════════════════════════════════════════
# UI HELPERS: Progress with ETA, tooltips, downloads
# ══════════════════════════════════════════════════════════════════════════════

class ProgressTracker:
    """Real-time progress with current folder/file, %, and ETA.

    Supports both serial (old-style) and parallel extraction callbacks.
    The parallel callback passes an extra `elapsed` argument so the ETA
    is computed from actual wall-clock time rather than cumulative per-file time.
    """
    def __init__(self, label="Processing"):
        self.label = label
        self.bar = st.progress(0)
        self.status = st.empty()
        self.start_time = time.time()

    def update(self, fraction, current_file, idx, total, elapsed=None):
        pct = min(fraction, 1.0)
        self.bar.progress(pct)
        if elapsed is None:
            elapsed = time.time() - self.start_time
        if pct > 0.02:
            eta_sec = (elapsed / pct) * (1 - pct)
            if eta_sec > 60:
                eta_str = f"{eta_sec / 60:.1f} min"
            else:
                eta_str = f"{eta_sec:.0f}s"
        else:
            eta_str = "calculating..."
        workers = ""
        try:
            n_w = int(os.environ.get("SLURM_CPUS_PER_TASK", "0"))
            if n_w > 1:
                workers = f" · ⚡ {max(1, n_w - 1)} parallel workers"
        except ValueError:
            pass
        self.status.markdown(
            f"**{self.label}** — File **{idx}** of **{total}**: `{current_file}` · "
            f"**{pct * 100:.0f}%** · ETA: **{eta_str}**{workers}"
        )

    def complete(self, msg="Complete!"):
        self.bar.progress(1.0)
        elapsed = time.time() - self.start_time
        self.status.markdown(f"✅ **{msg}** ({elapsed:.1f}s)")


def show_tooltip(key):
    """Render a small help tooltip next to a section."""
    text = TOOLTIPS.get(key, "")
    if text:
        st.caption(f"ℹ️ {text}")


def download_results(df, filename, key_prefix, show_merge=False, all_results=None):
    """Download helper with optional merge/separate toggle and code detail enrichment.

    Uses cached CSV conversion to avoid re-serialising on every Streamlit rerun.
    Automatically enriches data with lookup descriptions (Term, SNOMED, ICD descriptions, etc.)
    """
    # ── Enrichment toggle ──
    enrich = st.checkbox("📖 Include code descriptions in export",
                         value=True, key=f"{key_prefix}_enrich",
                         help="Merge lookup dictionaries to add human-readable descriptions "
                              "for medical codes, drug products, and ICD-10 diagnoses.")

    if show_merge and all_results and len(all_results) > 1:
        merge_mode = st.toggle("🔗 Merge all sources into one file", value=False,
                               key=f"{key_prefix}_merge", help=TOOLTIPS["merge_mode"])
    else:
        merge_mode = False

    if merge_mode and all_results:
        frames = []
        for src, src_df in all_results.items():
            sdf = src_df.copy()
            sdf["source"] = src
            if enrich:
                sdf = enrich_with_code_details(sdf)
            frames.append(sdf)
        merged = pd.concat(frames, ignore_index=True)
        # Cache key: use shape + column hash to detect changes
        cache_key = f"{key_prefix}_merged_{len(merged)}_{hash(tuple(merged.columns))}_{enrich}"
        csv = _df_to_csv_cached(cache_key, merged)
        dl_cols = st.columns([1, 1])
        with dl_cols[0]:
            st.download_button(f"📥 CSV ({len(merged):,} rows)",
                               csv, "merged_extraction.csv", "text/csv",
                               key=f"{key_prefix}_dl_merged")
        with dl_cols[1]:
            if HAS_PYARROW:
                pq_buf = io.BytesIO()
                _safe_df = merged.copy()
                for c in _safe_df.columns:
                    if _safe_df[c].dtype == object:
                        _safe_df[c] = _safe_df[c].astype(str)
                _safe_df.to_parquet(pq_buf, index=False, engine="pyarrow")
                st.download_button(f"📥 Parquet ({len(merged):,} rows)",
                                   pq_buf.getvalue(), "merged_extraction.parquet",
                                   "application/octet-stream",
                                   key=f"{key_prefix}_dl_merged_pq")
    else:
        export_df = df.copy()
        if enrich:
            export_df = enrich_with_code_details(export_df)
        cache_key = f"{key_prefix}_{len(export_df)}_{hash(tuple(export_df.columns))}_{enrich}"
        csv = _df_to_csv_cached(cache_key, export_df)
        pq_name = filename.replace(".csv", ".parquet")
        dl_cols = st.columns([1, 1])
        with dl_cols[0]:
            st.download_button(f"📥 CSV ({len(export_df):,} rows)",
                               csv, filename, "text/csv",
                               key=f"{key_prefix}_dl_single")
        with dl_cols[1]:
            if HAS_PYARROW:
                pq_buf = io.BytesIO()
                _safe_df = export_df.copy()
                for c in _safe_df.columns:
                    if _safe_df[c].dtype == object:
                        _safe_df[c] = _safe_df[c].astype(str)
                _safe_df.to_parquet(pq_buf, index=False, engine="pyarrow")
                st.download_button(f"📥 Parquet ({len(export_df):,} rows)",
                                   pq_buf.getvalue(), pq_name,
                                   "application/octet-stream",
                                   key=f"{key_prefix}_dl_single_pq")


def generate_plain_summary(all_results, code_classification):
    """Generate a plain-text summary explaining extraction results."""
    lines = []
    lines.append("═══ Extraction Summary ═══")
    lines.append("")

    total_records = sum(len(df) for df in all_results.values())
    all_patids = set()
    for df in all_results.values():
        if "patid" in df.columns:
            all_patids.update(df["patid"].unique())

    lines.append(f"Total records found: {total_records:,}")
    lines.append(f"Unique patients: {len(all_patids):,}")
    lines.append("")

    if code_classification["snomed"]:
        lines.append(f"SNOMED CT codes used: {len(code_classification['snomed'])} codes")
    if code_classification["icd10"]:
        lines.append(f"ICD-10 codes used: {len(code_classification['icd10'])} codes")
    lines.append("")

    src_map = {
        "CPRD Aurum": "GP/Primary Care records (practice-level observation files with SNOMED CT diagnoses)",
        "HES APC": "Hospital Admitted Patient Care (all NHS hospital admission diagnoses, ICD-10 coded)",
        "HES OP": "Hospital Outpatient (NHS outpatient clinic visits, ICD-10 coded without dots)",
        "ONS Death": "Office for National Statistics death certificates (underlying + contributory causes of death)",
    }

    lines.append("Data sources searched:")
    for src, df in all_results.items():
        desc = src_map.get(src, src)
        n = len(df)
        n_pat = df["patid"].nunique() if "patid" in df.columns else None
        pat_str = f"{n_pat:,}" if n_pat is not None else "N/A"
        lines.append(f"  • {src}: {n:,} records, {pat_str} patients")
        lines.append(f"    What this contains: {desc}")
    lines.append("")

    if len(all_results) > 1:
        lines.append("Note: A patient may appear in multiple sources (e.g. diagnosed in GP AND hospital).")
        lines.append("The 'source' column identifies where each record came from.")

    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
# UI STYLING
# ══════════════════════════════════════════════════════════════════════════════

def apply_custom_css():
    st.markdown("""
    <style>
    /* ── Theme-aware variables ── */
    :root {
        --cprd-accent: #2563eb;
        --cprd-accent-hover: #1d4ed8;
    }

    /* ── Sidebar: always dark ── */
    [data-testid="stSidebar"] { background-color: #1e293b; }
    [data-testid="stSidebar"] * { color: #e2e8f0 !important; }
    [data-testid="stSidebar"] .stSelectbox label,
    [data-testid="stSidebar"] .stRadio label,
    [data-testid="stSidebar"] .stCheckbox label { color: #e2e8f0 !important; }
    [data-testid="stSidebar"] input,
    [data-testid="stSidebar"] textarea { color: #e2e8f0 !important; }

    /* ── FIX: Last Extraction sidebar text readability ── */
    [data-testid="stSidebar"] [data-testid="stMetric"] {
        background: rgba(255,255,255,0.08);
        border-radius: 8px;
        padding: 12px;
        border: 1px solid rgba(255,255,255,0.15);
    }
    [data-testid="stSidebar"] [data-testid="stMetricValue"] {
        color: #93c5fd !important;
        font-weight: 700 !important;
        font-size: 1.4rem !important;
    }
    [data-testid="stSidebar"] [data-testid="stMetricLabel"] {
        color: #cbd5e1 !important;
        font-weight: 500 !important;
    }
    [data-testid="stSidebar"] [data-testid="stMetricDelta"] {
        color: #86efac !important;
    }

    /* ── Metric cards: adapt to theme ── */
    [data-testid="stMetric"] {
        border-radius: 12px; padding: 16px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    }

    /* ── Tabs ── */
    .stTabs [data-baseweb="tab"][aria-selected="true"] { color: var(--cprd-accent) !important; }

    /* ── DataFrame ── */
    .stDataFrame { border-radius: 8px; overflow: hidden; }

    /* ── Download button ── */
    .stDownloadButton button {
        background: var(--cprd-accent); color: white !important;
        border-radius: 8px; border: none;
    }
    .stDownloadButton button:hover { background: var(--cprd-accent-hover); }

    /* ── Patient scope banner ── */
    .scope-banner {
        border-radius: 8px; padding: 14px 18px; margin: 8px 0 16px 0;
        font-size: 14px; line-height: 1.5;
    }
    .scope-banner-linked {
        background: #eff6ff; border: 1px solid #93c5fd; color: #1e40af;
    }
    .scope-banner-all {
        background: #f0fdf4; border: 1px solid #86efac; color: #166534;
    }
    .scope-banner-custom {
        background: #fefce8; border: 1px solid #fcd34d; color: #92400e;
    }
    html[data-theme="dark"] .scope-banner-linked,
    .stApp[data-theme="dark"] .scope-banner-linked {
        background: #0c1e3a !important; border-color: #1e40af !important; color: #bfdbfe !important;
    }
    html[data-theme="dark"] .scope-banner-all,
    .stApp[data-theme="dark"] .scope-banner-all {
        background: #052e16 !important; border-color: #166534 !important; color: #bbf7d0 !important;
    }
    html[data-theme="dark"] .scope-banner-custom,
    .stApp[data-theme="dark"] .scope-banner-custom {
        background: #2a1f00 !important; border-color: #92400e !important; color: #fde68a !important;
    }

    /* ── Loading shimmer animation ── */
    @keyframes shimmer {
        0% { background-position: -200px 0; }
        100% { background-position: 200px 0; }
    }
    .loading-skeleton {
        background: linear-gradient(90deg, #f0f0f0 25%, #e0e0e0 50%, #f0f0f0 75%);
        background-size: 200px 100%;
        animation: shimmer 1.5s infinite;
        border-radius: 4px;
        height: 20px;
        margin: 8px 0;
    }

    /* ══════ LIGHT MODE ══════ */
    @media (prefers-color-scheme: light) {
        .success-box { background: #f0fdf4; border: 1px solid #86efac; border-radius: 8px; padding: 16px; margin: 8px 0; color: #166534; }
        .info-box { background: #eff6ff; border: 1px solid #93c5fd; border-radius: 8px; padding: 16px; margin: 8px 0; color: #1e40af; }
        .warn-box { background: #fffbeb; border: 1px solid #fcd34d; border-radius: 8px; padding: 16px; margin: 8px 0; color: #92400e; }
        .path-box { background: #f1f5f9; border: 1px solid #cbd5e1; border-radius: 8px; padding: 12px;
                    font-family: monospace; font-size: 13px; margin: 4px 0; color: #334155; }
        .summary-box { background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 16px; margin: 8px 0; color: #1e293b; white-space: pre-wrap; font-family: monospace; font-size: 13px; }
        [data-testid="stMetric"] { background: #ffffff; border: 1px solid #e2e8f0; }
    }

    /* ══════ DARK MODE ══════ */
    @media (prefers-color-scheme: dark) {
        .success-box { background: #052e16; border: 1px solid #166534; border-radius: 8px; padding: 16px; margin: 8px 0; color: #bbf7d0; }
        .info-box { background: #0c1e3a; border: 1px solid #1e40af; border-radius: 8px; padding: 16px; margin: 8px 0; color: #bfdbfe; }
        .warn-box { background: #2a1f00; border: 1px solid #92400e; border-radius: 8px; padding: 16px; margin: 8px 0; color: #fde68a; }
        .path-box { background: #1e293b; border: 1px solid #475569; border-radius: 8px; padding: 12px;
                    font-family: monospace; font-size: 13px; margin: 4px 0; color: #cbd5e1; }
        .summary-box { background: #1e293b; border: 1px solid #475569; border-radius: 8px; padding: 16px; margin: 8px 0; color: #e2e8f0; white-space: pre-wrap; font-family: monospace; font-size: 13px; }
        [data-testid="stMetric"] { background: #1e293b; border: 1px solid #334155; }
    }

    /* ══════ Streamlit dark fallback ══════ */
    html[data-theme="dark"] .success-box,
    .stApp[data-theme="dark"] .success-box { background: #052e16 !important; border-color: #166534 !important; color: #bbf7d0 !important; }
    html[data-theme="dark"] .info-box,
    .stApp[data-theme="dark"] .info-box { background: #0c1e3a !important; border-color: #1e40af !important; color: #bfdbfe !important; }
    html[data-theme="dark"] .warn-box,
    .stApp[data-theme="dark"] .warn-box { background: #2a1f00 !important; border-color: #92400e !important; color: #fde68a !important; }
    html[data-theme="dark"] .path-box,
    .stApp[data-theme="dark"] .path-box { background: #1e293b !important; border-color: #475569 !important; color: #cbd5e1 !important; }
    html[data-theme="dark"] .summary-box,
    .stApp[data-theme="dark"] .summary-box { background: #1e293b !important; border-color: #475569 !important; color: #e2e8f0 !important; }
    html[data-theme="dark"] [data-testid="stMetric"],
    .stApp[data-theme="dark"] [data-testid="stMetric"] { background: #1e293b !important; border-color: #334155 !important; }
    </style>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# EXTRACTION STATE MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════

def save_extraction(result, source_label="extraction"):
    """Store extraction results efficiently: save full data to a temp file and
    keep only lightweight summary stats in session_state.  This prevents Streamlit
    from serialising a multi-million-row DataFrame on every rerun."""
    if result is None or len(result) == 0:
        return

    # Pre-compute stats once (instant, never recomputed)
    stats = {
        "n_records": len(result),
        "n_patients": int(result["patid"].nunique()) if "patid" in result.columns else 0,
        "sources": list(result["source"].unique()) if "source" in result.columns else [source_label],
        "patient_ids": set(result["patid"].unique()) if "patid" in result.columns else set(),
        "columns": result.columns.tolist(),
        "source_label": source_label,
        "timestamp": time.strftime("%H:%M:%S"),
    }

    # Store the full DataFrame (needed for downloads / analytics) plus cached stats
    st.session_state["last_extraction"] = result
    st.session_state["extraction_stats"] = stats

    # ── Auto-save to disk (if enabled in Output Settings) ──
    saved = _auto_save_to_disk(result, source_label=source_label)
    if saved:
        st.session_state["last_saved_files"] = saved
        short = [os.path.basename(p) for p in saved]
        st.toast(f"💾 Saved: {', '.join(short)}", icon="✅")

    # Accumulate extraction history for patient-scope linking
    if "extraction_history" not in st.session_state:
        st.session_state["extraction_history"] = []
    st.session_state["extraction_history"].append(stats)


def get_extraction_stats():
    """Return cached stats dict or None. Never recomputes from DataFrame."""
    return st.session_state.get("extraction_stats", None)


def get_scope_patient_ids():
    """Return the set of patient IDs to filter by, or None for 'all patients'."""
    scope = st.session_state.get("patient_scope_mode", "all")
    if scope == "previous" and "extraction_stats" in st.session_state:
        return st.session_state["extraction_stats"].get("patient_ids", None)
    elif scope == "custom" and "custom_patient_ids" in st.session_state:
        return st.session_state["custom_patient_ids"]
    return None  # "all" → no filter


def apply_patient_scope(df):
    """Filter DataFrame by the current patient scope. Returns df unchanged if scope='all'."""
    if df is None or len(df) == 0 or "patid" not in df.columns:
        return df
    scope_ids = get_scope_patient_ids()
    if scope_ids is not None and len(scope_ids) > 0:
        filtered = df[df["patid"].isin(scope_ids)]
        return filtered if len(filtered) > 0 else df
    return df


def render_patient_scope(page_key="default"):
    """Show the Patient Scope selector at the top of extraction pages.

    Returns after rendering; the actual filtering is done by apply_patient_scope()
    which reads from session_state at extraction time.
    """
    stats = get_extraction_stats()
    has_previous = stats is not None and stats["n_patients"] > 0

    # Initialize scope mode
    if "patient_scope_mode" not in st.session_state:
        st.session_state["patient_scope_mode"] = "all"

    st.markdown("#### 🎯 Patient Scope")

    if has_previous:
        options = [
            f"🔗 Same patients from previous extraction ({stats['n_patients']:,} patients)",
            "🌍 All patients (whole population)",
            "📋 Custom patient list",
        ]
        choice = st.radio(
            "Which patients should this extraction cover?",
            options,
            index={"previous": 0, "all": 1, "custom": 2}.get(
                st.session_state.get("patient_scope_mode", "all"), 1
            ),
            key=f"scope_radio_{page_key}",
            horizontal=True,
        )
        if "Same patients" in choice:
            st.session_state["patient_scope_mode"] = "previous"
            st.markdown(
                f'<div class="scope-banner scope-banner-linked">'
                f'🔗 <strong>Linked mode:</strong> Extractions will be filtered to the '
                f'<strong>{stats["n_patients"]:,} patients</strong> found in your '
                f'{stats["source_label"]} extraction ({stats["timestamp"]})'
                f'</div>',
                unsafe_allow_html=True,
            )
        elif "All patients" in choice:
            st.session_state["patient_scope_mode"] = "all"
            st.markdown(
                '<div class="scope-banner scope-banner-all">'
                '🌍 <strong>Full population:</strong> Extractions will scan all patients '
                'across all practice folders — no patient filter applied.'
                '</div>',
                unsafe_allow_html=True,
            )
        elif "Custom" in choice:
            st.session_state["patient_scope_mode"] = "custom"
            _render_custom_patient_input(page_key)
    else:
        st.markdown(
            '<div class="scope-banner scope-banner-all">'
            '🌍 <strong>Full population mode:</strong> No previous extraction found. '
            'Extractions will scan all patients. Run a Quick Extract first to enable '
            'linked patient filtering.'
            '</div>',
            unsafe_allow_html=True,
        )
        st.session_state["patient_scope_mode"] = "all"

    st.markdown("---")


def _render_custom_patient_input(page_key):
    """Sub-component for entering custom patient IDs."""
    custom_method = st.radio(
        "How would you like to provide patient IDs?",
        ["Paste IDs", "Upload file"],
        horizontal=True,
        key=f"custom_method_{page_key}",
    )
    if custom_method == "Paste IDs":
        text = st.text_area(
            "Enter patient IDs (one per line or comma-separated)",
            placeholder="1234567890\n2345678901\n3456789012",
            key=f"custom_patids_{page_key}",
            height=100,
        )
        if text.strip():
            ids = set(c.strip() for c in re.split(r'[,\s\n]+', text.strip()) if c.strip())
            st.session_state["custom_patient_ids"] = ids
            st.markdown(
                f'<div class="scope-banner scope-banner-custom">'
                f'📋 <strong>Custom list:</strong> {len(ids):,} patient IDs loaded. '
                f'Extractions will be filtered to these patients only.'
                f'</div>',
                unsafe_allow_html=True,
            )
    else:
        uploaded = st.file_uploader(
            "Upload CSV/TXT with patient IDs",
            type=["csv", "txt", "xlsx"],
            key=f"custom_patid_file_{page_key}",
        )
        if uploaded:
            try:
                if uploaded.name.endswith(".xlsx"):
                    udf = pd.read_excel(uploaded, dtype=str)
                else:
                    udf = pd.read_csv(uploaded, dtype=str)
                col = st.selectbox(
                    "Select column containing patient IDs",
                    udf.columns.tolist(),
                    key=f"custom_patid_col_{page_key}",
                )
                ids = set(udf[col].dropna().astype(str).str.strip().tolist())
                st.session_state["custom_patient_ids"] = ids
                st.markdown(
                    f'<div class="scope-banner scope-banner-custom">'
                    f'📋 <strong>Custom list:</strong> {len(ids):,} patient IDs loaded from file.'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            except Exception as e:
                st.error(f"Error reading file: {e}")


def reset_all():
    """Clear all extraction data and reset app to initial state."""
    keys_to_clear = [
        "last_extraction", "extraction_stats", "extraction_history",
        "multi_source_results", "patient_scope_mode", "custom_patient_ids",
        "cohort_data", "cohort_config", "cohort_start_df", "cohort_steps", "cohort_patients",
    ]
    # Also clear CSV caches, persistent result keys, and drug library caches
    csv_keys = [k for k in st.session_state if k.startswith(("_csv_cache_", "_res_", "_drug_lib_", "_drug_tab_"))]
    keys_to_clear.extend(csv_keys)
    for key in keys_to_clear:
        if key in st.session_state:
            del st.session_state[key]


def _df_to_csv_cached(cache_key, df):
    """Cache CSV conversion in session_state to avoid re-serialising on every Streamlit rerun."""
    state_key = f"_csv_cache_{cache_key}"
    if state_key not in st.session_state:
        st.session_state[state_key] = df.to_csv(index=False).encode("utf-8")
    return st.session_state[state_key]


# ── Output-settings defaults ──────────────────────────────────────────────
_DEFAULT_OUTPUT_DIR = os.path.join(os.path.expanduser("~"), "cprd_outputs")

def _get_output_settings():
    """Return current output settings from session_state (with safe defaults)."""
    return {
        "auto_save":  st.session_state.get("output_auto_save", True),
        "format":     st.session_state.get("output_format", "Parquet"),
        "directory":  st.session_state.get("output_directory", _DEFAULT_OUTPUT_DIR),
        "enrich":     st.session_state.get("output_enrich", True),
    }


# ── Group & class abbreviations for auto-save filenames ───────────────────────
DISEASE_GROUP_ABBREV = {
    "Ischaemic Heart Disease": "IHD",
    "Valvular Heart Disease": "VHD",
    "Arrhythmia": "Arr",
    "Cardiomyopathy": "CMP",
    "Vascular Disease": "VascD",
    "Congenital Heart Disease": "CHD",
    "Infectious/Inflammatory Heart Disease": "InfHD",
    "Heart Failure": "HF",
}

DRUG_CLASS_ABBREV = {
    "Antiarrhythmics": "AntiArr",
    "Anticoagulants": "AntiCoag",
    "Antifibrinolytics": "AntiFib",
    "Antihypertensives": "AntiHTN",
    "Antiplatelets": "AntiPlt",
    "Cardiac Amyloidosis": "CardAmyl",
    "Critical Care & Vasoactive": "CritCare",
    "Glucose-Lowering": "GluLow",
    "Heart Failure": "HFDrug",
    "Lipid-Lowering": "LipLow",
    "Local Sclerosants": "Scler",
    "Nitrates & Antianginals": "Nitrate",
    "Obesity (CV-Relevant)": "Obesity",
    "Pericarditis & Inflammatory": "PeriInfl",
    "Peripheral Vascular Disease": "PVDDrug",
    "Potassium Management": "KMgmt",
    "Pulmonary Hypertension": "PulmHTN",
    "Thrombolytics": "Thrombo",
}


def _build_save_label(source, diseases=None, drug_classes=None, drugs=None):
    """Build a descriptive auto-save filename component.

    Examples:
        _build_save_label("Aurum_SNOMED", diseases=["Aortic Valve Stenosis"])
            → "AVS_Aurum_SNOMED"
        _build_save_label("Aurum_SNOMED", diseases=[...16 VHD diseases...])
            → "VHD_Aurum_SNOMED"
        _build_save_label("DrugIssue", drug_classes=["Antihypertensives","Lipid-Lowering"])
            → "AntiHTN_LipLow_DrugIssue"
    """
    prefix_parts = []

    # ── Disease context ──
    if diseases and len(diseases) > 0:
        if len(diseases) <= 3:
            for d in diseases:
                entry = DISEASE_CODE_LIBRARY.get(d, {})
                prefix_parts.append(entry.get("short", d[:4]))
        else:
            groups_seen = set()
            for d in diseases:
                for gname, glist in DISEASE_GROUPS.items():
                    if d in glist:
                        groups_seen.add(gname)
                        break
            if len(groups_seen) <= 3:
                for g in sorted(groups_seen):
                    prefix_parts.append(DISEASE_GROUP_ABBREV.get(g, g[:4]))
            else:
                prefix_parts.append(f"{len(diseases)}Diseases")

    # ── Drug context ──
    if drug_classes and len(drug_classes) > 0:
        if len(drug_classes) <= 3:
            for c in drug_classes:
                prefix_parts.append(DRUG_CLASS_ABBREV.get(c, c[:6]))
        else:
            prefix_parts.append(f"{len(drug_classes)}DrugClasses")
    elif drugs and len(drugs) > 0 and not drug_classes:
        if len(drugs) <= 3:
            prefix_parts.extend([d.replace(" ", "")[:8] for d in drugs])
        else:
            prefix_parts.append(f"{len(drugs)}Drugs")

    # ── Build final label ──
    if prefix_parts:
        prefix = "_".join(prefix_parts)
        return f"{prefix}_{source}"
    return source


def _auto_save_to_disk(df, source_label="extraction"):
    """Persist a DataFrame to the configured output directory.

    Formats supported:  CSV (.csv),  Parquet (.parquet),  Both.
    File naming:  <source_label>_<YYYYMMDD_HHMMSS>.<ext>
    All logic is self-contained — no external helper files needed.
    """
    cfg = _get_output_settings()
    if not cfg["auto_save"] or df is None or len(df) == 0:
        return None

    out_dir = cfg["directory"]
    fmt     = cfg["format"]   # "CSV", "Parquet", or "Both"

    # Ensure output directory exists
    try:
        os.makedirs(out_dir, exist_ok=True)
    except OSError as exc:
        st.warning(f"⚠️ Cannot create output folder `{out_dir}`: {exc}")
        return None

    # Build safe filename
    safe_label = re.sub(r"[^A-Za-z0-9_-]", "_", source_label).lower()
    ts = time.strftime("%Y%m%d_%H%M%S")
    base = f"{safe_label}_{ts}"

    # Optionally enrich with code descriptions before saving
    export_df = df.copy()
    if cfg["enrich"]:
        try:
            export_df = enrich_with_code_details(export_df)
        except Exception:
            pass  # silently skip enrichment if it fails

    saved_paths = []

    # ── CSV ──
    if fmt in ("CSV", "Both"):
        csv_path = os.path.join(out_dir, f"{base}.csv")
        try:
            export_df.to_csv(csv_path, index=False)
            saved_paths.append(csv_path)
        except Exception as exc:
            st.warning(f"⚠️ CSV save failed: {exc}")

    # ── Parquet ──
    if fmt in ("Parquet", "Both"):
        pq_path = os.path.join(out_dir, f"{base}.parquet")
        try:
            if not HAS_PYARROW:
                st.warning("⚠️ `pyarrow` not installed — falling back to CSV. "
                           "Run `pip install pyarrow --user` to enable Parquet.")
                if pq_path not in saved_paths:
                    csv_fb = os.path.join(out_dir, f"{base}.csv")
                    export_df.to_csv(csv_fb, index=False)
                    saved_paths.append(csv_fb)
            else:
                # Ensure all columns are Parquet-safe types
                for col in export_df.columns:
                    if export_df[col].dtype == object:
                        export_df[col] = export_df[col].astype(str)
                export_df.to_parquet(pq_path, index=False, engine="pyarrow")
                saved_paths.append(pq_path)
        except Exception as exc:
            st.warning(f"⚠️ Parquet save failed: {exc}")

    return saved_paths


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: HOME
# ══════════════════════════════════════════════════════════════════════════════

def page_home():
    # ── Hero section (modern, theme-compatible) ──
    st.markdown("""
    <div style="padding:16px 0 8px 0;">
        <h1 style="margin:0; font-size:2.1rem; font-weight:800; letter-spacing:-0.5px;
                    color:var(--text-color, inherit);">
            🗂️ CPRD Extractor
        </h1>
        <p style="margin:4px 0 0 0; font-size:0.92rem; opacity:0.65; font-weight:400;">
            Version 1.0 &nbsp;·&nbsp; Compatible with BMRC, generic Linux servers &amp; Windows
        </p>
    </div>
    <div style="padding:4px 0 16px 0; border-bottom:1px solid rgba(128,128,128,0.2); margin-bottom:20px;">
        <span style="font-size:0.82rem; opacity:0.5; letter-spacing:0.3px;">
            Developed by <b style="opacity:0.8;">Milad Nazarzadeh</b>
            &nbsp;·&nbsp; Nuffield Department of Women's &amp; Reproductive Health
            &nbsp;·&nbsp; University of Oxford
        </span>
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("#### 📂 Data Environment")
        mode_label = st.session_state.get("data_mode", "mock")
        mode_labels = {
            "mock": "🧪 Mock Data (Testing)",
            "bmrc": "🔬 Live CPRD Data — BMRC Cluster",
            "any_server": "🖥️ Live — Custom Server",
            "windows": "💻 Windows OS Mode",
        }
        st.info(f"**Mode:** {mode_labels.get(mode_label, mode_label)}")
        if st.session_state.engine.is_mock():
            mock = st.session_state.engine.mock_data
            st.metric("Patients", f"{len(mock['patient']):,}")
            st.metric("Observations", f"{len(mock['observation']):,}")
            st.metric("Drug Issues", f"{len(mock['drug']):,}")
            st.metric("HES APC Records", f"{len(mock['hes_apc']):,}")
        else:
            folders = st.session_state.engine.find_practice_folders()
            st.metric("Practice Folders Found", len(folders))

    with col2:
        st.markdown("#### 🗺 Data Paths")
        paths_to_show = {
            "CPRD Base": PATHS["aurum_base"],
            "EMIS Dictionary": PATHS["emis_dictionary"],
            "Linkage Eligibility": os.path.dirname(PATHS["linkage_eligibility"]),
        }
        for label, path in paths_to_show.items():
            exists = "✅" if os.path.exists(path) or st.session_state.engine.is_mock() else "❌"
            st.markdown(f"**{label}** {exists}")
            st.markdown(f'<div class="path-box">{path}</div>', unsafe_allow_html=True)

        st.markdown("**Linkage (HES/Death)**")
        for ldir in PATHS["linkage_base"]:
            exists = "✅" if os.path.exists(ldir) or st.session_state.engine.is_mock() else "❌"
            short = ldir.split("20_095_Results/")[-1] if "20_095_Results/" in ldir else ldir
            st.markdown(f'&nbsp;&nbsp;{exists} <span class="path-box">{short}</span>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("#### Available Modules")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown("**📋 Code List Development**\n\n6-stage pipeline from concept → clinician-reviewed code list")
    with c2:
        st.markdown("**🔬 CPRD Aurum**\n\nObservation, DrugIssue, Consultation, Patient, Problem, Referral")
    with c3:
        st.markdown("**🏥 Linkage + Multi-Source**\n\nHES APC, HES OP, ONS Death, IMD, or search all at once")
    with c4:
        st.markdown("**📊 Demographics + Analytics**\n\nSex, age, deprivation, ethnicity, Table 1, Kaplan-Meier")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: NEWBIE USER (Quick Extract) ★ NEW ★
# ══════════════════════════════════════════════════════════════════════════════

def page_newbie():
    st.title("🌟 Quick Extract — One-Click Workflow")
    st.markdown("**Just paste your codes. The system handles everything else.**")
    show_tooltip("nav_newbie")

    st.markdown("### Step 1: Provide your diagnostic codes")
    st.markdown("Paste any codes below — SNOMED CT (e.g. `60573004`), ICD-10 (e.g. `I05.0`), or a mix of both. "
                "The system will automatically detect the code type.")

    input_method = st.radio("How do you have your codes?",
                            ["Type / paste codes", "Upload a file", "Use built-in VHD codes"],
                            horizontal=True, key="newbie_input")

    raw_codes = []

    if input_method == "Type / paste codes":
        text = st.text_area(
            "Paste codes here (one per line, comma-separated, or space-separated)",
            placeholder="60573004\nI05.0\n836480008\nI35.0, I06.0",
            height=150, key="newbie_codes"
        )
        if text.strip():
            raw_codes = re.split(r'[,\s\n]+', text.strip())
            raw_codes = [c.strip() for c in raw_codes if c.strip()]

    elif input_method == "Upload a file":
        uploaded = st.file_uploader("Upload CSV, TXT, or XLSX", type=["csv", "txt", "xlsx"],
                                    key="newbie_upload")
        if uploaded:
            try:
                if uploaded.name.endswith(".xlsx"):
                    udf = pd.read_excel(uploaded, dtype=str)
                else:
                    udf = pd.read_csv(uploaded, dtype=str)
                st.dataframe(udf.head(5))
                col = st.selectbox("Which column contains the codes?", udf.columns.tolist(), key="newbie_col")
                raw_codes = udf[col].dropna().astype(str).str.strip().tolist()
                st.success(f"{len(raw_codes)} codes loaded from file")
            except Exception as e:
                st.error(f"Error reading file: {e}")

    elif input_method == "Use built-in VHD codes":
        all_s = []
        all_i = []
        for entry in DISEASE_CODE_LIBRARY.values():
            all_s.extend(entry.get("snomed", []))
            all_i.extend(entry.get("icd10", []))
        all_s = list(set(all_s))
        all_i = list(set(all_i))
        raw_codes = all_s + all_i
        st.success(f"Using {len(all_s)} SNOMED + {len(all_i)} ICD-10 built-in disease codes ({len(raw_codes)} total across 63 CVD diseases)")

    # ── Auto-classify codes ──
    if raw_codes:
        classified = classify_codes(raw_codes)
        n_s, n_i, n_u = len(classified["snomed"]), len(classified["icd10"]), len(classified["unknown"])

        col1, col2, col3 = st.columns(3)
        col1.metric("SNOMED CT detected", n_s)
        col2.metric("ICD-10 detected", n_i)
        col3.metric("Unrecognised", n_u)

        if n_u > 0:
            with st.expander(f"⚠️ {n_u} codes could not be classified"):
                st.write(", ".join(classified["unknown"]))
                st.markdown("These codes will be tried against all sources as-is.")
                # Add unknowns to both lists as fallback
                classified["snomed"].extend(classified["unknown"])
                classified["icd10"].extend(classified["unknown"])

        if n_s == 0 and n_i == 0 and n_u == 0:
            st.warning("No valid codes detected. Please check your input.")
            return

        st.markdown("### Step 2: Run Comprehensive Search")
        st.markdown("The system will search **all available data sources** simultaneously.")

        sources_to_search = []
        if classified["snomed"]:
            sources_to_search.append("CPRD Aurum (Primary Care)")
        if classified["icd10"]:
            sources_to_search.extend(["HES APC (Hospital)", "HES OP (Outpatient)", "ONS Death"])
        st.markdown(f"**Sources to search:** {' · '.join(sources_to_search)}")

        if st.button("🚀 Extract Everything", key="btn_newbie_go", type="primary", width='stretch'):
            all_results = {}

            # ── CPRD Aurum ──
            if classified["snomed"]:
                st.markdown("---")
                st.markdown("#### Searching CPRD Aurum (Primary Care)...")
                tracker = ProgressTracker("CPRD Aurum")
                aurum_result = st.session_state.engine.extract_observation_by_snomed(
                    list(set(classified["snomed"])),
                    progress_callback=tracker.update
                )
                if aurum_result is not None and len(aurum_result) > 0:
                    aurum_result["source"] = "CPRD Aurum"
                    all_results["CPRD Aurum"] = aurum_result
                    tracker.complete(f"Found {len(aurum_result):,} records")
                else:
                    tracker.complete("No records found")

            # ── HES APC ──
            if classified["icd10"]:
                st.markdown("#### Searching HES APC (Hospital)...")
                with st.spinner("Querying hospital diagnoses..."):
                    hes_result = st.session_state.engine.extract_hes_apc(list(set(classified["icd10"])))
                if len(hes_result) > 0:
                    hes_result["source"] = "HES APC"
                    all_results["HES APC"] = hes_result
                    st.markdown(f'<div class="success-box">✅ HES APC: {len(hes_result):,} records</div>', unsafe_allow_html=True)
                else:
                    st.info("HES APC: No matching records.")

            # ── HES OP ──
            if classified["icd10"]:
                st.markdown("#### Searching HES OP (Outpatient)...")
                with st.spinner("Querying outpatient diagnoses..."):
                    op_result = st.session_state.engine.extract_hes_op(list(set(classified["icd10"])))
                if len(op_result) > 0:
                    op_result["source"] = "HES OP"
                    all_results["HES OP"] = op_result
                    st.markdown(f'<div class="success-box">✅ HES OP: {len(op_result):,} records</div>', unsafe_allow_html=True)
                else:
                    st.info("HES OP: No matching records.")

            # ── ONS Death ──
            if classified["icd10"]:
                st.markdown("#### Searching ONS Death Records...")
                with st.spinner("Querying death certificates..."):
                    death_result = st.session_state.engine.extract_death_records(list(set(classified["icd10"])))
                if len(death_result) > 0:
                    death_result["source"] = "ONS Death"
                    all_results["ONS Death"] = death_result
                    st.markdown(f'<div class="success-box">✅ ONS Death: {len(death_result):,} records</div>', unsafe_allow_html=True)
                else:
                    st.info("ONS Death: No matching records.")

            # ── Store results ──
            if all_results:
                st.session_state["multi_source_results"] = all_results
                combined = pd.concat(all_results.values(), ignore_index=True)
                _qe_label = "AllCVD_Quick_Extract" if st.session_state.get("newbie_input") == "Use built-in VHD codes" else "Quick_Extract"
                save_extraction(combined, source_label=_qe_label)
                st.session_state["_res_newbie"] = combined
                st.session_state["_res_newbie_all"] = all_results
                st.session_state["_res_newbie_classified"] = classified
            else:
                st.session_state.pop("_res_newbie", None)
                st.session_state.pop("_res_newbie_all", None)
                st.session_state.pop("_res_newbie_classified", None)
                st.warning("No matching records found in any source.")

        # ── Persistent Quick Extract results ──
        if "_res_newbie" in st.session_state and st.session_state["_res_newbie"] is not None:
            all_results = st.session_state.get("_res_newbie_all", {})
            classified = st.session_state.get("_res_newbie_classified", {})
            combined = st.session_state["_res_newbie"]

            st.markdown("---")
            st.markdown("### 📋 Results")
            total_records = sum(len(df) for df in all_results.values())
            all_pats = set()
            for df in all_results.values():
                if "patid" in df.columns:
                    all_pats.update(df["patid"].unique())
            c1, c2, c3 = st.columns(3)
            c1.metric("Total Records", f"{total_records:,}")
            c2.metric("Unique Patients", f"{len(all_pats):,}")
            c3.metric("Sources with Matches", len(all_results))

            summary_text = generate_plain_summary(all_results, classified)
            st.markdown("#### What's in this extract:")
            st.markdown(f'<div class="summary-box">{summary_text}</div>', unsafe_allow_html=True)

            st.markdown("#### Download")
            download_results(combined, "quick_extract_merged.csv", "newbie",
                             show_merge=True, all_results=all_results)

            with st.expander("Preview individual source data"):
                for src_name, df in all_results.items():
                    st.markdown(f"**{src_name}** ({len(df):,} rows)")
                    st.dataframe(df.head(50), width='stretch')


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: CPRD AURUM EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def page_aurum_extraction():
    st.title("🔬 CPRD Aurum Primary Care Extraction")
    show_tooltip("nav_aurum")

    # ── Patient Scope selector ──
    render_patient_scope(page_key="aurum")

    tab_mc, tab_obs, tab_drug, tab_generic = st.tabs([
        "🆔 MedCode Extraction", "📋 Observation (SNOMED)", "💊 Drug Issue", "📦 Extract Any File Type (All Patients)"
    ])

    with tab_obs:
        st.markdown("### Extract Observations by SNOMED CT Codes")
        show_tooltip("snomed_codes")
        st.markdown("Maps SNOMED CT → MedCodeId via EMIS Dictionary, then scans all practice folders.")

        code_source = st.radio("SNOMED Code Source",
                               ["Built-in Disease Library (63 CVD)", "Enter Manually", "Upload Code List"],
                               horizontal=True, key="obs_snomed_source")
        snomed_codes = []

        if code_source == "Built-in Disease Library (63 CVD)":
            sel_groups = st.multiselect("Select disease groups", list(DISEASE_GROUPS.keys()),
                                         default=["Valvular Heart Disease"], key="obs_disease_groups")
            avail = []
            for g in sel_groups:
                avail.extend(DISEASE_GROUPS[g])
            if avail:
                sel = st.multiselect(f"Select diseases ({len(avail)} available)", avail, default=avail, key="obs_diseases")
                for d in sel:
                    snomed_codes.extend(DISEASE_CODE_LIBRARY.get(d, {}).get("snomed", []))
                snomed_codes = list(set(snomed_codes))
            if snomed_codes:
                st.success(f"**{len(snomed_codes)} unique SNOMED codes selected**")
                with st.expander("View selected codes"):
                    for d in sel:
                        entry = DISEASE_CODE_LIBRARY.get(d, {})
                        st.write(f"**{d}** ({entry.get('short','')}) — {', '.join(entry.get('snomed',[]))}")

        elif code_source == "Enter Manually":
            manual = st.text_area("Enter SNOMED CT codes (one per line or comma-separated)",
                                  placeholder="60573004\n836480008\n86466006", key="obs_manual_snomed")
            if manual.strip():
                snomed_codes = [c.strip() for c in manual.replace(",", "\n").split("\n") if c.strip()]
                st.info(f"{len(snomed_codes)} codes entered")

        elif code_source == "Upload Code List":
            uploaded = st.file_uploader("Upload CSV/TXT with SNOMED codes", type=["csv", "txt", "xlsx"],
                                        key="obs_upload_snomed")
            if uploaded:
                try:
                    if uploaded.name.endswith(".xlsx"):
                        udf = pd.read_excel(uploaded)
                    else:
                        udf = pd.read_csv(uploaded, dtype=str)
                    st.dataframe(udf.head())
                    col = st.selectbox("Select column containing SNOMED codes",
                                       udf.columns.tolist(), key="obs_upload_col")
                    snomed_codes = udf[col].dropna().astype(str).str.strip().tolist()
                    st.success(f"{len(snomed_codes)} codes loaded from file")
                except Exception as e:
                    st.error(f"Error reading file: {e}")

        if snomed_codes and st.button("🚀 Run SNOMED Extraction", key="btn_snomed_extract", type="primary"):
            tracker = ProgressTracker("CPRD Aurum SNOMED Scan")
            result = st.session_state.engine.extract_observation_by_snomed(
                list(set(snomed_codes)),
                progress_callback=tracker.update
            )
            tracker.complete()
            if result is not None and len(result) > 0:
                result = apply_patient_scope(result)
                _label = _build_save_label("Aurum_SNOMED",
                    diseases=st.session_state.get("obs_diseases"))
                save_extraction(result, source_label=_label)
                st.session_state["_res_aurum_snomed"] = result
            else:
                st.session_state.pop("_res_aurum_snomed", None)
                st.warning("No matching records found.")

        # ── Persistent results (survives download-triggered rerun) ──
        if "_res_aurum_snomed" in st.session_state and st.session_state["_res_aurum_snomed"] is not None:
            _r = st.session_state["_res_aurum_snomed"]
            scope_label = f" (scope: {st.session_state.get('patient_scope_mode', 'all')})" if st.session_state.get("patient_scope_mode") != "all" else ""
            st.markdown(f'<div class="success-box">✅ <strong>Found {len(_r):,} records</strong> across <strong>{_r["patid"].nunique():,} unique patients</strong>{scope_label}</div>', unsafe_allow_html=True)
            st.dataframe(_r.head(100), width='stretch')
            download_results(_r, "snomed_extraction.csv", "aurum_snomed")
    with tab_drug:
        st.markdown("### Extract Drug Issues by Product Code")
        st.markdown("Scans DrugIssue files across all practice folders.")

        # ── Auto-detect prodcodes from Drug Lookup page ──
        dl_codes = st.session_state.get("_drug_lookup_prodcodes")
        if dl_codes:
            st.success(f"💊 **{len(dl_codes):,} ProdCodeIds received from Drug Lookup**")
            if st.button("✅ Use Drug Lookup codes", key="btn_use_dl_codes", type="primary"):
                st.session_state["_drug_tab_from_lookup"] = dl_codes

        drug_source = st.radio("Product Code Source",
                               ["Built-in Drug Library (315 CV/Metabolic)", "Enter Manually", "Upload Code List"],
                               horizontal=True, key="drug_source")
        prodcodes = []

        # ── Built-in Drug Library ──
        if drug_source == "Built-in Drug Library (315 CV/Metabolic)":
            sel_groups = st.multiselect(
                "Select therapeutic classes",
                sorted(DRUG_GROUPS.keys()),
                default=["Antihypertensives"],
                key="drug_lib_groups",
            )
            # Gather drugs from selected classes
            avail_drugs_all = []
            for g in sel_groups:
                avail_drugs_all.extend(DRUG_GROUPS.get(g, []))

            # ── Pharmacological drug class filter ──
            avail_cats = sorted({
                DRUG_CODE_LIBRARY.get(d, {}).get("cat", "")
                for d in avail_drugs_all
            } - {""})

            if avail_cats:
                if "drug_lib_cats" in st.session_state:
                    st.session_state["drug_lib_cats"] = [
                        c for c in st.session_state["drug_lib_cats"] if c in avail_cats
                    ]
                sel_cats = st.multiselect(
                    f"Select pharmacological drug class ({len(avail_cats)} available)",
                    avail_cats, default=avail_cats, key="drug_lib_cats",
                )
            else:
                sel_cats = []

            # Filter drugs by selected pharmacological categories
            avail_drugs = [
                d for d in avail_drugs_all
                if DRUG_CODE_LIBRARY.get(d, {}).get("cat", "") in sel_cats
            ] if sel_cats else avail_drugs_all

            if avail_drugs:
                # Prune stale selections that are no longer in the options
                if "drug_lib_drugs" in st.session_state:
                    st.session_state["drug_lib_drugs"] = [
                        d for d in st.session_state["drug_lib_drugs"] if d in avail_drugs
                    ]
                sel_drugs = st.multiselect(
                    f"Select individual drugs ({len(avail_drugs)} available)",
                    avail_drugs, default=avail_drugs, key="drug_lib_drugs",
                )
                # Collect all search terms for selected drugs
                all_terms = []
                for d in sel_drugs:
                    all_terms.extend(DRUG_CODE_LIBRARY.get(d, {}).get("terms", []))
                all_terms = sorted(set(all_terms))
                if all_terms:
                    st.success(f"**{len(sel_drugs)} drugs selected** → **{len(all_terms)} search terms** (generic + brand names)")
                    with st.expander("View selected drugs & search terms"):
                        for d in sel_drugs:
                            entry = DRUG_CODE_LIBRARY.get(d, {})
                            st.caption(f"**{d}** ({entry.get('cat','')}, {entry.get('class','')}) — "
                                       f"terms: {', '.join(entry.get('terms', []))}")
                    # Auto-search Product Dictionary for matching ProdCodeIds
                    cache_key = "_drug_lib_cache"
                    last_terms_key = "_drug_lib_last_terms"
                    if st.session_state.get(last_terms_key) != all_terms:
                        with st.spinner(f"Searching Product Dictionary with {len(all_terms)} terms..."):
                            found_codes, match_df = _search_product_dict_by_terms(all_terms)
                            st.session_state[cache_key] = found_codes
                            st.session_state[last_terms_key] = all_terms
                            st.session_state["_drug_lib_match_df"] = match_df
                    found_codes = st.session_state.get(cache_key, [])
                    if found_codes:
                        st.info(f"📦 **{len(found_codes):,} ProdCodeIds** matched in Product Dictionary")
                        match_df = st.session_state.get("_drug_lib_match_df")
                        if match_df is not None and len(match_df) > 0:
                            with st.expander(f"Preview matched products ({len(match_df):,} rows)"):
                                show_cols = [c for c in ["ProdCodeId", "Term", "ProductName",
                                             "DrugSubstanceName", "BNFChapter"] if c in match_df.columns]
                                st.dataframe(match_df[show_cols].head(200) if show_cols else match_df.head(200),
                                             height=300, hide_index=True)
                        prodcodes = found_codes
                    else:
                        st.warning("No matching products found. Check Product Dictionary path in Configuration.")

        # Check if Drug Lookup codes were accepted (overrides radio selection)
        from_lookup = st.session_state.get("_drug_tab_from_lookup")
        if from_lookup:
            prodcodes = from_lookup
            st.info(f"Using {len(prodcodes):,} ProdCodeIds from Drug Lookup")

        elif drug_source == "Enter Manually":
            manual_drugs = st.text_area("Enter Product Codes (one per line or comma-separated)",
                                        placeholder="1234567\n2345678", key="drug_manual")
            if manual_drugs.strip():
                prodcodes = [c.strip() for c in manual_drugs.replace(",", "\n").split("\n") if c.strip()]
                st.info(f"{len(prodcodes)} product codes entered")

        elif drug_source == "Upload Code List":
            uploaded_drug = st.file_uploader("Upload CSV/TXT/XLSX with product codes",
                                             type=["csv", "txt", "xlsx"], key="drug_upload")
            if uploaded_drug:
                try:
                    if uploaded_drug.name.endswith(".xlsx"):
                        udf = pd.read_excel(uploaded_drug)
                    else:
                        udf = pd.read_csv(uploaded_drug, dtype=str)
                    st.dataframe(udf.head())
                    col = st.selectbox("Select column with product codes",
                                       udf.columns.tolist(), key="drug_upload_col")
                    prodcodes = udf[col].dropna().astype(str).str.strip().tolist()
                    st.success(f"{len(prodcodes)} product codes loaded")
                except Exception as e:
                    st.error(f"Error reading file: {e}")

        if st.session_state.engine.is_mock():
            if st.checkbox("Use all mock drug codes (for testing)", key="drug_use_all_mock"):
                prodcodes = st.session_state.engine.mock_data["drug"]["prodcodeid"].unique().tolist()
                st.info(f"Using all {len(prodcodes)} mock product codes")

        if prodcodes and st.button("🚀 Run Drug Extraction", key="btn_drug_extract", type="primary"):
            tracker = ProgressTracker("Drug Issue Scan")
            result = st.session_state.engine.extract_drugs_by_prodcode(
                prodcodes, progress_callback=tracker.update)
            tracker.complete()
            if result is not None and len(result) > 0:
                result = apply_patient_scope(result)
                _label = _build_save_label("Aurum_DrugIssue",
                    drug_classes=st.session_state.get("drug_lib_groups"))

                # ── Force Parquet save (drug data is large) ──
                cfg = _get_output_settings()
                out_dir = cfg["directory"]
                try:
                    os.makedirs(out_dir, exist_ok=True)
                    safe_label = re.sub(r"[^A-Za-z0-9_-]", "_", _label).lower()
                    ts = time.strftime("%Y%m%d_%H%M%S")
                    pq_path = os.path.join(out_dir, f"{safe_label}_{ts}.parquet")
                    export_df = result.copy()
                    if cfg["enrich"]:
                        try:
                            export_df = enrich_with_code_details(export_df)
                        except Exception:
                            pass
                    for col in export_df.columns:
                        if export_df[col].dtype == object:
                            export_df[col] = export_df[col].astype(str)
                    if HAS_PYARROW:
                        export_df.to_parquet(pq_path, index=False, engine="pyarrow")
                        st.toast(f"💾 Saved: {os.path.basename(pq_path)}", icon="✅")
                        st.session_state["last_saved_files"] = [pq_path]
                    else:
                        csv_fb = pq_path.replace(".parquet", ".csv")
                        export_df.to_csv(csv_fb, index=False)
                        st.warning("⚠️ pyarrow not installed — saved as CSV instead. "
                                   "Run `pip install pyarrow --user` for Parquet.")
                        st.session_state["last_saved_files"] = [csv_fb]
                except Exception as exc:
                    st.warning(f"⚠️ Save failed: {exc}")

                # Store stats (without re-saving via save_extraction)
                stats = {
                    "n_records": len(result),
                    "n_patients": int(result["patid"].nunique()) if "patid" in result.columns else 0,
                    "sources": list(result["source"].unique()) if "source" in result.columns else [_label],
                    "patient_ids": set(result["patid"].unique()) if "patid" in result.columns else set(),
                    "columns": result.columns.tolist(),
                    "source_label": _label,
                    "timestamp": time.strftime("%H:%M:%S"),
                }
                st.session_state["last_extraction"] = result
                st.session_state["extraction_stats"] = stats
                if "extraction_history" not in st.session_state:
                    st.session_state["extraction_history"] = []
                st.session_state["extraction_history"].append(stats)

                st.session_state["_res_aurum_drug"] = result
            else:
                st.session_state.pop("_res_aurum_drug", None)
                st.warning("No matching records found.")

        if "_res_aurum_drug" in st.session_state and st.session_state["_res_aurum_drug"] is not None:
            _r = st.session_state["_res_aurum_drug"]
            st.markdown(f'<div class="success-box">✅ <strong>Found {len(_r):,} drug records</strong> across <strong>{_r["patid"].nunique():,} unique patients</strong></div>', unsafe_allow_html=True)
            st.dataframe(_r.head(100), width='stretch')
            download_results(_r, "drug_extraction.csv", "aurum_drug")
    with tab_mc:
        st.markdown("### Extract Observations by MedCodeId")
        show_tooltip("medcode_ids")
        st.markdown("Direct extraction using CPRD MedCodeIds (no SNOMED mapping needed).")

        # ── Auto-detect medcodeids from Code List Development Stage 5 ──
        s5_ids = st.session_state.get("s5_medcodeids_for_extraction")
        if s5_ids:
            st.success(f"📋 **{len(s5_ids):,} medcodeids received from Code List Development → Stage 5**")
            if st.button("✅ Use Stage 5 medcodeids", key="btn_use_s5_medcodeids", type="primary"):
                st.session_state["_aurum_mc_from_s5"] = s5_ids

        medcode_source = st.radio("MedCode Source",
                                  ["Enter Manually", "Upload Code List"],
                                  horizontal=True, key="medcode_source")
        medcodes = []

        # Check if Stage 5 codes were accepted
        if st.session_state.get("_aurum_mc_from_s5"):
            medcodes = st.session_state["_aurum_mc_from_s5"]
            st.info(f"Using {len(medcodes):,} medcodeids from Stage 5 matching")

        elif medcode_source == "Enter Manually":
            manual_mc = st.text_area("Enter MedCodeIds (one per line or comma-separated)",
                                     placeholder="123456789012345\n234567890123456", key="medcode_manual")
            if manual_mc.strip():
                medcodes = [c.strip() for c in manual_mc.replace(",", "\n").split("\n") if c.strip()]
                st.info(f"{len(medcodes)} MedCodeIds entered")

        elif medcode_source == "Upload Code List":
            uploaded_mc = st.file_uploader("Upload CSV/TXT/XLSX with MedCodeIds",
                                           type=["csv", "txt", "xlsx"], key="medcode_upload")
            if uploaded_mc:
                try:
                    if uploaded_mc.name.endswith(".xlsx"):
                        udf = pd.read_excel(uploaded_mc)
                    else:
                        udf = pd.read_csv(uploaded_mc, dtype=str)
                    st.dataframe(udf.head())
                    col = st.selectbox("Select column with MedCodeIds",
                                       udf.columns.tolist(), key="medcode_upload_col")
                    medcodes = udf[col].dropna().astype(str).str.strip().tolist()
                    st.success(f"{len(medcodes)} MedCodeIds loaded")
                except Exception as e:
                    st.error(f"Error reading file: {e}")

        if st.session_state.engine.is_mock():
            if st.checkbox("Use all mock MedCodeIds (for testing)", key="mc_use_all_mock"):
                medcodes = st.session_state.engine.mock_data["observation"]["medcodeid"].unique().tolist()
                st.info(f"Using all {len(medcodes)} mock MedCodeIds")

        if medcodes and st.button("🚀 Run MedCode Extraction", key="btn_medcode_extract", type="primary"):
            tracker = ProgressTracker("MedCode Scan")
            result = st.session_state.engine.extract_observation_by_medcode(
                medcodes, progress_callback=tracker.update)
            tracker.complete()
            if result is not None and len(result) > 0:
                result = apply_patient_scope(result)
                save_extraction(result, source_label="Aurum MedCode")
                st.session_state["_res_aurum_mc"] = result
            else:
                st.session_state.pop("_res_aurum_mc", None)
                st.warning("No matching records found.")

        if "_res_aurum_mc" in st.session_state and st.session_state["_res_aurum_mc"] is not None:
            _r = st.session_state["_res_aurum_mc"]
            st.markdown(f'<div class="success-box">✅ <strong>Found {len(_r):,} records</strong> across <strong>{_r["patid"].nunique():,} unique patients</strong></div>', unsafe_allow_html=True)
            st.dataframe(_r.head(100), width='stretch')
            download_results(_r, "medcode_extraction.csv", "aurum_mc")
    with tab_generic:
        st.markdown("### Extract Any CPRD Aurum File Type For All Patients")
        file_type = st.selectbox("Select File Type", list(AURUM_FILE_TYPES.keys()), key="generic_ftype")
        finfo = AURUM_FILE_TYPES[file_type]
        st.markdown(f"**Key column:** `{finfo['key_col']}` · **Date column:** `{finfo['date_col'] or 'N/A'}`")
        st.markdown(f"**Columns:** {', '.join(finfo['cols'][:8])}{'...' if len(finfo['cols']) > 8 else ''}")

        use_filter = st.checkbox("Filter by specific codes?", key="generic_use_filter")
        filter_values = None
        if use_filter:
            filter_text = st.text_area(f"Enter {finfo['key_col']} values (one per line)", key="generic_filter_vals")
            if filter_text.strip():
                filter_values = [v.strip() for v in filter_text.strip().split("\n") if v.strip()]
                st.info(f"Filtering on {len(filter_values)} values")

        if st.button(f"🚀 Extract {file_type} Data", key="btn_generic_extract", type="primary"):
            tracker = ProgressTracker(f"{file_type} Scan")
            result = st.session_state.engine.extract_any_filetype(
                file_type,
                filter_col=finfo["key_col"] if use_filter else None,
                filter_values=filter_values,
                progress_callback=tracker.update
            )
            tracker.complete()
            if result is not None and len(result) > 0:
                result = apply_patient_scope(result)
                save_extraction(result, source_label=f"Aurum {file_type}")
                st.session_state["_res_aurum_generic"] = result
                st.session_state["_res_aurum_generic_ftype"] = file_type
            else:
                st.session_state.pop("_res_aurum_generic", None)
                st.warning("No records found.")

        if "_res_aurum_generic" in st.session_state and st.session_state["_res_aurum_generic"] is not None:
            _r = st.session_state["_res_aurum_generic"]
            _ft = st.session_state.get("_res_aurum_generic_ftype", "data")
            st.markdown(f'<div class="success-box">✅ <strong>Found {len(_r):,} records</strong> ({_r["patid"].nunique():,} patients)</div>', unsafe_allow_html=True)
            st.dataframe(_r.head(100), width='stretch')
            download_results(_r, f"{_ft.lower()}_extraction.csv", f"aurum_{_ft.lower()}")
# ══════════════════════════════════════════════════════════════════════════════
# PAGE: LINKAGE EXTRACTION (FIXED — was crashing)
# ══════════════════════════════════════════════════════════════════════════════

def page_linkage_extraction():
    st.title("🏥 Linkage Data Extraction")
    st.caption("HES APC · HES OP · HES A&E · ONS Death · IMD")
    show_tooltip("nav_linkage")

    # ── Patient Scope selector ──
    render_patient_scope(page_key="linkage")

    # ── Information Panel ──
    with st.expander("ℹ️ **What is each data source?** (click to expand)", expanded=False):
        info_cols = st.columns(2)
        with info_cols[0]:
            st.markdown("""
**🏥 HES APC (Admitted Patient Care)**
Records for all NHS hospital admissions — inpatient stays, day cases,
and surgical procedures. Diagnoses coded in ICD-10; procedures in OPCS-4.

**🏢 HES OP (Outpatient)**
Outpatient specialist appointments at NHS hospitals. Diagnoses coded in
ICD-10 *without* decimal points (e.g. I050 not I05.0).

**🚑 HES A&E (Accident & Emergency)**
Urgent and emergency care visit details including attendance type,
diagnosis, investigations performed, and treatments given.
            """)
        with info_cols[1]:
            st.markdown("""
**💀 ONS Death (Mortality Records)**
Official national mortality records from the Office for National Statistics.
Contains date of death and up to 16 cause-of-death ICD-10 codes
(1 underlying + 15 contributory).

**📊 IMD (Index of Multiple Deprivation)**
Neighbourhood-level socio-economic deprivation rankings (2015) mapped to
patient or practice LSOA. Quintile 1 = most deprived, Quintile 5 = least deprived.
            """)

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "🏥 HES APC (Hospital)", "🏢 HES OP (Outpatient)", "💀 ONS Death",
        "🔗 Linkage Eligibility", "📊 IMD (Deprivation)"
    ])

    def get_icd_codes(key_prefix):
        """Robust ICD code input helper with error protection."""
        try:
            source = st.radio("ICD-10 Code Source",
                              ["Built-in Disease Library (63 CVD)", "Enter Manually", "Upload Code List"],
                              horizontal=True, key=f"{key_prefix}_icd_src")
            codes = []
            if source == "Built-in Disease Library (63 CVD)":
                sel_groups = st.multiselect("Select disease groups", list(DISEASE_GROUPS.keys()),
                                             default=["Valvular Heart Disease"], key=f"{key_prefix}_groups")
                avail = []
                for g in sel_groups:
                    avail.extend(DISEASE_GROUPS[g])
                if avail:
                    sel = st.multiselect(f"Select diseases ({len(avail)} available)", avail, default=avail,
                                          key=f"{key_prefix}_diseases")
                    for d in sel:
                        codes.extend(DISEASE_CODE_LIBRARY.get(d, {}).get("icd10", []))
                if codes:
                    st.success(f"{len(set(codes))} ICD-10 codes selected")
                    with st.expander("View codes"):
                        for d in sel:
                            entry = DISEASE_CODE_LIBRARY.get(d, {})
                            st.write(f"**{d}** ({entry.get('short','')}) — {', '.join(entry.get('icd10',[]))}")
            elif source == "Enter Manually":
                text = st.text_area("Enter ICD-10 codes", placeholder="I05.0\nI06.0\nI35.0",
                                    key=f"{key_prefix}_manual")
                if text.strip():
                    codes = [c.strip() for c in text.replace(",", "\n").split("\n") if c.strip()]
            elif source == "Upload Code List":
                uploaded = st.file_uploader("Upload CSV/TXT/XLSX", type=["csv", "txt", "xlsx"],
                                            key=f"{key_prefix}_upload")
                if uploaded:
                    try:
                        if uploaded.name.endswith(".xlsx"):
                            udf = pd.read_excel(uploaded)
                        else:
                            udf = pd.read_csv(uploaded, dtype=str)
                        st.dataframe(udf.head())
                        col = st.selectbox("Select ICD-10 column", udf.columns.tolist(), key=f"{key_prefix}_col")
                        codes = udf[col].dropna().astype(str).str.strip().tolist()
                        st.success(f"{len(codes)} codes loaded")
                    except Exception as e:
                        st.error(f"Error reading file: {e}")
            return list(set(codes))
        except Exception as e:
            st.error(f"Error in code input: {e}")
            return []

    # ── TAB 1: HES APC ──
    with tab1:
        st.markdown("### HES Admitted Patient Care (Hospital Diagnoses)")
        show_tooltip("hes_apc")
        st.markdown(f'<div class="path-box">📂 {PATHS["hes_diagnosis_hosp"]}</div>', unsafe_allow_html=True)

        hes_source = st.selectbox("HES APC Source File", [
            "diagnosis_hosp (All hospital diagnoses)",
            "diagnosis_epi (Episode-level diagnoses)",
            "primary_diag (Primary diagnosis only)",
        ], key="hes_apc_file_select")
        source_key = hes_source.split(" ")[0]

        icd_codes = get_icd_codes("hes_apc")

        if icd_codes and st.button("🚀 Extract HES APC", key="btn_hes_apc", type="primary"):
            with st.spinner("Searching HES APC..."):
                try:
                    result = st.session_state.engine.extract_hes_apc(icd_codes, source=source_key)
                    if len(result) > 0:
                        result = add_icd10_descriptions(result)
                        result = apply_patient_scope(result)
                        _label = _build_save_label("HES_APC",
                            diseases=st.session_state.get("hes_apc_diseases"))
                        save_extraction(result, source_label=_label)
                        st.session_state["_res_hes_apc"] = result
                    else:
                        st.session_state.pop("_res_hes_apc", None)
                        st.warning("No matching records found.")
                except Exception as e:
                    st.error(f"HES APC extraction error: {e}")

        if "_res_hes_apc" in st.session_state and st.session_state["_res_hes_apc"] is not None:
            _r = st.session_state["_res_hes_apc"]
            st.markdown(f'<div class="success-box">✅ <strong>{len(_r):,} records</strong> · <strong>{_r["patid"].nunique():,} patients</strong></div>', unsafe_allow_html=True)
            st.dataframe(_r.head(100), width='stretch')
            download_results(_r, "hes_apc_extraction.csv", "link_hes_apc")
    # ── TAB 2: HES OP ──
    with tab2:
        st.markdown("### HES Outpatient Data")
        show_tooltip("hes_op")
        st.markdown(f'<div class="path-box">📂 {PATHS["hesop_clinical"]}</div>', unsafe_allow_html=True)
        st.info("Note: HES OP uses ICD codes **without dots** (e.g., I050 not I05.0). The tool handles this automatically.")

        icd_codes_op = get_icd_codes("hes_op")

        if icd_codes_op and st.button("🚀 Extract HES OP", key="btn_hes_op", type="primary"):
            with st.spinner("Searching HES OP..."):
                try:
                    result = st.session_state.engine.extract_hes_op(icd_codes_op)
                    if len(result) > 0:
                        result = add_icd10_descriptions(result)
                        result = apply_patient_scope(result)
                        _label = _build_save_label("HES_OP",
                            diseases=st.session_state.get("hes_op_diseases"))
                        save_extraction(result, source_label=_label)
                        st.session_state["_res_hes_op"] = result
                    else:
                        st.session_state.pop("_res_hes_op", None)
                        st.warning("No matching records found.")
                except Exception as e:
                    st.error(f"HES OP extraction error: {e}")

        if "_res_hes_op" in st.session_state and st.session_state["_res_hes_op"] is not None:
            _r = st.session_state["_res_hes_op"]
            st.markdown(f'<div class="success-box">✅ <strong>{len(_r):,} records</strong> · <strong>{_r["patid"].nunique():,} patients</strong></div>', unsafe_allow_html=True)
            st.dataframe(_r.head(100), width='stretch')
            download_results(_r, "hes_op_extraction.csv", "link_hes_op")
    # ── TAB 3: ONS Death ──
    with tab3:
        st.markdown("### ONS Death Records")
        show_tooltip("ons_death")
        st.markdown(f'<div class="path-box">📂 {PATHS["death_patient"]}</div>', unsafe_allow_html=True)
        st.markdown("Searches across all 16 cause-of-death columns (cause, cause1-cause15).")

        icd_codes_death = get_icd_codes("death")

        if icd_codes_death and st.button("🚀 Extract Death Records", key="btn_death", type="primary"):
            with st.spinner("Searching death records..."):
                try:
                    result = st.session_state.engine.extract_death_records(icd_codes_death)
                    if len(result) > 0:
                        result = add_icd10_descriptions(result)
                        result = apply_patient_scope(result)
                        _label = _build_save_label("ONS_Death",
                            diseases=st.session_state.get("death_diseases"))
                        save_extraction(result, source_label=_label)
                        st.session_state["_res_death"] = result
                    else:
                        st.session_state.pop("_res_death", None)
                        st.warning("No matching records found.")
                except Exception as e:
                    st.error(f"Death record extraction error: {e}")

        if "_res_death" in st.session_state and st.session_state["_res_death"] is not None:
            _r = st.session_state["_res_death"]
            st.markdown(f'<div class="success-box">✅ <strong>{len(_r):,} cause-of-death matches</strong> · <strong>{_r["patid"].nunique():,} patients</strong></div>', unsafe_allow_html=True)
            if "cause_position" in _r.columns:
                summary = _r.groupby("cause_position").size().reset_index(name="count")
                st.bar_chart(summary.set_index("cause_position"))
            st.dataframe(_r.head(100), width='stretch')
            download_results(_r, "death_extraction.csv", "link_death")
    # ── TAB 4: Linkage Eligibility ──
    with tab4:
        st.markdown("### Linkage Eligibility")
        show_tooltip("linkage_elig")
        st.markdown(f'<div class="path-box">📂 {PATHS["linkage_eligibility"]}</div>', unsafe_allow_html=True)

        if st.button("📋 Load Linkage Eligibility", key="btn_elig", type="primary"):
            with st.spinner("Loading..."):
                try:
                    result = st.session_state.engine.get_linkage_eligibility()
                    if len(result) > 0:
                        result = apply_patient_scope(result)
                        save_extraction(result, source_label="Linkage Eligibility")
                        st.session_state["_res_elig"] = result
                    else:
                        st.session_state.pop("_res_elig", None)
                        st.warning("No eligibility data found.")
                except Exception as e:
                    st.error(f"Eligibility load error: {e}")

        if "_res_elig" in st.session_state and st.session_state["_res_elig"] is not None:
            _r = st.session_state["_res_elig"]
            st.metric("Total Patients", f"{len(_r):,}")
            if "hes_e" in _r.columns:
                hes_elig = _r[_r["hes_e"].astype(str) == "1"]
                st.metric("HES Eligible", f"{len(hes_elig):,}")
            if "death_e" in _r.columns:
                death_elig = _r[_r["death_e"].astype(str) == "1"]
                st.metric("Death Eligible", f"{len(death_elig):,}")
            st.dataframe(_r.head(100), width='stretch')
            download_results(_r, "linkage_eligibility.csv", "link_elig")
    # ── TAB 5: IMD ──
    with tab5:
        st.markdown("### Index of Multiple Deprivation (IMD 2015)")
        show_tooltip("imd")
        st.markdown(f'<div class="path-box">📂 {PATHS["patient_imd"]}</div>', unsafe_allow_html=True)

        if st.button("📋 Load IMD Data", key="btn_imd", type="primary"):
            with st.spinner("Loading..."):
                try:
                    result = st.session_state.engine.get_patient_imd()
                    if len(result) > 0:
                        result = apply_patient_scope(result)
                        save_extraction(result, source_label="IMD")
                        st.session_state["_res_imd"] = result
                    else:
                        st.session_state.pop("_res_imd", None)
                        st.warning("No IMD data found.")
                except Exception as e:
                    st.error(f"IMD load error: {e}")

        if "_res_imd" in st.session_state and st.session_state["_res_imd"] is not None:
            _r = st.session_state["_res_imd"]
            st.metric("Patients with IMD", f"{len(_r):,}")
            if "imd2015_5" in _r.columns:
                fig = px.histogram(_r, x="imd2015_5", title="IMD Quintile Distribution",
                                   labels={"imd2015_5": "IMD Quintile (1=most deprived)"})
                fig.update_layout(template="plotly_white")
                st.plotly_chart(fig, width='stretch', key="plotly_1")
            st.dataframe(_r.head(100), width='stretch')
            download_results(_r, "imd_data.csv", "link_imd")
    # ── Footnote with all abbreviations ──
    st.markdown("---")
    st.markdown(LINKAGE_ABBREVIATIONS)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: MULTI-SOURCE SEARCH
# ══════════════════════════════════════════════════════════════════════════════

def page_multi_source():
    st.title("🔗 Multi-Source Extraction")
    st.markdown("Search across CPRD Aurum, HES APC, HES OP, and ONS Death simultaneously.")
    show_tooltip("nav_multi")

    # ── Patient Scope selector ──
    render_patient_scope(page_key="multi")

    st.markdown("### Step 1: Provide Your Code List")

    input_method = st.radio("How would you like to provide codes?",
                            ["Built-in Disease Library (63 CVD)", "Upload Code List File", "Enter Manually"],
                            horizontal=True, key="multi_input")
    snomed_codes = []
    icd_codes = []

    if input_method == "Built-in Disease Library (63 CVD)":
        # ── Step 1a: Select disease groups ──
        sel_groups = st.multiselect(
            "Select disease groups",
            list(DISEASE_GROUPS.keys()),
            default=["Ischaemic Heart Disease"],
            key="multi_disease_groups"
        )
        # Gather diseases from selected groups
        avail_diseases = []
        for g in sel_groups:
            avail_diseases.extend(DISEASE_GROUPS[g])
        # ── Step 1b: Select individual diseases ──
        if avail_diseases:
            sel_diseases = st.multiselect(
                f"Select diseases ({len(avail_diseases)} available)",
                avail_diseases,
                default=avail_diseases,
                key="multi_diseases"
            )
        else:
            sel_diseases = []

        # Collect codes from selected diseases
        for disease in sel_diseases:
            entry = DISEASE_CODE_LIBRARY.get(disease, {})
            snomed_codes.extend(entry.get("snomed", []))
            icd_codes.extend(entry.get("icd10", []))
        snomed_codes = list(set(snomed_codes))
        icd_codes = list(set(icd_codes))
        if sel_diseases:
            c1, c2, c3 = st.columns(3)
            c1.metric("Diseases selected", len(sel_diseases))
            c2.metric("SNOMED codes", len(snomed_codes))
            c3.metric("ICD-10 codes", len(icd_codes))
            with st.expander("View selected codes"):
                for disease in sel_diseases:
                    entry = DISEASE_CODE_LIBRARY.get(disease, {})
                    ns = len(entry.get("snomed", []))
                    ni = len(entry.get("icd10", []))
                    st.write(f"**{disease}** ({entry.get('short','')}) — {ns} SNOMED, {ni} ICD-10")

    elif input_method == "Upload Code List File":
        st.markdown("Upload a file containing your code list. The file should have columns for SNOMED and/or ICD-10 codes.")
        uploaded = st.file_uploader("Upload Code List (CSV/TXT/XLSX)",
                                    type=["csv", "txt", "xlsx"], key="multi_upload")
        if uploaded:
            try:
                if uploaded.name.endswith(".xlsx"):
                    udf = pd.read_excel(uploaded)
                else:
                    udf = pd.read_csv(uploaded, dtype=str)
                st.dataframe(udf.head(10))
                cols = udf.columns.tolist()
                col1, col2 = st.columns(2)
                with col1:
                    snomed_col = st.selectbox("SNOMED CT column (or None)", ["None"] + cols, key="multi_snomed_col")
                    if snomed_col != "None":
                        snomed_codes = udf[snomed_col].dropna().astype(str).str.strip().tolist()
                        st.success(f"{len(snomed_codes)} SNOMED codes loaded")
                with col2:
                    icd_col = st.selectbox("ICD-10 column (or None)", ["None"] + cols, key="multi_icd_col")
                    if icd_col != "None":
                        icd_codes = udf[icd_col].dropna().astype(str).str.strip().tolist()
                        st.success(f"{len(icd_codes)} ICD-10 codes loaded")
            except Exception as e:
                st.error(f"Error reading file: {e}")

    elif input_method == "Enter Manually":
        col1, col2 = st.columns(2)
        with col1:
            s_text = st.text_area("SNOMED CT codes (one per line)", key="multi_snomed_manual",
                                  placeholder="60573004\n86466006")
            if s_text.strip():
                snomed_codes = [c.strip() for c in s_text.strip().split("\n") if c.strip()]
        with col2:
            i_text = st.text_area("ICD-10 codes (one per line)", key="multi_icd_manual",
                                  placeholder="I05.0\nI06.0\nI35.0")
            if i_text.strip():
                icd_codes = [c.strip() for c in i_text.strip().split("\n") if c.strip()]

    st.markdown("### Step 2: Select Data Sources")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        use_aurum = st.checkbox("CPRD Aurum (SNOMED)", value=bool(snomed_codes), key="multi_use_aurum",
                                help=TOOLTIPS["snomed_codes"])
    with col2:
        use_hes_apc = st.checkbox("HES APC (ICD-10)", value=bool(icd_codes), key="multi_use_hes",
                                  help=TOOLTIPS["hes_apc"])
    with col3:
        use_hes_op = st.checkbox("HES OP (ICD-10)", value=bool(icd_codes), key="multi_use_op",
                                 help=TOOLTIPS["hes_op"])
    with col4:
        use_death = st.checkbox("ONS Death (ICD-10)", value=bool(icd_codes), key="multi_use_death",
                                help=TOOLTIPS["ons_death"])

    has_codes = bool(snomed_codes) or bool(icd_codes)
    has_sources = use_aurum or use_hes_apc or use_hes_op or use_death

    if not has_codes:
        st.warning("Please provide at least one SNOMED or ICD-10 code.")
    elif not has_sources:
        st.warning("Please select at least one data source.")

    if has_codes and has_sources and st.button("🚀 Run Multi-Source Extraction", key="btn_multi", type="primary"):
        all_results = {}

        if use_aurum and snomed_codes:
            st.markdown("#### CPRD Aurum")
            tracker = ProgressTracker("CPRD Aurum")
            aurum_result = st.session_state.engine.extract_observation_by_snomed(
                list(set(snomed_codes)), progress_callback=tracker.update)
            if aurum_result is not None and len(aurum_result) > 0:
                aurum_result["source"] = "CPRD Aurum"
                all_results["CPRD Aurum"] = aurum_result
                tracker.complete(f"Found {len(aurum_result):,} records")
            else:
                tracker.complete("No records")

        if use_hes_apc and icd_codes:
            with st.spinner("Searching HES APC..."):
                hes_result = st.session_state.engine.extract_hes_apc(list(set(icd_codes)))
                if len(hes_result) > 0:
                    hes_result["source"] = "HES APC"
                    all_results["HES APC"] = hes_result

        if use_hes_op and icd_codes:
            with st.spinner("Searching HES OP..."):
                op_result = st.session_state.engine.extract_hes_op(list(set(icd_codes)))
                if len(op_result) > 0:
                    op_result["source"] = "HES OP"
                    all_results["HES OP"] = op_result

        if use_death and icd_codes:
            with st.spinner("Searching ONS Death..."):
                death_result = st.session_state.engine.extract_death_records(list(set(icd_codes)))
                if len(death_result) > 0:
                    death_result["source"] = "ONS Death"
                    all_results["ONS Death"] = death_result

        if all_results:
            st.markdown("### Results Summary")
            summary_data = []
            all_patids = set()
            for src, df in all_results.items():
                n_records = len(df)
                n_patients = df["patid"].nunique()
                all_patids.update(df["patid"].unique())
                summary_data.append({"Source": src, "Records": n_records, "Unique Patients": n_patients})
            summary_data.append({"Source": "**TOTAL (unique)**",
                                 "Records": sum(len(df) for df in all_results.values()),
                                 "Unique Patients": len(all_patids)})
            st.dataframe(pd.DataFrame(summary_data), width='stretch', hide_index=True)

            if len(all_results) > 1:
                fig = go.Figure()
                for src, df in all_results.items():
                    fig.add_trace(go.Bar(name=src, x=[src], y=[df["patid"].nunique()]))
                fig.update_layout(title="Patients by Source", template="plotly_white", showlegend=True)
                st.plotly_chart(fig, width='stretch', key="plotly_2")

            st.session_state["multi_source_results"] = all_results
            combined = pd.concat(all_results.values(), ignore_index=True)
            combined = apply_patient_scope(combined)
            _label = _build_save_label("Multi_Source",
                diseases=st.session_state.get("multi_diseases"))
            save_extraction(combined, source_label=_label)
            st.session_state["_res_multi"] = combined
            st.session_state["_res_multi_all"] = all_results
        else:
            st.session_state.pop("_res_multi", None)
            st.session_state.pop("_res_multi_all", None)
            st.warning("No matching records found in any source.")

    # ── Persistent Multi-Source results ──
    if "_res_multi" in st.session_state and st.session_state["_res_multi"] is not None:
        all_results = st.session_state.get("_res_multi_all", {})
        combined = st.session_state["_res_multi"]

        st.markdown("### Results Summary")
        summary_data = []
        all_patids = set()
        for src, df in all_results.items():
            n_records = len(df)
            n_patients = df["patid"].nunique()
            all_patids.update(df["patid"].unique())
            summary_data.append({"Source": src, "Records": n_records, "Unique Patients": n_patients})
        summary_data.append({"Source": "**TOTAL (unique)**",
                             "Records": sum(len(df) for df in all_results.values()),
                             "Unique Patients": len(all_patids)})
        st.dataframe(pd.DataFrame(summary_data), width='stretch', hide_index=True)

        if len(all_results) > 1:
            fig = go.Figure()
            for src, df in all_results.items():
                fig.add_trace(go.Bar(name=src, x=[src], y=[df["patid"].nunique()]))
            fig.update_layout(title="Patients by Source", template="plotly_white", showlegend=True)
            st.plotly_chart(fig, width='stretch', key="plotly_3")

        st.markdown("### Download")
        download_results(combined, "multi_source_extraction.csv", "multi",
                         show_merge=True, all_results=all_results)

        if len(all_results) > 1:
            with st.expander("Download individual sources separately"):
                for src, df in all_results.items():
                    download_results(df, f"{src.lower().replace(' ', '_')}_extraction.csv",
                                     f"multi_{src.lower().replace(' ', '_')}")

        source_tabs = st.tabs(list(all_results.keys()))
        for tab, (src, df) in zip(source_tabs, all_results.items()):
            with tab:
                st.dataframe(df.head(100), width='stretch')
# ══════════════════════════════════════════════════════════════════════════════
# PAGE: COHORT BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def page_cohort_builder():
    st.title("👥 Cohort Builder")
    st.markdown("Build a patient cohort with inclusion/exclusion criteria, then extract data for that cohort.")
    show_tooltip("nav_cohort")

    if "cohort_steps" not in st.session_state:
        st.session_state.cohort_steps = []
    if "cohort_patients" not in st.session_state:
        st.session_state.cohort_patients = None
    if "cohort_start_df" not in st.session_state:
        st.session_state.cohort_start_df = None

    st.markdown("### Step 1: Define Starting Population")

    start_method = st.radio("Starting population",
                            ["All patients",
                             "From previous extraction",
                             "Upload patient list"],
                            key="cohort_start")

    start_df = st.session_state.cohort_start_df  # Persist across reruns

    if start_method == "All patients":
        if st.session_state.engine.is_mock():
            start_df = st.session_state.engine.mock_data["patient"]
            st.success(f"Mock data: {len(start_df):,} patients loaded with demographics (yob, gender, etc.)")
        else:
            st.markdown(
                "This will extract **all patient demographic data** from every practice folder. "
                "Patient files contain `yob`, `gender`, `regstartdate`, and other columns needed for cohort filtering."
            )
            if st.button("📥 Load All Patient Data", key="btn_load_all_patients", type="primary"):
                tracker = ProgressTracker("Loading Patient Demographics")
                result = st.session_state.engine.extract_any_filetype(
                    "Patient",
                    progress_callback=tracker.update,
                )
                tracker.complete()
                if result is not None and len(result) > 0:
                    start_df = result
                    st.session_state.cohort_start_df = start_df
                    st.success(f"✅ Loaded {len(start_df):,} patients with demographic data")
                else:
                    st.error("No patient data found. Check data paths in Configuration.")
            elif start_df is not None and "yob" in start_df.columns:
                st.success(f"✅ {len(start_df):,} patients already loaded (from previous load)")

    elif start_method == "From previous extraction":
        stats = get_extraction_stats()
        if "last_extraction" in st.session_state and st.session_state["last_extraction"] is not None:
            prev = st.session_state["last_extraction"]
            patids = prev["patid"].unique().tolist() if "patid" in prev.columns else []

            if not patids:
                st.warning("Previous extraction has no patient IDs.")
            else:
                st.success(f"Found {len(patids):,} unique patients from last extraction")

                # Check if previous extraction already has demographics
                has_demographics = all(c in prev.columns for c in ["yob", "gender"])
                if has_demographics:
                    start_df = prev.drop_duplicates(subset=["patid"]).copy()
                    st.info(f"Using demographics from previous extraction ({len(start_df):,} patients)")
                else:
                    st.markdown(
                        "⚠️ Previous extraction doesn't include demographic columns (`yob`, `gender`). "
                        "Loading patient demographics for these patients so you can filter by age, gender, etc."
                    )
                    if st.button("📥 Load Demographics for These Patients", key="btn_load_prev_demographics", type="primary"):
                        if st.session_state.engine.is_mock():
                            # Mock: get patient data and filter to matching patids
                            mock_patients = st.session_state.engine.mock_data["patient"]
                            start_df = mock_patients[mock_patients["patid"].isin(set(patids))].copy()
                        else:
                            tracker = ProgressTracker("Loading Patient Demographics")
                            result = st.session_state.engine.extract_any_filetype(
                                "Patient",
                                filter_col="patid",
                                filter_values=patids,
                                progress_callback=tracker.update,
                            )
                            tracker.complete()
                            if result is not None and len(result) > 0:
                                start_df = result
                            else:
                                # Fallback: create minimal DataFrame with just patids
                                start_df = pd.DataFrame({"patid": patids})
                                st.warning("Could not load demographics. Age/gender filters won't work.")

                        if start_df is not None:
                            st.session_state.cohort_start_df = start_df
                            has_yob = "yob" in start_df.columns
                            st.success(
                                f"✅ Loaded demographics for {len(start_df):,} patients"
                                + (" (includes yob, gender)" if has_yob else "")
                            )
                    elif start_df is not None and "yob" in start_df.columns:
                        st.success(f"✅ {len(start_df):,} patients with demographics (from previous load)")
        else:
            st.warning("No previous extraction available. Run an extraction first.")

    elif start_method == "Upload patient list":
        uploaded = st.file_uploader("Upload patient list (CSV/TXT)", type=["csv", "txt", "xlsx"],
                                    key="cohort_upload_patients")
        if uploaded:
            if uploaded.name.endswith(".xlsx"):
                start_df = pd.read_excel(uploaded, dtype=str)
            else:
                start_df = pd.read_csv(uploaded, dtype=str)
            st.session_state.cohort_start_df = start_df
            st.success(f"Loaded {len(start_df):,} rows")

            # Show available columns so user knows what's there
            has_yob = "yob" in start_df.columns
            has_gender = "gender" in start_df.columns
            col_info = []
            if has_yob:
                col_info.append("✅ yob (age filter available)")
            else:
                col_info.append("❌ yob (age filter unavailable)")
            if has_gender:
                col_info.append("✅ gender (gender filter available)")
            else:
                col_info.append("❌ gender (gender filter unavailable)")
            st.caption(" · ".join(col_info))
            st.dataframe(start_df.head())

    # ── Guard: no start population yet ──
    if start_df is None:
        st.info("Please select and load a starting population above to continue.")
        return

    # Show what we have
    cols_available = start_df.columns.tolist()
    has_yob = "yob" in cols_available
    has_gender = "gender" in cols_available
    has_regdate = "regstartdate" in cols_available

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Starting Population", f"{len(start_df):,} patients")
    c2.metric("Has YOB", "✅" if has_yob else "❌")
    c3.metric("Has Gender", "✅" if has_gender else "❌")
    c4.metric("Has Reg Date", "✅" if has_regdate else "❌")

    st.markdown("### Step 2: Apply Criteria")

    # Only show applicable criteria
    criteria_options = []
    if has_yob:
        criteria_options.append("Age Range")
    if has_gender:
        criteria_options.append("Gender")
    if has_regdate:
        criteria_options.append("Registration Period")
    criteria_options.extend([
        "Has Diagnosis (SNOMED codes in Observation)",
        "Has Drug (ProdCode in DrugIssue)",
        "Has HES APC Record (ICD-10)",
        "Linkage Eligible",
    ])

    if not has_yob and not has_gender:
        st.markdown(
            '<div class="warn-box">'
            '⚠️ <strong>Demographics not available.</strong> Age and gender filters require '
            'patient demographic data with <code>yob</code> and <code>gender</code> columns. '
            'Use "All patients" to load full demographics, or upload a file that includes these columns.'
            '</div>',
            unsafe_allow_html=True,
        )

    criteria_type = st.selectbox("Add Criterion", criteria_options, key="cohort_criteria_type")

    # Use current cohort (after previous filters) or start_df
    working_df = st.session_state.cohort_patients if st.session_state.cohort_patients is not None else start_df

    if criteria_type == "Age Range":
        col1, col2 = st.columns(2)
        with col1:
            min_age = st.number_input("Minimum Age", 0, 120, 18, key="cohort_min_age")
        with col2:
            max_age = st.number_input("Maximum Age", 0, 120, 100, key="cohort_max_age")
        ref_year = st.number_input("Reference Year", 2000, 2026, 2024, key="cohort_ref_year")
        if st.button("Apply Age Filter", key="btn_cohort_age"):
            df = working_df.copy()
            df["yob"] = pd.to_numeric(df["yob"], errors="coerce")
            df["age"] = ref_year - df["yob"]
            before = len(df)
            filtered = df[(df["age"] >= min_age) & (df["age"] <= max_age)]
            st.session_state.cohort_patients = filtered
            st.session_state.cohort_steps.append(f"Age {min_age}-{max_age} (ref {ref_year}): {len(filtered):,} patients")
            st.success(f"After age filter: {len(filtered):,} patients (removed {before - len(filtered):,})")

    elif criteria_type == "Gender":
        gender = st.selectbox("Select Gender", ["Male (1)", "Female (2)"], key="cohort_gender")
        gender_val = 1 if "Male" in gender else 2
        if st.button("Apply Gender Filter", key="btn_cohort_gender"):
            df = working_df.copy()
            df["gender"] = pd.to_numeric(df["gender"], errors="coerce")
            before = len(df)
            filtered = df[df["gender"] == gender_val]
            st.session_state.cohort_patients = filtered
            st.session_state.cohort_steps.append(f"Gender={gender}: {len(filtered):,} patients")
            st.success(f"After gender filter: {len(filtered):,} patients (removed {before - len(filtered):,})")

    elif criteria_type == "Registration Period":
        col1, col2 = st.columns(2)
        with col1:
            reg_start = st.date_input("Registered after", date(2000, 1, 1), key="cohort_reg_start")
        with col2:
            reg_end = st.date_input("Registered before", date(2020, 12, 31), key="cohort_reg_end")
        if st.button("Apply Registration Filter", key="btn_cohort_reg"):
            df = working_df.copy()
            df["regstartdate"] = pd.to_datetime(df["regstartdate"], errors="coerce")
            before = len(df)
            filtered = df[(df["regstartdate"] >= pd.Timestamp(reg_start)) &
                          (df["regstartdate"] <= pd.Timestamp(reg_end))]
            st.session_state.cohort_patients = filtered
            st.session_state.cohort_steps.append(f"Reg {reg_start} to {reg_end}: {len(filtered):,}")
            st.success(f"After registration filter: {len(filtered):,} patients (removed {before - len(filtered):,})")

    elif criteria_type == "Has Diagnosis (SNOMED codes in Observation)":
        diag_codes = st.text_area("Enter SNOMED codes (one per line)", key="cohort_snomed_filter")
        if diag_codes.strip() and st.button("Apply Diagnosis Filter", key="btn_cohort_diag"):
            codes = [c.strip() for c in diag_codes.strip().split("\n") if c.strip()]
            with st.spinner("Extracting observations..."):
                obs = st.session_state.engine.extract_observation_by_snomed(codes)
            if obs is not None and len(obs) > 0:
                dx_patids = set(obs["patid"].unique())
                before = len(working_df)
                filtered = working_df[working_df["patid"].isin(dx_patids)]
                st.session_state.cohort_patients = filtered
                st.session_state.cohort_steps.append(f"Has SNOMED dx ({len(codes)} codes): {len(filtered):,}")
                st.success(f"After diagnosis filter: {len(filtered):,} patients (removed {before - len(filtered):,})")
            else:
                st.warning("No matching observations found for these codes.")

    elif criteria_type == "Has Drug (ProdCode in DrugIssue)":
        drug_codes = st.text_area("Enter ProdCodeIds (one per line)", key="cohort_drug_filter")
        if st.session_state.engine.is_mock():
            if st.checkbox("Use all mock drug codes", key="cohort_drug_mock"):
                drug_codes = "\n".join(st.session_state.engine.mock_data["drug"]["prodcodeid"].unique()[:10])
        if drug_codes and drug_codes.strip() and st.button("Apply Drug Filter", key="btn_cohort_drug"):
            codes = [c.strip() for c in drug_codes.strip().split("\n") if c.strip()]
            with st.spinner("Extracting drug records..."):
                drugs = st.session_state.engine.extract_drugs_by_prodcode(codes)
            if drugs is not None and len(drugs) > 0:
                drug_patids = set(drugs["patid"].unique())
                before = len(working_df)
                filtered = working_df[working_df["patid"].isin(drug_patids)]
                st.session_state.cohort_patients = filtered
                st.session_state.cohort_steps.append(f"Has drug ({len(codes)} codes): {len(filtered):,}")
                st.success(f"After drug filter: {len(filtered):,} patients (removed {before - len(filtered):,})")
            else:
                st.warning("No matching drug records found.")

    elif criteria_type == "Has HES APC Record (ICD-10)":
        hes_codes = st.text_area("Enter ICD-10 codes (one per line)", key="cohort_hes_filter")
        if hes_codes.strip() and st.button("Apply HES Filter", key="btn_cohort_hes"):
            codes = [c.strip() for c in hes_codes.strip().split("\n") if c.strip()]
            with st.spinner("Extracting HES records..."):
                hes = st.session_state.engine.extract_hes_apc(codes)
            if len(hes) > 0:
                hes_patids = set(hes["patid"].unique())
                before = len(working_df)
                filtered = working_df[working_df["patid"].isin(hes_patids)]
                st.session_state.cohort_patients = filtered
                st.session_state.cohort_steps.append(f"Has HES dx ({len(codes)} codes): {len(filtered):,}")
                st.success(f"After HES filter: {len(filtered):,} patients (removed {before - len(filtered):,})")
            else:
                st.warning("No matching HES records found.")

    elif criteria_type == "Linkage Eligible":
        if st.button("Apply Linkage Eligibility", key="btn_cohort_linkage"):
            elig = st.session_state.engine.get_linkage_eligibility()
            if len(elig) > 0:
                elig_patids = set(elig["patid"].unique())
                before = len(working_df)
                filtered = working_df[working_df["patid"].isin(elig_patids)]
                st.session_state.cohort_patients = filtered
                st.session_state.cohort_steps.append(f"Linkage eligible: {len(filtered):,}")
                st.success(f"After linkage filter: {len(filtered):,} patients (removed {before - len(filtered):,})")
            else:
                st.warning("No linkage eligibility data found.")

    st.markdown("### Attrition Funnel")
    if st.session_state.cohort_steps:
        steps = [f"Starting: {len(start_df):,}"] + st.session_state.cohort_steps
        step_counts = [len(start_df)]
        for step in st.session_state.cohort_steps:
            try:
                count = int(step.split(": ")[-1].replace(",", "").split()[0])
                step_counts.append(count)
            except:
                step_counts.append(step_counts[-1])
        fig = go.Figure(go.Funnel(
            y=[s.split(":")[0] for s in steps],
            x=step_counts,
            textinfo="value+percent initial",
        ))
        fig.update_layout(title="Cohort Attrition", template="plotly_white", height=400)
        st.plotly_chart(fig, width='stretch', key="plotly_4")

    col_reset1, col_reset2 = st.columns(2)
    with col_reset1:
        if st.button("🔄 Reset Cohort Filters", key="btn_reset_cohort",
                     help="Clear all applied filters but keep the starting population"):
            st.session_state.cohort_steps = []
            st.session_state.cohort_patients = None
            st.rerun()
    with col_reset2:
        if st.button("🗑️ Reset Everything", key="btn_reset_cohort_all",
                     help="Clear starting population and all filters"):
            st.session_state.cohort_steps = []
            st.session_state.cohort_patients = None
            st.session_state.cohort_start_df = None
            st.rerun()

    if st.session_state.cohort_patients is not None and len(st.session_state.cohort_patients) > 0:
        st.markdown("### Export Cohort")
        cohort_df = st.session_state.cohort_patients
        st.metric("Final Cohort Size", f"{len(cohort_df):,} patients")
        download_results(cohort_df, "cohort_patients.csv", "cohort_export")
        save_extraction(cohort_df, source_label="Cohort Builder")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: CODE LIST DEVELOPMENT
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: DEMOGRAPHICS EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def page_demographics():
    st.title("👤 Demographics Extraction")
    st.markdown(
        "Extract a comprehensive demographic profile for **all patients** in the dataset. "
        "This combines data from CPRD Aurum Patient files, linked IMD deprivation scores, "
        "HES ethnicity, and practice-level region data."
    )

    st.markdown("### Select Variables")
    dem_vars = st.multiselect(
        "Variables to extract",
        ["Sex", "Year of Birth", "Age at Registration", "Registration Dates",
         "Practice Region", "Townsend / IMD Deprivation", "Ethnicity (HES)", "Acceptable Patient Flag"],
        default=["Sex", "Year of Birth", "Age at Registration", "Registration Dates",
                 "Practice Region", "Townsend / IMD Deprivation", "Ethnicity (HES)", "Acceptable Patient Flag"],
        key="demo_vars",
    )

    if st.button("🚀 Extract Demographics", key="btn_demo_extract", type="primary"):
        tracker = ProgressTracker("Demographics")

        # ── 1. Extract Patient file ──
        tracker.update(0.1, "Reading Patient files...")
        patient_df = st.session_state.engine.extract_file_type(
            "Patient", progress_callback=tracker.update
        )
        if patient_df is None or len(patient_df) == 0:
            st.session_state.pop("_res_demographics", None)
            st.error("No Patient data found. Check your CPRD data path.")
            return

        st.info(f"Loaded **{len(patient_df):,} patients** from {patient_df['pracid'].nunique()} practices")

        # ── 2. Build demographics table ──
        tracker.update(0.5, "Building demographics table...")
        demo = patient_df[["patid", "pracid"]].copy()

        # Sex
        if "Sex" in dem_vars and "gender" in patient_df.columns:
            demo["sex"] = patient_df["gender"].map({1: "Male", 2: "Female", 3: "Indeterminate", 0: "Unknown"})
            demo["sex"] = demo["sex"].fillna(patient_df["gender"].astype(str))

        # Year of birth
        if "Year of Birth" in dem_vars and "yob" in patient_df.columns:
            demo["year_of_birth"] = pd.to_numeric(patient_df["yob"], errors="coerce")

        # Registration dates
        if "Registration Dates" in dem_vars:
            for col in ["regstartdate", "regenddate"]:
                if col in patient_df.columns:
                    demo[col] = pd.to_datetime(patient_df[col], errors="coerce", dayfirst=True)

        # Age at registration
        if "Age at Registration" in dem_vars and "yob" in patient_df.columns and "regstartdate" in patient_df.columns:
            yob = pd.to_numeric(patient_df["yob"], errors="coerce")
            reg = pd.to_datetime(patient_df["regstartdate"], errors="coerce", dayfirst=True)
            demo["age_at_registration"] = reg.dt.year - yob
            demo.loc[demo["age_at_registration"] < 0, "age_at_registration"] = np.nan
            demo.loc[demo["age_at_registration"] > 120, "age_at_registration"] = np.nan

        # Acceptable flag
        if "Acceptable Patient Flag" in dem_vars and "acceptable" in patient_df.columns:
            demo["acceptable"] = patient_df["acceptable"].astype(str)

        # ── 3. Practice region ──
        if "Practice Region" in dem_vars:
            tracker.update(0.6, "Reading Practice files for region...")
            practice_df = st.session_state.engine.extract_file_type(
                "Practice", progress_callback=lambda p, m: None
            )
            if practice_df is not None and "region" in practice_df.columns:
                region_map = {
                    1: "North East", 2: "North West", 3: "Yorkshire & Humber",
                    4: "East Midlands", 5: "West Midlands", 6: "East of England",
                    7: "South West", 8: "South Central", 9: "London",
                    10: "South East Coast", 11: "Northern Ireland", 12: "Scotland", 13: "Wales",
                }
                practice_df["region_name"] = pd.to_numeric(practice_df["region"], errors="coerce").map(region_map)
                demo = demo.merge(
                    practice_df[["pracid", "region_name"]].drop_duplicates(),
                    on="pracid", how="left"
                )

        # ── 4. IMD / Townsend ──
        if "Townsend / IMD Deprivation" in dem_vars:
            tracker.update(0.7, "Reading IMD linkage data...")
            if st.session_state.engine.is_mock():
                demo["imd_quintile"] = np.random.choice([1, 2, 3, 4, 5], size=len(demo))
            else:
                imd_path = PATHS.get("patient_imd", "")
                if os.path.isfile(imd_path):
                    try:
                        imd_df = pd.read_csv(imd_path, sep="\t", dtype=str, usecols=["patid", "imd2015_5"])
                        imd_df = imd_df.rename(columns={"imd2015_5": "imd_quintile"})
                        demo = demo.merge(imd_df, on="patid", how="left")
                    except Exception as e:
                        st.warning(f"Could not read IMD data: {e}")
                else:
                    st.warning(f"IMD file not found: {imd_path}")

        # ── 5. Ethnicity (HES) ──
        if "Ethnicity (HES)" in dem_vars:
            tracker.update(0.8, "Reading HES ethnicity data...")
            if st.session_state.engine.is_mock():
                eth_codes = ["A", "B", "C", "D", "E", "F", "G", "H", "J", "K", "L", "M", "N", "P", "R", "S", "Z"]
                demo["ethnicity_hes"] = np.random.choice(eth_codes, size=len(demo))
            else:
                hes_patient_path = PATHS.get("hes_patient", "")
                if os.path.isfile(hes_patient_path):
                    try:
                        hes_pat = pd.read_csv(hes_patient_path, sep="\t", dtype=str, usecols=["patid", "gen_ethnicity"])
                        hes_pat = hes_pat.rename(columns={"gen_ethnicity": "ethnicity_hes"})
                        hes_pat = hes_pat.drop_duplicates(subset="patid")
                        demo = demo.merge(hes_pat, on="patid", how="left")
                    except Exception as e:
                        st.warning(f"Could not read HES patient data: {e}")
                else:
                    st.warning(f"HES patient file not found: {hes_patient_path}")

            # Map ethnicity codes
            eth_map = {
                "A": "White - British", "B": "White - Irish", "C": "White - Other",
                "D": "Mixed - White & Black Caribbean", "E": "Mixed - White & Black African",
                "F": "Mixed - White & Asian", "G": "Mixed - Other",
                "H": "Asian - Indian", "J": "Asian - Pakistani", "K": "Asian - Bangladeshi", "L": "Asian - Other",
                "M": "Black - Caribbean", "N": "Black - African", "P": "Black - Other",
                "R": "Chinese", "S": "Other", "Z": "Not stated", "99": "Not known",
            }
            if "ethnicity_hes" in demo.columns:
                demo["ethnicity_group"] = demo["ethnicity_hes"].map(eth_map)

        tracker.complete()

        # ── Save and display ──
        save_extraction(demo, source_label="Demographics")
        st.session_state["_res_demographics"] = demo

    # ── Persistent demographics results ──
    if "_res_demographics" in st.session_state and st.session_state["_res_demographics"] is not None:
        demo = st.session_state["_res_demographics"]

        st.markdown(f'<div class="success-box">✅ <strong>{len(demo):,} patients</strong> with '
                    f'<strong>{len(demo.columns)} variables</strong></div>', unsafe_allow_html=True)

        # ── Summary statistics ──
        st.markdown("### Summary")
        sc1, sc2, sc3, sc4 = st.columns(4)
        if "sex" in demo.columns:
            sc1.metric("Female", f"{(demo['sex']=='Female').sum():,}")
            sc1.metric("Male", f"{(demo['sex']=='Male').sum():,}")
        if "age_at_registration" in demo.columns:
            sc2.metric("Median Age at Reg.", f"{demo['age_at_registration'].median():.0f}")
            sc2.metric("Mean Age at Reg.", f"{demo['age_at_registration'].mean():.1f}")
        if "imd_quintile" in demo.columns:
            sc3.metric("IMD Available", f"{demo['imd_quintile'].notna().sum():,}")
        if "ethnicity_hes" in demo.columns:
            sc4.metric("Ethnicity Available", f"{demo['ethnicity_hes'].notna().sum():,}")

        if "acceptable" in demo.columns:
            n_acc = (demo["acceptable"] == "1").sum()
            st.info(f"Acceptable patients: **{n_acc:,}** / {len(demo):,} ({100*n_acc/len(demo):.1f}%)")

        # ── Distribution charts ──
        if "sex" in demo.columns:
            with st.expander("Sex Distribution", expanded=True):
                sex_counts = demo["sex"].value_counts()
                fig = go.Figure(go.Pie(labels=sex_counts.index, values=sex_counts.values,
                                       marker=dict(colors=["#4472C4", "#ED7D31", "#A5A5A5"])))
                fig.update_layout(height=300, margin=dict(t=20, b=20))
                st.plotly_chart(fig, use_container_width=True, key="plotly_5")

        if "age_at_registration" in demo.columns:
            with st.expander("Age at Registration Distribution", expanded=True):
                fig = go.Figure(go.Histogram(x=demo["age_at_registration"].dropna(),
                                             nbinsx=50, marker_color="#4472C4"))
                fig.update_layout(xaxis_title="Age (years)", yaxis_title="Count",
                                  height=300, margin=dict(t=20, b=40))
                st.plotly_chart(fig, use_container_width=True, key="plotly_6")

        if "imd_quintile" in demo.columns:
            with st.expander("IMD Deprivation Quintile Distribution"):
                imd_counts = demo["imd_quintile"].value_counts().sort_index()
                fig = go.Figure(go.Bar(x=imd_counts.index.astype(str), y=imd_counts.values,
                                       marker_color=["#2166ac", "#67a9cf", "#d1e5f0", "#fddbc7", "#ef8a62"]))
                fig.update_layout(xaxis_title="IMD Quintile (1=most deprived)", yaxis_title="Count",
                                  height=300, margin=dict(t=20, b=40))
                st.plotly_chart(fig, use_container_width=True, key="plotly_7")

        if "ethnicity_group" in demo.columns:
            with st.expander("Ethnicity Distribution"):
                eth_counts = demo["ethnicity_group"].value_counts().head(15)
                fig = go.Figure(go.Bar(x=eth_counts.values, y=eth_counts.index, orientation="h",
                                       marker_color="#4472C4"))
                fig.update_layout(xaxis_title="Count", height=400, margin=dict(t=20, l=200, b=40))
                st.plotly_chart(fig, use_container_width=True, key="plotly_8")

        if "region_name" in demo.columns:
            with st.expander("Practice Region Distribution"):
                reg_counts = demo["region_name"].value_counts()
                fig = go.Figure(go.Bar(x=reg_counts.values, y=reg_counts.index, orientation="h",
                                       marker_color="#002147"))
                fig.update_layout(xaxis_title="Patients", height=350, margin=dict(t=20, l=180, b=40))
                st.plotly_chart(fig, use_container_width=True, key="plotly_9")

        # ── Data table and download ──
        st.markdown("### Full Dataset")
        st.dataframe(demo.head(200), height=400)
        download_results(demo, "cprd_demographics.csv", "demo")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: CODE LIST DEVELOPMENT
# ══════════════════════════════════════════════════════════════════════════════

def page_code_list_dev():
    st.title("📋 Code List Development")
    st.markdown(
        "Systematic pipeline for developing validated clinical code lists for CPRD research, "
        "following [CPRD Guidance v7.1](https://cprd.com) and the "
        "[Matthewman et al. 10-step framework](https://doi.org/10.3310/nihropenres.13498.1)."
    )

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📘 Stage 1 — Define",
        "📗 Stage 2 — Synonyms & Existing Lists",
        "📙 Stage 3 — Search Code Browsers",
        "📕 Stage 4 — Review & Classify",
        "🔧 Stage 5 — Dictionary Matching",
        "🩺 Stage 6 — Clinical Review",
    ])

    # ── STAGE 1 ──────────────────────────────────────────────────────────────
    with tab1:
        st.header("Stage 1: Define the Clinical Feature of Interest")
        st.markdown("""
**Purpose:** Clearly define the clinical concept you want to identify in the data *before* searching for codes.

**Key principles (CPRD Guidance §1, §3):**

"Clinical code set construction is the process of assembling a set of clinical codes that represent a single
clinical concept such as a diagnosis, a procedure, an observation, or a medication" (Williams et al. 2017).
Errors in code lists — omitting important codes or including irrelevant ones — cause selection biases
that can have a major impact on analyses.

**What to define:**

- **Inclusion criteria:** Which diagnoses, symptoms, test results, or prescriptions define the condition?
  Should prescriptions for relevant medications count? Should abnormal test results count?
- **Exclusion criteria:** What should *not* be included? E.g., should "absence of cough" codes be included
  in a cough code list? Should anorexia nervosa be included in an appetite-loss code list?
- **Clinical involvement:** If possible, involve a clinician in defining the concept. They can identify edge
  cases and ensure the definition is clinically meaningful.

**Useful resources:**

- **ICPC-3 Browser** — provides definitions of symptoms and diagnoses with ICD-10 and SNOMED CT codes
- **Published definitions** — review how previous studies defined the same condition
- **QOF business rules** — contain well-validated code lists for common conditions

**Output of Stage 1:** A written definition of the clinical concept with explicit inclusion/exclusion criteria,
agreed with a clinician where possible. This becomes the reference document for all subsequent code searching.
        """)

    # ── STAGE 2 ──────────────────────────────────────────────────────────────
    with tab2:
        st.header("Stage 2: Create Synonyms & Identify Existing Lists")
        st.markdown("""
**Purpose:** Build a comprehensive list of synonyms for your condition, and search for existing code lists
that can serve as a starting point.

---

### 2A. Synonym Identification (CPRD Guidance §4)

Synonyms are alternative names for the same clinical concept. For example, "haemoptysis" has synonyms
including "coughing up blood", "blood in sputum", and "hemoptysis" (US spelling).

**Where to find synonyms:**

- **NHS SNOMED CT Browser** ([termbrowser.nhs.uk](https://termbrowser.nhs.uk)) — each concept shows its synonyms in the blue box
- **ICPC-3 Browser** — provides structured synonym lists
- **UMLS Metathesaurus Browser** ([nih.gov](https://uts.nlm.nih.gov/uts/umls)) — comprehensive cross-terminology synonym database
- **CPRD Code Browser** — browsing related terms in the Read Code hierarchy reveals synonyms
- **Existing code lists** — reviewing what others included identifies terms you may have missed

**Tip:** Have a clinician review the synonym list to catch clinical jargon and abbreviations.

---

### 2B. Identify Existing Code Lists (CPRD Guidance §5)

Search for published code lists in these repositories:

| Repository | URL | Notes |
|---|---|---|
| LSHTM Data Compass | [datacompass.lshtm.ac.uk](https://datacompass.lshtm.ac.uk) | Large collection of CPRD code lists |
| HDR UK Phenotype Library | [phenotypes.healthdatagateway.org](https://phenotypes.healthdatagateway.org) | National phenotype repository |
| ClinicalCodes | [clinicalcodes.org](https://clinicalcodes.org) | Manchester-based repository |
| CPRD @ Cambridge | [phpc.cam.ac.uk/pcu/cprd](https://www.phpc.cam.ac.uk/pcu/cprd_cam/codelists/) | Gold code lists from Cambridge PCU |
| OpenCodelists | [opencodelists.org](https://www.opencodelists.org) | OpenSAFELY / Bennett Institute |
| QOF Business Rules | [digital.nhs.uk](https://digital.nhs.uk) | Well-validated for QOF conditions |
| OHDSI ATLAS | [atlas-demo.ohdsi.org](https://atlas-demo.ohdsi.org) | SNOMED-based concept sets |

**Important caveats about existing lists:**

- Code lists for the same condition can vary *widely* depending on how broadly the condition was defined
- Published lists may contain errors (e.g., codes for *absence* of a condition)
- Best practice: **review multiple lists**, check all codes, and supplement with your own searches
        """)

    # ── STAGE 3 ──────────────────────────────────────────────────────────────
    with tab3:
        st.header("Stage 3: Search Code Browsers")
        st.markdown("""
**Purpose:** Systematically search the CPRD code browsers and relevant classification systems using your
synonym list to find all candidate codes.

---

### 3A. Coding Systems in CPRD (CPRD Guidance §6)

| System | Used in | Identifier | Notes |
|---|---|---|---|
| **SNOMED CT** | CPRD Aurum (Observation) | `medcodeid` | Current NHS standard. Concepts have multiple descriptions. |
| **Read V2** | CPRD Gold (Clinical, Referral) | `medcode` | Legacy system, superseded by SNOMED CT ~2018. |
| **ICD-10** | HES APC, HES OP, ONS Deaths | ICD-10 code | Hospital diagnoses & death certificates. HES uses dots, cancer registry doesn't. |
| **BNF** | Drug prescribing | BNF code | Chapter/section/paragraph structure for medications. |
| **Dm+d** | Medicines dictionary | Dm+d code | NHS standard for medicines and devices. |

---

### 3B. Searching the CPRD Code Browser (CPRD Guidance §7)

**Key tips:**

- Select the correct **search field** (Read Term, Read Code, or medcode for Gold; SNOMED for Aurum)
- Select the correct **dictionary** (Medical for diagnoses, Product for drugs)
- **Use wildcards:** Enclose search terms with `*` for partial matching (e.g., `*haemoptysis*`)
- **Search hierarchies:** If "R063" appears in results for haemoptysis, search `R063*` for related codes
- **Read hierarchies include negation codes** — e.g., under "Cough" hierarchy you'll find "No Cough". Exclude these.

**For medications (CPRD Guidance §7.3):**

- Use BNF codes: In Aurum Browser, omit the first `0` and use `*` for the last two characters (e.g., `403030*`)
- In Gold Browser, keep the first `0` (e.g., `0403030*`)
- Also search by **Drug Substance Name** and **Product Name** — not all medications have BNF codes

---

### 3C. Searching Dictionary Files Directly (CPRD Guidance §7.5)

The dictionary files (`CPRDAurumMedical.txt` / `medical.txt`) can be read directly in R or Python
and searched programmatically. Example R code for this is in CPRD Guidance Appendices 2-4.

**⚡ This app's Stage 5 tab automates this process** — you provide your candidate SNOMED CT codes,
and the app matches them against the EMIS dictionary to find all corresponding medcodeids.

---

### 3D. ICD-10 Code Format Warning (CPRD Guidance §6.1)

> **HES data includes the decimal point** (e.g., `C34.1`)
> **Cancer registry data does NOT include the decimal point** (e.g., `C341`)
> This app's extraction engine handles this automatically.

**Output of Stage 3:** A list of candidate codes in each relevant coding system (SNOMED CT, ICD-10,
Read V2, BNF) generated from systematic searches using your synonym list.
        """)

    # ── STAGE 4 ──────────────────────────────────────────────────────────────
    with tab4:
        st.header("Stage 4: Review, Classify & Cross-Map Codes")
        st.markdown("""
**Purpose:** Review all candidate codes for clinical appropriateness, classify them by subgroup, and
create cross-system mappings. This is the final manual curation step before dictionary matching.

---

### 4A. Reviewing Code Lists (CPRD Guidance §9)

Every candidate code must be reviewed to ensure it is suitable:

- **Remove negation codes:** Searches for `*cough*` will find "Night cough absent" — exclude these
- **Remove screening codes:** Searches for `*diabetes*` may find non-diagnostic screening codes
- **Check Read code hierarchies:** Related codes may appear that are off-topic
- **Involve a clinician:** They can identify codes that are clinically inappropriate or missing

---

### 4B. Classify by Clinical Subgroup

Organise your codes into clinically meaningful subgroups. For example, for VHD:

- Aortic Stenosis, Aortic Regurgitation, Mixed Aortic, Aortic Valve Disease NOS
- Mitral Stenosis, Mitral Regurgitation, Mixed Mitral, Mitral Valve Prolapse
- Tricuspid Stenosis, Tricuspid Regurgitation
- Pulmonary Stenosis, Pulmonary Regurgitation
- Multi-valve Disease, General VHD, Prosthetic Valve, Congenital

Tag each code with: **etiology** (rheumatic, non-rheumatic, congenital, any),
**severity** (mild, moderate, severe, ungraded), and **decision flags** for codes needing clinical adjudication.

---

### 4C. Cross-System Mapping

Map codes between classification systems using:

- **SNOMED CT to ICD-10:** NHS TRUD crosswalk tables, or SNOMED International mapping files
- **Read V2 to SNOMED CT:** UK Data Migration Reference Set (DMRS)
- **Manual semantic matching:** Where automated maps don't exist, match by clinical description equivalence

**Document one-to-many mappings** — e.g., SNOMED "Aortic valve stenosis" (etiology-agnostic) maps to
*both* ICD-10 I06.0 (rheumatic) and I35.0 (non-rheumatic).

---

### 4D. Consider Event Frequency (CPRD Guidance §10)

The CPRD Gold code browser shows total event counts per medcode. This is useful because:

- Codes with **zero or very few events** can be deprioritised (saves review time)
- Codes with **very high frequency** have high potential to introduce bias if erroneously included

---

### 4E. Validation Planning (CPRD Guidance §11)

Plan validation approaches:

- **Sensitivity analyses:** Compare patient counts with broader vs. narrower code lists against known prevalence (QOF data, BHF statistics, Cancer Research UK)
- **Concordance checks:** Compare cases identified in primary care vs. HES vs. ONS
- **Association checks:** Verify expected associations (e.g., smoking ↔ CHD, age ↔ chronic conditions)

**Output of Stage 4:** A curated, classified candidate code list ready for dictionary matching.
This is the input to **Stage 5** →
        """)

    # ── STAGE 5 ──────────────────────────────────────────────────────────────
    with tab5:
        _render_stage5_matching()

    # ── STAGE 6 ──────────────────────────────────────────────────────────────
    with tab6:
        _render_stage6_clinical_review()


def _render_stage5_matching():
    """Stage 5: Interactive CPRD dictionary matching (Python port of R script)."""

    st.header("Stage 5: CPRD Dictionary Matching")
    st.markdown(
        "Match your candidate SNOMED CT codes from Stage 4 to CPRD Aurum **medcodeids** using the EMIS "
        "Medical Dictionary, then build a unified master code list combining SNOMED CT, ICD-10, and Read V2. "
        "This replaces the manual R script workflow with an automated, interactive process."
    )

    # ── State initialisation ──
    for key in ["s5_dict_df", "s5_snomed_input", "s5_icd_input", "s5_matched",
                "s5_unmatched", "s5_master", "s5_log", "s5_expansion"]:
        if key not in st.session_state:
            st.session_state[key] = None

    # ══════════════════════════════════════════════════════════════════════
    # STEP A: Load EMIS Medical Dictionary
    # ══════════════════════════════════════════════════════════════════════
    st.markdown("### Step A: Load EMIS Medical Dictionary")
    st.markdown(
        "The dictionary file (`CPRDAurumMedical.txt` or `EMISMedicalDictionary.txt`) maps "
        "SNOMED CT concept IDs to CPRD Aurum medcodeids. It is typically located in the lookups folder."
    )

    dict_source = st.radio(
        "Dictionary source",
        ["Auto-detect from CPRD data path", "Upload dictionary file", "Use mock dictionary (testing)"],
        key="s5_dict_source",
    )

    dict_df = st.session_state.s5_dict_df

    if dict_source == "Auto-detect from CPRD data path":
        base = st.session_state.get("cprd_base_path", "/gpfs3/well/rahimi/projects/CPRD/")
        lookup_patterns = [
            os.path.join(base, "202102_lookups", "202102_EMISMedicalDictionary.txt"),
            os.path.join(base, "202102_lookups", "CPRDAurumMedical.txt"),
            os.path.join(base, "lookups", "EMISMedicalDictionary.txt"),
            os.path.join(base, "lookups", "CPRDAurumMedical.txt"),
        ]
        found_path = None
        for p in lookup_patterns:
            if os.path.exists(p):
                found_path = p
                break

        if found_path:
            st.success(f"Found dictionary: `{found_path}`")
            if st.button("📥 Load Dictionary", key="btn_s5_load_dict", type="primary"):
                with st.spinner("Loading EMIS Medical Dictionary (this may take a moment for large files)..."):
                    dict_df = pd.read_csv(found_path, sep="\t", dtype=str, low_memory=False)
                    st.session_state.s5_dict_df = dict_df
        elif dict_df is not None:
            st.success(f"Dictionary already loaded: {len(dict_df):,} rows")
        else:
            st.warning(
                f"No dictionary file found in `{base}`. "
                "Try uploading it manually or check your data path in Configuration."
            )

    elif dict_source == "Upload dictionary file":
        uploaded_dict = st.file_uploader(
            "Upload EMIS Medical Dictionary (.txt, tab-separated)",
            type=["txt", "csv", "tsv"], key="s5_dict_upload",
        )
        if uploaded_dict:
            with st.spinner("Parsing dictionary file..."):
                dict_df = pd.read_csv(uploaded_dict, sep="\t", dtype=str, low_memory=False)
                st.session_state.s5_dict_df = dict_df

    elif dict_source == "Use mock dictionary (testing)":
        if st.button("Generate Mock Dictionary", key="btn_s5_mock_dict"):
            # Create a realistic mock dictionary with known SNOMED CT codes
            import random
            mock_rows = []
            # Add some real VHD SNOMED concepts with multiple medcodeids each
            vhd_concepts = [
                ("60573004", "Aortic valve stenosis"), ("60573004", "Aortic stenosis"),
                ("60234000", "Aortic valve regurgitation"), ("60234000", "Aortic regurgitation"),
                ("60234000", "Aortic valve incompetence"),
                ("79619009", "Mitral valve stenosis"), ("79619009", "Mitral stenosis"),
                ("48724000", "Mitral valve regurgitation"), ("48724000", "Mitral regurgitation"),
                ("48724000", "Mitral valve incompetence"), ("48724000", "Mitral insufficiency"),
                ("409712001", "Mitral valve prolapse"),
                ("56786000", "Pulmonary valve stenosis"),
                ("91434003", "Pulmonary valve regurgitation"),
                ("111287006", "Tricuspid valve regurgitation"), ("111287006", "Tricuspid regurgitation"),
                ("49915006", "Tricuspid valve stenosis"),
                ("72011007", "Rheumatic aortic stenosis"),
                ("78031003", "Rheumatic aortic regurgitation"),
                ("86466006", "Rheumatic mitral stenosis"),
                ("31085000", "Rheumatic mitral regurgitation"),
                ("8722008", "Aortic valve disorder"), ("8722008", "Aortic valve disease"),
                ("83916000", "Rheumatic mitral valve disease"),
                ("472847005", "Mixed aortic valve disease"),
                ("24211005", "Prosthetic heart valve present"),
                ("368009", "Heart valve disorder"), ("368009", "Heart valve disease"),
                ("62067003", "Hypoplastic left heart syndrome"),
                ("204357006", "Bicuspid aortic valve"),
                ("424031003", "Aortic valve sclerosis"),
                ("250978003", "Calcification of aortic valve"),
                ("787001", "Rheumatic mitral stenosis with regurgitation"),
                ("194733006", "Rheumatic mitral and aortic valve disease"),
                ("17759006", "Rheumatic aortic stenosis with regurgitation"),
                ("194984004", "Non-rheumatic aortic regurgitation"),
                ("708966001", "Non-rheumatic mitral valve disease"),
                ("82458004", "Congenital mitral stenosis"),
                ("253545000", "Congenital aortic stenosis"),
                ("78495000", "Congenital mitral regurgitation"),
                ("63042009", "Ebstein anomaly"),
                ("233917008", "Pulmonary atresia"),
                ("11851006", "Mitral annular calcification"),
                ("16440002", "Rheumatic aortic valve disease"),
                ("398995000", "Valvular heart disease"),
            ]
            for i, (sctid, term) in enumerate(vhd_concepts):
                mock_rows.append({
                    "MedCodeId": str(1000000 + i),
                    "CleansedReadCode": f"G5{i%10}{i//10}.00" if i % 3 == 0 else "",
                    "SnomedCTConceptId": sctid,
                    "SnomedCTDescriptionId": str(2000000 + i),
                    "Release": "20241001",
                    "Term": term,
                    "OriginalReadCode": "",
                    "EmisCodeCategoryId": "1",
                })
            # Add some non-VHD filler rows
            for j in range(200):
                mock_rows.append({
                    "MedCodeId": str(2000000 + j),
                    "CleansedReadCode": "",
                    "SnomedCTConceptId": str(random.randint(100000000, 999999999)),
                    "SnomedCTDescriptionId": str(random.randint(100000000, 999999999)),
                    "Release": "20241001",
                    "Term": f"Mock condition {j}",
                    "OriginalReadCode": "",
                    "EmisCodeCategoryId": "1",
                })
            dict_df = pd.DataFrame(mock_rows)
            st.session_state.s5_dict_df = dict_df
            st.success(f"Mock dictionary generated: {len(dict_df):,} rows ({len(vhd_concepts)} VHD entries)")

    # Show dictionary info if loaded
    if dict_df is not None:
        c1, c2, c3 = st.columns(3)
        c1.metric("Dictionary Rows", f"{len(dict_df):,}")
        c2.metric("Unique MedCodeIds", f"{dict_df['MedCodeId'].nunique():,}" if "MedCodeId" in dict_df.columns else "N/A")
        c3.metric("Unique SNOMED Concepts", f"{dict_df['SnomedCTConceptId'].nunique():,}" if "SnomedCTConceptId" in dict_df.columns else "N/A")

        with st.expander("Preview dictionary columns"):
            st.dataframe(dict_df.head(5))

        # Validate expected columns
        expected = ["MedCodeId", "SnomedCTConceptId", "Term"]
        missing = [c for c in expected if c not in dict_df.columns]
        if missing:
            st.error(f"⚠️ Missing expected columns: {', '.join(missing)}. Available: {', '.join(dict_df.columns.tolist())}")
            st.info("The column names in your dictionary may differ. Common alternatives: `medcodeid` → `MedCodeId`, `snomedctconceptid` → `SnomedCTConceptId`")
            # Attempt case-insensitive column matching
            col_map = {}
            for exp in missing:
                for actual in dict_df.columns:
                    if actual.lower().replace("_", "") == exp.lower().replace("_", ""):
                        col_map[actual] = exp
            if col_map:
                st.info(f"Auto-detected column mappings: {col_map}. Renaming...")
                dict_df = dict_df.rename(columns=col_map)
                st.session_state.s5_dict_df = dict_df
    else:
        st.info("Load the EMIS Medical Dictionary above to proceed.")

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════════
    # STEP B: Input Candidate Codes from Stage 4
    # ══════════════════════════════════════════════════════════════════════
    st.markdown("### Step B: Input Candidate Codes from Stage 4")

    input_tabs = st.tabs(["📤 Upload Master CSV", "🧬 SNOMED CT Codes", "🏥 ICD-10 Codes"])

    # ── SNOMED CT input ──
    with input_tabs[1]:
        st.markdown("Enter your SNOMED CT candidate codes. Each row needs: **SNOMED CT Concept ID**, **Category**, and optionally **Severity** and **Source**.")
        snomed_method = st.radio("Input method", ["Paste CSV/table", "Upload CSV file", "Use built-in VHD codes (55 concepts)"], key="s5_snomed_method")

        if snomed_method == "Paste CSV/table":
            snomed_text = st.text_area(
                "Paste SNOMED codes (CSV format: SnomedCTConceptId,category,Severity,Source)",
                height=200, key="s5_snomed_text",
                placeholder="SnomedCTConceptId,category,Severity,Source\n60573004,Aortic Stenosis,Ungraded parent,Existing\n836480008,Aortic Stenosis - Mild,DMRS Mild,NEW-DMRS",
            )
            if snomed_text.strip() and st.button("Parse SNOMED codes", key="btn_s5_parse_snomed"):
                try:
                    snomed_df = pd.read_csv(io.StringIO(snomed_text.strip()), dtype=str)
                    # Normalise column names
                    snomed_df.columns = [c.strip() for c in snomed_df.columns]
                    if "SnomedCTConceptId" not in snomed_df.columns:
                        # Try to find it
                        for col in snomed_df.columns:
                            if "snomed" in col.lower() or "concept" in col.lower() or "code" in col.lower():
                                snomed_df = snomed_df.rename(columns={col: "SnomedCTConceptId"})
                                break
                    st.session_state.s5_snomed_input = snomed_df
                    st.success(f"Parsed {len(snomed_df)} SNOMED CT codes")
                except Exception as e:
                    st.error(f"Failed to parse: {e}")

        elif snomed_method == "Upload CSV file":
            snomed_file = st.file_uploader("Upload SNOMED candidate CSV", type=["csv", "xlsx", "txt"], key="s5_snomed_file")
            if snomed_file:
                if snomed_file.name.endswith(".xlsx"):
                    snomed_df = pd.read_excel(snomed_file, dtype=str)
                else:
                    snomed_df = pd.read_csv(snomed_file, dtype=str)
                snomed_df.columns = [c.strip() for c in snomed_df.columns]
                st.session_state.s5_snomed_input = snomed_df
                st.success(f"Loaded {len(snomed_df)} codes from {snomed_file.name}")
                st.dataframe(snomed_df.head())

        elif snomed_method == "Use built-in VHD codes (55 concepts)":
            if st.button("Load built-in VHD SNOMED codes", key="btn_s5_builtin_snomed"):
                builtin = _get_builtin_vhd_snomed()
                st.session_state.s5_snomed_input = builtin
                st.success(f"Loaded {len(builtin)} built-in VHD SNOMED CT codes")

        if st.session_state.s5_snomed_input is not None:
            sdf = st.session_state.s5_snomed_input
            st.metric("SNOMED CT Codes Loaded", len(sdf))
            with st.expander("View loaded SNOMED codes"):
                st.dataframe(sdf, height=300)

    # ── ICD-10 input ──
    with input_tabs[2]:
        st.markdown("Enter your ICD-10 candidate codes. Each row needs: **code_value**, **description**, **category**, and optionally **etiology**.")
        icd_method = st.radio("Input method", ["Paste CSV/table", "Upload CSV file", "Use built-in VHD ICD-10 codes (73 codes)"], key="s5_icd_method")

        if icd_method == "Paste CSV/table":
            icd_text = st.text_area(
                "Paste ICD-10 codes (CSV format: code_value,description,category,etiology)",
                height=200, key="s5_icd_text",
                placeholder="code_value,description,category,etiology\nI35.0,Nonrheumatic aortic stenosis,Aortic Stenosis,Non-rheumatic",
            )
            if icd_text.strip() and st.button("Parse ICD-10 codes", key="btn_s5_parse_icd"):
                try:
                    icd_df = pd.read_csv(io.StringIO(icd_text.strip()), dtype=str)
                    icd_df.columns = [c.strip() for c in icd_df.columns]
                    st.session_state.s5_icd_input = icd_df
                    st.success(f"Parsed {len(icd_df)} ICD-10 codes")
                except Exception as e:
                    st.error(f"Failed to parse: {e}")

        elif icd_method == "Upload CSV file":
            icd_file = st.file_uploader("Upload ICD-10 candidate CSV", type=["csv", "xlsx", "txt"], key="s5_icd_file")
            if icd_file:
                if icd_file.name.endswith(".xlsx"):
                    icd_df = pd.read_excel(icd_file, dtype=str)
                else:
                    icd_df = pd.read_csv(icd_file, dtype=str)
                icd_df.columns = [c.strip() for c in icd_df.columns]
                st.session_state.s5_icd_input = icd_df
                st.success(f"Loaded {len(icd_df)} ICD-10 codes")

        elif icd_method == "Use built-in VHD ICD-10 codes (73 codes)":
            if st.button("Load built-in VHD ICD-10 codes", key="btn_s5_builtin_icd"):
                builtin_icd = _get_builtin_vhd_icd10()
                st.session_state.s5_icd_input = builtin_icd
                st.success(f"Loaded {len(builtin_icd)} built-in VHD ICD-10 codes")

        if st.session_state.s5_icd_input is not None:
            idf = st.session_state.s5_icd_input
            st.metric("ICD-10 Codes Loaded", len(idf))

    # ── Upload Master CSV ──
    with input_tabs[0]:
        st.markdown(
            "Upload a CSV that contains both SNOMED CT and ICD-10 codes (like the Stage 5 master file)."
        )
        with st.expander("📐 Required CSV structure", expanded=False):
            st.markdown("""
**Minimum required column:**

| Column name | Purpose | Example values |
|---|---|---|
| `code_system` | Identifies which coding system each row belongs to | `SNOMED CT`, `ICD-10`, `Read V2` |
| `code_value` | The actual code | `60573004`, `I35.0`, `G541.00` |

The app splits rows by `code_system`: rows containing **"SNOMED"** go to SNOMED input,
rows containing **"ICD"** go to ICD-10 input.

**Recommended additional columns** (any name, auto-detected):

| Column | Example | Purpose |
|---|---|---|
| `category` | Aortic Stenosis | Clinical category |
| `description` | Aortic valve stenosis | Human-readable term |
| `etiology` | Rheumatic | Aetiology classification |
| `Severity` | Ungraded | Severity grading |
| `Source` | Existing | Where the code came from |

**Column names are flexible** — the app will try to auto-detect columns if they
don't match exactly (e.g., `Code_System` → `code_system`, `snomed_code` → `code_value`).

**Example rows:**
```
code_system,code_value,description,category,etiology,Severity
SNOMED CT,60573004,Aortic valve stenosis,Aortic Stenosis,Any,Ungraded parent
ICD-10,I35.0,Nonrheumatic aortic stenosis,Aortic Stenosis,Non-rheumatic,Ungraded
```
            """)

        master_file = st.file_uploader("Upload Master Code List CSV", type=["csv", "xlsx"], key="s5_master_upload")
        if master_file:
            if master_file.name.endswith(".xlsx"):
                mdf = pd.read_excel(master_file, dtype=str)
            else:
                mdf = pd.read_csv(master_file, dtype=str)
            mdf.columns = [c.strip() for c in mdf.columns]

            # ── Auto-detect code_system column ──
            code_system_col = None
            for col in mdf.columns:
                if col.lower().replace("_", "").replace(" ", "") in ("codesystem", "code_system", "coding_system", "system", "codingsystem"):
                    code_system_col = col
                    break
            # Fallback: check if any column contains "SNOMED" and "ICD" values
            if code_system_col is None:
                for col in mdf.columns:
                    vals = mdf[col].dropna().str.upper().unique()
                    has_snomed = any("SNOMED" in v for v in vals)
                    has_icd = any("ICD" in v for v in vals)
                    if has_snomed or has_icd:
                        code_system_col = col
                        break

            # ── Auto-detect code_value column ──
            code_value_col = None
            for col in mdf.columns:
                cl = col.lower().replace("_", "").replace(" ", "")
                if cl in ("codevalue", "code_value", "code", "snomedctconceptid", "snomedcode", "icdcode", "icd10code"):
                    code_value_col = col
                    break
            if code_value_col is None:
                # Pick the first column that looks like codes (numeric or Ixx.x patterns)
                for col in mdf.columns:
                    sample = mdf[col].dropna().head(20)
                    if sample.str.match(r'^[\dA-Z]').mean() > 0.8 and col != code_system_col:
                        code_value_col = col
                        break

            if code_system_col:
                # Normalise column names for downstream
                rename_map = {}
                if code_system_col != "code_system":
                    rename_map[code_system_col] = "code_system"
                if code_value_col and code_value_col != "code_value":
                    rename_map[code_value_col] = "code_value"
                if rename_map:
                    mdf = mdf.rename(columns=rename_map)
                    st.info(f"Auto-mapped columns: {rename_map}")

                snomed_part = mdf[mdf["code_system"].str.contains("SNOMED", case=False, na=False)].copy()
                icd_part = mdf[mdf["code_system"].str.contains("ICD", case=False, na=False)].copy()

                if len(snomed_part) > 0:
                    if "code_value" in snomed_part.columns and "SnomedCTConceptId" not in snomed_part.columns:
                        snomed_part = snomed_part.rename(columns={"code_value": "SnomedCTConceptId"})
                    st.session_state.s5_snomed_input = snomed_part
                if len(icd_part) > 0:
                    st.session_state.s5_icd_input = icd_part

                st.success(f"Loaded master file: **{len(snomed_part)}** SNOMED CT + **{len(icd_part)}** ICD-10 rows")
                with st.expander("Preview loaded data"):
                    st.dataframe(mdf.head(10))
            else:
                st.error(
                    "❌ Could not find a `code_system` column (or equivalent) in your file. "
                    f"Your columns are: **{', '.join(mdf.columns.tolist())}**"
                )
                st.markdown(
                    "The app needs a column that contains values like `SNOMED CT` and `ICD-10` "
                    "to know which rows are which. Click the **Required CSV structure** expander above for details."
                )
                with st.expander("Preview your file"):
                    st.dataframe(mdf.head(10))

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════════
    # STEP C: Run Dictionary Matching
    # ══════════════════════════════════════════════════════════════════════
    st.markdown("### Step C: Run Dictionary Matching")

    can_run = (dict_df is not None and st.session_state.s5_snomed_input is not None)

    if not can_run:
        missing_items = []
        if dict_df is None:
            missing_items.append("EMIS Medical Dictionary (Step A)")
        if st.session_state.s5_snomed_input is None:
            missing_items.append("SNOMED CT candidate codes (Step B)")
        st.warning(f"To run matching, please load: {', '.join(missing_items)}")
    else:
        if st.button("🚀 Run SNOMED → medcodeid Matching", key="btn_s5_run_match", type="primary"):
            log_lines = []
            def log(msg):
                log_lines.append(msg)

            log(f"═══════════════════════════════════════════════════")
            log(f"Stage 5: CPRD Dictionary Matching — Audit Log")
            log(f"Run date: {time.strftime('%Y-%m-%d %H:%M:%S')}")
            log(f"═══════════════════════════════════════════════════")

            snomed_input = st.session_state.s5_snomed_input.copy()
            # Ensure SnomedCTConceptId column exists
            sctid_col = "SnomedCTConceptId"
            if sctid_col not in snomed_input.columns:
                for col in snomed_input.columns:
                    if "snomed" in col.lower() or "concept" in col.lower() or col == "code_value":
                        snomed_input = snomed_input.rename(columns={col: sctid_col})
                        break

            if sctid_col not in snomed_input.columns:
                st.error(f"Cannot find SNOMED CT concept ID column. Available: {snomed_input.columns.tolist()}")
            else:
                snomed_input[sctid_col] = snomed_input[sctid_col].astype(str).str.strip()
                candidate_ids = set(snomed_input[sctid_col].unique())

                log(f"\nSNOMED CT concepts to search: {len(candidate_ids)}")
                log(f"Dictionary size: {len(dict_df):,} rows")

                # ── Primary match ──
                dict_df_local = dict_df.copy()
                dict_df_local["SnomedCTConceptId"] = dict_df_local["SnomedCTConceptId"].astype(str).str.strip()

                matched = dict_df_local[dict_df_local["SnomedCTConceptId"].isin(candidate_ids)].copy()

                # Merge with candidate metadata
                merge_cols = [sctid_col]
                meta_cols = [c for c in snomed_input.columns if c != sctid_col]
                # De-duplicate candidate input for merge
                snomed_dedup = snomed_input.drop_duplicates(subset=[sctid_col])
                matched = matched.merge(snomed_dedup, on=sctid_col, how="left")

                # ── Statistics ──
                matched_concepts = set(matched["SnomedCTConceptId"].unique())
                unmatched_concepts = candidate_ids - matched_concepts
                n_searched = len(candidate_ids)
                n_matched = len(matched_concepts)
                n_unmatched = len(unmatched_concepts)
                n_medcodeids = matched["MedCodeId"].nunique() if "MedCodeId" in matched.columns else 0
                match_rate = (n_matched / n_searched * 100) if n_searched > 0 else 0
                expansion_ratio = (n_medcodeids / n_matched) if n_matched > 0 else 0

                log(f"\n┌──────────────────────────────────────────┐")
                log(f"│       MATCHING RESULTS SUMMARY           │")
                log(f"├──────────────────────────────────────────┤")
                log(f"│ SNOMED concepts searched:  {n_searched:>5}         │")
                log(f"│ SNOMED concepts MATCHED:   {n_matched:>5}         │")
                log(f"│ SNOMED concepts UNMATCHED: {n_unmatched:>5}         │")
                log(f"│ Total unique medcodeids:   {n_medcodeids:>5}         │")
                log(f"│ Match rate:                {match_rate:>5.1f}%        │")
                log(f"│ Expansion ratio:           {expansion_ratio:>5.2f}x        │")
                log(f"└──────────────────────────────────────────┘")

                # ── Expansion analysis ──
                if "MedCodeId" in matched.columns:
                    expansion_df = (
                        matched.groupby("SnomedCTConceptId")
                        .agg(
                            n_medcodeids=("MedCodeId", "nunique"),
                            category=("category", "first") if "category" in matched.columns else ("SnomedCTConceptId", "first"),
                            Term=("Term", "first") if "Term" in matched.columns else ("SnomedCTConceptId", "first"),
                        )
                        .reset_index()
                        .sort_values("n_medcodeids", ascending=False)
                    )
                    st.session_state.s5_expansion = expansion_df

                    log(f"\nOne-to-many expansions:")
                    log(f"  Concepts with 1 medcodeid:  {(expansion_df['n_medcodeids'] == 1).sum()}")
                    log(f"  Concepts with 2 medcodeids: {(expansion_df['n_medcodeids'] == 2).sum()}")
                    log(f"  Concepts with 3+ medcodeids: {(expansion_df['n_medcodeids'] >= 3).sum()}")

                # ── Unmatched codes ──
                unmatched_df = snomed_input[snomed_input[sctid_col].isin(unmatched_concepts)].copy()

                # ── DMRS assessment ──
                if "Source" in snomed_input.columns:
                    dmrs_searched = snomed_input[snomed_input["Source"] == "NEW-DMRS"]
                    dmrs_matched_ids = matched_concepts & set(dmrs_searched[sctid_col].unique())
                    log(f"\nDMRS severity-graded codes searched: {len(dmrs_searched)}")
                    log(f"DMRS codes found in dictionary: {len(dmrs_matched_ids)}")
                    if len(dmrs_matched_ids) == 0 and len(dmrs_searched) > 0:
                        log(f"⚠ NOTE: No DMRS codes found. This is expected if dictionary predates Oct 2024 SNOMED CT UK release.")

                # ── Read code mappings ──
                if "CleansedReadCode" in matched.columns:
                    read_mapped = matched[matched["CleansedReadCode"].notna() & (matched["CleansedReadCode"] != "")]
                    n_read = read_mapped["CleansedReadCode"].nunique()
                    log(f"\nVHD medcodeids with CleansedReadCode: {len(read_mapped)} ({n_read} unique Read codes)")

                # ── Build master file ──
                master_parts = []

                # SNOMED CT rows
                snomed_master = matched.copy()
                snomed_master["code_system"] = "SNOMED CT"
                snomed_master["code_value"] = snomed_master["SnomedCTConceptId"]
                snomed_master["description"] = snomed_master.get("Term", "")
                snomed_master["medcodeid_aurum"] = snomed_master.get("MedCodeId", "")
                snomed_master["medcode_gold"] = ""
                snomed_master["data_source"] = "CPRD Aurum"
                for col in ["category", "Severity", "Source", "etiology"]:
                    if col not in snomed_master.columns:
                        snomed_master[col] = ""
                master_parts.append(snomed_master)

                # ICD-10 rows
                if st.session_state.s5_icd_input is not None:
                    icd_master = st.session_state.s5_icd_input.copy()
                    icd_master["code_system"] = "ICD-10"
                    if "code_value" not in icd_master.columns:
                        for col in icd_master.columns:
                            if "code" in col.lower() and "system" not in col.lower():
                                icd_master = icd_master.rename(columns={col: "code_value"})
                                break
                    icd_master["medcodeid_aurum"] = ""
                    icd_master["medcode_gold"] = ""
                    if "data_source" not in icd_master.columns:
                        icd_master["data_source"] = "HES APC, HES OP, ONS"
                    if "Severity" not in icd_master.columns:
                        icd_master["Severity"] = "Ungraded"
                    if "Source" not in icd_master.columns:
                        icd_master["Source"] = "Stage 4"
                    for col in ["category", "etiology", "description"]:
                        if col not in icd_master.columns:
                            icd_master[col] = ""
                    master_parts.append(icd_master)

                # Combine
                master_cols = ["category", "code_system", "code_value", "description",
                               "medcode_gold", "medcodeid_aurum", "data_source", "etiology",
                               "Severity", "Source"]
                all_parts = []
                for part in master_parts:
                    for col in master_cols:
                        if col not in part.columns:
                            part[col] = ""
                    all_parts.append(part[master_cols])

                master_df = pd.concat(all_parts, ignore_index=True)
                master_df = master_df.fillna("").astype(str)
                master_df = master_df.sort_values(["category", "code_system", "code_value"])

                n_snomed_rows = (master_df["code_system"] == "SNOMED CT").sum()
                n_icd_rows = (master_df["code_system"] == "ICD-10").sum()
                log(f"\n┌──────────────────────────────────────────┐")
                log(f"│       MASTER FILE SUMMARY               │")
                log(f"├──────────────────────────────────────────┤")
                log(f"│ Total rows:          {len(master_df):>6}            │")
                log(f"│ SNOMED CT entries:   {n_snomed_rows:>6}            │")
                log(f"│ ICD-10 entries:      {n_icd_rows:>6}            │")
                log(f"│ Unique categories:    {master_df['category'].nunique():>6}            │")
                log(f"└──────────────────────────────────────────┘")

                # Store results
                st.session_state.s5_matched = matched
                st.session_state.s5_unmatched = unmatched_df
                st.session_state.s5_master = master_df
                st.session_state.s5_log = "\n".join(log_lines)

                st.success(f"✅ Matching complete! {n_matched}/{n_searched} concepts matched → {n_medcodeids} medcodeids")

    # ══════════════════════════════════════════════════════════════════════
    # STEP D: Results Display
    # ══════════════════════════════════════════════════════════════════════
    if st.session_state.s5_matched is not None:
        st.markdown("---")
        st.markdown("### Step D: Results")

        matched = st.session_state.s5_matched
        unmatched_df = st.session_state.s5_unmatched
        master_df = st.session_state.s5_master
        expansion_df = st.session_state.s5_expansion

        # Summary metrics
        c1, c2, c3, c4 = st.columns(4)
        n_searched = st.session_state.s5_snomed_input["SnomedCTConceptId"].nunique() if "SnomedCTConceptId" in st.session_state.s5_snomed_input.columns else 0
        c1.metric("Concepts Matched", f"{matched['SnomedCTConceptId'].nunique()}/{n_searched}")
        c2.metric("Unique medcodeids", f"{matched['MedCodeId'].nunique():,}" if "MedCodeId" in matched.columns else "0")
        c3.metric("Master File Rows", f"{len(master_df):,}")
        match_rate = matched["SnomedCTConceptId"].nunique() / n_searched * 100 if n_searched > 0 else 0
        c4.metric("Match Rate", f"{match_rate:.1f}%")

        result_tabs = st.tabs(["✅ Matched", "❌ Unmatched", "📊 Expansion", "📋 Master File", "📝 Audit Log"])

        with result_tabs[0]:
            st.markdown(f"**{len(matched):,} matched rows** ({matched['SnomedCTConceptId'].nunique()} unique SNOMED concepts → {matched['MedCodeId'].nunique() if 'MedCodeId' in matched.columns else 0} medcodeids)")
            show_cols = [c for c in ["MedCodeId", "SnomedCTConceptId", "Term", "CleansedReadCode", "category", "Severity"] if c in matched.columns]
            st.dataframe(matched[show_cols].sort_values(show_cols[0] if show_cols else matched.columns[0]), height=400)

        with result_tabs[1]:
            if unmatched_df is not None and len(unmatched_df) > 0:
                st.warning(f"**{len(unmatched_df)} SNOMED CT codes NOT found** in the dictionary")
                st.dataframe(unmatched_df, height=300)
                st.markdown("""
**Why codes may be unmatched:**
- **DMRS severity codes** (836xxxxxx/838xxxxxx): Only present if dictionary includes October 2024 UK SNOMED CT release
- **Concept retired/inactive:** The SNOMED CT concept may have been retired or merged into another concept
- **Dictionary version mismatch:** Older dictionary versions may not contain newer SNOMED CT concepts

**What to do:** Check if a newer dictionary is available, or try searching by term/description instead.
                """)
            else:
                st.success("🎉 All SNOMED CT codes matched successfully!")

        with result_tabs[2]:
            if expansion_df is not None:
                st.markdown("Each SNOMED CT concept may expand to multiple medcodeids (one per description/term variant).")
                st.dataframe(expansion_df, height=300)

                # Expansion chart
                if len(expansion_df) > 0:
                    bins = expansion_df["n_medcodeids"].value_counts().sort_index()
                    fig = go.Figure(go.Bar(x=[f"{k} medcodeid{'s' if k > 1 else ''}" for k in bins.index],
                                           y=bins.values, text=bins.values, textposition="auto"))
                    fig.update_layout(title="SNOMED Concept → medcodeid Expansion Distribution",
                                      xaxis_title="medcodeids per concept", yaxis_title="Number of concepts",
                                      template="plotly_white", height=350)
                    st.plotly_chart(fig, use_container_width=True, key="plotly_10")

        with result_tabs[3]:
            st.markdown(f"**Unified Master Code List:** {len(master_df):,} rows across {master_df['code_system'].nunique()} coding systems")

            # Filter options
            fc1, fc2 = st.columns(2)
            with fc1:
                filter_system = st.multiselect("Filter by code system", master_df["code_system"].unique().tolist(),
                                               default=master_df["code_system"].unique().tolist(), key="s5_filter_system")
            with fc2:
                filter_subgroup = st.multiselect("Filter by category", sorted(master_df["category"].unique().tolist()),
                                                 key="s5_filter_subgroup")
            display_df = master_df[master_df["code_system"].isin(filter_system)]
            if filter_subgroup:
                display_df = display_df[display_df["category"].isin(filter_subgroup)]
            st.dataframe(display_df, height=400)

            # Summary by subgroup
            summary = master_df.groupby(["category", "code_system"]).size().unstack(fill_value=0)
            st.markdown("**Code counts by category and system:**")
            st.dataframe(summary)

        with result_tabs[4]:
            if st.session_state.s5_log:
                st.code(st.session_state.s5_log, language="text")

        # ── Downloads ──
        st.markdown("### Step E: Download Outputs")
        dl_cols = st.columns(4)

        with dl_cols[0]:
            csv_matched = matched.to_csv(index=False)
            st.download_button("📥 Matched medcodeids", csv_matched,
                               "matched_medcodeids.csv", "text/csv", key="dl_s5_matched")
        with dl_cols[1]:
            if unmatched_df is not None and len(unmatched_df) > 0:
                csv_unmatched = unmatched_df.to_csv(index=False)
                st.download_button("📥 Unmatched codes", csv_unmatched,
                                   "unmatched_codes.csv", "text/csv", key="dl_s5_unmatched")
        with dl_cols[2]:
            csv_master = master_df.to_csv(index=False)
            st.download_button("📥 Master code list", csv_master,
                               "master_code_list_stage5.csv", "text/csv", key="dl_s5_master")
        with dl_cols[3]:
            if "MedCodeId" in matched.columns:
                medcodeid_list = matched["MedCodeId"].unique().tolist()
                st.download_button("📥 medcodeid list (.txt)", "\n".join(medcodeid_list),
                                   "medcodeids_for_extraction.txt", "text/plain", key="dl_s5_medcodeids")

        # ── Integration with extraction engine ──
        st.markdown("### Step F: Use in Extraction")
        st.markdown("Push matched medcodeids directly to the extraction engine for immediate data extraction.")
        if "MedCodeId" in matched.columns:
            medcodeid_list = matched["MedCodeId"].unique().tolist()
            st.info(f"Ready to extract with {len(medcodeid_list):,} unique medcodeids")
            if st.button("🚀 Send to Aurum Extraction page", key="btn_s5_to_extraction"):
                st.session_state["s5_medcodeids_for_extraction"] = medcodeid_list
                st.success(f"✅ {len(medcodeid_list)} medcodeids stored. Go to **CPRD Aurum Extraction** → **MedCode Extraction** tab to use them.")


def _render_stage6_clinical_review():
    """Stage 6: Generate professional clinical review Excel for clinician adjudication."""

    st.header("Stage 6: Clinical Review Questionnaire")

    if not HAS_OPENPYXL:
        st.error("⚠️ **openpyxl** is required for Excel generation but is not installed.")
        st.code("pip install openpyxl", language="bash")
        st.markdown("On BMRC, run this in your terminal before launching the app:")
        st.code("module load Python/3.11.3-GCCcore-12.3.0\npip install --user openpyxl", language="bash")
        return

    st.markdown(
        "Generate a professionally formatted Excel workbook pre-populated with your matched codes "
        "from Stage 5. This workbook is designed to be sent to clinicians for review — each code "
        "has columns for **Include / Exclude / Uncertain** decisions, clinical comments, and "
        "suggested alternatives. The clinician returns the completed file and you import their "
        "decisions to finalise the code list."
    )

    master_df = st.session_state.get("s5_master")
    matched_df = st.session_state.get("s5_matched")

    if master_df is None:
        st.warning(
            "No matched codes found. Please complete **Stage 5 — Dictionary Matching** first, "
            "then return here to generate the clinical review workbook."
        )
        return

    st.success(f"✅ Master code list available: **{len(master_df):,} codes** across **{master_df['category'].nunique()} subgroups**")

    # ── Study metadata ──
    st.markdown("### Study Details")
    st.markdown("These details appear on the cover sheet of the review workbook.")
    mc1, mc2 = st.columns(2)
    with mc1:
        study_name = st.text_input("Study / project name", value="Valvular Heart Disease (VHD) Code List", key="s6_study")
        researcher = st.text_input("Researcher name", value="Milad Nazarzadeh", key="s6_researcher")
        institution = st.text_input("Institution", value="University of Oxford", key="s6_institution")
    with mc2:
        reviewer_name = st.text_input("Clinician reviewer name", value="", key="s6_reviewer",
                                       placeholder="Dr. Jane Smith")
        review_date = st.date_input("Review deadline", key="s6_date")
        database = st.text_input("Database", value="CPRD Aurum + HES + ONS", key="s6_database")

    # ── Options ──
    st.markdown("### Workbook Options")
    oc1, oc2 = st.columns(2)
    with oc1:
        include_instructions = st.checkbox("Include instructions sheet", value=True, key="s6_instructions")
        include_summary = st.checkbox("Include summary statistics sheet", value=True, key="s6_summary")
    with oc2:
        include_decision_flags = st.checkbox("Pre-flag codes needing clinical decision", value=True, key="s6_flags")
        group_by_subgroup = st.checkbox("Group codes by category (separate sections)", value=True, key="s6_group")

    st.markdown("---")

    # ── Generate ──
    if st.button("📋 Generate Clinical Review Workbook", key="btn_s6_generate", type="primary"):
        with st.spinner("Building Excel workbook..."):
            excel_bytes = _build_clinical_review_excel(
                master_df=master_df,
                study_name=study_name,
                researcher=researcher,
                institution=institution,
                reviewer_name=reviewer_name,
                review_deadline=str(review_date),
                database=database,
                include_instructions=include_instructions,
                include_summary=include_summary,
                include_decision_flags=include_decision_flags,
                group_by_subgroup=group_by_subgroup,
            )
            st.session_state["s6_excel_bytes"] = excel_bytes

    if st.session_state.get("s6_excel_bytes"):
        st.success("✅ Clinical review workbook generated!")
        st.download_button(
            "📥 Download Clinical Review Workbook (.xlsx)",
            st.session_state["s6_excel_bytes"],
            file_name=f"Clinical_Review_{time.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_s6_excel",
            type="primary",
        )

        st.markdown("### What happens next?")
        st.markdown("""
1. **Send** the Excel file to your clinician reviewer
2. The clinician fills in the **Include / Exclude / Uncertain** column and adds **comments** for each code
3. They return the completed file to you
4. **Import** the completed file back here (below) to apply their decisions and finalise the code list
        """)

        # ── Import completed review ──
        st.markdown("### Import Completed Review")
        completed_file = st.file_uploader(
            "Upload the clinician's completed review workbook",
            type=["xlsx"], key="s6_import",
        )
        if completed_file:
            try:
                reviewed = pd.read_excel(completed_file, sheet_name="Code Review", dtype=str)
                reviewed.columns = [c.strip() for c in reviewed.columns]

                if "Clinician Decision" in reviewed.columns:
                    decision_counts = reviewed["Clinician Decision"].fillna("").value_counts()
                    st.markdown("**Review Summary:**")
                    dc1, dc2, dc3, dc4 = st.columns(4)
                    dc1.metric("Include", int(decision_counts.get("Include", 0)))
                    dc2.metric("Exclude", int(decision_counts.get("Exclude", 0)))
                    dc3.metric("Uncertain", int(decision_counts.get("Uncertain", 0)))
                    dc4.metric("Not reviewed", int(decision_counts.get("", 0)) + int(decision_counts.get("NOT REVIEWED", 0)))

                    with st.expander("View reviewed codes"):
                        st.dataframe(reviewed, height=400)

                    # Export final list
                    included = reviewed[reviewed["Clinician Decision"].str.strip().str.lower() == "include"]
                    if len(included) > 0:
                        st.success(f"**{len(included)} codes approved** for inclusion in final code list")
                        csv_final = included.to_csv(index=False)
                        st.download_button("📥 Download Final Approved Code List", csv_final,
                                           "final_approved_code_list.csv", "text/csv", key="dl_s6_final")
                else:
                    st.warning("Could not find 'Clinician Decision' column. Make sure the clinician used the correct workbook.")
            except Exception as e:
                st.error(f"Error reading file: {e}")


def _build_clinical_review_excel(master_df, study_name, researcher, institution,
                                  reviewer_name, review_deadline, database,
                                  include_instructions, include_summary,
                                  include_decision_flags, group_by_subgroup):
    """Build a professionally formatted Excel workbook for clinical code review."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel generation")

    from openpyxl import Workbook

    wb = Workbook()

    # ── Colour palette ──
    OXFORD_BLUE = "002147"
    HEADER_FILL = PatternFill("solid", fgColor=OXFORD_BLUE)
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    SUBHEADER_FILL = PatternFill("solid", fgColor="4472C4")
    SUBHEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    TITLE_FONT = Font(name="Arial", bold=True, size=16, color=OXFORD_BLUE)
    SUBTITLE_FONT = Font(name="Arial", bold=True, size=12, color=OXFORD_BLUE)
    BODY_FONT = Font(name="Arial", size=10)
    BOLD_FONT = Font(name="Arial", bold=True, size=10)
    INCLUDE_FILL = PatternFill("solid", fgColor="C6EFCE")  # green
    EXCLUDE_FILL = PatternFill("solid", fgColor="FFC7CE")  # red
    UNCERTAIN_FILL = PatternFill("solid", fgColor="FFEB9C")  # amber
    DMRS_FILL = PatternFill("solid", fgColor="E2EFDA")  # light green for DMRS severity codes
    DECISION_FILL = PatternFill("solid", fgColor="FFF2CC")  # light amber for decision-needed codes
    LIGHT_GREY = PatternFill("solid", fgColor="F2F2F2")
    THIN_BORDER = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1: COVER PAGE
    # ═══════════════════════════════════════════════════════════════════════
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.sheet_properties.tabColor = OXFORD_BLUE

    ws_cover.column_dimensions["A"].width = 5
    ws_cover.column_dimensions["B"].width = 30
    ws_cover.column_dimensions["C"].width = 50

    ws_cover.merge_cells("B2:C2")
    ws_cover["B2"] = "Clinical Code List Review"
    ws_cover["B2"].font = TITLE_FONT

    ws_cover.merge_cells("B3:C3")
    ws_cover["B3"] = study_name
    ws_cover["B3"].font = SUBTITLE_FONT

    details = [
        ("", ""),
        ("Study / Project:", study_name),
        ("Database:", database),
        ("Researcher:", researcher),
        ("Institution:", institution),
        ("", ""),
        ("Clinician Reviewer:", reviewer_name if reviewer_name else "[To be completed]"),
        ("Review Deadline:", review_deadline),
        ("Date Generated:", time.strftime("%d %B %Y")),
        ("", ""),
        ("INSTRUCTIONS:", ""),
    ]
    row = 5
    for label, value in details:
        ws_cover[f"B{row}"] = label
        ws_cover[f"B{row}"].font = BOLD_FONT
        ws_cover[f"C{row}"] = value
        ws_cover[f"C{row}"].font = BODY_FONT
        row += 1

    instructions_text = [
        "1. Go to the 'Code Review' sheet",
        "2. For each code, select Include / Exclude / Uncertain in column G",
        "3. Add comments in column H explaining your decision",
        "4. If you suggest an alternative code, enter it in column I",
        "5. Codes highlighted in AMBER require your specific attention",
        "6. Codes highlighted in GREEN are DMRS severity-graded (new in 2024)",
        "7. Return the completed file to the researcher",
    ]
    for line in instructions_text:
        ws_cover[f"B{row}"] = line
        ws_cover[f"B{row}"].font = BODY_FONT
        row += 1

    # Summary box
    row += 1
    ws_cover[f"B{row}"] = "CODE LIST SUMMARY"
    ws_cover[f"B{row}"].font = SUBTITLE_FONT
    row += 1
    n_snomed = (master_df["code_system"] == "SNOMED CT").sum() if "code_system" in master_df.columns else 0
    n_icd = (master_df["code_system"] == "ICD-10").sum() if "code_system" in master_df.columns else 0
    n_subgroups = master_df["category"].nunique() if "category" in master_df.columns else 0
    for label, val in [("Total codes:", len(master_df)), ("SNOMED CT codes:", n_snomed),
                        ("ICD-10 codes:", n_icd), ("Clinical categories:", n_subgroups)]:
        ws_cover[f"B{row}"] = label
        ws_cover[f"B{row}"].font = BOLD_FONT
        ws_cover[f"C{row}"] = str(val)
        ws_cover[f"C{row}"].font = BODY_FONT
        row += 1

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 2: CODE REVIEW (main sheet)
    # ═══════════════════════════════════════════════════════════════════════
    ws_review = wb.create_sheet("Code Review")
    ws_review.sheet_properties.tabColor = "4472C4"

    # Column definitions
    review_columns = [
        ("A", "#", 5),
        ("B", "Category", 25),
        ("C", "Code System", 14),
        ("D", "Code Value", 18),
        ("E", "Description", 45),
        ("F", "Etiology", 16),
        ("G", "Clinician Decision", 20),
        ("H", "Clinician Comments", 40),
        ("I", "Suggested Alternative", 25),
        ("J", "Severity", 14),
        ("K", "Data Source", 28),
        ("L", "Source (Stage 4)", 16),
    ]

    # Header row
    for col_letter, col_name, width in review_columns:
        cell = ws_review[f"{col_letter}1"]
        cell.value = col_name
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws_review.column_dimensions[col_letter].width = width

    # Freeze header
    ws_review.freeze_panes = "A2"

    # Auto-filter
    ws_review.auto_filter.ref = f"A1:L{len(master_df) + 1}"

    # Sort data
    if group_by_subgroup and "category" in master_df.columns:
        sorted_df = master_df.sort_values(["category", "code_system", "code_value"]).reset_index(drop=True)
    else:
        sorted_df = master_df.sort_values(["code_system", "code_value"]).reset_index(drop=True)

    # Data rows
    prev_subgroup = ""
    row_num = 2
    for idx, row_data in sorted_df.iterrows():
        subgroup = str(row_data.get("category", ""))
        code_sys = str(row_data.get("code_system", ""))
        code_val = str(row_data.get("code_value", ""))
        desc = str(row_data.get("description", ""))
        etiol = str(row_data.get("etiology", ""))
        severity = str(row_data.get("Severity", ""))
        source = str(row_data.get("Source", row_data.get("source_stage4", "")))
        data_src = str(row_data.get("data_source", ""))
        decision_flag = str(row_data.get("decision_flag", ""))

        # Subgroup separator row
        if group_by_subgroup and subgroup != prev_subgroup and subgroup:
            ws_review.merge_cells(f"A{row_num}:L{row_num}")
            ws_review[f"A{row_num}"] = f"▸ {subgroup}"
            ws_review[f"A{row_num}"].font = SUBHEADER_FONT
            ws_review[f"A{row_num}"].fill = SUBHEADER_FILL
            ws_review[f"A{row_num}"].alignment = Alignment(horizontal="left", vertical="center")
            row_num += 1
            prev_subgroup = subgroup

        # Write data
        ws_review[f"A{row_num}"] = idx + 1
        ws_review[f"B{row_num}"] = subgroup
        ws_review[f"C{row_num}"] = code_sys
        ws_review[f"D{row_num}"] = code_val
        ws_review[f"E{row_num}"] = desc
        ws_review[f"F{row_num}"] = etiol
        ws_review[f"G{row_num}"] = "NOT REVIEWED"  # Clinician fills this
        ws_review[f"H{row_num}"] = ""  # Clinician comments
        ws_review[f"I{row_num}"] = ""  # Suggested alternative
        ws_review[f"J{row_num}"] = severity
        ws_review[f"K{row_num}"] = data_src
        ws_review[f"L{row_num}"] = source

        # Apply formatting to all cells
        for col_letter, _, _ in review_columns:
            cell = ws_review[f"{col_letter}{row_num}"]
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT_WRAP if col_letter in ("E", "H", "K") else CENTER

        # Row highlighting
        is_dmrs = "DMRS" in severity or "DMRS" in str(source)
        is_decision = decision_flag.upper() in ("Y", "YES", "TRUE", "1")

        if include_decision_flags and is_decision:
            for col_letter, _, _ in review_columns:
                ws_review[f"{col_letter}{row_num}"].fill = DECISION_FILL
        elif is_dmrs:
            for col_letter, _, _ in review_columns:
                ws_review[f"{col_letter}{row_num}"].fill = DMRS_FILL
        elif row_num % 2 == 0:
            for col_letter, _, _ in review_columns:
                ws_review[f"{col_letter}{row_num}"].fill = LIGHT_GREY

        # Data validation for clinician decision column
        dv = DataValidation(type="list", formula1='"Include,Exclude,Uncertain,NOT REVIEWED"', allow_blank=True)
        dv.error = "Please select Include, Exclude, or Uncertain"
        dv.errorTitle = "Invalid Decision"
        dv.prompt = "Select your clinical decision"
        dv.promptTitle = "Clinician Decision"
        ws_review.add_data_validation(dv)
        dv.add(ws_review[f"G{row_num}"])

        row_num += 1

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 3: INSTRUCTIONS (optional)
    # ═══════════════════════════════════════════════════════════════════════
    if include_instructions:
        ws_instr = wb.create_sheet("Instructions")
        ws_instr.sheet_properties.tabColor = "70AD47"
        ws_instr.column_dimensions["A"].width = 5
        ws_instr.column_dimensions["B"].width = 80

        ws_instr["B2"] = "Instructions for Clinical Code Review"
        ws_instr["B2"].font = TITLE_FONT

        instr_lines = [
            "",
            "BACKGROUND",
            f"This workbook contains {len(master_df)} candidate clinical codes for the study:",
            f'"{study_name}"',
            f"Database: {database}",
            "",
            "These codes were identified through a systematic code list development process",
            "following CPRD Guidance v7.1 and the Matthewman et al. (2024) 10-step framework.",
            "",
            "YOUR TASK",
            "Review each code in the 'Code Review' sheet and provide your clinical judgement:",
            "",
            "  Include  — This code correctly identifies the condition of interest",
            "  Exclude  — This code does NOT identify the condition (wrong concept, too broad, etc.)",
            "  Uncertain — You are unsure; please add a comment explaining why",
            "",
            "COLUMN GUIDE",
            "  Column G (Clinician Decision) — Use the dropdown to select Include / Exclude / Uncertain",
            "  Column H (Clinician Comments) — Explain your reasoning, especially for Exclude/Uncertain",
            "  Column I (Suggested Alternative) — If you know a better code, enter it here",
            "",
            "COLOUR CODING",
            "  AMBER rows — These codes have been flagged as needing your specific clinical input",
            "  GREEN rows — These are new DMRS severity-graded SNOMED CT codes (Oct 2024 UK release)",
            "  WHITE/GREY rows — Standard codes, review as normal",
            "",
            "CODING SYSTEMS",
            "  SNOMED CT — Used in CPRD Aurum primary care records (via medcodeids)",
            "  ICD-10   — Used in Hospital Episode Statistics (HES) and ONS death records",
            "  Read V2  — Used in CPRD Gold primary care records (legacy)",
            "",
            "IMPORTANT NOTES",
            "  • A single clinical concept may have multiple codes (e.g., SNOMED + ICD-10 + Read)",
            "  • SNOMED codes with 'DMRS' severity may not yet be in all practice systems",
            "  • ICD-10 does not have severity sub-categories for valvular heart disease",
            "  • Some codes appear under both rheumatic and non-rheumatic categories",
            "",
            "REFERENCES",
            "  CPRD. Developing a code list for research using CPRD primary care data. v7.1",
            "  Matthewman J et al. NIHR Open Research 2024. doi:10.3310/nihropenres.13498.1",
            "  Benchimol EI et al. The RECORD Statement. PLoS Med 2015;12(10):e1001885",
        ]
        for i, line in enumerate(instr_lines, start=4):
            ws_instr[f"B{i}"] = line
            if line in ("BACKGROUND", "YOUR TASK", "COLUMN GUIDE", "COLOUR CODING",
                        "CODING SYSTEMS", "IMPORTANT NOTES", "REFERENCES"):
                ws_instr[f"B{i}"].font = SUBTITLE_FONT
            else:
                ws_instr[f"B{i}"].font = BODY_FONT

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 4: SUMMARY STATISTICS (optional)
    # ═══════════════════════════════════════════════════════════════════════
    if include_summary:
        ws_sum = wb.create_sheet("Summary")
        ws_sum.sheet_properties.tabColor = "FFC000"

        ws_sum.column_dimensions["A"].width = 5
        ws_sum.column_dimensions["B"].width = 30
        ws_sum.column_dimensions["C"].width = 16
        ws_sum.column_dimensions["D"].width = 16
        ws_sum.column_dimensions["E"].width = 16
        ws_sum.column_dimensions["F"].width = 16

        ws_sum["B2"] = "Code List Summary Statistics"
        ws_sum["B2"].font = TITLE_FONT

        # Table 1: By code system
        ws_sum["B4"] = "Codes by Classification System"
        ws_sum["B4"].font = SUBTITLE_FONT

        for col_letter, label in [("B", "Code System"), ("C", "Count"), ("D", "Unique Concepts")]:
            cell = ws_sum[f"{col_letter}5"]
            cell.value = label
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        r = 6
        if "code_system" in master_df.columns:
            for system in sorted(master_df["code_system"].unique()):
                subset = master_df[master_df["code_system"] == system]
                ws_sum[f"B{r}"] = system
                ws_sum[f"C{r}"] = len(subset)
                ws_sum[f"D{r}"] = subset["code_value"].nunique() if "code_value" in subset.columns else len(subset)
                for cl in "BCD":
                    ws_sum[f"{cl}{r}"].font = BODY_FONT
                    ws_sum[f"{cl}{r}"].border = THIN_BORDER
                    ws_sum[f"{cl}{r}"].alignment = CENTER
                r += 1
        # Total row
        ws_sum[f"B{r}"] = "TOTAL"
        ws_sum[f"B{r}"].font = BOLD_FONT
        ws_sum[f"C{r}"] = len(master_df)
        ws_sum[f"C{r}"].font = BOLD_FONT
        for cl in "BCD":
            ws_sum[f"{cl}{r}"].border = THIN_BORDER
            ws_sum[f"{cl}{r}"].alignment = CENTER

        # Table 2: By category
        r += 2
        ws_sum[f"B{r}"] = "Codes by Clinical Subgroup"
        ws_sum[f"B{r}"].font = SUBTITLE_FONT
        r += 1

        t2_headers = [("B", "Category"), ("C", "SNOMED CT"), ("D", "ICD-10"), ("E", "Other"), ("F", "Total")]
        for col_letter, label in t2_headers:
            cell = ws_sum[f"{col_letter}{r}"]
            cell.value = label
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        r += 1

        if "category" in master_df.columns and "code_system" in master_df.columns:
            for sg in sorted(master_df["category"].unique()):
                subset = master_df[master_df["category"] == sg]
                n_snomed_sg = (subset["code_system"] == "SNOMED CT").sum()
                n_icd_sg = (subset["code_system"] == "ICD-10").sum()
                n_other = len(subset) - n_snomed_sg - n_icd_sg
                ws_sum[f"B{r}"] = sg
                ws_sum[f"C{r}"] = n_snomed_sg
                ws_sum[f"D{r}"] = n_icd_sg
                ws_sum[f"E{r}"] = n_other
                ws_sum[f"F{r}"] = len(subset)
                for cl in "BCDEF":
                    ws_sum[f"{cl}{r}"].font = BODY_FONT
                    ws_sum[f"{cl}{r}"].border = THIN_BORDER
                    ws_sum[f"{cl}{r}"].alignment = CENTER
                r += 1

    # ═══════════════════════════════════════════════════════════════════════
    # SAVE TO BYTES
    # ═══════════════════════════════════════════════════════════════════════
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _get_builtin_vhd_snomed():
    """Return the 55 built-in VHD SNOMED CT candidate codes."""
    data = [
        ("60573004","Aortic Stenosis","Ungraded parent","Existing"),
        ("836480008","Aortic Stenosis - Mild","DMRS Mild","NEW-DMRS"),
        ("836481007","Aortic Stenosis - Moderate","DMRS Moderate","NEW-DMRS"),
        ("836482000","Aortic Stenosis - Severe","DMRS Severe","NEW-DMRS"),
        ("72011007","Aortic Stenosis (Rheumatic)","Ungraded","Existing"),
        ("253545000","Aortic Stenosis (Congenital)","Ungraded","NEW-Hierarchy"),
        ("60234000","Aortic Regurgitation","Ungraded parent","Existing"),
        ("838544003","Aortic Regurgitation - Mild","DMRS Mild","NEW-DMRS"),
        ("838545002","Aortic Regurgitation - Moderate","DMRS Moderate","NEW-DMRS"),
        ("838546001","Aortic Regurgitation - Severe","DMRS Severe","NEW-DMRS"),
        ("78031003","Aortic Regurgitation (Rheumatic)","Ungraded","Existing"),
        ("194984004","Aortic Regurgitation (Non-rheumatic)","Ungraded","Existing"),
        ("472847005","Mixed Aortic Valve Disease","Ungraded","Existing"),
        ("17759006","Rheumatic AS with AR","Ungraded","Existing"),
        ("8722008","Aortic Valve Disease","Ungraded","Existing"),
        ("16440002","Rheumatic Aortic Valve Disease","Ungraded","Existing"),
        ("424031003","Aortic Valve Sclerosis","Pre-clinical","NEW-Hierarchy"),
        ("250978003","Calcification of Aortic Valve","Finding","NEW-Hierarchy"),
        ("204357006","Bicuspid Aortic Valve","Congenital","NEW-Hierarchy"),
        ("79619009","Mitral Stenosis","Ungraded parent","Existing"),
        ("838448003","Mitral Stenosis - Mild","DMRS Mild","NEW-DMRS"),
        ("838449006","Mitral Stenosis - Moderate","DMRS Moderate","NEW-DMRS"),
        ("838450006","Mitral Stenosis - Severe","DMRS Severe","NEW-DMRS"),
        ("86466006","Mitral Stenosis (Rheumatic)","Ungraded","Existing"),
        ("82458004","Mitral Stenosis (Congenital)","Ungraded","Existing"),
        ("48724000","Mitral Regurgitation","Ungraded parent","Existing"),
        ("838451005","Mitral Regurgitation - Mild","DMRS Mild","NEW-DMRS"),
        ("838452003","Mitral Regurgitation - Moderate","DMRS Moderate","NEW-DMRS"),
        ("838453008","Mitral Regurgitation - Severe","DMRS Severe","NEW-DMRS"),
        ("31085000","Mitral Regurgitation (Rheumatic)","Ungraded","Existing"),
        ("78495000","Mitral Regurgitation (Congenital)","Ungraded","Existing"),
        ("787001","Mixed Mitral (Rheumatic MS+MR)","Ungraded","Existing"),
        ("409712001","Mitral Valve Prolapse","Ungraded","Existing"),
        ("83916000","Mitral Valve Disease (Rheumatic)","Ungraded","Existing"),
        ("708966001","Mitral Valve Disease (Non-rheumatic)","Ungraded","Existing"),
        ("11851006","Mitral Annular Calcification","Finding","NEW-Hierarchy"),
        ("49915006","Tricuspid Stenosis","Ungraded parent","NEW-Gap2"),
        ("838535004","Tricuspid Stenosis - Mild","DMRS Mild","NEW-DMRS"),
        ("838536003","Tricuspid Stenosis - Moderate","DMRS Moderate","NEW-DMRS"),
        ("838537007","Tricuspid Stenosis - Severe","DMRS Severe","NEW-DMRS"),
        ("111287006","Tricuspid Regurgitation","Ungraded parent","NEW-Gap2"),
        ("838454002","Tricuspid Regurgitation - Mild","DMRS Mild","NEW-DMRS"),
        ("838455001","Tricuspid Regurgitation - Moderate","DMRS Moderate","NEW-DMRS"),
        ("838456000","Tricuspid Regurgitation - Severe","DMRS Severe","NEW-DMRS"),
        ("63042009","Ebstein Anomaly / Congenital Tricuspid","Congenital","Existing"),
        ("56786000","Pulmonary Stenosis","Ungraded parent","Existing"),
        ("838538002","Pulmonary Stenosis - Mild","DMRS Mild","NEW-DMRS"),
        ("838540007","Pulmonary Stenosis - Severe","DMRS Severe","NEW-DMRS"),
        ("91434003","Pulmonary Regurgitation","Ungraded","Existing"),
        ("233917008","Pulmonary Atresia","Congenital","Existing"),
        ("194733006","Multi-valve (Mitral + Aortic)","Ungraded","NEW-Hierarchy"),
        ("368009","Heart Valve Disease (General)","Ungraded","NEW-Hierarchy"),
        ("398995000","Heart Valve Disease (Synonym)","Ungraded","NEW-Hierarchy"),
        ("24211005","Prosthetic Heart Valve Present","Status","NEW-Gap6"),
        ("62067003","Hypoplastic Left Heart Syndrome","Congenital","Existing"),
    ]
    return pd.DataFrame(data, columns=["SnomedCTConceptId", "category", "Severity", "Source"])


def _get_builtin_vhd_icd10():
    """Return the 73 built-in VHD ICD-10 codes."""
    data = [
        ("I05.0","Rheumatic mitral stenosis","Mitral Stenosis","Rheumatic"),
        ("I05.1","Rheumatic mitral insufficiency","Mitral Regurgitation","Rheumatic"),
        ("I05.2","Mitral stenosis with insufficiency","Mixed Mitral","Rheumatic"),
        ("I05.8","Other rheumatic mitral valve diseases","Mitral Valve Disease","Rheumatic"),
        ("I05.9","Rheumatic mitral valve disease, unspecified","Mitral Valve Disease","Rheumatic"),
        ("I06.0","Rheumatic aortic stenosis","Aortic Stenosis","Rheumatic"),
        ("I06.1","Rheumatic aortic insufficiency","Aortic Regurgitation","Rheumatic"),
        ("I06.2","Rheumatic aortic stenosis with insufficiency","Mixed Aortic","Rheumatic"),
        ("I06.8","Other rheumatic aortic valve diseases","Aortic Valve Disease","Rheumatic"),
        ("I06.9","Rheumatic aortic valve disease, unspecified","Aortic Valve Disease","Rheumatic"),
        ("I07.0","Rheumatic tricuspid stenosis","Tricuspid Stenosis","Rheumatic"),
        ("I07.1","Rheumatic tricuspid insufficiency","Tricuspid Regurgitation","Rheumatic"),
        ("I07.2","Rheumatic tricuspid stenosis with insufficiency","Mixed Tricuspid","Rheumatic"),
        ("I07.8","Other rheumatic tricuspid valve diseases","Tricuspid Valve Disease","Rheumatic"),
        ("I07.9","Rheumatic tricuspid valve disease, unspecified","Tricuspid Valve Disease","Rheumatic"),
        ("I08.0","Rheumatic disorders of both mitral and aortic valves","Multi-valve Disease","Rheumatic"),
        ("I08.1","Rheumatic disorders of both mitral and tricuspid valves","Multi-valve Disease","Rheumatic"),
        ("I08.2","Rheumatic disorders of both aortic and tricuspid valves","Multi-valve Disease","Rheumatic"),
        ("I08.3","Combined rheumatic disorders of mitral, aortic, and tricuspid valves","Multi-valve Disease","Rheumatic"),
        ("I08.8","Other rheumatic multiple valve diseases","Multi-valve Disease","Rheumatic"),
        ("I08.9","Rheumatic multiple valve disease, unspecified","Multi-valve Disease","Rheumatic"),
        ("I09.1","Rheumatic diseases of endocardium, valve unspecified","VHD General","Rheumatic"),
        ("I09.9","Rheumatic heart disease, unspecified","VHD General","Rheumatic"),
        ("I34.0","Nonrheumatic mitral (valve) insufficiency","Mitral Regurgitation","Non-rheumatic"),
        ("I34.1","Nonrheumatic mitral (valve) prolapse","Mitral Valve Prolapse","Non-rheumatic"),
        ("I34.2","Nonrheumatic mitral (valve) stenosis","Mitral Stenosis","Non-rheumatic"),
        ("I34.8","Other nonrheumatic mitral valve disorders","Mitral Valve Disease","Non-rheumatic"),
        ("I34.9","Nonrheumatic mitral valve disorder, unspecified","Mitral Valve Disease","Non-rheumatic"),
        ("I35.0","Nonrheumatic aortic (valve) stenosis","Aortic Stenosis","Non-rheumatic"),
        ("I35.1","Nonrheumatic aortic (valve) insufficiency","Aortic Regurgitation","Non-rheumatic"),
        ("I35.2","Nonrheumatic aortic (valve) stenosis with insufficiency","Mixed Aortic","Non-rheumatic"),
        ("I35.8","Other nonrheumatic aortic valve disorders","Aortic Valve Disease","Non-rheumatic"),
        ("I35.9","Nonrheumatic aortic valve disorder, unspecified","Aortic Valve Disease","Non-rheumatic"),
        ("I36.0","Nonrheumatic tricuspid (valve) stenosis","Tricuspid Stenosis","Non-rheumatic"),
        ("I36.1","Nonrheumatic tricuspid (valve) insufficiency","Tricuspid Regurgitation","Non-rheumatic"),
        ("I36.2","Nonrheumatic tricuspid (valve) stenosis with insufficiency","Mixed Tricuspid","Non-rheumatic"),
        ("I36.8","Other nonrheumatic tricuspid valve disorders","Tricuspid Valve Disease","Non-rheumatic"),
        ("I36.9","Nonrheumatic tricuspid valve disorder, unspecified","Tricuspid Valve Disease","Non-rheumatic"),
        ("I37.0","Nonrheumatic pulmonary valve stenosis","Pulmonary Stenosis","Non-rheumatic"),
        ("I37.1","Nonrheumatic pulmonary valve insufficiency","Pulmonary Regurgitation","Non-rheumatic"),
        ("I37.2","Nonrheumatic pulmonary valve stenosis with insufficiency","Mixed Pulmonary","Non-rheumatic"),
        ("I37.8","Other nonrheumatic pulmonary valve disorders","Pulmonary Valve Disease","Non-rheumatic"),
        ("I37.9","Nonrheumatic pulmonary valve disorder, unspecified","Pulmonary Valve Disease","Non-rheumatic"),
        ("I33.0","Acute and subacute infective endocarditis","Endocarditis","Infective"),
        ("I33.9","Acute and subacute endocarditis, unspecified","Endocarditis","Unspecified"),
        ("I38","Endocarditis, valve unspecified","Endocarditis","Non-rheumatic"),
        ("I39","Endocarditis in diseases classified elsewhere","Endocarditis","Secondary"),
        ("Q22.0","Pulmonary valve atresia","Pulmonary Atresia","Congenital"),
        ("Q22.1","Congenital pulmonary valve stenosis","Pulmonary Stenosis","Congenital"),
        ("Q22.2","Congenital pulmonary valve insufficiency","Pulmonary Regurgitation","Congenital"),
        ("Q22.3","Other congenital malformations of pulmonary valve","Pulmonary Valve Disease","Congenital"),
        ("Q22.4","Congenital tricuspid stenosis","Tricuspid Stenosis","Congenital"),
        ("Q22.5","Ebstein anomaly","Tricuspid Valve Disease","Congenital"),
        ("Q22.6","Hypoplastic right heart syndrome","Complex Congenital","Congenital"),
        ("Q22.8","Other congenital malformations of tricuspid valve","Tricuspid Valve Disease","Congenital"),
        ("Q22.9","Congenital malformation of tricuspid valve, unspecified","Tricuspid Valve Disease","Congenital"),
        ("Q23.0","Congenital stenosis of aortic valve","Aortic Stenosis","Congenital"),
        ("Q23.1","Congenital insufficiency of aortic valve","Aortic Regurgitation","Congenital"),
        ("Q23.2","Congenital mitral stenosis","Mitral Stenosis","Congenital"),
        ("Q23.3","Congenital mitral insufficiency","Mitral Regurgitation","Congenital"),
        ("Q23.4","Hypoplastic left heart syndrome","Complex Congenital","Congenital"),
        ("Q23.8","Other congenital malformations of aortic and mitral valves","Aortic/Mitral Valve","Congenital"),
        ("Q23.9","Congenital malformation of aortic and mitral valves, unspecified","Aortic/Mitral Valve","Congenital"),
        ("Z95.2","Presence of prosthetic heart valve","Prosthetic Valve","Status"),
        ("Z95.3","Presence of xenogenic heart valve","Prosthetic Valve","Status"),
        ("Z95.4","Presence of other heart-valve replacement","Prosthetic Valve","Status"),
        ("T82.01","Breakdown (mechanical) of heart valve prosthesis","Prosthetic Complication","Complication"),
        ("T82.02","Displacement of heart valve prosthesis","Prosthetic Complication","Complication"),
        ("T82.03","Leakage of heart valve prosthesis","Prosthetic Complication","Complication"),
        ("T82.09","Other mechanical complication of heart valve prosthesis","Prosthetic Complication","Complication"),
        ("T82.6","Infection and inflammatory reaction due to cardiac valve prosthesis","Prosthetic Complication","Complication"),
    ]
    return pd.DataFrame(data, columns=["code_value", "description", "category", "etiology"])


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: DRUG LOOKUP — BNF/Drug Name → ProdCodeId
# ══════════════════════════════════════════════════════════════════════════════

def page_drug_lookup():
    st.title("💊 Drug Lookup")
    st.markdown("Search the EMIS Product Dictionary by **BNF code**, **drug name**, or **free text** to find ProdCodeIds for Drug Issue extraction.")

    tab_search, tab_library, tab_results = st.tabs([
        "🔍 Search Product Dictionary", "📚 Built-in Drug Libraries", "📋 My Drug Code List"
    ])

    # ── Load Product Dictionary ──
    prod_dict = _load_emis_product_dict()
    if prod_dict is None and not st.session_state.engine.is_mock():
        st.warning("⚠️ EMIS Product Dictionary not found. Check Configuration page for correct path.")

    # ═══════════════════════════════════════════════════════════════════
    # TAB 1: Free-text / Custom Search
    # ═══════════════════════════════════════════════════════════════════
    with tab_search:
        if prod_dict is None:
            st.info("Product Dictionary not available. Switch to Mock Data mode to test, or check your data paths.")
            return

        st.markdown("### Search by BNF Code, Drug Name, or Free Text")
        st.markdown("The search looks across **all available columns** in the Product Dictionary: "
                     "ProductName, DrugSubstanceName, Term, BNFChapter, and more.")

        search_method = st.radio("Search method",
                                  ["Free text search", "BNF code prefix", "Drug substance name"],
                                  horizontal=True, key="drug_search_method")

        available_cols = prod_dict.columns.tolist()

        if search_method == "Free text search":
            query = st.text_input("Enter search term (searches ProductName, DrugSubstanceName, Term)",
                                   placeholder="e.g. amlodipine, ramipril, statin",
                                   key="drug_free_search")
            if query and len(query) >= 2:
                q = query.strip().lower()
                # Search across all text columns
                search_cols = [c for c in ["Term", "ProductName", "DrugSubstanceName",
                                           "FormulationName"] if c in available_cols]
                if not search_cols:
                    search_cols = available_cols[:3]
                mask = pd.Series(False, index=prod_dict.index)
                for col in search_cols:
                    mask = mask | prod_dict[col].fillna("").str.lower().str.contains(q, regex=False)
                matches = prod_dict[mask].copy()
                _show_drug_matches(matches, f"free_search_{q}")

        elif search_method == "BNF code prefix":
            bnf_input = st.text_area("Enter BNF code prefixes (one per line)",
                                      placeholder="0205051R0\n0212000AC\n0204000",
                                      key="drug_bnf_input", height=120)
            if bnf_input.strip():
                prefixes = [p.strip() for p in bnf_input.strip().split("\n") if p.strip()]
                st.info(f"{len(prefixes)} BNF prefixes entered")
                if st.button("🔍 Search by BNF", key="btn_bnf_search"):
                    if "BNFChapter" in available_cols:
                        mask = pd.Series(False, index=prod_dict.index)
                        for prefix in prefixes:
                            p_lower = prefix.lower()
                            # Also try without leading zero
                            p_no_zero = p_lower.lstrip("0") if p_lower.startswith("0") else p_lower
                            mask = mask | (
                                prod_dict["BNFChapter"].fillna("").str.lower().str.contains(p_lower, regex=False) |
                                prod_dict["BNFChapter"].fillna("").str.lower().str.contains(p_no_zero, regex=False)
                            )
                        matches = prod_dict[mask].copy()
                        _show_drug_matches(matches, "bnf_search")
                    else:
                        st.error("BNFChapter column not found in Product Dictionary.")

        elif search_method == "Drug substance name":
            name_input = st.text_area("Enter drug substance names (one per line)",
                                       placeholder="amlodipine\nramipril\nbisoprolol",
                                       key="drug_name_input", height=120)
            if name_input.strip():
                names = [n.strip().lower() for n in name_input.strip().split("\n") if n.strip()]
                st.info(f"{len(names)} drug names entered")
                if st.button("🔍 Search by Name", key="btn_name_search"):
                    search_cols = [c for c in ["DrugSubstanceName", "ProductName", "Term"]
                                   if c in available_cols]
                    if search_cols:
                        mask = pd.Series(False, index=prod_dict.index)
                        for name in names:
                            for col in search_cols:
                                mask = mask | prod_dict[col].fillna("").str.lower().str.contains(name, regex=False)
                        matches = prod_dict[mask].copy()
                        _show_drug_matches(matches, "name_search")
                    else:
                        st.error("No searchable name columns found in Product Dictionary.")

    # ═══════════════════════════════════════════════════════════════════
    # TAB 2: Built-in Drug Libraries
    # ═══════════════════════════════════════════════════════════════════
    with tab_library:
        st.markdown("### Built-in Drug Code Libraries")

        lib_mode = st.radio("Library type",
                             ["📦 CV/Metabolic Drug Library (315 drugs, 18 classes)",
                              "📋 Legacy BNF-based Libraries"],
                             key="drug_lib_mode")

        if lib_mode.startswith("📦"):
            st.markdown("Search the Product Dictionary using **comprehensive generic + brand name** "
                         "matching from the curated CV/Cardiometabolic code list (v3).")
            sel_groups = st.multiselect("Select therapeutic classes",
                                         sorted(DRUG_GROUPS.keys()),
                                         default=list(DRUG_GROUPS.keys())[:3],
                                         key="dlk_drug_groups")
            avail_all = []
            for g in sel_groups:
                avail_all.extend(DRUG_GROUPS.get(g, []))

            # ── Pharmacological drug class filter ──
            dlk_cats = sorted({
                DRUG_CODE_LIBRARY.get(d, {}).get("cat", "")
                for d in avail_all
            } - {""})
            if dlk_cats:
                if "dlk_drug_cats" in st.session_state:
                    st.session_state["dlk_drug_cats"] = [
                        c for c in st.session_state["dlk_drug_cats"] if c in dlk_cats
                    ]
                sel_cats = st.multiselect(
                    f"Select pharmacological drug class ({len(dlk_cats)} available)",
                    dlk_cats, default=dlk_cats, key="dlk_drug_cats")
            else:
                sel_cats = []

            avail = [
                d for d in avail_all
                if DRUG_CODE_LIBRARY.get(d, {}).get("cat", "") in sel_cats
            ] if sel_cats else avail_all

            if avail:
                # Prune stale selections that are no longer in the options
                if "dlk_drug_sel" in st.session_state:
                    st.session_state["dlk_drug_sel"] = [
                        d for d in st.session_state["dlk_drug_sel"] if d in avail
                    ]
                sel = st.multiselect(f"Select drugs ({len(avail)} available)", avail,
                                      default=avail, key="dlk_drug_sel")
                all_terms = []
                for d in sel:
                    all_terms.extend(DRUG_CODE_LIBRARY.get(d, {}).get("terms", []))
                all_terms = sorted(set(all_terms))
                st.info(f"**{len(sel)} drugs** → **{len(all_terms)} search terms** (generic + brand names)")

                with st.expander(f"View all {len(all_terms)} search terms"):
                    for d in sorted(sel):
                        e = DRUG_CODE_LIBRARY.get(d, {})
                        st.caption(f"**{d}** ({e.get('cat','')}, {e.get('class','')}) → "
                                   f"{', '.join(e.get('terms', []))}")

                if prod_dict is None:
                    st.warning("Product Dictionary not available.")
                elif all_terms and st.button("🚀 Search Product Dictionary", key="btn_dcl_search", type="primary"):
                    search_cols = [c for c in ["Term", "ProductName", "DrugSubstanceName",
                                                "FormulationName"] if c in prod_dict.columns]
                    pid_col = "ProdCodeId" if "ProdCodeId" in prod_dict.columns else prod_dict.columns[0]

                    with st.spinner(f"Searching {len(all_terms)} terms across {len(prod_dict):,} products..."):
                        mask = pd.Series(False, index=prod_dict.index)
                        term_to_drug = {}
                        for d in sel:
                            for t in DRUG_CODE_LIBRARY.get(d, {}).get("terms", []):
                                term_to_drug[t.lower()] = d
                        type_map = {}
                        for term in all_terms:
                            t_lower = term.lower()
                            hit = pd.Series(False, index=prod_dict.index)
                            for col in search_cols:
                                hit = hit | prod_dict[col].fillna("").str.lower().str.contains(t_lower, regex=False)
                            mask = mask | hit
                            drug = term_to_drug.get(t_lower, "")
                            for idx in prod_dict.index[hit]:
                                if idx not in type_map:
                                    type_map[idx] = drug

                    matches = prod_dict[mask].copy()
                    matches["matched_drug"] = matches.index.map(lambda i: type_map.get(i, ""))
                    matches["drug_type"] = matches["matched_drug"].map(
                        lambda d: DRUG_CODE_LIBRARY.get(d, {}).get("class", ""))
                    matches = matches.reset_index(drop=True)

                    # Store for persistence across reruns
                    st.session_state["_res_dlk_library"] = matches

                    # ── Auto-save as Parquet ──
                    if len(matches) > 0:
                        _label = _build_save_label("DrugLookup",
                            drug_classes=st.session_state.get("dlk_drug_groups"))
                        cfg = _get_output_settings()
                        out_dir = cfg["directory"]
                        try:
                            os.makedirs(out_dir, exist_ok=True)
                            safe_label = re.sub(r"[^A-Za-z0-9_-]", "_", _label).lower()
                            ts = time.strftime("%Y%m%d_%H%M%S")
                            pq_path = os.path.join(out_dir, f"{safe_label}_{ts}.parquet")
                            export_df = matches.copy()
                            for col in export_df.columns:
                                if export_df[col].dtype == object:
                                    export_df[col] = export_df[col].astype(str)
                            if HAS_PYARROW:
                                export_df.to_parquet(pq_path, index=False, engine="pyarrow")
                                st.toast(f"💾 Saved: {os.path.basename(pq_path)}", icon="✅")
                            else:
                                csv_fb = pq_path.replace(".parquet", ".csv")
                                export_df.to_csv(csv_fb, index=False)
                                st.warning("⚠️ pyarrow not installed — saved as CSV. "
                                           "Run `pip install pyarrow --user` for Parquet.")
                        except Exception as exc:
                            st.warning(f"⚠️ Auto-save failed: {exc}")

                # ── Persistent Drug Lookup library results ──
                if "_res_dlk_library" in st.session_state and st.session_state["_res_dlk_library"] is not None:
                    _dlk_matches = st.session_state["_res_dlk_library"]
                    st.markdown(f'<div class="success-box">✅ Found <strong>{len(_dlk_matches):,}</strong> products '
                                f'across <strong>{_dlk_matches["matched_drug"].nunique()}</strong> drugs</div>',
                                unsafe_allow_html=True)
                    _show_drug_matches(_dlk_matches, "dcl_search")

        else:
            # ── Legacy BNF-based Libraries ──
            st.markdown("Select a pre-built drug library to search the Product Dictionary. "
                         "Libraries combine **BNF code prefix matching** and **drug substance name matching** "
                         "for comprehensive coverage.")

            lib_name = st.selectbox("Select Library", list(DRUG_LIBRARIES.keys()),
                                     key="drug_lib_select")
            lib = DRUG_LIBRARIES[lib_name]

            # Show library contents
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("**BNF Code Prefixes**")
                bnf_data = []
                for drug_type, codes in lib["bnf_codes"].items():
                    for code in codes:
                        bnf_data.append({"Drug Type": drug_type, "BNF Code": code})
                bnf_df = pd.DataFrame(bnf_data)
                st.dataframe(bnf_df, height=300, hide_index=True)
                n_bnf = len(bnf_df)
                st.caption(f"{n_bnf} BNF codes across {len(lib['bnf_codes'])} drug types")

            with col2:
                st.markdown("**Drug Substance Names**")
                name_data = [{"Drug Name": name, "Drug Type": dtype}
                             for name, dtype in lib["drug_names"].items()]
                name_df = pd.DataFrame(name_data)
                st.dataframe(name_df, height=300, hide_index=True)
                st.caption(f"{len(name_df)} drug names")

            # Select which drug types to include
            all_types = sorted(set(list(lib["bnf_codes"].keys()) +
                                   list(set(lib["drug_names"].values()))))
            selected_types = st.multiselect("Filter by drug type (or leave empty for all)",
                                             all_types, key="drug_lib_types")

            if prod_dict is None:
                st.warning("Product Dictionary not available. Cannot run search.")
            elif st.button(f"🚀 Search Product Dictionary with {lib_name}", key="btn_lib_search", type="primary"):
                available_cols = prod_dict.columns.tolist()
                all_matches = []

                # ── BNF code matching ──
                if "BNFChapter" in available_cols:
                    bnf_prefixes = []
                    for drug_type, codes in lib["bnf_codes"].items():
                        if not selected_types or drug_type in selected_types:
                            for code in codes:
                                bnf_prefixes.append((code, drug_type))

                    if bnf_prefixes:
                        with st.spinner(f"Searching {len(bnf_prefixes)} BNF prefixes..."):
                            mask = pd.Series(False, index=prod_dict.index)
                            type_map = {}
                            for prefix, dtype in bnf_prefixes:
                                p_lower = prefix.lower()
                                p_no_zero = p_lower.lstrip("0") if p_lower.startswith("0") else p_lower
                                hit = (
                                    prod_dict["BNFChapter"].fillna("").str.lower().str.contains(p_lower, regex=False) |
                                    prod_dict["BNFChapter"].fillna("").str.lower().str.contains(p_no_zero, regex=False)
                                )
                                mask = mask | hit
                                # Track drug type for matched rows
                                for idx in prod_dict.index[hit]:
                                    if idx not in type_map:
                                        type_map[idx] = dtype

                            bnf_matches = prod_dict[mask].copy()
                            bnf_matches["drug_type"] = bnf_matches.index.map(lambda i: type_map.get(i, ""))
                            bnf_matches["match_source"] = "BNF"
                            all_matches.append(bnf_matches)

                # ── Drug name matching ──
                drug_names_to_search = {name: dtype for name, dtype in lib["drug_names"].items()
                                        if not selected_types or dtype in selected_types}
                if drug_names_to_search:
                    search_cols = [c for c in ["DrugSubstanceName", "ProductName", "Term"]
                                   if c in available_cols]
                    if search_cols:
                        with st.spinner(f"Searching {len(drug_names_to_search)} drug names..."):
                            mask = pd.Series(False, index=prod_dict.index)
                            type_map = {}
                            for name, dtype in drug_names_to_search.items():
                                name_lower = name.lower()
                                hit = pd.Series(False, index=prod_dict.index)
                                for col in search_cols:
                                    hit = hit | prod_dict[col].fillna("").str.lower().str.contains(name_lower, regex=False)
                                mask = mask | hit
                                for idx in prod_dict.index[hit]:
                                    if idx not in type_map:
                                        type_map[idx] = dtype

                            name_matches = prod_dict[mask].copy()
                            name_matches["drug_type"] = name_matches.index.map(lambda i: type_map.get(i, ""))
                            name_matches["match_source"] = "DrugName"
                            all_matches.append(name_matches)

                # ── Combine and deduplicate ──
                if all_matches:
                    combined = pd.concat(all_matches, ignore_index=False)
                    # BNF match takes priority
                    combined = combined.sort_values("match_source")  # "BNF" sorts before "DrugName"
                    pid_col = "ProdCodeId" if "ProdCodeId" in combined.columns else combined.columns[0]
                    combined = combined.drop_duplicates(subset=[pid_col], keep="first")
                    combined = combined.reset_index(drop=True)

                    n_bnf_only = (combined["match_source"] == "BNF").sum()
                    n_name_only = (combined["match_source"] == "DrugName").sum()

                    st.markdown(f'<div class="success-box">✅ Found <strong>{len(combined):,}</strong> unique products '
                                f'({n_bnf_only:,} via BNF, {n_name_only:,} via name only)</div>',
                                unsafe_allow_html=True)
                    if n_name_only > 0:
                        st.info(f"💡 {n_name_only:,} products were found **only** by drug name (no BNF match) — "
                                f"these would be missed without name-based searching.")

                    # Store for persistence across reruns
                    st.session_state["_res_dlk_legacy"] = combined

                    # ── Auto-save as Parquet ──
                    _label = _build_save_label("DrugLookup_Legacy",
                        drug_classes=[lib_name])
                    cfg = _get_output_settings()
                    out_dir = cfg["directory"]
                    try:
                        os.makedirs(out_dir, exist_ok=True)
                        safe_label = re.sub(r"[^A-Za-z0-9_-]", "_", _label).lower()
                        ts = time.strftime("%Y%m%d_%H%M%S")
                        pq_path = os.path.join(out_dir, f"{safe_label}_{ts}.parquet")
                        export_df = combined.copy()
                        for col in export_df.columns:
                            if export_df[col].dtype == object:
                                export_df[col] = export_df[col].astype(str)
                        if HAS_PYARROW:
                            export_df.to_parquet(pq_path, index=False, engine="pyarrow")
                            st.toast(f"💾 Saved: {os.path.basename(pq_path)}", icon="✅")
                        else:
                            csv_fb = pq_path.replace(".parquet", ".csv")
                            export_df.to_csv(csv_fb, index=False)
                            st.warning("⚠️ pyarrow not installed — saved as CSV.")
                    except Exception as exc:
                        st.warning(f"⚠️ Auto-save failed: {exc}")

                else:
                    st.warning("No matches found.")

            # ── Persistent legacy library results ──
            if "_res_dlk_legacy" in st.session_state and st.session_state["_res_dlk_legacy"] is not None:
                _leg = st.session_state["_res_dlk_legacy"]
                st.markdown(f'<div class="success-box">✅ Found <strong>{len(_leg):,}</strong> unique products</div>',
                            unsafe_allow_html=True)
                _show_drug_matches(_leg, "lib_search")


        # ═══════════════════════════════════════════════════════════════════
        # TAB 3: My Drug Code List (accumulated results)
        # ═══════════════════════════════════════════════════════════════════
    with tab_results:
        st.markdown("### My Drug Code List")
        drug_list = st.session_state.get("drug_lookup_results")
        if drug_list is not None and len(drug_list) > 0:
            pid_col = "ProdCodeId" if "ProdCodeId" in drug_list.columns else drug_list.columns[0]
            st.markdown(f"**{len(drug_list):,} products** in your current list "
                        f"({drug_list[pid_col].nunique():,} unique ProdCodeIds)")

            # Filter by drug type if available
            if "drug_type" in drug_list.columns:
                types = sorted(drug_list["drug_type"].dropna().unique())
                if types:
                    selected = st.multiselect("Filter by drug type", types, default=types,
                                               key="drug_results_filter")
                    drug_list = drug_list[drug_list["drug_type"].isin(selected)]

            st.dataframe(drug_list, height=400, hide_index=True)

            # Downloads
            col1, col2, col3 = st.columns(3)
            with col1:
                csv_data = drug_list.to_csv(index=False).encode("utf-8")
                st.download_button("📥 Download Full List (CSV)", csv_data,
                                    "drug_codelist.csv", "text/csv",
                                    key="btn_dl_drug_full")
            with col2:
                prodcodes_only = drug_list[[pid_col]].drop_duplicates()
                csv_pc = prodcodes_only.to_csv(index=False).encode("utf-8")
                st.download_button(f"📥 Download ProdCodeIds Only ({len(prodcodes_only):,})",
                                    csv_pc, "prodcodeids.csv", "text/csv",
                                    key="btn_dl_prodcodes")
            with col3:
                # Push to Drug Issue extraction
                if st.button("🚀 Push to Drug Extraction", key="btn_push_drug", type="primary"):
                    codes = drug_list[pid_col].dropna().astype(str).unique().tolist()
                    st.session_state["_drug_lookup_prodcodes"] = codes
                    st.success(f"✅ {len(codes):,} ProdCodeIds ready! Go to **CPRD Aurum Extraction → Drug Issue** tab.")

            st.markdown("---")
            if st.button("🗑️ Clear Drug Code List", key="btn_clear_drug_list"):
                st.session_state.pop("drug_lookup_results", None)
                st.rerun()
        else:
            st.info("No drug codes yet. Use the **Search** or **Built-in Libraries** tab to find products, "
                     "then click 'Add to My Drug Code List'.")


def _show_drug_matches(matches, key_prefix):
    """Display drug search results with add-to-list and download buttons."""
    if matches is None or len(matches) == 0:
        st.warning("No matching products found.")
        return

    pid_col = "ProdCodeId" if "ProdCodeId" in matches.columns else matches.columns[0]
    st.markdown(f"**{len(matches):,} products found** ({matches[pid_col].nunique():,} unique ProdCodeIds)")

    # Show summary by drug type if available
    if "drug_type" in matches.columns:
        type_summary = matches.groupby("drug_type").agg(
            Products=(pid_col, "count"),
        ).reset_index().sort_values("Products", ascending=False)
        st.dataframe(type_summary, hide_index=True)

    # Show data
    st.dataframe(matches.head(500), height=350, hide_index=True)
    if len(matches) > 500:
        st.caption(f"Showing first 500 of {len(matches):,} rows")

    col1, col2 = st.columns(2)
    with col1:
        csv_data = matches.to_csv(index=False).encode("utf-8")
        st.download_button("📥 Download Results (CSV)", csv_data,
                            f"drug_search_{key_prefix}.csv", "text/csv",
                            key=f"btn_dl_{key_prefix}")
    with col2:
        if st.button("➕ Add to My Drug Code List", key=f"btn_add_{key_prefix}", type="primary"):
            existing = st.session_state.get("drug_lookup_results")
            if existing is not None and len(existing) > 0:
                combined = pd.concat([existing, matches], ignore_index=True)
                combined = combined.drop_duplicates(subset=[pid_col], keep="first")
            else:
                combined = matches.copy()
            st.session_state["drug_lookup_results"] = combined
            st.success(f"✅ Added! Total: {len(combined):,} products in your list.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: ANALYTICS
# ══════════════════════════════════════════════════════════════════════════════

def page_analytics():
    st.title("📊 Analytics")
    show_tooltip("nav_analytics")

    data_source = st.radio("Data Source", [
        "Generate Mock Analytics Data",
        "Use Last Extraction",
        "Upload Data File",
    ], horizontal=True, key="analytics_source")

    df = None

    if data_source == "Generate Mock Analytics Data":
        if st.button("Generate Mock Data", key="btn_gen_mock_analytics"):
            mock = generate_mock_data()
            patients = mock["patient"].copy()
            patients["yob"] = pd.to_numeric(patients["yob"], errors="coerce")
            patients["age"] = 2020 - patients["yob"]
            patients["gender"] = pd.to_numeric(patients["gender"], errors="coerce")
            patients["gender_label"] = patients["gender"].map({1: "Male", 2: "Female"})
            patients["hypertension"] = np.random.choice([0, 1], size=len(patients), p=[0.4, 0.6])
            patients["diabetes"] = np.random.choice([0, 1], size=len(patients), p=[0.7, 0.3])
            patients["af"] = np.random.choice([0, 1], size=len(patients), p=[0.6, 0.4])
            patients["heart_failure"] = np.random.choice([0, 1], size=len(patients), p=[0.75, 0.25])
            patients["systolic_bp"] = np.random.normal(135, 20, size=len(patients)).round(0)
            patients["bmi"] = np.random.normal(27, 5, size=len(patients)).round(1)
            patients["hba1c"] = np.random.normal(48, 12, size=len(patients)).round(1)
            patients["followup_days"] = np.random.exponential(1000, size=len(patients)).astype(int)
            patients["event"] = np.random.choice([0, 1], size=len(patients), p=[0.7, 0.3])
            patients["treatment_group"] = np.random.choice(["Surgery", "Medical"], size=len(patients))
            patients["imd_quintile"] = np.random.randint(1, 6, size=len(patients))
            df = patients
            st.session_state["analytics_data"] = df
            st.success(f"Generated analytics dataset: {len(df):,} patients")

    elif data_source == "Use Last Extraction":
        if "last_extraction" in st.session_state and st.session_state["last_extraction"] is not None:
            df = st.session_state["last_extraction"]
            st.success(f"Using last extraction: {len(df):,} rows")
        else:
            st.warning("No previous extraction. Run an extraction first, or generate mock data.")

    elif data_source == "Upload Data File":
        uploaded = st.file_uploader("Upload CSV/XLSX", type=["csv", "xlsx"], key="analytics_upload")
        if uploaded:
            if uploaded.name.endswith(".xlsx"):
                df = pd.read_excel(uploaded)
            else:
                df = pd.read_csv(uploaded)
            st.success(f"Loaded: {len(df):,} rows, {len(df.columns)} columns")

    if df is None and "analytics_data" in st.session_state:
        df = st.session_state["analytics_data"]

    if df is None:
        st.info("Please select or generate data above to begin analytics.")
        return

    st.dataframe(df.head(20), width='stretch')

    tab1, tab2, tab3, tab4 = st.tabs([
        "📋 Table 1 (Baseline)", "📈 Kaplan-Meier", "📊 Distributions", "🔍 Custom Summary"
    ])

    with tab1:
        st.markdown("### Baseline Characteristics (Table 1)")
        numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        categorical_cols = [c for c in df.columns if df[c].nunique() <= 10 and c in df.select_dtypes(include=[np.number, "object", "category"]).columns]
        col1, col2 = st.columns(2)
        with col1:
            selected_numeric = st.multiselect("Continuous Variables", numeric_cols,
                                               default=[c for c in ["age", "systolic_bp", "bmi", "hba1c", "followup_days"] if c in numeric_cols],
                                               key="t1_numeric")
        with col2:
            selected_categorical = st.multiselect("Categorical Variables", categorical_cols,
                                                   default=[c for c in ["gender", "hypertension", "diabetes", "af", "heart_failure", "imd_quintile"] if c in categorical_cols],
                                                   key="t1_categorical")
        group_col = st.selectbox("Group By (optional)", ["None"] + categorical_cols, key="t1_group")

        if st.button("Generate Table 1", key="btn_table1", type="primary"):
            try:
                table_rows = []
                if group_col == "None":
                    table_rows.append({"Variable": "**N**", "Value": f"{len(df):,}"})
                    for col in selected_numeric:
                        vals = pd.to_numeric(df[col], errors="coerce").dropna()
                        table_rows.append({"Variable": col, "Value": f"{vals.mean():.1f} ± {vals.std():.1f}"})
                    for col in selected_categorical:
                        for val in sorted(df[col].dropna().unique()):
                            count = (df[col] == val).sum()
                            pct = count / len(df) * 100
                            table_rows.append({"Variable": f"{col} = {val}", "Value": f"{count:,} ({pct:.1f}%)"})
                else:
                    groups = sorted(df[group_col].dropna().unique())
                    header = {"Variable": "**N**"}
                    for g in groups:
                        gdf = df[df[group_col] == g]
                        header[f"{group_col}={g}"] = f"{len(gdf):,}"
                    table_rows.append(header)
                    for col in selected_numeric:
                        row = {"Variable": col}
                        for g in groups:
                            vals = pd.to_numeric(df[df[group_col] == g][col], errors="coerce").dropna()
                            row[f"{group_col}={g}"] = f"{vals.mean():.1f} ± {vals.std():.1f}"
                        table_rows.append(row)
                    for col in selected_categorical:
                        for val in sorted(df[col].dropna().unique()):
                            row = {"Variable": f"{col} = {val}"}
                            for g in groups:
                                gdf = df[df[group_col] == g]
                                count = (gdf[col] == val).sum()
                                pct = count / len(gdf) * 100 if len(gdf) > 0 else 0
                                row[f"{group_col}={g}"] = f"{count:,} ({pct:.1f}%)"
                            table_rows.append(row)
                t1_df = pd.DataFrame(table_rows)
                st.dataframe(t1_df, width='stretch', hide_index=True)
                st.download_button("📥 Download Table 1", t1_df.to_csv(index=False), "table1.csv")
            except Exception as e:
                st.error(f"Error generating Table 1: {e}")

    with tab2:
        st.markdown("### Kaplan-Meier Survival Analysis")
        time_col = st.selectbox("Time Variable (days)",
                                [c for c in numeric_cols if "time" in c.lower() or "day" in c.lower() or "followup" in c.lower()] or numeric_cols, key="km_time")
        event_col = st.selectbox("Event Variable (0/1)",
                                 [c for c in numeric_cols if "event" in c.lower() or "death" in c.lower() or "censor" in c.lower()] or numeric_cols, key="km_event")
        strata_col = st.selectbox("Stratify By (optional)", ["None"] + categorical_cols, key="km_strata")

        if st.button("Generate KM Curve", key="btn_km", type="primary"):
            try:
                from lifelines import KaplanMeierFitter
                fig = go.Figure()
                colors = ["#2563eb", "#dc2626", "#16a34a", "#d97706", "#7c3aed"]
                if strata_col == "None":
                    kmf = KaplanMeierFitter()
                    T = pd.to_numeric(df[time_col], errors="coerce").dropna()
                    E = pd.to_numeric(df[event_col], errors="coerce").fillna(0)
                    T, E = T.align(E, join="inner")
                    kmf.fit(T, E, label="Overall")
                    fig.add_trace(go.Scatter(
                        x=kmf.survival_function_.index,
                        y=kmf.survival_function_["Overall"],
                        mode="lines", name="Overall", line=dict(color=colors[0], width=2)
                    ))
                else:
                    for idx, group in enumerate(sorted(df[strata_col].dropna().unique())):
                        mask = df[strata_col] == group
                        T = pd.to_numeric(df.loc[mask, time_col], errors="coerce").dropna()
                        E = pd.to_numeric(df.loc[mask, event_col], errors="coerce").fillna(0)
                        T, E = T.align(E, join="inner")
                        if len(T) > 0:
                            kmf = KaplanMeierFitter()
                            kmf.fit(T, E, label=str(group))
                            fig.add_trace(go.Scatter(
                                x=kmf.survival_function_.index,
                                y=kmf.survival_function_[str(group)],
                                mode="lines", name=str(group),
                                line=dict(color=colors[idx % len(colors)], width=2)
                            ))
                fig.update_layout(title="Kaplan-Meier Survival Curve",
                                  xaxis_title="Time (days)", yaxis_title="Survival Probability",
                                  yaxis=dict(range=[0, 1.05]), template="plotly_white", height=500)
                st.plotly_chart(fig, width='stretch', key="plotly_11")
            except ImportError:
                st.error("lifelines not installed. Run: pip install lifelines")
            except Exception as e:
                st.error(f"Error: {e}")

    with tab3:
        st.markdown("### Variable Distributions")
        dist_col = st.selectbox("Select Variable", numeric_cols, key="dist_col")
        dist_group = st.selectbox("Group By (optional)", ["None"] + categorical_cols, key="dist_group")
        if dist_col:
            vals = pd.to_numeric(df[dist_col], errors="coerce").dropna()
            if dist_group == "None":
                fig = px.histogram(vals, nbins=50, title=f"Distribution of {dist_col}",
                                   labels={"value": dist_col})
            else:
                plot_df = df[[dist_col, dist_group]].copy()
                plot_df[dist_col] = pd.to_numeric(plot_df[dist_col], errors="coerce")
                fig = px.histogram(plot_df.dropna(), x=dist_col, color=dist_group, nbins=50,
                                   title=f"Distribution of {dist_col} by {dist_group}",
                                   barmode="overlay", opacity=0.7)
            fig.update_layout(template="plotly_white")
            st.plotly_chart(fig, width='stretch', key="plotly_12")

    with tab4:
        st.markdown("### Custom Summary Statistics")
        summary_cols = st.multiselect("Select columns to summarize", df.columns.tolist(), key="summary_cols")
        if summary_cols:
            summary = df[summary_cols].describe(include="all").T
            st.dataframe(summary, width='stretch')
            st.markdown("#### Missing Data")
            missing = df[summary_cols].isnull().sum().reset_index()
            missing.columns = ["Column", "Missing"]
            missing["Pct"] = (missing["Missing"] / len(df) * 100).round(1)
            st.dataframe(missing, width='stretch', hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: CONFIGURATION (cleaned up — removed obsolete "Rerun R")
# ══════════════════════════════════════════════════════════════════════════════

def page_config():
    st.title("⚙️ Configuration & Data Paths")
    show_tooltip("nav_config")

    tab_paths, tab_output, tab_ssh = st.tabs(["📂 Data Paths", "💾 Output Settings", "🔌 SSH Connection"])

    with tab_paths:
        st.markdown("### CPRD Aurum Practice Folders")
    st.markdown(f'<div class="path-box">📂 Base: {PATHS["aurum_base"]}</div>', unsafe_allow_html=True)

    folders = st.session_state.engine.find_practice_folders()
    st.metric("Practice Folders Found", len(folders))
    if folders:
        with st.expander(f"View all {len(folders)} practice folders"):
            for f in folders:
                st.text(f)

    st.markdown("### Linkage Data Files")
    linkage_files = {
        "HES Diagnosis (Hospital)": PATHS["hes_diagnosis_hosp"],
        "HES Diagnosis (Episode)": PATHS["hes_diagnosis_epi"],
        "HES Primary Diagnosis": PATHS["hes_primary_diag"],
        "HES Episodes": PATHS["hes_episodes"],
        "HES Hospital": PATHS["hes_hospital"],
        "HES Procedures": PATHS["hes_procedures_epi"],
        "HES OP Clinical": PATHS["hesop_clinical"],
        "HES OP Patient": PATHS["hesop_patient"],
        "HES A&E Attendance": PATHS["hesae_attendance"],
        "HES A&E Diagnosis": PATHS["hesae_diagnosis"],
        "ONS Death": PATHS["death_patient"],
        "HES Patient": PATHS["hes_patient"],
        "HES Maternity": PATHS["hes_maternity"],
        "HES Critical Care": PATHS["hes_ccare"],
        "Patient IMD": PATHS["patient_imd"],
        "Practice IMD": PATHS["practice_imd"],
        "Linkage Eligibility (main)": PATHS["linkage_eligibility"],
        "Linkage Eligibility (Aurum)": PATHS["linkage_eligibility_aurum"],
        "EMIS Dictionary": PATHS["emis_dictionary"],
    }
    status_data = []
    for name, path in linkage_files.items():
        exists = os.path.exists(path) if not st.session_state.engine.is_mock() else True
        status_data.append({"File": name, "Path": path, "Status": "✅" if exists else "❌"})
    st.dataframe(pd.DataFrame(status_data), width='stretch', hide_index=True)

    st.markdown("### Linkage Coverage")
    coverage_data = [
        {"Source": "HES APC", "Start": "01/04/1997", "End": "31/10/2020"},
        {"Source": "HES OP", "Start": "01/04/2003", "End": "31/10/2020"},
        {"Source": "HES A&E", "Start": "01/04/2007", "End": "31/03/2020"},
        {"Source": "ONS Death", "Start": "02/01/1998", "End": "16/11/2020"},
        {"Source": "HES Critical Care", "Start": "01/04/2008", "End": "31/10/2020"},
    ]
    st.dataframe(pd.DataFrame(coverage_data), width='stretch', hide_index=True)

    with tab_output:
        st.markdown("### Auto-Save Configuration")
        st.markdown("Control how and where extraction results are automatically saved to disk.")

        col_o1, col_o2 = st.columns(2)
        with col_o1:
            auto_on = st.toggle(
                "Enable auto-save",
                value=st.session_state.get("output_auto_save", True),
                key="_cfg_auto_save",
                help="Automatically save every extraction result to disk.",
            )
            st.session_state["output_auto_save"] = auto_on

            fmt = st.selectbox(
                "Default file format",
                ["Parquet", "CSV", "Both"],
                index=["Parquet", "CSV", "Both"].index(
                    st.session_state.get("output_format", "Parquet")),
                key="_cfg_format",
                help="Parquet is compact and fast. CSV is universally readable. "
                     "'Both' writes one file in each format.",
            )
            st.session_state["output_format"] = fmt

            if not HAS_PYARROW and fmt in ("Parquet", "Both"):
                st.warning("⚠️ `pyarrow` is not installed — Parquet will fall back to CSV. "
                           "Install it with:  `pip install pyarrow --user`")

        with col_o2:
            enrich_on = st.toggle(
                "Include code descriptions in saved files",
                value=st.session_state.get("output_enrich", True),
                key="_cfg_enrich",
                help="Merge human-readable code descriptions before saving.",
            )
            st.session_state["output_enrich"] = enrich_on

        st.markdown("### Output Directory")
        cur_dir = st.session_state.get("output_directory", _DEFAULT_OUTPUT_DIR)
        new_dir = st.text_input(
            "Path (absolute)", value=cur_dir, key="_cfg_dir",
            help="Full path where files are saved. Created automatically if missing.",
        )
        st.session_state["output_directory"] = new_dir

        # Validate & show status
        if os.path.isdir(new_dir):
            files_in = [f for f in os.listdir(new_dir) if f.endswith((".csv", ".parquet"))]
            total_size = sum(os.path.getsize(os.path.join(new_dir, f)) for f in files_in)
            sz_str = f"{total_size / 1_048_576:.1f} MB" if total_size > 1_048_576 else f"{total_size / 1024:.1f} KB"
            st.success(f"✅ Directory exists — {len(files_in)} output file(s), {sz_str} total")

            if files_in:
                with st.expander(f"View {len(files_in)} saved files"):
                    file_info = []
                    for fn in sorted(files_in, reverse=True):
                        fp = os.path.join(new_dir, fn)
                        sz = os.path.getsize(fp)
                        mt = time.strftime("%Y-%m-%d %H:%M", time.localtime(os.path.getmtime(fp)))
                        ext = os.path.splitext(fn)[1].upper().replace(".", "")
                        file_info.append({"File": fn, "Format": ext, "Size": f"{sz / 1024:.1f} KB", "Modified": mt})
                    st.dataframe(pd.DataFrame(file_info), width='stretch', hide_index=True)

                if st.button("🗑️ Clear all output files", key="_cfg_clear"):
                    for fn in files_in:
                        os.remove(os.path.join(new_dir, fn))
                    st.success(f"Deleted {len(files_in)} file(s)")
                    st.rerun()
        elif new_dir:
            st.info(f"📁 Directory will be created on first save: `{new_dir}`")

        st.markdown("### Summary")
        cfg = _get_output_settings()
        status_icon = "🟢" if cfg["auto_save"] else "🔴"
        st.markdown(f"""
        | Setting | Value |
        |---------|-------|
        | Auto-save | {status_icon} **{'ON' if cfg['auto_save'] else 'OFF'}** |
        | Format | **{cfg['format']}** |
        | Enrichment | **{'Yes' if cfg['enrich'] else 'No'}** |
        | Directory | `{cfg['directory']}` |
        """)

    with tab_ssh:
        render_ssh_connection_panel()

def page_definitions():
    st.title("📖 Term Definitions & Reference")
    st.markdown("Quick reference for all CPRD data tables, code systems, and linked datasets used in this platform.")

    # ── Main definitions table ──
    st.markdown("### CPRD & Linkage Term Reference")
    def_df = pd.DataFrame(CPRD_DEFINITIONS)
    st.dataframe(def_df, width='stretch', hide_index=True, height=700)

    # ── Download as CSV ──
    csv = def_df.to_csv(index=False)
    st.download_button("📥 Download Definitions Table", csv, "cprd_definitions.csv",
                       "text/csv", key="dl_definitions")

    # ── Embedded ICD-10 code reference ──
    st.markdown("---")
    st.markdown("### ICD-10 Code Quick Reference (VHD & Comorbidities)")
    icd_rows = [{"ICD-10 Code": k, "Description": v} for k, v in ICD10_DESCRIPTIONS.items()]
    icd_df = pd.DataFrame(icd_rows)
    st.dataframe(icd_df, width='stretch', hide_index=True, height=400)
    icd_csv = icd_df.to_csv(index=False)
    st.download_button("📥 Download ICD-10 Reference", icd_csv, "icd10_reference.csv",
                       "text/csv", key="dl_icd10_ref")

    # ── Abbreviations ──
    st.markdown("---")
    st.markdown("### Abbreviations")
    st.markdown(LINKAGE_ABBREVIATIONS)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: SSH CONNECTION (embedded in Configuration)
# ══════════════════════════════════════════════════════════════════════════════

def render_ssh_connection_panel():
    """Render SSH connection info with auto-detected username and cross-platform snippets."""
    st.markdown("### 🔌 SSH Tunnel Connection")
    st.markdown("Connect to the BMRC cluster Streamlit app from your local machine.")

    username = get_current_username()
    user_override = st.text_input("Your BMRC username", value=username, key="ssh_username",
                                  help="Auto-detected from your session. Edit if needed.")
    port = st.number_input("Streamlit port", value=SSH_CONFIG["default_port"],
                           min_value=1024, max_value=65535, key="ssh_port")

    snippets, compute_node, remote_port = generate_ssh_snippets(user_override, port)

    st.markdown(f"**Detected compute node:** `{compute_node}`")
    if compute_node == "<compute_node>":
        st.warning("Not running on BMRC — replace `<compute_node>` with your actual SLURM node name (e.g. `comp001`).")

    # ── Cross-platform snippets ──
    for key in ["windows_powershell", "macos", "linux"]:
        snip = snippets[key]
        with st.expander(f"**{snip['title']}**"):
            for cmd in snip["commands"]:
                st.code(cmd, language="bash")
            st.caption(snip["note"])
            st.markdown(f"**Browser URL:** `http://localhost:{port}`")

    # ── Keepalive / stability explanation ──
    st.markdown("---")
    st.markdown("#### 🛡️ Preventing 'Connection Refused' Crashes")
    st.info(snippets["keepalive_info"])
    st.markdown("""
**Why does the tunnel drop every 10–15 minutes?**

SSH tunnels are killed by intermediate firewalls/NAT routers when no traffic flows for their idle timeout
(often 10–15 min on institutional networks). The keepalive flags above send periodic packets to prevent this.

**Recommended `.ssh/config` entry** (add to `~/.ssh/config` on your local machine):
    """)
    ssh_config_block = f"""Host bmrc-tunnel
    HostName {detect_login_node()}
    User {user_override}
    LocalForward {port} {compute_node}:{remote_port}
    ServerAliveInterval {SSH_CONFIG["heartbeat_interval"]}
    ServerAliveCountMax 3
    TCPKeepAlive yes
    Compression yes"""
    st.code(ssh_config_block, language="text")
    st.caption("After adding this, connect with: `ssh bmrc-tunnel`")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN APPLICATION
# ══════════════════════════════════════════════════════════════════════════════

def main():
    st.set_page_config(
        page_title="CPRD Extractor v1.0",
        page_icon="🏥",
        layout="wide",
        initial_sidebar_state="expanded"
    )

    apply_custom_css()

    if "engine" not in st.session_state:
        st.session_state.engine = CPRDEngine()
    if "data_mode" not in st.session_state:
        st.session_state.data_mode = "mock" if st.session_state.engine.is_mock() else "bmrc"

    with st.sidebar:
        # ── Modern sidebar branding ──
        st.markdown("""
        <div style="text-align:center; padding:8px 0 4px 0;">
            <span style="font-size:1.2rem; font-weight:700; letter-spacing:-0.3px;">🗂️ CPRD Extractor</span><br/>
            <span style="font-size:0.7rem; opacity:0.5; letter-spacing:0.4px;">v1.0 · University of Oxford</span>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════
        # DATA MODE SELECTOR
        # ══════════════════════════════════════════════════════════════════
        st.markdown("### Data Mode")
        data_mode = st.radio(
            "Select environment",
            ["🧪 Mock Data (Testing)", "🔬 Live — BMRC Cluster",
             "🖥️ Live — Any Server", "💻 Windows OS"],
            key="data_mode_radio",
            label_visibility="collapsed",
        )

        if "Mock" in data_mode:
            if not st.session_state.engine.is_mock():
                st.session_state.engine.use_mock_data()
                st.session_state.data_mode = "mock"
                st.rerun()
            st.session_state.data_mode = "mock"

        elif "BMRC" in data_mode:
            if st.session_state.engine.is_mock():
                _reconfigure_paths(_DEFAULT_CPRD_BASE)
                st.session_state.engine = CPRDEngine()
                st.session_state.data_mode = "bmrc"
            st.session_state.data_mode = "bmrc"

        elif "Any Server" in data_mode:
            st.session_state.data_mode = "any_server"
            custom_path = st.text_input(
                "CPRD data root folder",
                value=st.session_state.get("custom_cprd_path", "/path/to/CPRD"),
                key="custom_cprd_path_input",
                help="Top-level folder containing practice_* directories, lookups, and linkage data"
            )
            if st.button("📂 Set Path", key="btn_set_custom_path"):
                st.session_state.custom_cprd_path = custom_path
                _reconfigure_paths(custom_path)
                st.session_state.engine = CPRDEngine()
                st.success(f"Path set to: {custom_path}")
                st.rerun()
            # Auto-detect folder structure
            if os.path.isdir(custom_path):
                detected = []
                for item in ["practice_*", "202102_lookups", "linkage"]:
                    if glob.glob(os.path.join(custom_path, item)):
                        detected.append(item.replace("_*", "s"))
                if detected:
                    st.caption(f"Found: {', '.join(detected)}")

        elif "Windows" in data_mode:
            st.session_state.data_mode = "windows"
            win_path = st.text_input(
                "CPRD data folder (Windows path)",
                value=st.session_state.get("windows_cprd_path", r"C:\CPRD_Data"),
                key="windows_cprd_path_input",
                help="e.g., C:\\Users\\milad\\CPRD_Data or D:\\Research\\CPRD"
            )
            if st.button("📂 Set Path", key="btn_set_win_path"):
                # Convert Windows backslashes for Python compatibility
                norm_path = win_path.replace("\\", "/") if platform.system() != "Windows" else win_path
                st.session_state.windows_cprd_path = win_path
                _reconfigure_paths(norm_path)
                st.session_state.engine = CPRDEngine()
                st.success(f"Path set to: {win_path}")
                st.rerun()
            st.caption("ℹ️ Zip extraction uses Python — no Unix tools needed.")

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════
        # OUTPUT SETTINGS
        # ══════════════════════════════════════════════════════════════════
        with st.expander("💾 Output Settings", expanded=False):
            st.session_state["output_auto_save"] = st.toggle(
                "Auto-save extractions to disk",
                value=st.session_state.get("output_auto_save", True),
                key="_output_auto_save_toggle",
                help="When enabled, every extraction is automatically saved "
                     "to the output directory in your chosen format.",
            )

            st.session_state["output_format"] = st.selectbox(
                "File format",
                ["Parquet", "CSV", "Both"],
                index=["Parquet", "CSV", "Both"].index(
                    st.session_state.get("output_format", "Parquet")),
                key="_output_format_sel",
                help="Parquet is ~5× smaller and ~10× faster to read than CSV. "
                     "'Both' saves one file in each format.",
            )
            if not HAS_PYARROW and st.session_state["output_format"] in ("Parquet", "Both"):
                st.caption("⚠️ `pyarrow` not found — will fall back to CSV. "
                           "Run `pip install pyarrow --user`.")

            default_dir = st.session_state.get("output_directory", _DEFAULT_OUTPUT_DIR)
            new_dir = st.text_input(
                "Output directory",
                value=default_dir,
                key="_output_dir_input",
                help="Full path where extraction files are saved.",
            )
            st.session_state["output_directory"] = new_dir

            st.session_state["output_enrich"] = st.toggle(
                "Include code descriptions",
                value=st.session_state.get("output_enrich", True),
                key="_output_enrich_toggle",
                help="Merge human-readable descriptions into saved files.",
            )

            # Show recently saved files
            last_saved = st.session_state.get("last_saved_files")
            if last_saved:
                st.markdown("**Last saved:**")
                for fp in last_saved:
                    sz = os.path.getsize(fp) if os.path.exists(fp) else 0
                    if sz > 1_048_576:
                        sz_str = f"{sz / 1_048_576:.1f} MB"
                    elif sz > 1024:
                        sz_str = f"{sz / 1024:.1f} KB"
                    else:
                        sz_str = f"{sz} B"
                    st.caption(f"📄 `{os.path.basename(fp)}` ({sz_str})")

            # Quick folder stats
            if os.path.isdir(new_dir):
                files_in_dir = [f for f in os.listdir(new_dir)
                                if f.endswith((".csv", ".parquet"))]
                st.caption(f"📂 {len(files_in_dir)} file(s) in output folder")

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════════
        # NAVIGATION (reordered: Code List Dev comes early)
        # ══════════════════════════════════════════════════════════════════
        page = st.radio("Navigation", [
            "🏠 Home",
            "📋 Code List Development",
            "💊 Drug Lookup",
            "🌟 Quick Extract (Newbie)",
            "👤 Demographics",
            "🔬 CPRD Aurum Extraction",
            "🏥 Linkage Extraction",
            "🔗 Multi-Source Search",
            "👥 Cohort Builder",
            "📊 Analytics",
            "📖 Definitions",
            "⚙️ Configuration",
        ], key="nav",
        help="Select a module. Hover over each for details.")

        st.markdown("---")

        # ── Last Extraction summary ──
        stats = get_extraction_stats()
        if stats is not None:
            st.markdown("### 📋 Last Extraction")
            st.metric("Records", f"{stats['n_records']:,}")
            if stats["n_patients"] > 0:
                st.metric("Patients", f"{stats['n_patients']:,}")
            if stats["sources"]:
                st.caption(f"Sources: {', '.join(stats['sources'])}")
            st.caption(f"🕐 {stats.get('timestamp', '')}")

            scope = st.session_state.get("patient_scope_mode", "all")
            if scope == "previous":
                st.caption("🔗 Scope: Linked patients")
            elif scope == "custom":
                n_custom = len(st.session_state.get("custom_patient_ids", set()))
                st.caption(f"📋 Scope: Custom ({n_custom:,} IDs)")

        # ── Reset button ──
        st.markdown("---")
        if st.button("🔄 Reset All", key="btn_reset_all",
                     help="Clear all extraction results, patient scope, and start fresh."):
            reset_all()
            st.success("✅ All data cleared!")
            st.rerun()

    # ═══════════════════════════════════════════════════════════════════
    # PAGE ROUTING
    # ═══════════════════════════════════════════════════════════════════
    if "Home" in page:
        page_home()
    elif "Code List" in page:
        page_code_list_dev()
    elif "Drug Lookup" in page:
        page_drug_lookup()
    elif "Quick Extract" in page or "Newbie" in page:
        page_newbie()
    elif "Demographics" in page:
        page_demographics()
    elif "Aurum" in page:
        page_aurum_extraction()
    elif "Linkage" in page:
        page_linkage_extraction()
    elif "Multi-Source" in page:
        page_multi_source()
    elif "Cohort" in page:
        page_cohort_builder()
    elif "Analytics" in page:
        page_analytics()
    elif "Definitions" in page:
        page_definitions()
    elif "Config" in page:
        page_config()


# ══════════════════════════════════════════════════════════════════════════════
# SLURM TASK ARRAY — CLI MODE
# ══════════════════════════════════════════════════════════════════════════════

def _patch_streamlit_for_cli():
    """Monkey-patch Streamlit UI calls so they print to stderr in CLI mode."""
    import streamlit as _st
    _st.error   = lambda msg, **kw: print(f"[ERROR]   {msg}", file=sys.stderr)
    _st.warning = lambda msg, **kw: print(f"[WARNING] {msg}", file=sys.stderr)
    _st.info    = lambda msg, **kw: print(f"[INFO]    {msg}", file=sys.stderr)
    _st.success = lambda msg, **kw: print(f"[OK]      {msg}", file=sys.stderr)
    _st.caption = lambda msg, **kw: None
    _st.markdown = lambda msg, **kw: None
    _st.toast   = lambda msg, **kw: None


def _load_codes(args):
    """Parse codes from --codes (comma-separated) or --codes_file (one per line)."""
    codes = []
    if args.codes:
        codes = [c.strip() for c in args.codes.split(",") if c.strip()]
    if args.codes_file:
        with open(args.codes_file, "r") as f:
            for line in f:
                c = line.strip()
                if c and not c.startswith("#"):
                    codes.append(c)
    if not codes:
        print("[ERROR] No codes provided. Use --codes or --codes_file.", file=sys.stderr)
        sys.exit(1)
    return codes


def cli_extract(args):
    """Run a partitioned extraction for one Slurm array task.

    Each task processes a subset of practice folders (folders[task_id::total_tasks])
    and writes results to a Parquet shard file.
    """
    _patch_streamlit_for_cli()

    task_id     = args.task_id          # 0-based internally
    total_tasks = args.total_tasks
    extract_type = args.extract_type    # snomed | medcode | prodcode | filetype
    file_type    = args.file_type       # only for 'filetype' mode
    output_dir   = args.output_dir
    job_name     = args.job_name or f"cprd_{extract_type}"

    os.makedirs(output_dir, exist_ok=True)

    codes = _load_codes(args)

    print(f"╔══════════════════════════════════════════════════════════════╗")
    print(f"║  CPRD Slurm Task Array — Task {task_id+1}/{total_tasks}                        ║")
    print(f"╚══════════════════════════════════════════════════════════════╝")
    print(f"  Extract type : {extract_type}")
    print(f"  Codes        : {len(codes)} codes")
    print(f"  File type    : {file_type or 'N/A'}")
    print(f"  Output dir   : {output_dir}")
    print(f"  Job name     : {job_name}")

    # ── Initialise engine ──
    engine = CPRDEngine()

    # ── Discover & partition folders ──
    all_folders = engine.find_practice_folders()
    my_folders = all_folders[task_id::total_tasks]

    print(f"  Total folders: {len(all_folders)}")
    print(f"  My folders   : {len(my_folders)}  (indices {task_id}, {task_id+total_tasks}, ...)")
    try:
        _con = duckdb.connect()
        n_threads = _con.execute("SELECT current_setting('threads')").fetchone()[0]
        _con.close()
        print(f"  DuckDB threads: {n_threads}")
    except Exception:
        print(f"  DuckDB threads: (unable to detect)")
    print()

    if not my_folders:
        print("[WARN] No folders assigned to this task — nothing to do.")
        # Write empty parquet with at least one column so merge doesn't fail
        out_path = os.path.join(output_dir, f"{job_name}_task_{task_id:04d}.parquet")
        pd.DataFrame({"_empty": pd.Series(dtype="str")}).iloc[:0].to_parquet(
            out_path, engine="pyarrow", index=False
        )
        print(f"  Wrote empty shard: {out_path}")
        return

    # ── Run extraction on assigned folders ──
    t_start = time.time()
    results = []

    def cli_progress(frac, folder_name, idx, total):
        elapsed = time.time() - t_start
        eta = (elapsed / max(frac, 0.001)) * (1 - frac) if frac > 0 else 0
        print(f"  [{idx+1}/{total}]  {folder_name:<30}  "
              f"{frac*100:5.1f}%  elapsed={elapsed:.0f}s  ETA={eta:.0f}s", flush=True)

    if extract_type == "snomed":
        # SNOMED → lookup medcodes → extract from Observation files
        emis_path = PATHS["emis_dictionary"]
        if not os.path.exists(emis_path):
            print(f"[ERROR] EMIS Dictionary not found: {emis_path}", file=sys.stderr)
            sys.exit(1)
        emis_df = pd.read_csv(emis_path, sep='\t', dtype=str)
        matching = emis_df[emis_df["SnomedCTConceptId"].isin(codes)]
        medcode_list = matching["MedCodeId"].unique().tolist()
        if not medcode_list:
            print("[WARN] No MedCodeIds found for given SNOMED codes.")
        else:
            print(f"  Mapped {len(codes)} SNOMED codes → {len(medcode_list)} MedCodeIds")
            total_f = len(my_folders)
            for i, folder in enumerate(my_folders):
                cli_progress(i / total_f, os.path.basename(folder), i, total_f)
                zips = engine.find_zip_files(folder, "Observation")
                for zp in zips:
                    df = engine.extract_from_zip(
                        zp, "Observation",
                        filter_col="medcodeid", filter_values=medcode_list,
                        select_cols=["patid", "obsid", "obsdate", "medcodeid",
                                     "value", "numunitid", "pracid"]
                    )
                    if df is not None:
                        df["practice_folder"] = os.path.basename(folder)
                        results.append(df)
            # Merge EMIS terms
            if results:
                combined = pd.concat(results, ignore_index=True)
                combined = combined.merge(
                    matching[["MedCodeId", "Term", "SnomedCTConceptId"]],
                    left_on="medcodeid", right_on="MedCodeId", how="left"
                )
                # Drop redundant MedCodeId (duplicate of medcodeid)
                combined.drop(columns=["MedCodeId"], errors="ignore", inplace=True)
                results = [combined]

    elif extract_type == "medcode":
        total_f = len(my_folders)
        for i, folder in enumerate(my_folders):
            cli_progress(i / total_f, os.path.basename(folder), i, total_f)
            zips = engine.find_zip_files(folder, "Observation")
            for zp in zips:
                df = engine.extract_from_zip(
                    zp, "Observation", filter_col="medcodeid",
                    filter_values=codes,
                    select_cols=["patid", "obsid", "obsdate", "medcodeid",
                                 "value", "numunitid", "pracid"]
                )
                if df is not None:
                    df["practice_folder"] = os.path.basename(folder)
                    results.append(df)

    elif extract_type == "prodcode":
        total_f = len(my_folders)
        for i, folder in enumerate(my_folders):
            cli_progress(i / total_f, os.path.basename(folder), i, total_f)
            zips = engine.find_zip_files(folder, "DrugIssue")
            for zp in zips:
                df = engine.extract_from_zip(
                    zp, "DrugIssue", filter_col="prodcodeid",
                    filter_values=codes,
                    select_cols=["patid", "issueid", "issuedate", "prodcodeid",
                                 "quantity", "duration", "pracid"]
                )
                if df is not None:
                    df["practice_folder"] = os.path.basename(folder)
                    results.append(df)

    elif extract_type == "filetype":
        if not file_type or file_type not in AURUM_FILE_TYPES:
            print(f"[ERROR] --file_type must be one of: {list(AURUM_FILE_TYPES.keys())}",
                  file=sys.stderr)
            sys.exit(1)
        ftype_info = AURUM_FILE_TYPES[file_type]
        fc = ftype_info.get("key_col")
        # codes are optional for filetype — if empty, extract all records
        filter_vals = codes if codes and codes != ["ALL"] else None
        total_f = len(my_folders)
        for i, folder in enumerate(my_folders):
            cli_progress(i / total_f, os.path.basename(folder), i, total_f)
            zips = engine.find_zip_files(folder, file_type)
            for zp in zips:
                df = engine.extract_from_zip(
                    zp, file_type, filter_col=fc if filter_vals else None,
                    filter_values=filter_vals
                )
                if df is not None:
                    df["practice_folder"] = os.path.basename(folder)
                    results.append(df)
    else:
        print(f"[ERROR] Unknown extract_type: {extract_type}. "
              f"Use: snomed, medcode, prodcode, filetype", file=sys.stderr)
        sys.exit(1)

    # ── Write Parquet shard ──
    if results:
        df_out = pd.concat(results, ignore_index=True)
    else:
        df_out = pd.DataFrame()

    out_path = os.path.join(output_dir, f"{job_name}_task_{task_id:04d}.parquet")
    df_out.to_parquet(out_path, engine="pyarrow", index=False)

    elapsed = time.time() - t_start
    print()
    print(f"  ✅ Task {task_id+1}/{total_tasks} complete in {elapsed:.1f}s")
    print(f"  Records: {len(df_out):,}")
    if "patid" in df_out.columns:
        print(f"  Patients: {df_out['patid'].nunique():,}")
    print(f"  Output : {out_path}  ({os.path.getsize(out_path) / 1024 / 1024:.1f} MB)")


def cli_merge(args):
    """Merge all Parquet shard files from a completed task array into one master file."""
    _patch_streamlit_for_cli()

    output_dir = args.output_dir
    job_name   = args.job_name or "cprd_*"
    out_file   = args.merge_output or os.path.join(output_dir, f"{job_name}_MERGED.parquet")

    # Find all shard files
    pattern = os.path.join(output_dir, f"{job_name}_task_*.parquet")
    shards = sorted(glob.glob(pattern))

    if not shards:
        print(f"[ERROR] No shard files found matching: {pattern}", file=sys.stderr)
        sys.exit(1)

    print(f"╔══════════════════════════════════════════════════════════════╗")
    print(f"║  CPRD Slurm Merge — {len(shards)} shards                              ║")
    print(f"╚══════════════════════════════════════════════════════════════╝")
    print(f"  Pattern : {pattern}")
    print(f"  Output  : {out_file}")

    t0 = time.time()

    # Use DuckDB for fast union of all Parquet files
    con = duckdb.connect()

    # Configure DuckDB for performance
    con.execute("SET threads TO 8")
    con.execute("SET memory_limit = '24GB'")

    # Filter out empty/broken shards
    valid_shards = []
    for s in shards:
        try:
            row_count = con.execute(
                f"SELECT COUNT(*) FROM read_parquet('{s}')"
            ).fetchone()[0]
            if row_count > 0:
                valid_shards.append(s)
            else:
                print(f"  ⏭ Skipping empty shard: {os.path.basename(s)}")
        except Exception:
            print(f"  ⏭ Skipping unreadable shard: {os.path.basename(s)}")

    if not valid_shards:
        print("\n  ⚠️ No non-empty shards found — nothing to merge.")
        con.close()
        return

    total_bytes = sum(os.path.getsize(s) for s in valid_shards)
    print(f"  Valid shards: {len(valid_shards)}/{len(shards)}")
    print(f"  Total shard size: {total_bytes / 1024 / 1024:.1f} MB")

    # Read & union all valid shards in one query
    shard_list_sql = ", ".join(f"'{s}'" for s in valid_shards)
    con.execute(f"""
        COPY (
            SELECT * FROM read_parquet([{shard_list_sql}],
                                       union_by_name = true)
        ) TO '{out_file}' (FORMAT PARQUET, COMPRESSION ZSTD)
    """)
    con.close()

    # Report
    elapsed = time.time() - t0
    merged_size = os.path.getsize(out_file) / 1024 / 1024

    # Quick stats
    con2 = duckdb.connect()
    n_records = con2.execute(f"SELECT COUNT(*) FROM read_parquet('{out_file}')").fetchone()[0]
    # patid may not exist in all extraction types
    try:
        n_patients = con2.execute(
            f"SELECT COUNT(DISTINCT patid) FROM read_parquet('{out_file}')"
        ).fetchone()[0]
    except Exception:
        n_patients = "N/A"
    con2.close()

    print()
    print(f"  ✅ Merge complete in {elapsed:.1f}s")
    print(f"  Records  : {n_records:,}")
    print(f"  Patients : {n_patients:,}" if isinstance(n_patients, int) else f"  Patients : {n_patients}")
    print(f"  File size: {merged_size:.1f} MB (ZSTD compressed)")
    print(f"  Output   : {out_file}")

    # Optionally export to CSV too
    if args.also_csv:
        csv_path = out_file.replace(".parquet", ".csv")
        con3 = duckdb.connect()
        con3.execute(f"""
            COPY (SELECT * FROM read_parquet('{out_file}'))
            TO '{csv_path}' (HEADER, DELIMITER ',')
        """)
        con3.close()
        csv_size = os.path.getsize(csv_path) / 1024 / 1024
        print(f"  CSV copy : {csv_path}  ({csv_size:.1f} MB)")


def cli_generate_slurm(args):
    """Generate a ready-to-submit Slurm .sh script for the extraction."""
    output_dir = args.output_dir
    n_tasks    = args.total_tasks or 50
    extract_type = args.extract_type or "snomed"
    job_name   = args.job_name or f"cprd_{extract_type}"
    file_type  = args.file_type or ""

    codes_arg = ""
    if args.codes:
        codes_arg = f'--codes "{args.codes}"'
    elif args.codes_file:
        codes_arg = f'--codes_file "{args.codes_file}"'
    else:
        codes_arg = '--codes_file "/path/to/your/codes.txt"  # ← EDIT THIS'

    ft_arg = f'--file_type "{file_type}"' if extract_type == "filetype" else ""

    # Resolve the absolute path of THIS script at generation time
    app_path = os.path.abspath(sys.argv[0])

    # Build the python command args list (avoid empty lines)
    cmd_args = [
        f'python3 {app_path}',
        '    --task_id $TASK_ID',
        f'    --total_tasks {n_tasks}',
        f'    --extract_type {extract_type}',
    ]
    if ft_arg:
        cmd_args.append(f'    {ft_arg}')
    cmd_args.append(f'    {codes_arg}')
    cmd_args.append(f'    --output_dir {output_dir}/shards')
    cmd_args.append(f'    --job_name {job_name}')
    python_cmd = " \\\n".join(cmd_args)

    script = f'''#!/bin/bash
#══════════════════════════════════════════════════════════════════
# CPRD Slurm Task Array Extraction
# Generated by app_v9.py on {datetime.now().strftime("%Y-%m-%d %H:%M")}
#══════════════════════════════════════════════════════════════════
#
# USAGE:
#   1. Edit the code list (--codes or --codes_file) below
#   2. Submit:   sbatch {job_name}.sh
#   3. Monitor:  squeue -u $USER
#   4. Merge:    sbatch {job_name}_merge.sh   (after all tasks finish)
#
#SBATCH --job-name={job_name}
#SBATCH --partition=short
#SBATCH --array=1-{n_tasks}
#SBATCH --cpus-per-task=8
#SBATCH --mem=32G
#SBATCH --time=06:00:00
#SBATCH --output={output_dir}/logs/{job_name}_%A_%a.out
#SBATCH --error={output_dir}/logs/{job_name}_%A_%a.err

# ── Setup ──
module load Python/3.11.3-GCCcore-12.3.0 2>/dev/null || true
module load Anaconda3 2>/dev/null || true

# Create output directories
mkdir -p {output_dir}/logs
mkdir -p {output_dir}/shards

# ── Convert Slurm 1-based array ID to 0-based task ID ──
TASK_ID=$(( $SLURM_ARRAY_TASK_ID - 1 ))

echo "═══════════════════════════════════════════════════════"
echo "  Job: $SLURM_JOB_NAME  |  Array Task: $SLURM_ARRAY_TASK_ID  |  Node: $(hostname)"
echo "  CPUs: $SLURM_CPUS_PER_TASK  |  Memory: $SLURM_MEM_PER_NODE MB"
echo "═══════════════════════════════════════════════════════"

# ── Run extraction for this task's folder subset ──
{python_cmd}

echo "Task $SLURM_ARRAY_TASK_ID finished with exit code $?"
'''

    merge_script = f'''#!/bin/bash
#══════════════════════════════════════════════════════════════════
# CPRD Merge — Run AFTER all array tasks complete
#══════════════════════════════════════════════════════════════════
#
# USAGE: sbatch --dependency=afterok:${{ARRAY_JOB_ID}} {job_name}_merge.sh
#        Or just: sbatch {job_name}_merge.sh  (if array is done)
#
#SBATCH --job-name={job_name}_merge
#SBATCH --partition=short
#SBATCH --cpus-per-task=8
#SBATCH --mem=64G
#SBATCH --time=02:00:00
#SBATCH --output={output_dir}/logs/{job_name}_merge_%j.out
#SBATCH --error={output_dir}/logs/{job_name}_merge_%j.err

module load Python/3.11.3-GCCcore-12.3.0 2>/dev/null || true
module load Anaconda3 2>/dev/null || true

echo "Merging {n_tasks} shards..."

python3 {app_path} \\
    --merge \\
    --output_dir {output_dir}/shards \\
    --merge_output {output_dir}/{job_name}_FINAL.parquet \\
    --job_name {job_name} \\
    --also_csv

echo "Merge complete. Output: {output_dir}/{job_name}_FINAL.parquet"
'''

    # Convenience all-in-one launch script
    launch_script = f'''#!/bin/bash
#══════════════════════════════════════════════════════════════════
# CPRD One-Command Launch — Submit extraction + auto-merge
#══════════════════════════════════════════════════════════════════
# USAGE:  bash {job_name}_launch.sh

mkdir -p {output_dir}/logs {output_dir}/shards

# Submit the array job
ARRAY_JOB_ID=$(sbatch --parsable {job_name}.sh)
echo "Submitted array job: $ARRAY_JOB_ID ({n_tasks} tasks)"

# Submit merge job — runs only after ALL array tasks succeed
MERGE_JOB_ID=$(sbatch --parsable --dependency=afterok:$ARRAY_JOB_ID {job_name}_merge.sh)
echo "Submitted merge job: $MERGE_JOB_ID (depends on $ARRAY_JOB_ID)"

echo ""
echo "Monitor with:  squeue -u $USER"
echo "Cancel with:   scancel $ARRAY_JOB_ID $MERGE_JOB_ID"
echo "Final output:  {output_dir}/{job_name}_FINAL.parquet"
'''

    # Write scripts
    os.makedirs(output_dir, exist_ok=True)

    script_path = os.path.join(output_dir, f"{job_name}.sh")
    with open(script_path, "w") as f:
        f.write(script)
    os.chmod(script_path, 0o755)

    merge_path = os.path.join(output_dir, f"{job_name}_merge.sh")
    with open(merge_path, "w") as f:
        f.write(merge_script)
    os.chmod(merge_path, 0o755)

    launch_path = os.path.join(output_dir, f"{job_name}_launch.sh")
    with open(launch_path, "w") as f:
        f.write(launch_script)
    os.chmod(launch_path, 0o755)

    print(f"╔══════════════════════════════════════════════════════════════╗")
    print(f"║  Slurm Scripts Generated                                    ║")
    print(f"╚══════════════════════════════════════════════════════════════╝")
    print(f"  1. {script_path}")
    print(f"     → Array job ({n_tasks} tasks × 8 CPUs × 32GB each)")
    print(f"  2. {merge_path}")
    print(f"     → Merge job (runs after array completes)")
    print(f"  3. {launch_path}")
    print(f"     → One-command launch (submits both with dependency)")
    print()
    print(f"  Quick start:")
    print(f"    cd {output_dir}")
    print(f"    bash {job_name}_launch.sh")
    print()
    print(f"  Or step-by-step:")
    print(f"    sbatch {job_name}.sh")
    print(f"    # Wait for completion, then:")
    print(f"    sbatch {job_name}_merge.sh")


def parse_cli_args():
    """Parse command-line arguments for Slurm task array mode."""
    parser = argparse.ArgumentParser(
        description="CPRD RWE Platform v9.0 — CLI mode for Slurm task arrays",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
EXAMPLES:

  # Generate Slurm scripts for a SNOMED extraction:
  python app_v9.py --generate_slurm --extract_type snomed \\
      --codes "60573004,60234000,86466006" \\
      --total_tasks 50 --output_dir /well/rahimi/projects/CPRD/results/vhd_extract

  # Run a single task (Slurm calls this):
  python app_v9.py --task_id 0 --total_tasks 50 --extract_type snomed \\
      --codes_file /path/to/snomed_codes.txt \\
      --output_dir /well/rahimi/projects/CPRD/results/vhd_extract/shards

  # Merge all shards after completion:
  python app_v9.py --merge --output_dir /path/to/shards \\
      --merge_output /path/to/final_result.parquet --also_csv

  # Extract all Patient records (no filter codes):
  python app_v9.py --task_id 0 --total_tasks 50 --extract_type filetype \\
      --file_type Patient --codes "ALL" --output_dir /path/to/output

  # Launch Streamlit UI (default — no CLI args):
  python app_v9.py
  streamlit run app_v9.py
        """
    )

    # Mode flags
    parser.add_argument("--merge", action="store_true",
                        help="Merge mode: combine Parquet shards into one file")
    parser.add_argument("--generate_slurm", action="store_true",
                        help="Generate Slurm .sh scripts for submission")

    # Task array params
    parser.add_argument("--task_id", type=int, default=None,
                        help="0-based task ID (from SLURM_ARRAY_TASK_ID - 1)")
    parser.add_argument("--total_tasks", type=int, default=50,
                        help="Total number of tasks in array (default: 50)")

    # Extraction params
    parser.add_argument("--extract_type", type=str, default="snomed",
                        choices=["snomed", "medcode", "prodcode", "filetype"],
                        help="Type of extraction")
    parser.add_argument("--file_type", type=str, default=None,
                        help="Aurum file type for --extract_type filetype "
                             f"(choices: {list(AURUM_FILE_TYPES.keys())})")
    parser.add_argument("--codes", type=str, default=None,
                        help="Comma-separated code list")
    parser.add_argument("--codes_file", type=str, default=None,
                        help="Path to file with one code per line")

    # Output
    parser.add_argument("--output_dir", type=str,
                        default=os.path.join(CPRD_BASE, "results", "slurm_extractions"),
                        help="Directory for output Parquet shards")
    parser.add_argument("--job_name", type=str, default=None,
                        help="Job name prefix for output files (default: cprd_<extract_type>)")
    parser.add_argument("--merge_output", type=str, default=None,
                        help="Path for merged output file (--merge mode)")
    parser.add_argument("--also_csv", action="store_true",
                        help="Also export CSV when merging")

    return parser.parse_args()


if __name__ == "__main__":
    # ── Detect CLI mode: any known argument present → CLI, otherwise → Streamlit ──
    cli_flags = {"--task_id", "--merge", "--generate_slurm", "--help", "-h"}
    if any(arg in cli_flags for arg in sys.argv[1:]):
        args = parse_cli_args()

        if args.merge:
            cli_merge(args)
        elif args.generate_slurm:
            cli_generate_slurm(args)
        elif args.task_id is not None:
            cli_extract(args)
        else:
            parse_cli_args()  # will print help
    else:
        # No CLI arguments → launch Streamlit UI
        main()
